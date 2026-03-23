from decimal import Decimal
from .config import ROUND_PERCENTAGES, ROUND_CURRENCY 
from django.http import JsonResponse, HttpResponse
from django.db import connections
import io
from django.views.decorators.csrf import csrf_exempt
from django.middleware.csrf import get_token
from django.utils.decorators import method_decorator
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from django.http import JsonResponse, StreamingHttpResponse
from django.core.files.storage import default_storage
from rest_framework_simplejwt.tokens import RefreshToken
from typing import Dict, Any, List, Tuple, Optional
from openpyxl.utils import get_column_letter
from urllib.parse import quote_plus
import uuid
from rest_framework import status, generics
from django.contrib.auth import authenticate, get_user_model
from django.db import connection
from rest_framework.decorators import api_view , parser_classes, permission_classes
from rest_framework.parsers import MultiPartParser, FormParser
from datetime import datetime, timedelta, date
import numpy as np
import json
from langchain.embeddings import HuggingFaceEmbeddings
from langchain.vectorstores import FAISS
from langchain.text_splitter import CharacterTextSplitter
from langchain.docstore.document import Document
from langchain.chains import RetrievalQA
from loguru import logger  # Make sure logger is configured
import os
from .llm_runner.llm_utils import llm_generate_chart_config
from .llm_runner.db_introspect import extract_schema_from_sqlalchemy
from .langgraph_logic.db_embedding import embed_schema_for_user
from .langgraph_logic.langgraph_runner import run_sql_generation_graph
import re
from urllib.parse import quote_plus
from sqlalchemy import create_engine, text
 
from langchain.chains import RetrievalQA

from urllib.parse import quote_plus
from sqlalchemy import create_engine
from PyPDF2 import PdfReader
# from .langgraph_logic.langgraph_runner import normalize_sql_text
# from .langgraph_logic.langgraph_runner import sanitize_and_wrap_sql
import math
import pandas as pd
import os,sys
import requests
from .serializers import UserSerializer
from .models import UploadedFile , CustomUser, UserRole, Page, UserPageAccess
from asgiref.sync import async_to_sync
import torch
from transformers import AutoModelForCausalLM, AutoTokenizer, pipeline
import time, threading
import tiktoken
import csv  # ✅ Ensure this import exists
import logging
import calendar
import difflib
import hashlib
import matplotlib.pyplot as plt
from difflib import SequenceMatcher
from transformers import pipeline
from django.core.cache import cache
from fuzzywuzzy import fuzz
User = get_user_model()
logger = logging.getLogger(__name__)
from django.conf import settings
from openpyxl import load_workbook
from .tasks import  async_validate_excel
# from api.utils import store_dataframe_in_postgres
import psycopg2
import warnings
import subprocess
from io import BytesIO
from sqlalchemy import create_engine, text, inspect
import seaborn as sns
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from langchain.prompts import PromptTemplate
from dataclasses import dataclass
from concurrent.futures import ThreadPoolExecutor
from sqlalchemy.orm import sessionmaker
from sqlalchemy.exc import SQLAlchemyError
import openpyxl
import pyxlsb
import traceback
import logging
from dotenv  import load_dotenv
from pathlib import Path
from rest_framework.decorators import api_view
from django.conf import settings
import re
import base64
from django.db.models import Count
from django.db.models.functions import TruncDate
from .utils import encrypt_value
from azure.storage.blob import BlobServiceClient
from langchain.vectorstores import FAISS
from langchain.embeddings import HuggingFaceEmbeddings
from langchain_openai import ChatOpenAI
from sentence_transformers import CrossEncoder
from sklearn.metrics.pairwise import cosine_similarity
from sklearn.feature_extraction.text import TfidfVectorizer

from langchain_core.prompts import PromptTemplate
#from langchain.docstore.document import Document
from langchain.docstore.document import Document
from psycopg2 import sql
from psycopg2.extras import RealDictCursor
from rest_framework_simplejwt.views import TokenObtainPairView
from .serializers import UserSerializer, UpdateUserSerializer,MyTokenObtainPairSerializer, CreateUserSerializer, UserPageAccessSerializer, PageSerializer,UserPageAccessNestedSerializer
from rest_framework import serializers
from django.shortcuts import get_object_or_404
from django.http import FileResponse, Http404
from django.utils.timezone import now

from rest_framework.views import APIView
from rest_framework.response import Response
import psycopg2
from langchain_openai import AzureChatOpenAI, ChatOpenAI
from langchain_azure_ai.chat_models.inference import AzureAIChatCompletionsModel
from django.views import View
session_store = {}
schema_context_store = {}
conversation_memory_store = {}

GROQ_API_KEY = "REMOVEDx4sevu1Zrp096Df8YkMUWGdyb3FYWRAZROV7i2sFCSb8sRkb0dtH"


import uuid
import traceback
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from datetime import datetime
from sqlalchemy import text
import traceback
from django.conf import settings
import json
import sqlparse
import csv
from django.http import HttpResponse
import requests
import os
from django.utils.timezone import now
import re
import time
import requests
from loguru import logger  # Make sure logger is configured
import os
from api.llm_runner.llm_utils import llm_generate_chart_config
from api.llm_runner.db_introspect import extract_schema_from_sqlalchemy
from .langgraph_logic.db_embedding import embed_schema_for_user
from .langgraph_logic.langgraph_runner import run_sql_generation_graph
from .langgraph_logic.langgraph_runner import generate_summary_from_rows
import re
from urllib.parse import quote_plus
from sqlalchemy import create_engine, text
 
from langchain.chains import RetrievalQA
from django.http import StreamingHttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json, traceback, time
from .humanizer import humanize_narrative,safe_generate_narrative  # you already have this




import os
import time
import logging
from typing import Optional

from azure.ai.inference import ChatCompletionsClient
from azure.ai.inference.models import SystemMessage, UserMessage
from azure.core.credentials import AzureKeyCredential
from azure.core.exceptions import HttpResponseError, ServiceRequestError, ServiceResponseError
CHUNK_SIZE = 1000 
CHUNK_OVERLAP = 100
 # Base chunk size
MAX_CHUNKS_PER_QUERY = 15  # Increased for large files
VECTOR_DIR = getattr(settings, 'VECTOR_DIR', './vectornew')
os.makedirs(VECTOR_DIR, exist_ok=True)

logger = logging.getLogger(__name__)

AZURE_ENDPOINT = os.getenv("AZURE_INFERENCE_ENDPOINT")  # e.g., https://genaiprochurn.services.ai.azure.com/models
AZURE_API_KEY = os.getenv("AZURE_INFERENCE_API_KEY")
AZURE_MODEL = os.getenv("AZURE_INFERENCE_MODEL", "Llama-4-Maverick-17B-128E-Instruct-FP8-prochurn-demo")
AZURE_API_VERSION = "2024-05-01-preview"
MAX_PROMPT_TOKENS = int(os.getenv("AZURE_MAX_PROMPT_TOKENS", "120000"))  # 128k-context model -> ~120k safety

_SYSTEM_PROMPT = os.getenv("AZURE_SYSTEM_PROMPT", "You are a helpful SQL assistant.")

from azure.ai.inference import ChatCompletionsClient
from azure.core.credentials import AzureKeyCredential
from azure.core.pipeline.transport import RequestsTransport
from .humanizer import humanize_narrative,safe_generate_narrative,_derive_contextual_metrics,_generate_dynamic_next_steps  # you already have this

def _ndjson(obj):               # one JSON object per line
    return json.dumps(obj, default=str) + "\n"

def _ev(event, **data):         # event wrapper
    return _ndjson({"event": event, **data})

# Configure custom transport with shorter timeouts
transport = RequestsTransport(
    connection_timeout=30,   # wait max 30s to establish a connection
    read_timeout=120         # wait max 120s for a response body
)

_client: Optional[ChatCompletionsClient] = None

AGG_RE = re.compile(r'''
    \b (sum|count|avg|min|max) \s*      # aggregate name
    \( \s*
        (?:
            \*                          # COUNT(*)
            | "([A-Za-z_][\w\.]*)"      # "quoted.identifier"
            | ([A-Za-z_][\w\.]*)        # unquoted identifier
        )
    \s* \)
''', re.IGNORECASE | re.VERBOSE)



import os, json, math, traceback, base64
from datetime import datetime
import numpy as np
import pandas as pd
import psycopg2
from psycopg2.extras import RealDictCursor
from django.utils import timezone
from rest_framework.views import APIView
from rest_framework.response import Response
from pathlib import Path
import logging

import joblib
logger = logging.getLogger(__name__)
# Optional: Groq for email drafting (falls back if not configured)
# try:
#     from langchain_groq import ChatGroq
# except Exception:
#     ChatGroq = None

# # ------------ Model assets: lazy singletons ------------
# GROQ_API_KEY_Churn = "REMOVEDhqk7OIprKrMafXvFTRxIWGdyb3FYf19I66DpucmiC21LG1zwxrI8"
# GROQ_MODEL_Churn   = "meta-llama/llama-4-maverick-17b-128e-instruct"

# ---------------- Azure chat helper ----------------
from azure.ai.inference import ChatCompletionsClient
from azure.ai.inference.models import SystemMessage, UserMessage
from azure.core.credentials import AzureKeyCredential
from azure.core.exceptions import HttpResponseError, ServiceRequestError, ServiceResponseError

_client: Optional[ChatCompletionsClient] = None

def _get_azure_client() -> ChatCompletionsClient:
    global _client
    if _client is None:
        if not AZURE_ENDPOINT or not AZURE_API_KEY:
            raise RuntimeError("Azure endpoint/key are not configured.")
        _client = ChatCompletionsClient(
            endpoint=AZURE_ENDPOINT,
            credential=AzureKeyCredential(AZURE_API_KEY),
            api_version=AZURE_API_VERSION,
            transport=transport,  # your shorter timeouts
        )
    return _client

def azure_chat(prompt: str,
               *,
               system: str = _SYSTEM_PROMPT,
               temperature: float = 0.2) -> str:
    """
    Call Azure AI Inference chat completion and return the text content.
    (No max tokens argument; relies on server-side defaults.)
    """
    client = _get_azure_client()
    try:
        resp = client.complete(
            model=AZURE_MODEL,
            messages=[
                SystemMessage(content=system or ""),
                UserMessage(content=prompt or "")
            ],
            temperature=float(temperature),
            top_p=0.9
            # NOTE: do not pass max_output_tokens / max_tokens here for this SDK build
        )
        if not resp or not getattr(resp, "choices", None):
            return ""
        msg = getattr(resp.choices[0], "message", None)
        return (getattr(msg, "content", "") or "").strip()
    except (HttpResponseError, ServiceRequestError, ServiceResponseError) as e:
        logger.error("Azure chat error: %s", e)
        return ""


BASE_DIR_Churn = Path(__file__).resolve().parents[2]
DEFAULT_MODEL_DIR = BASE_DIR_Churn / "models"

MODEL_DIR     = Path(os.getenv("MODEL_DIR", str(DEFAULT_MODEL_DIR)))
MODEL_PATH    = Path(os.getenv("MODEL_PATH", str(MODEL_DIR / "gbm_model.pkl")))
ENCODERS_PATH = Path(os.getenv("ENCODERS_PATH", str(MODEL_DIR / "label_encoders_gbm.pkl")))
FEATURES_PATH = Path(os.getenv("FEATURES_PATH", str(MODEL_DIR / "model_features_gbm.pkl")))

_g_model = None
_g_encoders = None
_g_features = None

MV_SEGMENT_POLICY = '"Prediction"."mv_not_renewed_policies_for_filter_CS_1"'

TABLE = '"Prediction"."GBM1_predictions_JFMAMJ(Final)_top3_reasons"'



def _must_exist(p: Path, name: str):
    if not p.exists():
        raise FileNotFoundError(f"{name} not found at {p}")

FRONTEND_KEY_MAP = {
    "customerid": "customerid",
    "policy no": "policy no",
    "Customer Segment": "customersegment",
    "biztype": "biztype",
    "tie up": "tie up",
    "Cleaned Zone 2": "Cleaned Zone 2",
    "Cleaned State2": "Cleaned State2",
    "Cleaned Branch Name 2": "Cleaned Branch Name 2",
    "make_clean": "make_clean",
    "model_clean": "model_clean",
    "variant": "variant",
    "total od premium": "total od premium",
    "total tp premium": "total tp premium",
    "before gst add-on gwp": "before gst add-on gwp",
    "applicable discount with ncb": "applicable discount with ncb",
    "ncb % previous year": "ncb % previous year",
    "vehicle idv": "vehicle idv",
    "Predicted Status": "Predicted Status",
    "Churn Probability": "Churn Probability",
    "Top 3 Reasons": "Top 3 Reasons",
    # keep policy_end_date etc. if needed downstream
    "policy end date": "policy_end_date",
    "policy end date_MONTH": "policy_end_date_month",
    "policy end date_YEAR": "policy_end_date_year",
}

FALLBACK_PARAM_RANGES = {
    "discount": {"min": 0.0, "max": 90.0},
    "od_premium": {"min": 10000, "max": 50000},
    "tp_premium": {"min": 1500, "max": 8000},
    "idv": {"min": 200000, "max": 2000000},
    "add_on_premium": {"min": 1000, "max": 10000},
    "ncb": {"min": 0.0, "max": 50.0},
}

def _get_conn():
    return psycopg2.connect(
        dbname=DB_New_Pred["dbname"],
        user=DB_New_Pred["user"],
        password=DB_New_Pred["password"],
        host=DB_New_Pred["host"],
        port=DB_New_Pred["port"],
    )

def _row_to_frontend(row: dict) -> dict:
    """Rename DB columns to exactly what the React component expects."""
    out = {}
    for db_key, fe_key in FRONTEND_KEY_MAP.items():
        if db_key in row:
            val = row[db_key]
            if hasattr(val, "isoformat"):
                val = val.isoformat()
            out[fe_key] = val
    return out


class ChurnSegments(APIView):
    """
    GET /api/churn/segments
    Returns distinct segments among Not Renewed from the MV.
    """
    def get(self, request):
        sql = f"""
            SELECT DISTINCT customer_segment
            FROM {MV_SEGMENT_POLICY}
            WHERE customer_segment IS NOT NULL
            ORDER BY customer_segment
        """
        try:
            conn = _get_conn()
            cur = conn.cursor()
            cur.execute(sql)
            segments = [r[0] for r in cur.fetchall() if r[0]]
            return Response({"segments": segments})
        except Exception as e:
            return Response({"error": str(e)}, status=500)
        finally:
            try:
                cur.close(); conn.close()
            except Exception:
                pass

class ChurnProbability(APIView):
    """
    GET /api/churn/policies?segment=Platinum
    Returns policy numbers for the selected segment from the MV.
    """
    def get(self, request):
        segment = request.GET.get("segment")
        if not segment:
            return Response({"error": "segment is required"}, status=400)

        sql = f"""
            SELECT distinct probability_range
            FROM {MV_SEGMENT_POLICY}
            WHERE customer_segment = %s
              AND probability_range IS NOT NULL
            ORDER BY probability_range
        """
        try:
            conn = _get_conn()
            cur = conn.cursor()
            cur.execute(sql, [segment])
            probability = [r[0] for r in cur.fetchall() if r[0]]
            return Response({"probability": probability})
        except Exception as e:
            return Response({"error": str(e)}, status=500)
        finally:
            try:
                cur.close(); conn.close()
            except Exception:
                pass



class ChurnPolicies(APIView):
    """
    GET /api/churn/policies?segment=Platinum&probability_range=60-80&page=1&page_size=50&q=...
    Returns a paged list of policy numbers for the selected segment+probability from the MV.
    """
    def get(self, request):
        segment = request.GET.get("segment")
        probability = request.GET.get("probability_range")

        if not segment:
            return Response({"error": "segment is required"}, status=400)
        if not probability:
            return Response({"error": "probability is required"}, status=400)

        # pagination & search parameters coming from React
        try:
            page = int(request.GET.get("page", 1))
            page_size = int(request.GET.get("page_size", 50))
        except ValueError:
            return Response({"error": "invalid page or page_size"}, status=400)

        q = request.GET.get("q")  # optional text search on policy_no
        offset = (page - 1) * page_size

        base_sql = f"""
            SELECT policy_no
            FROM {MV_SEGMENT_POLICY}
            WHERE customer_segment = %s
              AND probability_range = %s
              AND policy_no IS NOT NULL
        """
        params = [segment, probability]

        if q:
            base_sql += " AND policy_no ILIKE %s"
            params.append(f"%{q}%")

        # limit+offset for paging
        sql = base_sql + " ORDER BY policy_no LIMIT %s OFFSET %s"
        params.extend([page_size, offset])

        try:
            conn = _get_conn()
            cur = conn.cursor()
            cur.execute(sql, params)
            rows = cur.fetchall()
            policies = [r[0] for r in rows if r[0]]

            # simple "hasMore": if we filled the page, assume more may exist
            has_more = len(policies) == page_size

            return Response({"policies": policies, "hasMore": has_more})
        except Exception as e:
            return Response({"error": str(e)}, status=500)
        finally:
            try:
                cur.close()
                conn.close()
            except Exception:
                pass



class ChurnPolicyDetail(APIView):
    """
    GET /api/churn/policy?policy_no=POL001  (or "Policy 100007")
    Returns a single policy row with keys renamed for the React UI.
    """
    def get(self, request):
        policy_no = request.GET.get("policy_no")
        if not policy_no:
            return Response({"error": "policy_no is required"}, status=400)

        # Be forgiving about whitespace / URL-decoding
        policy_no = str(policy_no).strip()

        sql = f"""
            SELECT
                "customerid",
                "policy no",
                "Customer Segment",
                "biztype",
                "tie up",
                "Cleaned Zone 2",
                "Cleaned State2",
                "Cleaned Branch Name 2",
                "make_clean",
                "model_clean",
                "variant",
                "total od premium",
                "total tp premium",
                "before gst add-on gwp",      -- NOTE the hyphen
                "applicable discount with ncb",
                "ncb %% previous year",        -- NOTE the percent sign
                "vehicle idv",
                "Predicted Status",
                "Churn Probability",
                "Top 3 Reasons",
                "policy end date",
                EXTRACT(MONTH FROM "policy end date")::int AS "policy end date_MONTH",
                EXTRACT(YEAR  FROM "policy end date")::int AS "policy end date_YEAR"
            FROM {TABLE}
            WHERE "policy no" = %s
            LIMIT 1
        """
        try:
            conn = _get_conn()
            try:
                cur = conn.cursor(cursor_factory=RealDictCursor)
                cur.execute(sql, (policy_no,))
                row = cur.fetchone()     # <— never index into fetchall()
            finally:
                try:
                    cur.close()
                except Exception:
                    pass
                try:
                    conn.close()
                except Exception:
                    pass

            if not row:
                # Helpful log + a 404 (not a 500)
                logger.info("ChurnPolicyDetail: policy not found: %r", policy_no)
                return Response({"error": "policy not found"}, status=404)

            # Map DB keys to UI keys safely (no indexing anywhere)
            data = _row_to_frontend(row)
            return Response({"data": data})

        except Exception as e:
            # Log the full traceback so you can see exactly where things fail
            logger.error("ChurnPolicyDetail error for %r: %s\n%s",
                         policy_no, e, traceback.format_exc())
            return Response({"error": str(e)}, status=500)


class ChurnParamRanges(APIView):
    """
    GET /api/churn/param-ranges
    Returns min/max for adjustable parameters (derived from DB; fallback to defaults).
    Maps to your UI parameters:
      discount -> applicable_discount_with_ncb
      od_premium -> total_od_premium
      tp_premium -> total_tp_premium
      idv -> vehicle_idv
      add_on_premium -> "before_gst_add-on_gwp"
      ncb ->"ncb_%_previous_year"
    """

    

    def get(self, request):

        num =  _sql_num_ident
        sql = f"""
            SELECT
      MIN({num('applicable discount with ncb')}) AS discount_min,
      MAX({num('applicable discount with ncb')}) AS discount_max,
      MIN({num('total od premium')})            AS od_min,
      MAX({num('total od premium')})            AS od_max,
      MIN({num('total tp premium')})            AS tp_min,
      MAX({num('total tp premium')})            AS tp_max,
      MIN({num('vehicle idv')})                 AS idv_min,
      MAX({num('vehicle idv')})                 AS idv_max,
      MIN({num('before gst add-on gwp')})       AS addon_min,
      MAX({num('before gst add-on gwp')})       AS addon_max,
      MIN({num('ncb % previous year')})         AS ncb_min,
      MAX({num('ncb % previous year')})         AS ncb_max
    FROM {TABLE}
    WHERE "Predicted Status" = 'Not Renewed'
        """
        try:
            conn = _get_conn()
            cur = conn.cursor(cursor_factory=RealDictCursor)
            cur.execute(sql)
            agg = cur.fetchone() or {}

            # Build response with graceful fallbacks
            def _fallback(k_min, k_max, fb_key):
                mn = agg.get(k_min)
                mx = agg.get(k_max)
                if mn is None or mx is None or mn >= mx:
                    return FALLBACK_PARAM_RANGES[fb_key]
                return {"min": float(mn), "max": float(mx)}

            ranges = {
                "discount": _fallback("discount_min", "discount_max", "discount"),
                "od_premium": _fallback("od_min", "od_max", "od_premium"),
                "tp_premium": _fallback("tp_min", "tp_max", "tp_premium"),
                "idv": _fallback("idv_min", "idv_max", "idv"),
                "add_on_premium": _fallback("addon_min", "addon_max", "add_on_premium"),
                "ncb": _fallback("ncb_min", "ncb_max", "ncb"),
            }
            return Response({"ranges": ranges})
        except Exception as e:
            return Response({"error": str(e)}, status=500)
        finally:
            try:
                cur.close(); conn.close()
            except Exception:
                pass


# --- imports (add if not present) ---


def _load_assets():
    global _g_model, _g_encoders, _g_features
    try:
        _must_exist(MODEL_PATH,    "MODEL_PATH")
        _must_exist(ENCODERS_PATH, "ENCODERS_PATH")
        _must_exist(FEATURES_PATH, "FEATURES_PATH")

        logger.info("Loading assets from: MODEL=%s, ENCODERS=%s, FEATURES=%s",
                    MODEL_PATH, ENCODERS_PATH, FEATURES_PATH)

        if _g_model is None:
            _g_model = joblib.load(str(MODEL_PATH))
        if _g_encoders is None:
            _g_encoders = joblib.load(str(ENCODERS_PATH))
        if _g_features is None:
            _g_features = joblib.load(str(FEATURES_PATH))

        logger.info("Assets loaded OK: features=%s encoders=%s model=%s",
                    len(_g_features) if _g_features is not None else "None",
                    type(_g_encoders).__name__ if _g_encoders is not None else "None",
                    type(_g_model).__name__ if _g_model is not None else "None")
    except Exception as e:
        logger.exception("Asset load failed: %s", e)
        raise RuntimeError(f"Asset load failed: {e}")





# ------------ Col maps & helpers (ported from Streamlit) ------------
PARAM_TO_COL = {
    "discount":       "applicable discount with ncb",
    "od_premium":     "total od premium",
    "tp_premium":     "total tp premium",
    "idv":            "vehicle idv",
    "add_on_premium": "before gst add-on gwp",
    "ncb":            "ncb % previous year"
}

_EPS = 1e-9
IDV_MAX_INCREASE_FACTOR = 1.30

def _fetch_policy_row(policy_no: str) -> dict | None:
    sql = f"""
        SELECT
            "customerid","policy no","Customer Segment","biztype","tie up",
            "Cleaned Zone 2","Cleaned State2","Cleaned Branch Name 2",
            "make_clean","model_clean","variant",
            "total od premium","total tp premium","before gst add-on gwp",
            "applicable discount with ncb","ncb %% previous year","vehicle idv",
            "Predicted Status","Churn Probability","Top 3 Reasons",
            "policy end date",
            EXTRACT(MONTH FROM "policy end date")::int AS "policy end date_MONTH",
            EXTRACT(YEAR  FROM "policy end date")::int AS "policy end date_YEAR"
        FROM {TABLE}
        WHERE "policy no" = %s
        LIMIT 1
    """
    conn = _get_conn(); cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute(sql, (policy_no,))
        row = cur.fetchone()
        return row
    finally:
        try: cur.close(); conn.close()
        except Exception: pass

def _encode_for_model(row_series: pd.Series, features: list[str], label_encoders) -> pd.DataFrame:
    """
    Build a single-row DataFrame in the model's feature order, with robust imputations.
    - Categorical (in label_encoders): fill missing with "__UNK__" and map to new code.
    - Numeric (not in label_encoders): coerce to float, fill invalid/missing with 0.0.
    Also logs which features were imputed and warns if any NaNs survive before final fillna.
    """
    is_cat = set(label_encoders.keys()) if isinstance(label_encoders, dict) else set()

    vals = {}
    filled_unknowns: list[str] = []
    filled_zeroes: list[str] = []

    for f in features:
        v = row_series.get(f, None)
        missing = (
            v is None
            or (isinstance(v, float) and np.isnan(v))
            or (isinstance(v, str) and v.strip() == "")
        )
        if missing:
            if f in is_cat:
                v = "__UNK__"
                filled_unknowns.append(f)
            else:
                v = 0.0
                filled_zeroes.append(f)
        vals[f] = v

    s = pd.Series(vals, dtype="object")

    # Encode categoricals
    if isinstance(label_encoders, dict):
        for col in is_cat:
            if col in s.index:
                enc = label_encoders[col]
                classes = getattr(enc, "classes_", None)
                if classes is not None:
                    mapping = {lab: i for i, lab in enumerate(classes)}
                    unk_code = len(mapping)  # unseen label bucket
                    s[col] = mapping.get(s[col], unk_code)

    # Coerce numerics
    for col in s.index:
        if col not in is_cat:
            try:
                s[col] = float(str(s[col]).replace("₹", "").replace(",", "").strip() or "0")
            except Exception:
                s[col] = 0.0

    X = pd.DataFrame([s[features]])

    # ⬇️ Your requested snippet: warn if anything is still NaN pre-final-fill
    missing = [c for c in X.columns if pd.isna(X.iloc[0][c])]
    if missing:
        logger.warning("Missing features filled at predict time: %s", missing)

    # Final safety net
    X = X.replace([np.inf, -np.inf], 0.0).fillna(0.0)

    # Helpful audit log (what we imputed)
    if filled_unknowns or filled_zeroes:
        logger.info(
            "Imputed features -> unknown(cats): %s | zero(nums): %s",
            filled_unknowns, filled_zeroes
        )

    return X



def _recompute_totals(row: dict, gst_rate=0.18):
    od = _num(row.get("total od premium"))
    tp = _num(row.get("total tp premium"))
    gst = (od + tp) * float(gst_rate)
    row["gst"] = gst
    row["total premium payable"] = od + tp + gst


def _apply_discount_coupling(original_row: dict, test_row: dict, param_ranges: dict, gst_rate=0.18, stack_on_top=True) -> bool:
    """Return True if would zero-out (stop)."""
    old_d = float(original_row["applicable discount with ncb"]) / 100.0
    new_d = float(test_row["applicable discount with ncb"]) / 100.0
    extra = max(0.0, new_d - old_d)

    od0 = float(original_row["total od premium"])
    tp0 = float(original_row["total tp premium"])

    if extra <= 0.0:
        _recompute_totals(test_row, gst_rate); return False

    base_od = float(test_row["total od premium"]) if stack_on_top else od0
    base_tp = float(test_row["total tp premium"]) if stack_on_top else tp0

    new_od = base_od * (1.0 - extra)
    new_tp = base_tp * (1.0 - extra)

    if od0 <= _EPS: new_od = 0.0
    elif new_od <= _EPS: return True
    if tp0 <= _EPS: new_tp = 0.0
    elif new_tp <= _EPS: return True

    mn_od, mx_od = float(param_ranges["od_premium"]["min"]), float(param_ranges["od_premium"]["max"])
    mn_tp, mx_tp = float(param_ranges["tp_premium"]["min"]), float(param_ranges["tp_premium"]["max"])
    test_row["total od premium"] = 0.0 if od0 <= _EPS else min(max(new_od, mn_od), mx_od)
    test_row["total tp premium"] = 0.0 if tp0 <= _EPS else min(max(new_tp, mn_tp), mx_tp)

    _recompute_totals(test_row, gst_rate)
    return False

def _infer_discount_from_odtp(original_row: dict, stepped_row: dict, param_ranges: dict) -> float:
    od0 = float(original_row["total od premium"])
    tp0 = float(original_row["total tp premium"])
    ods = float(stepped_row["total od premium"])
    tps = float(stepped_row["total tp premium"])
    base = od0 + tp0
    if base <= _EPS:
        return float(original_row["applicable discount with ncb"])
    extra = max(0.0, 1.0 - (ods + tps) / base)
    disc0 = float(original_row["applicable discount with ncb"])
    mn_d = float(param_ranges["discount"]["min"])
    mx_d = float(param_ranges["discount"]["max"])
    return min(max(disc0 + 100.0 * extra, mn_d), mx_d)

def _apply_manual_two_way(original_row: dict, param_ranges: dict, sliders: dict) -> dict:
    row = {**original_row}

    disc_col = PARAM_TO_COL["discount"]; od_col = PARAM_TO_COL["od_premium"]
    tp_col   = PARAM_TO_COL["tp_premium"]; idv_col = PARAM_TO_COL["idv"]
    add_col  = PARAM_TO_COL["add_on_premium"]; ncb_col = PARAM_TO_COL["ncb"]

    # current numeric (null-safe)
    od0 = _num(original_row.get(od_col))
    tp0 = _num(original_row.get(tp_col))
    add0 = _num(original_row.get(add_col))
    disc0 = _num(original_row.get(disc_col))
    idv0 = _num(original_row.get(idv_col))

    # apply sliders into row (null-safe)
    if "discount" in sliders: row[disc_col] = float(sliders["discount"])
    if "od_premium" in sliders: row[od_col] = float(sliders["od_premium"])
    if "tp_premium" in sliders: row[tp_col] = float(sliders["tp_premium"])
    if "idv" in sliders:
        row[idv_col] = min(float(sliders["idv"]), idv0 * IDV_MAX_INCREASE_FACTOR if idv0 else float(sliders["idv"]))
    if "add_on_premium" in sliders: row[add_col] = float(sliders["add_on_premium"])
    if "ncb" in sliders: row[ncb_col] = float(sliders["ncb"])

    has_disc = "discount" in sliders
    has_od   = "od_premium" in sliders
    has_tp   = "tp_premium" in sliders

    if has_disc and (has_od or has_tp):
        new_d = _num(row.get(disc_col))
        old_d = disc0
        extra = max(0.0, (new_d - old_d)/100.0)

        def _fill(comp_key, col, base0):
            if base0 <= _EPS: return 0.0
            cand = base0 * (1.0 - extra) if extra > 0.0 else base0
            mn = float(param_ranges[comp_key]["min"]); mx = float(param_ranges[comp_key]["max"])
            if extra > 0.0 and cand <= _EPS: cand = base0
            return min(max(cand, mn), mx)

        if not has_od: row[od_col] = _fill("od_premium", od_col, od0)
        if not has_tp: row[tp_col] = _fill("tp_premium", tp_col, tp0)
        _recompute_totals(row, 0.18)
        return row

    if has_disc and not (has_od or has_tp):
        _apply_discount_coupling(original_row, row, param_ranges, 0.18, stack_on_top=False)
        return row

    if (has_od or has_tp) and not has_disc:
        row[disc_col] = _infer_discount_from_odtp(original_row, row, param_ranges)
        _recompute_totals(row, 0.18)
        return row

    _recompute_totals(row, 0.18)
    return row


def _get_param_ranges() -> dict:
    """Use your range endpoint’s logic; here we call the same aggregate as ChurnParamRanges."""
    num = _sql_num_ident
    sql = f"""
      SELECT
        MIN({num('applicable discount with ncb')}) AS discount_min,
        MAX({num('applicable discount with ncb')}) AS discount_max,
        MIN({num('total od premium')})            AS od_min,
        MAX({num('total od premium')})            AS od_max,
        MIN({num('total tp premium')})            AS tp_min,
        MAX({num('total tp premium')})            AS tp_max,
        MIN({num('vehicle idv')})                 AS idv_min,
        MAX({num('vehicle idv')})                 AS idv_max,
        MIN({num('before gst add-on gwp')})       AS addon_min,
        MAX({num('before gst add-on gwp')})       AS addon_max,
        MIN({num('ncb % previous year')})         AS ncb_min,
        MAX({num('ncb % previous year')})         AS ncb_max
      FROM {TABLE}
      WHERE "Predicted Status" = 'Not Renewed'
    """
    conn = _get_conn(); cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute(sql)
        agg = cur.fetchone() or {}
    finally:
        try: cur.close(); conn.close()
        except Exception: pass

    def _mk(mn, mx, fb_min, fb_max):
        a, b = agg.get(mn), agg.get(mx)
        if a is None or b is None or a >= b: return {"min": fb_min, "max": fb_max}
        return {"min": float(a), "max": float(b)}

    return {
        "discount":       _mk("discount_min","discount_max", 0.0, 90.0),
        "od_premium":     _mk("od_min","od_max", 10000.0, 50000.0),
        "tp_premium":     _mk("tp_min","tp_max", 1500.0, 8000.0),
        "idv":            _mk("idv_min","idv_max", 200000.0, 2000000.0),
        "add_on_premium": _mk("addon_min","addon_max", 1000.0, 10000.0),
        "ncb":            _mk("ncb_min","ncb_max", 0.0, 50.0),
    }

# -------------------------- /churn/simulate-manual --------------------------
# views.py
import traceback
import numpy as np
import pandas as pd

from django.conf import settings
from rest_framework.views import APIView
from rest_framework.response import Response

# Expected globals loaded by _load_assets():
#   _g_model, _g_features (list/Index), _g_encoders (dict or Pipeline)
# Expected helpers you already have:
#   _load_assets(), _fetch_policy_row(policy_no), _get_param_ranges(),
#   _encode_for_model(series, features, encoders),
#   _apply_manual_two_way(base_row_dict, ranges, sliders)
#
# Also expects: PARAM_TO_COL (mapping from FE param key -> column name)

def _num(x):
    """Parse numbers robustly: handles None, '₹', commas, blanks."""
    if x is None:
        return 0.0
    try:
        # stringify and strip currency/commas/spaces
        s = str(x).strip().replace('₹', '').replace(',', '')
        if s == '' or s.lower() == 'nan':
            return 0.0
        return float(s)
    except Exception:
        return 0.0
def _sql_num(col: str) -> str:
    # Strips everything except digits, dot, minus; empty -> NULL; cast to double precision
    return f"""CAST(NULLIF(regexp_replace({col}::text, '[^0-9.\\-]', '', 'g'), '') AS double precision)"""

def _sql_num_ident(col: str) -> str:
    # Safely quote an identifier like app/TP columns that contain spaces or %.
    # Produces: CAST(NULLIF(regexp_replace("total od premium"::text, '[^0-9.\-]', '', 'g'), '') AS double precision)
    return f'''CAST(NULLIF(regexp_replace("{col}"::text, '[^0-9.\\-]', '', 'g'), '') AS double precision)'''

def _predict_pct(model, X):
    """
    Returns churn probability (0-100).
    Works for models that return:
      - predict_proba -> (n,2) or (n,) or (n,1)
      - predict -> (n,) logits/score (fallback, treated as prob if [0,1])
    """
    # Try predict_proba first
    if hasattr(model, "predict_proba"):
        proba = model.predict_proba(X)
        proba = np.asarray(proba)
        # (n,2) -> take column 1; (n,1) or (n,) -> take that
        if proba.ndim == 2 and proba.shape[1] >= 2:
            p1 = proba[:, 1]
        else:
            p1 = proba.ravel()
        return float(np.clip(p1[0] * 100.0, 0.0, 100.0))

    # Fallback: predict; assume already a prob in [0,1]
    pred = np.asarray(getattr(model, "predict")(X)).ravel()
    p = float(pred[0])
    # if it looks like a logit/score, squash (very defensive)
    if p < 0.0 or p > 1.0:
        p = 1.0 / (1.0 + np.exp(-p))
    return float(np.clip(p * 100.0, 0.0, 100.0))

class ChurnSimulateManual(APIView):
    """
    POST /api/churn/simulate-manual
    Body:
      {
        "policy_no": "...",
        "sliders": {discount?, od_premium?, tp_premium?, idv?, add_on_premium?, ncb?}
      }
    Returns:
      { baseline_pct, updated_pct, row, deltas }
    """
    def post(self, request):
        stage = "start"
        try:
            # ── 1) Load assets with strong diagnostics ───────────────────────────
            stage = "load_assets"
            try:
                _load_assets()
            except Exception as e:
                logger.exception(
                    "simulate-manual: asset load failed. "
                    "MODEL=%s (exists=%s)  ENCODERS=%s (exists=%s)  FEATURES=%s (exists=%s)",
                    MODEL_PATH, MODEL_PATH.exists(),
                    ENCODERS_PATH, ENCODERS_PATH.exists(),
                    FEATURES_PATH, FEATURES_PATH.exists()
                )
                return Response({
                    "error": "Model assets missing or unreadable",
                    "stage": stage,
                    "paths": {
                        "MODEL_PATH": str(MODEL_PATH),
                        "ENCODERS_PATH": str(ENCODERS_PATH),
                        "FEATURES_PATH": str(FEATURES_PATH),
                    }
                }, status=503)

            # Sanity: ensure globals are populated
            if any(x is None for x in (_g_model, _g_encoders, _g_features)):
                logger.error(
                    "simulate-manual: assets are None after _load_assets(). "
                    "model=%s encoders=%s features=%s",
                    type(_g_model).__name__ if _g_model is not None else None,
                    type(_g_encoders).__name__ if _g_encoders is not None else None,
                    (len(_g_features) if hasattr(_g_features, "__len__") else None)
                )
                return Response({
                    "error": "Model assets not initialized after load",
                    "stage": stage
                }, status=503)

            # ── 2) Inputs ───────────────────────────────────────────────────────
            stage = "read_inputs"
            policy_no = str(request.data.get("policy_no") or request.data.get("policy no") or "").strip()
            sliders = request.data.get("sliders") or {}
            if not policy_no:
                return Response({"error": "policy_no is required"}, status=422)

            # ── 3) Fetch baseline row ───────────────────────────────────────────
            stage = "fetch_base_row"
            base = _fetch_policy_row(policy_no)
            if not base:
                return Response({"error": "policy not found"}, status=404)
            base = dict(base)  # ensure plain dict

            # ── 4) Param ranges (DB with graceful fallbacks) ────────────────────
            stage = "get_ranges"
            ranges = _get_param_ranges()

            # ── 5) Baseline encode + predict ────────────────────────────────────
            stage = "encode_baseline"
            base_series = pd.Series(base)
            X_base = _encode_for_model(base_series, _g_features, _g_encoders)
            if isinstance(X_base, pd.Series):
                X_base = X_base.to_frame().T
            elif not isinstance(X_base, pd.DataFrame):
                X_base = np.asarray(X_base).reshape(1, -1)

            stage = "predict_baseline"
            baseline_pct = _predict_pct(_g_model, X_base)
            logger.info(
                "simulate-manual: baseline ok. X_base.shape=%s  features=%s  model=%s  baseline_pct=%.4f",
                getattr(X_base, "shape", None),
                len(_g_features) if hasattr(_g_features, "__len__") else None,
                type(_g_model).__name__,
                baseline_pct
            )

            # ── 6) Apply manual coupling + predict ──────────────────────────────
            stage = "apply_manual_two_way"
            test_row = _apply_manual_two_way(dict(base), ranges, sliders)

            stage = "encode_updated"
            X_t = _encode_for_model(pd.Series(test_row), _g_features, _g_encoders)
            if isinstance(X_t, pd.Series):
                X_t = X_t.to_frame().T
            elif not isinstance(X_t, pd.DataFrame):
                X_t = np.asarray(X_t).reshape(1, -1)

            stage = "predict_updated"
            updated_pct = _predict_pct(_g_model, X_t)

            # ── 7) Deltas ───────────────────────────────────────────────────────
            stage = "compute_deltas"
            deltas = {}
            for fe_key, col in (PARAM_TO_COL or {}).items():
                bv = _num(base.get(col))
                tv = _num(test_row.get(col))
                deltas[fe_key] = round(tv - bv, 6)

            deltas["gst"] = round(_num(test_row.get("gst")) - _num(base.get("gst")), 6)
            deltas["total"] = round(
                _num(test_row.get("total premium payable")) - _num(base.get("total premium payable")), 6
            )

            # ── 8) Respond ─────────────────────────────────────────────────────
            stage = "respond"
            out_row = {**test_row}
            return Response({
                "baseline_pct": round(baseline_pct, 4),
                "updated_pct": round(updated_pct, 4),
                "row": out_row,
                "deltas": deltas
            })

        except Exception as e:
            # When DEBUG=True, surface the last few traceback lines to help pin the stage.
            if settings.DEBUG:
                tb_last_lines = traceback.format_exc().splitlines()[-12:]
                logger.error("simulate-manual: error at stage=%s: %s", stage, e)
                return Response(
                    {"error": str(e), "stage": stage, "trace": tb_last_lines},
                    status=500
                )
            return Response({"error": f"{e.__class__.__name__} at {stage}"}, status=500)

# -------------------------- auto-suggest core --------------------------
def _get_param_direction(reasons: list[str]) -> dict:
    direction = {}
    for r in reasons:
        if "High Own-Damage Premium" in r: direction["od_premium"] = "decrease"
        if "High Third-Party Premium" in r: direction["tp_premium"] = "decrease"
        if "High Add-On Premium" in r: direction["add_on_premium"] = "decrease"
        if "Low Vehicle IDV" in r: direction["idv"] = "increase"
        if "Low Discount" in r or "Low Discount with NCB" in r: direction["discount"] = "increase"
        if "Low No Claim Bonus" in r or "Low No Claim Bonus Percentage" in r: direction["ncb"] = "increase"
    return direction

def _step_size(param: str, value: float) -> float:
    if param in ("od_premium","tp_premium","add_on_premium"): return 1000.0 if value >= 2000.0 else 100.0
    if param == "idv": return 10000.0
    if param in ("discount","ncb"): return 10.0
    return 1.0

def _build_combo_for_k(original_row: dict, direction: dict, param_ranges: dict, k_moves: int):
    row = dict(original_row); moved=set(); stop=False
    od0=float(original_row["total od premium"]); tp0=float(original_row["total tp premium"])
    add0=float(original_row["before gst add-on gwp"]); orig_idv=float(original_row["vehicle idv"])

    def _apply(param_key, col_name):
        nonlocal stop, row, moved
        cur=float(original_row[col_name]); step=_step_size(param_key, cur)
        mn=float(param_ranges[param_key]["min"]); mx=float(param_ranges[param_key]["max"])
        dirn=direction.get(param_key)
        if not dirn: return
        if dirn=="decrease" and cur<=_EPS: return
        if param_key=="idv" and dirn=="increase":
            pct=min(k_moves*0.10, IDV_MAX_INCREASE_FACTOR-1.0)  # 10% per k, capped at +30%
            cand = orig_idv*(1.0+pct)
            newv=min(max(cand,mn), min(mx, orig_idv*IDV_MAX_INCREASE_FACTOR))
        else:
            cand = cur + (step*k_moves if dirn=="increase" else -step*k_moves)
            newv=min(max(cand,mn), mx)
        if param_key=="od_premium" and od0<=_EPS: newv=0.0
        if param_key=="tp_premium" and tp0<=_EPS: newv=0.0
        if param_key=="add_on_premium" and add0<=_EPS: newv=0.0
        if abs(newv-cur)>1e-9: row[col_name]=newv; moved.add(param_key)

    # priority order: idv/addon/ncb first, then discount/od/tp w/ coupling
    if "idv" in direction: _apply("idv", PARAM_TO_COL["idv"])
    if "add_on_premium" in direction: _apply("add_on_premium", PARAM_TO_COL["add_on_premium"])
    if "ncb" in direction: _apply("ncb", PARAM_TO_COL["ncb"])

    if "discount" in direction:
        _apply("discount", PARAM_TO_COL["discount"])
        if abs(float(row[PARAM_TO_COL["discount"]]) - float(original_row[PARAM_TO_COL["discount"]])) > 1e-9:
            stop = _apply_discount_coupling(original_row, row, param_ranges, 0.18, True)
            if stop: return row, moved, True
        else:
            _recompute_totals(row, 0.18)
    else:
        if "od_premium" in direction: _apply("od_premium", PARAM_TO_COL["od_premium"])
        if "tp_premium" in direction: _apply("tp_premium", PARAM_TO_COL["tp_premium"])
        row[PARAM_TO_COL["discount"]] = _infer_discount_from_odtp(original_row, row, param_ranges)
        _recompute_totals(row, 0.18)

    return row, moved, False

def _run_auto(original_row: dict, features: list[str], encoders: dict, model, param_ranges: dict, target=50.0, max_k=50):
    raw = str(original_row.get("Top 3 Reasons",""))
    reasons = [r.strip() for r in re.split(r",|\band\b", raw) if r.strip()]
    direction = _get_param_direction(reasons)
    Xb = _encode_for_model(pd.Series(original_row), features, encoders)
    baseline_pct = float(model.predict_proba(Xb)[0][1] * 100.0)

    trials=[]; best_idx=None; best=baseline_pct; prev_sig=None
    for k in range(1, max_k+1):
        row, moved, stop = _build_combo_for_k(original_row, direction, param_ranges, k)
        if stop: break
        sig=(float(row["applicable discount with ncb"]), float(row["vehicle idv"]),
             float(row["before gst add-on gwp"]), float(row["total od premium"]), float(row["total tp premium"]))
        if prev_sig==sig: break
        prev_sig=sig
        Xt=_encode_for_model(pd.Series(row), features, encoders)
        churn=float(model.predict_proba(Xt)[0][1]*100.0)
        _rec = {
            "k_moves": k, "moved": sorted(list(moved)) if moved else [],
            "row": row, "churn_pct": churn,
            "od": float(row["total od premium"]), "tp": float(row["total tp premium"]),
            "gst": float(row.get("gst", (row["total od premium"]+row["total tp premium"])*0.18)),
            "total_premium": float(row.get("total premium payable", row["total od premium"]+row["total tp premium"]+row.get("gst",0))),
            "discount": float(row["applicable discount with ncb"]),
            "idv": float(row["vehicle idv"]), "addon": float(row["before gst add-on gwp"]),
            "ncb": float(row.get("ncb % previous year", 0.0)),   # ✅ correct key

        }
        trials.append(_rec)
        if churn < best: best, best_idx = churn, len(trials)-1
        if churn < target: break
    best_obj = trials[best_idx] if best_idx is not None else (trials[-1] if trials else None)
    return baseline_pct, trials, best_obj

# -------------------------- /churn/auto-suggest --------------------------
import re
class ChurnAutoSuggest(APIView):
    """
    POST /api/churn/auto-suggest
    Body: { "policy_no": "...", "target"?:50, "max_k"?:50 }
    Returns: { baseline_pct, trials:[...], best:{row, churn_pct, ...} }
    """
    def post(self, request):
        try:
            _load_assets()
            policy_no = str(request.data.get("policy_no") or request.data.get("policy no") or "").strip()
            target = float(request.data.get("target", 50.0))
            max_k = int(request.data.get("max_k", 50))

            base = _fetch_policy_row(policy_no)
            if not base: return Response({"error": "policy not found"}, status=404)
            ranges = _get_param_ranges()

            baseline_pct, trials, best_obj = _run_auto(base, _g_features, _g_encoders, _g_model, ranges, target, max_k)
            out_trials = [{
                "k_moves": t["k_moves"], "moved": t["moved"],
                "discount": t["discount"], "idv": t["idv"], "addon": t["addon"], "ncb": t["ncb"],
                "od": t["od"], "tp": t["tp"], "gst": t["gst"], "total_premium": t["total_premium"],
                "churn_pct": t["churn_pct"],
            } for t in trials]
            best_light = None
            if best_obj:
                best_light = { k: best_obj.get(k) for k in ("k_moves","moved","churn_pct") }
                best_light["row"] = best_obj.get("row")
            return Response({"baseline_pct": round(baseline_pct,4), "trials": out_trials, "best": best_light})
        except Exception as e:
            return Response({"error": str(e)}, status=500)

# -------------------------- save/show Selected changes --------------------------
def _norm_policy(s: str) -> str:
    s = str(s or "").strip()
    if s.startswith(("'", '"')): s = s[1:]
    return s.replace(" ", "").upper()

SAVE_SCHEMA = "Prediction"
SAVE_TABLE  = "selected_changes"
FULL_SAVE_FQN = f'"{SAVE_SCHEMA}"."{SAVE_TABLE}"'

def _ensure_save_table():
    ddl = f"""
    CREATE TABLE IF NOT EXISTS {FULL_SAVE_FQN} (
        policy_no TEXT,
        policy_no_norm TEXT,
        customerid TEXT,
        selection_type TEXT,
        business_type TEXT,
        tie_up_type TEXT,
        zone TEXT,
        state TEXT,
        branch TEXT,
        vehicle TEXT,
        old_discount DOUBLE PRECISION,
        old_ncb DOUBLE PRECISION,
        old_idv DOUBLE PRECISION,
        old_add_on_premium DOUBLE PRECISION,
        old_od DOUBLE PRECISION,
        old_tp DOUBLE PRECISION,
        old_gst DOUBLE PRECISION,
        old_total_premium DOUBLE PRECISION,
        new_discount DOUBLE PRECISION,
        new_ncb DOUBLE PRECISION,
        new_idv DOUBLE PRECISION,
        new_add_on_premium DOUBLE PRECISION,
        new_od DOUBLE PRECISION,
        new_tp DOUBLE PRECISION,
        new_gst DOUBLE PRECISION,
        new_total_premium DOUBLE PRECISION,
        churn_risk_pct DOUBLE PRECISION,
        top_3_reasons TEXT,
        created_at TIMESTAMPTZ DEFAULT NOW()
    );
    """
    idx = f'CREATE INDEX IF NOT EXISTS idx_selected_changes_policy_norm ON {FULL_SAVE_FQN}(policy_no_norm);'
    conn = _get_conn(); cur = conn.cursor()
    try:
        cur.execute(ddl); cur.execute(idx); conn.commit()
    finally:
        try: cur.close(); conn.close()
        except Exception: pass

class ChurnSaveSelected(APIView):
    """
    POST /api/churn/save-selected
    Body: { policy_no, selection_type, selected_row:{...} }
    """
    def post(self, request):
        try:
            policy_no = str(request.data.get("policy_no") or request.data.get("policy no") or "").strip()
            selection_type = str(request.data.get("selection_type","")).strip() or "Manual Simulation"
            selected_row = request.data.get("selected_row") or {}
            if not policy_no or not selected_row: return Response({"error":"policy_no & selected_row required"}, status=400)

            base = _fetch_policy_row(policy_no)
            if not base: return Response({"error":"policy not found"}, status=404)

            # ensure computed fields on selected
            if "gst" not in selected_row or "total premium payable" not in selected_row:
                _recompute_totals(selected_row, 0.18)

            # base stats
            old_discount = float(base.get("applicable discount with ncb", 0.0))
            old_ncb      = float(base.get("ncb % previous year", 0.0))
            old_idv      = float(base.get("vehicle idv", 0.0))
            old_addon    = float(base.get("before gst add-on gwp", 0.0))
            old_od       = float(base.get("total od premium", 0.0))
            old_tp       = float(base.get("total tp premium", 0.0))
            old_gst      = float(base.get("gst", (old_od + old_tp) * 0.18))
            old_total    = float(base.get("total premium payable", old_od + old_tp + old_gst))

            new_discount = float(selected_row.get("applicable discount with ncb", selected_row.get("discount", 0.0)))
            new_ncb      = float(selected_row.get("ncb % previous year", selected_row.get("ncb", 0.0)))
            new_idv      = float(selected_row.get("vehicle idv", selected_row.get("idv", 0.0)))
            new_addon    = float(selected_row.get("before gst add-on gwp", selected_row.get("add_on_premium", 0.0)))
            new_od       = float(selected_row.get("total od premium", selected_row.get("od", 0.0)))
            new_tp       = float(selected_row.get("total tp premium", selected_row.get("tp", 0.0)))
            new_gst      = float(selected_row.get("gst", (new_od + new_tp) * 0.18))
            new_total    = float(selected_row.get("total premium payable", selected_row.get("total_premium", new_od + new_tp + new_gst)))

            # baseline churn risk in %
            _load_assets()
            Xb = _encode_for_model(pd.Series(base), _g_features, _g_encoders)
            churn_pct = float(_g_model.predict_proba(Xb)[0][1] * 100.0)

            policy_norm = _norm_policy(policy_no)
            vehicle = f"{base.get('make_clean','')} {base.get('model_clean','')} ({base.get('variant','')})".strip()

            _ensure_save_table()
            conn = _get_conn(); cur = conn.cursor()
            try:
                # upsert = delete old then insert new snapshot
                cur.execute(f'DELETE FROM {FULL_SAVE_FQN} WHERE policy_no_norm=%s;', (policy_norm,))
                cur.execute(f"""
                    INSERT INTO {FULL_SAVE_FQN} (
                      policy_no, policy_no_norm, customerid, selection_type, business_type, tie_up_type,
                      zone, state, branch, vehicle,
                      old_discount, old_ncb, old_idv, old_add_on_premium, old_od, old_tp, old_gst, old_total_premium,
                      new_discount, new_ncb, new_idv, new_add_on_premium, new_od, new_tp, new_gst, new_total_premium,
                      churn_risk_pct, top_3_reasons
                    ) VALUES (
                      %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
                      %s,%s,%s,%s,%s,%s,%s,%s,
                      %s,%s,%s,%s,%s,%s,%s,%s,
                      %s,%s
                    );
                """, (
                    base.get("policy no"), policy_norm, str(base.get("customerid","")),
                    selection_type, str(base.get("biztype","")), str(base.get("tie up","")),
                    str(base.get("Cleaned Zone 2","")), str(base.get("Cleaned State2","")),
                    str(base.get("Cleaned Branch Name 2","")), vehicle,
                    old_discount, old_ncb, old_idv, old_addon, old_od, old_tp, old_gst, old_total,
                    new_discount, new_ncb, new_idv, new_addon, new_od, new_tp, new_gst, new_total,
                    churn_pct, str(base.get("Top 3 Reasons",""))
                ))
                conn.commit()
            finally:
                try: cur.close(); conn.close()
                except Exception: pass

            return Response({"ok": True})
        except Exception as e:
            return Response({"error": str(e)}, status=500)

class ChurnLatestSelected(APIView):
    """
    GET /api/churn/latest-selected?policy_no=...
    """
    def get(self, request):
        policy_no = str(request.GET.get("policy_no") or request.GET.get("policy no") or "").strip()
        if not policy_no: return Response({"error":"policy_no required"}, status=400)
        policy_norm = _norm_policy(policy_no)
        sql = f"""
            SELECT * FROM {FULL_SAVE_FQN}
            WHERE policy_no_norm = %s
            ORDER BY created_at DESC
            LIMIT 1;
        """
        conn = _get_conn(); cur = conn.cursor(cursor_factory=RealDictCursor)
        try:
            cur.execute(sql, (policy_norm,))
            row = cur.fetchone()
            if not row: return Response({"error":"not found"}, status=404)
            return Response(row)
        finally:
            try: cur.close(); conn.close()
            except Exception: pass

# -------------------------- Draft Email (optional Groq) --------------------------
# GROQ_API_KEY_Churn = "REMOVEDhqk7OIprKrMafXvFTRxIWGdyb3FYf19I66DpucmiC21LG1zwxrI8"

SPECIAL_REASONS = {
    "Young Vehicle Age","Old Vehicle Age","Claims Happened",
    "Multiple Claims on Record","Minimal Policies Purchased","Tie Up with Non-OEM"
}


# # ---------------------- NEW: Hidden rules prompt ----------------------
# RULES_PROMPT = """
# You are an expert retention-email writer for car insurance.

# DO NOT reveal the following instructions in the output—use them only to guide the draft.

# Rules (must follow):
# - First line must start with "Subject: ".
# - Mention their policy details like vehicle names.
# - Use figures in ctx to show discount Δ (pp) and savings (₹).
# - Mention OD/TP impact with numbers from ctx.
# - Address each reason in top_3_reasons implicitly; if any of these appear:
#   Young Vehicle Age, Old Vehicle Age, Claims Happened, Multiple Claims on Record,
#   Minimal Policies Purchased, Tie Up with Non-OEM
#   then add a short reassurance line (but do not name the reason).
# - Do not mention the exact phrases from top_3_reasons in the draft.
# - Do not use the phrase "we noticed".
# - Do not use phrases like "renewal approaches".
# - Friendly, persuasive, action-oriented. Finish with a clear renewal CTA.
# """.strip()
# # ---------------------------------------------------------------------


# _SUBJ_RE = re.compile(r"^\s*subject\s*:\s*(.+)$", re.I)

# def _split_subj_body(text: str) -> tuple[str, str]:
#     lines = (text or "").splitlines()
#     subj, body = None, []
#     for ln in lines:
#         m = _SUBJ_RE.match(ln)
#         if m and subj is None:
#             subj = m.group(1).strip()
#         else:
#             body.append(ln)
#     return (subj or "").strip()[:78], "\n".join(body).strip()


# def _draft_email_text(latest_row: dict, fallback_title="Policy Renewal Options"):
#     vehicle    = str(latest_row.get("vehicle", "")).strip()

#     # safe float parser
#     def f(x, default=0.0):
#         try:
#             return float(x if x is not None else default)
#         except (TypeError, ValueError):
#             return default

#     old_total  = f(latest_row.get("old_total_premium"))
#     new_total  = f(latest_row.get("new_total_premium"))
#     savings    = old_total - new_total

#     old_disc   = f(latest_row.get("old_discount"))
#     new_disc   = f(latest_row.get("new_discount"))
#     disc_pp    = new_disc - old_disc  # Δ in percentage points

#     old_od     = f(latest_row.get("old_od"))
#     new_od     = f(latest_row.get("new_od"))
#     od_delta   = new_od - old_od

#     old_tp     = f(latest_row.get("old_tp"))
#     new_tp     = f(latest_row.get("new_tp"))
#     tp_delta   = new_tp - old_tp

#     reasons     = str(latest_row.get("top_3_reasons", "") or "")
#     reassurance = ""
#     if 'SPECIAL_REASONS' in globals() and any(r in reasons for r in SPECIAL_REASONS):
#         reassurance = (
#             "We’ve cross-checked your cover and eligibility so the protection remains appropriate, "
#             "with no gaps on critical risks."
#         )

#     # Subject tailored to vehicle (no "we noticed")
#     subj = f"Renew your {vehicle}: extra savings with aligned coverage" if vehicle else fallback_title

#     # Precompute reassurance line to avoid backslashes in f-string expressions
#     reassurance_line = (reassurance + "\n") if reassurance else ""

#     # Body — follows all rules; does not echo reason phrases
#     body = (
#         f"Subject: {subj}\n\n"
#         f"Hi,\n\n"
#         f"I’ve reviewed your {('policy for ' + vehicle) if vehicle else 'car insurance policy'} and prepared a renewal option that keeps your coverage aligned while lowering the total cost.\n\n"
#         f"• Discount improvement: {disc_pp:+.1f} pp\n"
#         f"• Estimated savings: ₹{int(round(savings)):,} vs. your previous total\n"
#         f"• OD impact: ₹{int(round(od_delta)):,} (new OD ₹{int(round(new_od)):,})\n"
#         f"• TP impact: ₹{int(round(tp_delta)):,} (new TP ₹{int(round(new_tp)):,})\n"
#         f"{reassurance_line}"
#         f"Everything else—including key protections and add-ons—remains consistent so you don’t lose the benefits you rely on.\n\n"
#         f"If you’d like, I can share the revised schedule right away and complete the renewal in a few clicks. "
#         f"Shall I proceed?\n\n"
#         f"Regards,\n"
#         f"Team"
#     )
#     return body


# class ChurnDraftEmail(APIView):
#     def post(self, request):
#         try:
#             policy_no = str(request.data.get("policy_no") or request.data.get("policy no") or "").strip()
#             if not policy_no:
#                 return Response({"error": "policy_no required"}, status=400)

#             policy_norm = _norm_policy(policy_no)
#             sql = f"""SELECT * FROM {FULL_SAVE_FQN}
#                       WHERE policy_no_norm=%s ORDER BY created_at DESC LIMIT 1;"""
#             conn = _get_conn(); cur = conn.cursor(cursor_factory=RealDictCursor)
#             try:
#                 cur.execute(sql, (policy_norm,))
#                 latest = cur.fetchone()
#             finally:
#                 try: cur.close(); conn.close()
#                 except Exception: pass

#             if not latest:
#                 return Response({"error": "No saved Selected changes. Save first."}, status=400)

#             ctx = "; ".join(
#                 f"{k}={latest.get(k)}" for k in [
#                     "vehicle","old_discount","new_discount","old_total_premium","new_total_premium",
#                     "old_od","new_od","old_tp","new_tp","old_add_on_premium","new_add_on_premium",
#                     "old_idv","new_idv","top_3_reasons"
#                 ] if k in latest
#             )

#             prompt = f"""{RULES_PROMPT}

# OUTPUT FORMAT (strict):
# - First line: "Subject: <your subject here>"
# - Then a blank line, then the body only.

# Additional guidance:
# - Make the subject short, compelling, and specific to the vehicle (if present).
# - Mention the vehicle/policy in the body.
# - Use numbers from ctx to explain premium change and component deltas (OD/TP, discount Δ in pp, total ₹ savings).
# - Address reasons implicitly; never print the exact reason strings.
# - Do not use the phrase "we noticed".
# - End with a clear renewal CTA.

# ctx: {ctx}
# """.strip()

#             txt = azure_chat(prompt, system="You are a helpful assistant for drafting renewal emails.", temperature=0.2)
#             subject, body = _split_subj_body(txt)

#             if not subject or not body:
#                 # Deterministic fallback obeying rules
#                 fb = _draft_email_text(latest)
#                 subject, body = _split_subj_body(fb)

#             return Response({"subject": subject, "body": body})
#         except Exception as e:
#             logger.exception("draft-email failed: %s", e)
#             return Response({"error": str(e)}, status=500)




_SUBJ_RE = re.compile(r"^\s*subject\s*:\s*(.+)$", re.I)

def _split_subj_body(text: str) -> tuple[str, str]:
    lines = (text or "").splitlines()
    subj, body = None, []
    for ln in lines:
        m = _SUBJ_RE.match(ln)
        if m and subj is None:
            subj = m.group(1).strip()
        else:
            body.append(ln)
    return (subj or "").strip()[:78], "\n".join(body).strip()


def _draft_email_text(latest_row: dict, fallback_title="Policy Renewal Options"):
    vehicle = latest_row.get("vehicle","")
    old_total = latest_row.get("old_total_premium",0.0); new_total = latest_row.get("new_total_premium",0.0)
    delta = float(old_total) - float(new_total)
    reasons = str(latest_row.get("top_3_reasons",""))
    reassurance = ""
    if any(r in reasons for r in SPECIAL_REASONS):
        reassurance = "\nWe’ve reviewed your policy specifics and ensured coverage remains appropriate for your situation."
    body = (
        f"Subject: {fallback_title}\n\n"
        f"Hi,\n\n"
        f"Thanks for being with us. We reviewed your {vehicle} policy and identified a way to optimize your premium.\n"
        f"- Estimated premium change: ₹{int(round(delta)):,} savings compared to your previous total.\n"
        f"- Revised components reflect adjusted OD/TP and discounts while keeping coverage intact.\n"
        f"{reassurance}\n\n"
        f"Would you like me to proceed with these updates and share the revised policy schedule?\n"
        f"Regards,\nTeam"
    )
    return body

class ChurnDraftEmail(APIView):
    def post(self, request):
        try:
            policy_no = str(request.data.get("policy_no") or request.data.get("policy no") or "").strip()
            if not policy_no:
                return Response({"error": "policy_no required"}, status=400)

            # if not (GROQ_API_KEY_Churn and ChatGroq):
            #     return Response({"error": "Groq is required for drafting; configure GROQ_API_KEY_Churn."}, status=503)

            # fetch latest Selected changes
            policy_norm = _norm_policy(policy_no)
            sql = f"""SELECT * FROM {FULL_SAVE_FQN}
                      WHERE policy_no_norm=%s ORDER BY created_at DESC LIMIT 1;"""
            conn = _get_conn(); cur = conn.cursor(cursor_factory=RealDictCursor)
            try:
                cur.execute(sql, (policy_norm,))
                latest = cur.fetchone()
            finally:
                try: cur.close(); conn.close()
                except Exception: pass

            if not latest:
                return Response({"error": "No saved Selected changes. Save first."}, status=400)

            # llm = ChatGroq(model=GROQ_MODEL_Churn, api_key=GROQ_API_KEY_Churn, temperature=0.2)

            ctx = "; ".join(
                f"{k}={latest.get(k)}" for k in [
                    "vehicle","old_discount","new_discount","old_total_premium","new_total_premium",
                    "old_od","new_od","old_tp","new_tp","old_add_on_premium","new_add_on_premium",
                    "old_idv","new_idv","top_3_reasons"
                ] if k in latest
            )

            prompt = f"""
You are drafting a car-insurance renewal email.

OUTPUT FORMAT (strict):
- First line: "Subject: <your subject here>"
- Then a blank line, then the body.

Guidance (not mandatory style, just hints):
- Make the subject short, compelling, and specific to the vehicle (if present).
- In the body, mention the vehicle/policy context.
- Use numbers from ctx to explain premium change and component deltas (OD/TP, discount, etc.).
- If any of these appear in top_3_reasons: Young Vehicle Age, Old Vehicle Age, Claims Happened,
  Multiple Claims on Record, Minimal Policies Purchased, Tie Up with Non-OEM — include a brief reassurance line.
- Friendly, persuasive, action-oriented; end with a clear renewal CTA.

ctx: {ctx}
""".strip()

            txt = azure_chat(prompt, system="You are a helpful assistant for drafting renewal emails.", temperature=0.2)
            subject, body = _split_subj_body(txt)

            # Groq-only: if Groq didn't produce a valid subject or body, return error (no deterministic fallback)
            if not subject or not body:
                return Response({"error": "Groq returned an incomplete draft."}, status=502)

            return Response({"subject": subject, "body": body})
        except Exception as e:
            logger.exception("draft-email failed: %s", e)
            return Response({"error": str(e)}, status=500)


import re
from email.utils import getaddresses, formataddr

_EMAIL_RE = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")

def _dedup_preserve_order(items):
    seen = set(); out = []
    for x in items:
        if x not in seen:
            seen.add(x); out.append(x)
    return out

def _normalize_recipients(value) -> list[str]:
    """Accept str or list[str]; return clean RFC-friendly list (no CR/LF)."""
    if not value:
        return []
    parts = value if isinstance(value, (list, tuple, set)) else [value]
    # Parse names and comma-separated lists
    pairs = getaddresses([", ".join(str(p) for p in parts if p is not None)])
    cleaned = []
    for name, addr in pairs:
        s = (addr or "").replace("\r", " ").replace("\n", " ").strip().strip("<>").strip()
        if _EMAIL_RE.fullmatch(s):
            name = (name or "").strip()
            cleaned.append(formataddr((name, s)) if name else s)
    # Fallback: regex-scan if nothing parsed
    if not cleaned:
        for p in parts:
            s = str(p or "").replace("\r", " ").replace("\n", " ")
            for m in _EMAIL_RE.finditer(s):
                cleaned.append(m.group(0))
    return _dedup_preserve_order([x.strip().strip(",") for x in cleaned if x.strip().strip(",")])

def _safe_header(text: str) -> str:
    """Strip CR/LF to avoid invalid headers / injection."""
    return str(text or "").replace("\r", "").replace("\n", "").strip()

# --- Gmail send: mirrors Streamlit flow (OAuth + Gmail API) -------------------
import base64 as _b64
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.utils import formataddr
from email import encoders as _enc
from googleapiclient.discovery import build as _gbuild
from googleapiclient.errors import HttpError as _GHttpError
from google.auth.transport.requests import Request as _GRequest
from google.oauth2.credentials import Credentials as _GCreds
import mimetypes

GMAIL_SCOPES = ["https://www.googleapis.com/auth/gmail.send"]



# Defaults (can be overridden with env vars)
GMAIL_CREDENTIALS_FILE = os.getenv("GMAIL_CREDENTIALS_FILE", str(BASE_DIR_Churn / "gmail" / "credentials.json"))
GMAIL_TOKEN_FILE       = os.getenv("GMAIL_TOKEN_FILE",       str(BASE_DIR_Churn / "gmail" / "token.json"))
GMAIL_DEFAULT_FROM_EMAIL = os.getenv("GMAIL_DEFAULT_FROM_EMAIL", "")
GMAIL_DEFAULT_FROM_NAME  = os.getenv("GMAIL_DEFAULT_FROM_NAME", "Retention Team")

def _load_gmail_service():
    """
    Load/refresh Gmail OAuth credentials and return a Gmail service client.
    Matches Streamlit pattern: use a saved token.json; refresh if expired.
    """
    if not os.path.exists(GMAIL_CREDENTIALS_FILE):
        raise FileNotFoundError(
            f"GMAIL_CREDENTIALS_FILE missing at {GMAIL_CREDENTIALS_FILE}. "
            f"Download your OAuth client (installed app) JSON and set env var."
        )
    if not os.path.exists(GMAIL_TOKEN_FILE):
        raise FileNotFoundError(
            f"GMAIL_TOKEN_FILE missing at {GMAIL_TOKEN_FILE}. "
            f"Run your Streamlit/OAuth consent once to create a token.json with a refresh_token."
        )

    creds = _GCreds.from_authorized_user_file(GMAIL_TOKEN_FILE, GMAIL_SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(_GRequest())
            # persist refresh result (important on servers)
            with open(GMAIL_TOKEN_FILE, "w") as f:
                f.write(creds.to_json())
        else:
            raise RuntimeError(
                "Gmail credentials invalid and no refresh_token available. "
                "Re-run OAuth consent (like your Streamlit flow) to capture a refresh token."
            )
    return _gbuild("gmail", "v1", credentials=creds, cache_discovery=False)

def _make_mime_message(
    to_email: str | list[str],
    subject: str,
    body: str,
    *,
    content_type: str = "plain",  # or "html"
    cc: list[str] | str | None = None,
    bcc: list[str] | str | None = None,
    reply_to: str | None = None,
    from_email: str | None = None,
    from_name: str | None = None,
    attachments: list[dict] | None = None,
) -> str:
    """
    Build RFC 2822 MIME message.
    attachments: list of dicts with either
      - {"filename": "x.pdf", "b64": "<base64 string>"}
      - {"filename": "x.csv", "content": "<bytes|str>", "mime": "text/csv"}
    """
    from_email = (from_email or GMAIL_DEFAULT_FROM_EMAIL or "").strip()
    from_name  = (from_name  or GMAIL_DEFAULT_FROM_NAME or "").strip()

    # --- Body container (alternative + optional mixed for attachments) ---
    if attachments:
        outer = MIMEMultipart("mixed")
        alt = MIMEMultipart("alternative")
        outer.attach(alt)
    else:
        alt = MIMEMultipart("alternative")
        outer = alt

    if str(content_type).lower() == "html":
        alt.attach(MIMEText(body or "", "html", "utf-8"))
    else:
        alt.attach(MIMEText(body or "", "plain", "utf-8"))

    # --- Attachments ---
    for att in (attachments or []):
        fname = att.get("filename") or "attachment"
        mime = att.get("mime") or (mimetypes.guess_type(fname)[0] or "application/octet-stream")
        maintype, subtype = mime.split("/", 1)
        part = MIMEBase(maintype, subtype)

        if "b64" in att and att["b64"]:
            part.set_payload(_b64.b64decode(att["b64"]))
        elif "content" in att and att["content"] is not None:
            content = att["content"]
            if isinstance(content, str):
                content = content.encode("utf-8")
            part.set_payload(content)
        else:
            continue

        _enc.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="{fname}"')
        outer.attach(part)

    # --- Headers (normalized & injection-safe) ---
    tos  = _normalize_recipients(to_email)
    ccs  = _normalize_recipients(cc)
    bccs = _normalize_recipients(bcc)

    if not tos:
        raise ValueError("No valid recipient emails in 'to'")

    outer["To"] = ", ".join(tos)
    if ccs:
        outer["Cc"] = ", ".join(ccs)
    if bccs:
        # Include Bcc so Gmail will deliver to them; recipients won't see it.
        outer["Bcc"] = ", ".join(bccs)

    outer["Subject"] = _safe_header(subject)
    if from_email:
        outer["From"] = formataddr((from_name, from_email)) if from_name else from_email
    if reply_to:
        outer["Reply-To"] = _safe_header(reply_to)

    # Gmail expects base64url-encoded raw
    raw = _b64.urlsafe_b64encode(outer.as_bytes()).decode("utf-8")
    return raw


class ChurnSendEmail(APIView):
    def post(self, request):
        try:
            payload   = request.data or {}
            to_email  = str(payload.get("to", "")).strip()
            subject   = payload.get("subject")
            body      = payload.get("body")
            policy_no = (payload.get("policy_no") or payload.get("policy no") or "").strip()

            if not to_email:
                return Response({"error": "Recipient 'to' is required"}, status=422)

            # Draft via Azure if subject/body missing
            if not subject or not body:
                if not policy_no:
                    return Response({"error": "Missing subject/body. Provide policy_no to auto-draft."}, status=422)

                policy_norm = _norm_policy(policy_no)
                sql = f"""SELECT * FROM {FULL_SAVE_FQN}
                          WHERE policy_no_norm=%s
                          ORDER BY created_at DESC
                          LIMIT 1;"""
                conn = _get_conn(); cur = conn.cursor(cursor_factory=RealDictCursor)
                try:
                    cur.execute(sql, (policy_norm,))
                    latest = cur.fetchone()
                finally:
                    try:
                        cur.close(); conn.close()
                    except Exception:
                        pass

                if not latest:
                    return Response({"error": "No saved Selected changes for the given policy_no. Save first."}, status=400)

                ctx = "; ".join(
                    f"{k}={latest.get(k)}" for k in [
                        "vehicle","old_discount","new_discount","old_total_premium","new_total_premium",
                        "old_od","new_od","old_tp","new_tp","old_add_on_premium","new_add_on_premium",
                        "old_idv","new_idv","top_3_reasons"
                    ] if k in latest
                )

                prompt = f"""
{RULES_PROMPT}

OUTPUT FORMAT (strict):
- First line: "Subject: <your subject here>"
- Then a blank line, then the body.

Guidance:
- Make the subject concise, benefit-oriented, and vehicle-aware if possible.
- In the body, mention the vehicle/policy context and quantify the premium change (OD/TP/discount).
- If any of these appear in top_3_reasons: Young Vehicle Age, Old Vehicle Age, Claims Happened,
  Multiple Claims on Record, Minimal Policies Purchased, Tie Up with Non-OEM — add one short reassurance line.
- Friendly, persuasive, action-oriented; clear renewal CTA.

ctx: {ctx}
""".strip()

                txt = azure_chat(
                    prompt,
                    system="You are a helpful assistant for drafting renewal emails.",
                    temperature=0.2
                )
                s, b = _split_subj_body(txt)

                if not s or not b:
                    # fallback to deterministic builder
                    fb = _draft_email_text(latest)
                    s, b = _split_subj_body(fb)

                subject, body = s, b

            # ... proceed with Gmail send exactly as you already do ...
            service = _load_gmail_service()
            tos  = _normalize_recipients(payload.get("to"))
            ccs  = _normalize_recipients(payload.get("cc") or None)
            bccs = _normalize_recipients(payload.get("bcc") or None)
            if not tos:
                return Response({"error": "No valid 'to' address after normalization"}, status=422)

            raw = _make_mime_message(
                to_email=tos,
                subject=subject,
                body=body,
                content_type=str(payload.get("content_type", "plain")).lower(),
                cc=ccs, bcc=bccs,
                reply_to=payload.get("reply_to") or None,
                from_email=payload.get("from_email") or None,
                from_name=payload.get("from_name") or None,
                attachments=payload.get("attachments") or None,
            )
            resp = service.users().messages().send(userId="me", body={"raw": raw}).execute()
            return Response({"ok": True, "gmail_id": resp.get("id")})

        except _GHttpError as he:
            content = getattr(he, "error_details", None) or getattr(he, "content", None) or str(he)
            logger.error("Gmail API error: %s", content)
            return Response({"error": "Gmail API error", "details": str(content)}, status=502)
        except FileNotFoundError as fe:
            return Response({"error": str(fe)}, status=503)
        except Exception as e:
            logger.exception("send-email failed: %s", e)
            return Response({"error": str(e)}, status=500)







import json, time, datetime, traceback
import chromadb
from django.http import JsonResponse, StreamingHttpResponse
from django.views.decorators.csrf import csrf_exempt
from sqlalchemy import text
from .llm_runner.llm_config import get_llama_maverick_llm

def _clean_and_extract_json(raw_text: str) -> dict:
    """Robustly extract JSON from LLM response"""
    import re, json
    
    # Clean the text
    text = (raw_text or "").strip()
    text = re.sub(r'[\U00010000-\U0010ffff]', '', text)  # Remove emojis
    
    # Try to find JSON block
    json_patterns = [
        r'\{.*\}',  # Basic JSON
        r'```json\s*(\{.*\})\s*```',  # Markdown code block
        r'```\s*(\{.*\})\s*```',  # Generic code block
    ]
    
    for pattern in json_patterns:
        match = re.search(pattern, text, re.DOTALL)
        if match:
            json_text = match.group(1) if match.groups() else match.group(0)
            try:
                # Clean common JSON issues
                json_text = re.sub(r',\s*}', '}', json_text)
                json_text = re.sub(r',\s*]', ']', json_text)
                return json.loads(json_text)
            except:
                continue
    
    # If no valid JSON found, return empty structure
    return {}

from django.http import JsonResponse
import chromadb, json

def list_corpus_entries(request):
    chroma_client = chromadb.PersistentClient(path="corpus_db")
    collection = chroma_client.get_or_create_collection("corpus_liberty")
    items = collection.get()
    entries = []
    for doc_id, doc, meta in zip(items["ids"], items["documents"], items["metadatas"]):
        try:
            parsed = json.loads(doc)
        except:
            parsed = {"summary": doc}
        entries.append({"id": doc_id, "metadata": meta, "content": parsed})
    return JsonResponse(entries, safe=False)

from django.http import JsonResponse
import chromadb
import json
import re
import calendar
from django.http import JsonResponse
import chromadb
import json
import re
import calendar

# -----------------------------
# Entities and Keywords
# -----------------------------
BOOST_KEYWORDS = [
    "date", "time", "year", "month",
    "zone", "state", "branch", "city",
    "country", "district", "region"
]

# Month mapping - FIXED: Complete month mapping
MONTHS = {
    "jan": "January", "january": "January",
    "feb": "February", "february": "February", 
    "mar": "March", "march": "March",
    "apr": "April", "april": "April",
    "may": "May",
    "jun": "June", "june": "June",
    "jul": "July", "july": "July",
    "aug": "August", "august": "August",
    "sep": "September", "september": "September",
    "oct": "October", "october": "October", 
    "nov": "November", "november": "November",
    "dec": "December", "december": "December"
}

# Reverse mapping for proper replacement
MONTH_REVERSE = {v.lower(): v for v in MONTHS.values()}
for short, full in MONTHS.items():
    MONTH_REVERSE[short] = full

# -----------------------------
# Utils
# -----------------------------
def clean_question_text(q: str) -> str:
    """Filter out greetings, placeholders, empty strings."""
    if not q:
        return ""

    q = q.strip()
    q_lower = q.lower()

    greetings = {"hi", "hello", "hey", "good morning", "good evening"}
    if q_lower in greetings:
        return ""

    # Skip if contains placeholders like <MONTH>, <STATE>
    if re.search(r'<[A-Z_]+>', q):
        return ""

    return q


# -----------------------------
# Helpers
# -----------------------------
import re, json, chromadb
from django.http import JsonResponse


# ✅ ADD THIS FUNCTION 
# def remove_unwanted_business_filters(sql: str, question: str) -> str:
#     """Remove main_churn_reason filter if not explicitly requested"""
#     print("🚨 remove_unwanted_business_filters CALLED")
#     try:
#         q_lower = question.lower()
        
#         # Check if question explicitly asks for churn-related data
#         churn_keywords = ['churn', 'not renew', 'cancel', 'left', 'attrition', 'lost']
#         asks_for_churn = any(kw in q_lower for kw in churn_keywords)
        
#         if asks_for_churn:
#             print("✅ Question asks for churn data - keeping filters")
#             return sql
        
#         # Simple string replacement instead of regex
#         original_sql = sql
        
#         # Pattern 1: AND main_churn_reason IS NOT NULL
#         sql = sql.replace("AND main_churn_reason IS NOT NULL", "")
#         sql = sql.replace("and main_churn_reason is not null", "")
        
#         # Pattern 2: WHERE main_churn_reason IS NOT NULL AND
#         if "WHERE main_churn_reason IS NOT NULL AND" in sql:
#             sql = sql.replace("WHERE main_churn_reason IS NOT NULL AND", "WHERE")
        
#         if sql != original_sql:
#             print("🧹 Removed main_churn_reason filter")
#             print(f"📝 Original SQL had: main_churn_reason IS NOT NULL")
        
#         return sql
        
#     except Exception as e:
#         print(f"❌ Error in filter removal: {e}")
#         return sql  # Return original if error
def remove_unwanted_business_filters(sql: str, question: str) -> str:
    """Remove main_churn_reason filter only when NOT asking about reasons"""
    print("🚨 remove_unwanted_business_filters CALLED")
    try:
        q_lower = question.lower()
        
        # ✅ NEW: Check if question is about churn REASONS
        asking_about_reasons = any(keyword in q_lower for keyword in [
            'reason', 'why', 'cause', 'factor', 'main_churn_reason', 'explanation'
        ])
        
        # ✅ NEW: Check if SQL selects main_churn_reason column
        sql_upper = sql.upper()
        if 'FROM' in sql_upper:
            select_part = sql_upper.split('FROM')[0]
            selects_churn_reason = 'MAIN_CHURN_REASON' in select_part
        else:
            selects_churn_reason = False
        
        # If asking about reasons OR selecting that column, KEEP the filter
        if asking_about_reasons or selects_churn_reason:
            print("✅ Question about reasons OR SQL selects main_churn_reason - keeping filter")
            return sql
        
        # Check if question explicitly asks for churn-related data
        churn_keywords = ['churn', 'not renew', 'cancel', 'left', 'attrition', 'lost']
        asks_for_churn = any(kw in q_lower for kw in churn_keywords)
        
        if asks_for_churn:
            print("✅ Question asks for churn data - keeping filters")
            return sql
        
        # Simple listing query - safe to remove filter
        original_sql = sql
        
        # Pattern 1: AND main_churn_reason IS NOT NULL
        sql = sql.replace("AND main_churn_reason IS NOT NULL", "")
        sql = sql.replace("and main_churn_reason is not null", "")
        
        # Pattern 2: WHERE main_churn_reason IS NOT NULL AND
        if "WHERE main_churn_reason IS NOT NULL AND" in sql:
            sql = sql.replace("WHERE main_churn_reason IS NOT NULL AND", "WHERE")
        
        if sql != original_sql:
            print("🧹 Removed main_churn_reason filter (simple list query)")
        
        return sql
        
    except Exception as e:
        print(f"❌ Error in filter removal: {e}")
        return sql  # Return original if error


def validate_customer_query(sql: str, question: str) -> str:
    """Fix customer queries"""
    
    # Check if question is about customers
    if any(word in question.lower() for word in ['customer', 'customers', 'client', 'clients']):
        
        # ✅ FIX 1: Add customerid if missing
        if 'GROUP BY' in sql.upper() and 'customerid' not in sql.lower():
            if 'insured_client_name' in sql.lower():
                print("⚠️ Adding missing customerid")
                sql = sql.replace(
                    'SELECT \n  insured_client_name',
                    'SELECT \n  customerid,\n  insured_client_name'
                )
                sql = sql.replace(
                    'GROUP BY insured_client_name',
                    'GROUP BY customerid, insured_client_name'
                )
        
        # ✅ FIX 2: Remove ROW_NUMBER filter (keeps only top 1)
        if 'WHERE churn_rank = 1' in sql or 'WHERE.*rank.*=.*1' in sql:
            print("⚠️ Removing rank=1 filter to show all top customers")
            
            # Remove the WHERE churn_rank = 1 line
            sql = re.sub(r'\s*WHERE\s+\w*rank\w*\s*=\s*1', '', sql, flags=re.IGNORECASE)
            
            # If it's a CTE query, simplify it
            if 'WITH' in sql.upper() and 'ROW_NUMBER' in sql.upper():
                print("⚠️ Simplifying CTE query - removing ranking logic")
                
                # Extract the base query from first CTE
                match = re.search(
                    r'SELECT\s+(.*?)\s+FROM\s+"bi_dwh"\."main_cai_lib".*?GROUP BY\s+(.*?)(?=\)|$)',
                    sql,
                    re.DOTALL | re.IGNORECASE
                )
                
                if match:
                    select_cols = match.group(1).strip()
                    group_cols = match.group(2).strip()
                    
                    # Rebuild simple query
                    sql = f"""SELECT 
  {select_cols}
FROM "bi_dwh"."main_cai_lib"
WHERE is_churn IS NOT NULL AND is_churn ILIKE 'Not Renewed'
  AND customerid IS NOT NULL
  AND policy_no IS NOT NULL
GROUP BY {group_cols}
ORDER BY policy_end_date_year, policy_end_date_month, churn_count DESC"""
    
    return sql

# def fix_percentage_formatting(sql: str) -> str:
#     """
#     Auto-fix percentage calculations to match ROUND_PERCENTAGES config
#     Catches queries that LLM generates incorrectly
#     """
#     from .config import ROUND_PERCENTAGES
#     import re
    
#     # Pattern: percentage calculation without ROUND
#     # Matches: COUNT(...) / NULLIF(...) * 100 AS some_percentage
#     pattern = r'''
#         (COUNT\(CASE[^)]+\)::numeric\s*/\s*NULLIF\([^)]+\)\s*\*\s*100)
#         \s+AS\s+(\w*(?:percentage|rate)\w*)
#     '''
    
#     def replace_fn(match):
#         calculation = match.group(1)
#         alias = match.group(2)
        
#         if ROUND_PERCENTAGES:
#             # Add ROUND(..., 0) ::INTEGER
#             return f"ROUND(\n    {calculation},\n    0\n  ) ::INTEGER AS {alias}"
#         else:
#             # Add ROUND(..., 2)
#             return f"ROUND(\n    {calculation},\n    2\n  ) AS {alias}"
    
#     # Fix unrounded percentages
#     sql_fixed = re.sub(pattern, replace_fn, sql, flags=re.VERBOSE | re.IGNORECASE)
    
#     if sql_fixed != sql:
#         print("🔧 Auto-fixed percentage formatting in SQL")
    
#     return sql_fixed

def fix_percentage_formatting(sql: str) -> str:
    """Force ROUND in percentage calculations - simple string approach"""
    from .config import ROUND_PERCENTAGES
    
    try:
        # Check if percentage calculation exists without ROUND
        if ('* 100 AS' in sql or '* 100\n' in sql) and 'ROUND(' not in sql:
            print("🔧 Fixing unrounded percentage calculation")
            
            # Find the percentage calculation pattern
            import re
            
            # Pattern matches the entire calculation on multiple lines
            pattern = r'(COUNT\(CASE\s+WHEN[^)]+\)::numeric\s*/\s*NULLIF\([^)]+\)\s*\*\s*100)\s+(AS\s+\w+_(?:percentage|rate))'
            
            if ROUND_PERCENTAGES:
                replacement = r'ROUND(\1, 0) ::INTEGER \2'
            else:
                replacement = r'ROUND(\1, 2) \2'
            
            sql_fixed = re.sub(pattern, replacement, sql, flags=re.DOTALL | re.IGNORECASE)
            
            if sql_fixed != sql:
                print("✅ Successfully added ROUND to SQL")
                return sql_fixed
            else:
                print("⚠️ Pattern did not match - trying fallback")
                
                # Fallback: manual string manipulation
                lines = sql.split('\n')
                for i, line in enumerate(lines):
                    if '* 100 AS' in line and 'churn_rate_percentage' in line:
                        # Found the line - need to wrap previous lines too
                        # Find where calculation starts
                        calc_start = i
                        for j in range(i, -1, -1):
                            if 'COUNT(CASE' in lines[j]:
                                calc_start = j
                                break
                        
                        # Extract calculation
                        calc_lines = lines[calc_start:i+1]
                        calc_text = '\n'.join(calc_lines)
                        
                        # Split at AS
                        parts = calc_text.split(' AS ')
                        if len(parts) == 2:
                            calc_part = parts[0].strip()
                            alias_part = parts[1].strip()
                            
                            if ROUND_PERCENTAGES:
                                fixed_calc = f"  ROUND({calc_part}, 0) ::INTEGER AS {alias_part}"
                            else:
                                fixed_calc = f"  ROUND({calc_part}, 2) AS {alias_part}"
                            
                            # Replace in original
                            lines[calc_start:i+1] = [fixed_calc]
                            print("✅ Fallback method successful")
                            return '\n'.join(lines)
        
        return sql
        
    except Exception as e:
        print(f"❌ Error in fix_percentage_formatting: {e}")
        import traceback
        traceback.print_exc()
        return sql

def month_in_text(text, month):
    return re.search(rf"\b{month}\b", text, re.I)

def clean_question_text(q: str) -> str:
    if not q:
        return ""
    q = q.strip().lower()

    # Skip greetings
    if q in ["hi", "hello", "hey", "good morning", "good evening"]:
        return ""

    # Skip placeholder-only questions - FIXED: Better placeholder detection
    if re.search(r'<[A-Z_]+>', q):
        return ""

    return q


def resolve_query(prev, new):
    """
    Replace entities in prev query with new ones if user says 'in Delhi' / 'in 2026'.
    """
    if not prev:
        return new

    # FIXED: Replace month correctly
    for month_key, month_full in MONTHS.items():
        if re.search(rf"\b{month_key}\b", new, re.I):
            # Replace any existing month in prev with the new one
            prev_replaced = prev
            for existing_key, existing_full in MONTHS.items():
                prev_replaced = re.sub(rf"\b{existing_key}\b", month_key, prev_replaced, flags=re.I)
                prev_replaced = re.sub(rf"\b{existing_full}\b", month_full, prev_replaced, flags=re.I)
            return prev_replaced

    # Replace year
    years = re.findall(r"\b(20\d{2}|19\d{2})\b", new)
    if years:
        return re.sub(r"\b(20\d{2}|19\d{2})\b", years[0], prev)

    # Replace generic entities (state, city, branch, etc.)
    tokens = [t for t in new.split() if t.lower() not in ["in", "at", "on", "the"]]
    if tokens:
        return re.sub(r"\b\w+\b", tokens[-1], prev, count=1)

    return new or prev

def rewrite_question(cleaned_q, query):
    """
    Smart rewrite: force-align months/years/entities in suggestion with the user query.
    """
    if not cleaned_q:
        return ""

    # FIXED: Rewrite month correctly
    for month_key, month_full in MONTHS.items():
        if re.search(rf"\b{month_key}\b", query, re.I):
            # Replace any existing month with the correct one from query
            for existing_key, existing_full in MONTHS.items():
                cleaned_q = re.sub(rf"\b{existing_key}\b", month_key, cleaned_q, flags=re.I)
                cleaned_q = re.sub(rf"\b{existing_full}\b", month_full, cleaned_q, flags=re.I)
            break

    # Rewrite year
    years = re.findall(r"\b(20\d{2}|19\d{2})\b", query)
    if years:
        cleaned_q = re.sub(r"\b(20\d{2}|19\d{2})\b", years[0], cleaned_q)

    # Rewrite other entities (state, branch, city, etc.)
    tokens = [t for t in query.split() if t.lower() not in ["in", "at", "on", "the"]]
    if tokens:
        last_entity = tokens[-1]
        # Replace the first non-month, non-year word
        words = cleaned_q.split()
        for i, word in enumerate(words):
            if not any(m in word.lower() for m in MONTHS.keys()) and not re.match(r"\b(20\d{2}|19\d{2})\b", word):
                words[i] = last_entity
                break
        cleaned_q = " ".join(words)

    return cleaned_q

def extract_entities11(text):
    """
    Extract month, year, state, branch, etc. from user query.
    """
    entities = {}
    text_lower = text.lower()

    # FIXED: Month extraction
    for month_key, month_full in MONTHS.items():
        if re.search(rf"\b{month_key}\b", text_lower):
            entities["<MONTH>"] = month_full
            break

    # year
    m = re.search(r"\b(20\d{2}|19\d{2})\b", text)
    if m:
        entities["<YEAR>"] = m.group(1)

    # simple state/branch detection
    words = text.split()
    if len(words) >= 2:
        for i, word in enumerate(words):
            if word.lower() == "in" and i + 1 < len(words):
                entity_value = words[i + 1].capitalize()
                entities["<STATE>"] = entity_value
                entities["<BRANCH>"] = entity_value
                entities["<CITY>"] = entity_value
                entities["<REGION>"] = entity_value
                entities["<COUNTRY>"] = entity_value
                entities["<DISTRICT>"] = entity_value
                entities["<ZONE>"] = entity_value
                break

    return entities

def replace_placeholders(question, entities):
    """
    Replace placeholders (<MONTH>, <YEAR>, <STATE>, <ZONE>, <BRANCH>, etc.)
    with the detected entity values dynamically.
    """
    if not question:
        return ""

    out = question
    for placeholder, value in entities.items():
        val = value

        # ✅ Normalize months (jan → January, january → January)
        if placeholder == "<MONTH>":
            month_key = value.lower()[:3]
            for short, full in MONTHS.items():
                if month_key.startswith(short):
                    val = full
                    break

        # ✅ Normalize year (e.g. '24 → 2024)
        elif placeholder == "<YEAR>":
            if re.match(r"^\d{2}$", value):  # '24 style
                val = "20" + value

        # ✅ Normalize state/zone/branch/city/etc. (capitalize properly)
        elif placeholder in ["<STATE>", "<ZONE>", "<BRANCH>", "<CITY>", "<REGION>", "<COUNTRY>", "<DISTRICT>"]:
            val = value.strip().title()

        # ✅ Normalize date (keep as is but can be extended for formats)
        elif placeholder == "<DATE>":
            # Try to enforce DD-MM-YYYY format if possible
            try:
                import datetime
                parsed = datetime.datetime.strptime(value, "%d-%m-%Y")
                val = parsed.strftime("%d-%b-%Y")
            except Exception:
                val = value  # leave unchanged if not parseable

        # Replace placeholder with normalized value
        out = re.sub(re.escape(placeholder), val, out, flags=re.I)

    return out


def rewrite_with_entities(question, query):
    """
    Replace placeholders (<MONTH>, <YEAR>, <STATE>, etc.) with values from user query.
    """
    if not question:
        return ""

    out = question
    query_lower = query.lower()

    # FIXED: Replace month correctly
    for month_key, month_full in MONTHS.items():
        if re.search(rf"\b{month_key}\b", query_lower):
            out = re.sub(r"<MONTH>", month_full, out, flags=re.I)
            break

    # Replace year
    m = re.search(r"\b(20\d{2}|19\d{2})\b", query)
    if m:
        out = re.sub(r"<YEAR>", m.group(1), out, flags=re.I)

    # Replace state/branch/city/country/etc.
    words = query.split()
    for i, word in enumerate(words):
        if word.lower() == "in" and i + 1 < len(words):
            ent = words[i + 1].capitalize()
            for ph in ["<STATE>", "<BRANCH>", "<CITY>", "<REGION>", "<COUNTRY>", "<DISTRICT>", "<ZONE>"]:
                out = re.sub(re.escape(ph), ent, out, flags=re.I)
            break

    return out

def expand_query_with_entities(query: str, entities: dict) -> list:
    """
    Expand query with variants for months, years, states, zones, branches, etc.
    to improve recall from corpus.
    """
    query_lower = query.lower()
    queries = [query]

    # --- Months expansion ---
    for short, full in MONTHS.items():
        if re.search(rf"\b{short}\b", query_lower):
            queries.append(query_lower.replace(short, full.lower()))
            queries.append(query_lower.replace(short, full))
        elif full.lower() in query_lower:
            queries.append(query_lower.replace(full.lower(), short))
            queries.append(query_lower.replace(full.lower(), full))

    # --- Year expansion ---
    year_match = re.findall(r"\b(19\d{2}|20\d{2})\b", query_lower)
    if year_match:
        for y in year_match:
            # allow '2025' <-> '25'
            if len(y) == 4:
                queries.append(y[2:])  # short '25'
            elif len(y) == 2:
                queries.append("20" + y)  # expand '25' -> '2025'

    # --- State / Branch / Zone expansion ---
    for ph in ["<STATE>", "<BRANCH>", "<ZONE>", "<CITY>", "<REGION>", "<COUNTRY>", "<DISTRICT>"]:
        if ph in entities:
            val = entities[ph].lower()
            queries.append(val)
            # add variants
            queries.append(f"in {val}")
            queries.append(f"{val} branch")
            queries.append(f"{val} zone")
            queries.append(f"{val} state")

    # --- Date expansion (basic) ---
    date_match = re.findall(r"\b(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})\b", query_lower)
    for d in date_match:
        queries.append(d.replace("-", "/"))
        queries.append(d.replace("/", "-"))

    return list(set(q.strip() for q in queries if q.strip()))

import json, re
import chromadb
from django.http import JsonResponse

# Example month dictionary
MONTHS = {
    "jan": "January", "feb": "February", "mar": "March", "apr": "April",
    "may": "May", "jun": "June", "jul": "July", "aug": "August",
    "sep": "September", "oct": "October", "nov": "November", "dec": "December"
}

BOOST_KEYWORDS = [
    "date", "time", "year", "month",
    "zone", "state", "branch", "city", "country", "district", "region"
]

# -----------------------------------
# Dynamic Entity Extractor
# -----------------------------------
"""
Helper functions for entity extraction, validation, and corpus matching.
These ensure 100% accurate entity matching for production.
"""

import re
import json
from typing import Dict, List, Optional

# Complete month mapping
MONTHS = {
    "jan": "January", "january": "January",
    "feb": "February", "february": "February", 
    "mar": "March", "march": "March",
    "apr": "April", "april": "April",
    "may": "May",
    "jun": "June", "june": "June",
    "jul": "July", "july": "July",
    "aug": "August", "august": "August",
    "sep": "September", "september": "September",
    "oct": "October", "october": "October", 
    "nov": "November", "november": "November",
    "dec": "December", "december": "December"
}

# Month name to number mapping
MONTH_TO_NUMBER = {
    "January": 1, "February": 2, "March": 3, "April": 4,
    "May": 5, "June": 6, "July": 7, "August": 8,
    "September": 9, "October": 10, "November": 11, "December": 12
}


def extract_entities(text: str) -> Dict[str, str]:
    """
    Extract ALL entities from text with high accuracy.
    Returns dict with placeholder keys: <MONTH>, <YEAR>, <STATE>, <ZONE>, <BRANCH>, etc.
    """
    entities = {}
    text_lower = text.lower()
    
    # ============ MONTH EXTRACTION ============
    for month_key, month_full in MONTHS.items():
        if re.search(rf"\b{month_key}\b", text_lower):
            entities["<MONTH>"] = month_full
            break
    
    # ============ YEAR EXTRACTION ============
    year_match = re.search(r"\b(20\d{2}|19\d{2})\b", text)
    if year_match:
        entities["<YEAR>"] = year_match.group(1)
    
    # ============ DATE EXTRACTION ============
    date_match = re.search(r"\b(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})\b", text)
    if date_match:
        entities["<DATE>"] = date_match.group(1)
    
    # ============ GEOGRAPHIC ENTITIES (from cached DB) ============
    try:
        from .entity_cache import get_entities
        cached_entities = get_entities()
        
        # State detection
        for state in cached_entities.get("state", []):
            if re.search(rf"\b{re.escape(state)}\b", text, re.IGNORECASE):
                entities["<STATE>"] = state
                break
        
        # Zone detection
        for zone in cached_entities.get("zone", []):
            if re.search(rf"\b{re.escape(zone)}\b", text, re.IGNORECASE):
                entities["<ZONE>"] = zone
                break
        
        # Branch detection
        for branch in cached_entities.get("branch", []):
            if re.search(rf"\b{re.escape(branch)}\b", text, re.IGNORECASE):
                entities["<BRANCH>"] = branch
                break
        
        # City detection
        for city in cached_entities.get("city", []):
            if re.search(rf"\b{re.escape(city)}\b", text, re.IGNORECASE):
                entities["<CITY>"] = city
                break
        
        # Region detection
        for region in cached_entities.get("region", []):
            if re.search(rf"\b{re.escape(region)}\b", text, re.IGNORECASE):
                entities["<REGION>"] = region
                break
        
        # District detection
        for district in cached_entities.get("district", []):
            if re.search(rf"\b{re.escape(district)}\b", text, re.IGNORECASE):
                entities["<DISTRICT>"] = district
                break
        
        # Country detection
        for country in cached_entities.get("country", []):
            if re.search(rf"\b{re.escape(country)}\b", text, re.IGNORECASE):
                entities["<COUNTRY>"] = country
                break
    except Exception as e:
        print(f"⚠️ Warning: Could not load cached entities: {e}")
    
    return entities

def entities_exact_match(user_entities: Dict[str, str], corpus_entities: Dict[str, str]) -> tuple:
    """
    Check if ALL entities match exactly between user question and corpus entry.
    Returns (match_status, list_of_mismatches)
    """
    mismatches = []
    
    # Define all possible entity types
    entity_types = ["<MONTH>", "<YEAR>", "<DATE>", "<STATE>", "<ZONE>", "<BRANCH>", 
                    "<CITY>", "<REGION>", "<DISTRICT>", "<COUNTRY>"]
    
    for entity_type in entity_types:
        user_val = user_entities.get(entity_type)
        corpus_val = corpus_entities.get(entity_type)
        
        # Case 1: User has entity, corpus must have SAME value
        if user_val:
            if not corpus_val:
                mismatches.append(f"{entity_type}: user has '{user_val}' but corpus missing")
            elif user_val.lower().strip() != corpus_val.lower().strip():
                mismatches.append(f"{entity_type}: user '{user_val}' != corpus '{corpus_val}'")
        
        # Case 2: Corpus has entity but user doesn't - also a mismatch
        elif corpus_val:
            mismatches.append(f"{entity_type}: corpus has '{corpus_val}' but user missing")
    
    # If any mismatch, return False
    return (len(mismatches) == 0, mismatches)


def normalize_question_entities(question: str) -> str:
    """
    Normalize question by replacing specific entities with placeholders.
    This creates a template for corpus matching.
    """
    q = question.lower()
    
    # Normalize months
    for month_key, month_full in MONTHS.items():
        q = re.sub(rf"\b{month_key}\b", "<MONTH>", q, flags=re.IGNORECASE)
        q = re.sub(rf"\b{month_full}\b", "<MONTH>", q, flags=re.IGNORECASE)
    
    # Normalize years
    q = re.sub(r"\b(20\d{2}|19\d{2})\b", "<YEAR>", q)
    
    # Normalize dates
    q = re.sub(r"\b\d{1,2}[-/]\d{1,2}[-/]\d{2,4}\b", "<DATE>", q)
    
    # Normalize geographic entities
    try:
        from .entity_cache import get_entities
        entities = get_entities()
        
        for state in entities.get("state", []):
            q = re.sub(rf"\b{re.escape(state)}\b", "<STATE>", q, flags=re.IGNORECASE)
        
        for zone in entities.get("zone", []):
            q = re.sub(rf"\b{re.escape(zone)}\b", "<ZONE>", q, flags=re.IGNORECASE)
        
        for branch in entities.get("branch", []):
            q = re.sub(rf"\b{re.escape(branch)}\b", "<BRANCH>", q, flags=re.IGNORECASE)
        
        for city in entities.get("city", []):
            q = re.sub(rf"\b{re.escape(city)}\b", "<CITY>", q, flags=re.IGNORECASE)
        
        for region in entities.get("region", []):
            q = re.sub(rf"\b{re.escape(region)}\b", "<REGION>", q, flags=re.IGNORECASE)
        
        for district in entities.get("district", []):
            q = re.sub(rf"\b{re.escape(district)}\b", "<DISTRICT>", q, flags=re.IGNORECASE)
        
        for country in entities.get("country", []):
            q = re.sub(rf"\b{re.escape(country)}\b", "<COUNTRY>", q, flags=re.IGNORECASE)
    except Exception as e:
        print(f"⚠️ Warning: Could not normalize geographic entities: {e}")
    
    return q


def replace_placeholders(question: str, entities: Dict[str, str]) -> str:
    """
    Replace placeholders in normalized question with actual entity values.
    """
    if not question:
        return ""
    
    out = question
    for placeholder, value in entities.items():
        val = value
        
        # Normalize month format
        if placeholder == "<MONTH>":
            month_key = value.lower()[:3]
            for short, full in MONTHS.items():
                if month_key == short:
                    val = full
                    break
        
        # Normalize year format
        elif placeholder == "<YEAR>":
            if re.match(r"^\d{2}$", value):  # '24 style
                val = "20" + value
        
        # Capitalize geographic entities properly
        elif placeholder in ["<STATE>", "<ZONE>", "<BRANCH>", "<CITY>", "<REGION>", "<DISTRICT>", "<COUNTRY>"]:
            val = value.strip().title()
        
        # Replace placeholder with normalized value
        out = re.sub(re.escape(placeholder), val, out, flags=re.IGNORECASE)
    
    return out


def expand_query_with_entities(query: str, entities: Dict[str, str]) -> List[str]:
    """
    Expand query with entity variants for better semantic search recall.
    """
    query_lower = query.lower()
    queries = [query_lower]
    
    # Month expansion
    if "<MONTH>" in entities:
        month_val = entities["<MONTH>"]
        for short, full in MONTHS.items():
            if full == month_val:
                queries.append(query_lower.replace(month_val.lower(), short))
                queries.append(query_lower.replace(month_val.lower(), full.lower()))
                break
    
    # Year expansion (2025 <-> '25)
    if "<YEAR>" in entities:
        year_val = entities["<YEAR>"]
        if len(year_val) == 4:
            queries.append(query_lower.replace(year_val, year_val[-2:]))
    
    # Geographic entity expansion
    for entity_type in ["<STATE>", "<ZONE>", "<BRANCH>", "<CITY>", "<REGION>", "<DISTRICT>", "<COUNTRY>"]:
        if entity_type in entities:
            val = entities[entity_type].lower()
            queries.append(f"in {val}")
            queries.append(f"{val} {entity_type.lower().strip('<>')}")
    
    return list(set(q.strip() for q in queries if q.strip()))


def validate_sql_entity_filters(sql: str, user_entities: Dict[str, str]) -> tuple[bool, List[str]]:
    """
    Validate that SQL filters match user's entities.
    Returns (is_valid, list_of_issues)
    """
    issues = []
    sql_lower = sql.lower()
    
    # Validate MONTH filter
    if "<MONTH>" in user_entities:
        user_month = user_entities["<MONTH>"]
        expected_month_num = MONTH_TO_NUMBER.get(user_month)
        
        month_pattern = r"policy_end_date_month\s*=\s*(\d+)"
        month_match = re.search(month_pattern, sql_lower)
        
        if month_match:
            sql_month_num = int(month_match.group(1))
            if sql_month_num != expected_month_num:
                issues.append(f"Month mismatch: SQL has {sql_month_num}, user asked for {user_month} ({expected_month_num})")
        else:
            issues.append(f"Missing month filter: user asked for {user_month}")
    
    # Validate YEAR filter
    if "<YEAR>" in user_entities:
        user_year = user_entities["<YEAR>"]
        year_pattern = r"policy_end_date_year\s*=\s*(\d{4})"
        year_match = re.search(year_pattern, sql_lower)
        
        if year_match:
            sql_year = year_match.group(1)
            if sql_year != user_year:
                issues.append(f"Year mismatch: SQL has {sql_year}, user asked for {user_year}")
        else:
            issues.append(f"Missing year filter: user asked for {user_year}")
    
    # Validate STATE filter (if applicable)
    if "<STATE>" in user_entities:
        user_state = user_entities["<STATE>"]
        state_pattern = rf"state\s*=\s*['\"]([^'\"]+)['\"]"
        state_match = re.search(state_pattern, sql, re.IGNORECASE)
        
        if state_match:
            sql_state = state_match.group(1)
            if sql_state.lower() != user_state.lower():
                issues.append(f"State mismatch: SQL has {sql_state}, user asked for {user_state}")
        else:
            issues.append(f"Missing state filter: user asked for {user_state}")
    
    # Validate ZONE filter
    if "<ZONE>" in user_entities:
        user_zone = user_entities["<ZONE>"]
        zone_pattern = rf"zone\s*=\s*['\"]([^'\"]+)['\"]"
        zone_match = re.search(zone_pattern, sql, re.IGNORECASE)
        
        if zone_match:
            sql_zone = zone_match.group(1)
            if sql_zone.lower() != user_zone.lower():
                issues.append(f"Zone mismatch: SQL has {sql_zone}, user asked for {user_zone}")
    
    # Validate BRANCH filter
    if "<BRANCH>" in user_entities:
        user_branch = user_entities["<BRANCH>"]
        branch_pattern = rf"branch\s*=\s*['\"]([^'\"]+)['\"]"
        branch_match = re.search(branch_pattern, sql, re.IGNORECASE)
        
        if branch_match:
            sql_branch = branch_match.group(1)
            if sql_branch.lower() != user_branch.lower():
                issues.append(f"Branch mismatch: SQL has {sql_branch}, user asked for {user_branch}")
    
    return (len(issues) == 0, issues)


def print_entity_comparison(user_entities: Dict[str, str], corpus_entities: Dict[str, str]):
    """
    Debug helper to print entity comparison clearly.
    """
    print("=" * 80)
    print("ENTITY COMPARISON")
    print("=" * 80)
    
    all_entity_types = set(list(user_entities.keys()) + list(corpus_entities.keys()))
    
    for entity_type in sorted(all_entity_types):
        user_val = user_entities.get(entity_type, "NOT PRESENT")
        corpus_val = corpus_entities.get(entity_type, "NOT PRESENT")
        
        match_symbol = "✅" if user_val == corpus_val else "❌"
        
        print(f"{match_symbol} {entity_type:15s}: User='{user_val:20s}' | Corpus='{corpus_val:20s}'")
    
    print("=" * 80)

# -----------------------------------
# Expand queries with entities (dynamic)
# -----------------------------------
def expand_query_with_entities(query: str, entities: dict) -> list:
    """
    Expand user query with entity synonyms for better recall.
    Works for month, year, state, zone, branch, etc.
    """
    query_lower = query.lower()
    queries = [query_lower]

    for ph, val in entities.items():
        val_lower = val.lower()

        # Month expansion
        if ph == "<MONTH>":
            for short, full in MONTHS.items():
                if val_lower.startswith(short):
                    queries.append(query_lower.replace(short, full.lower()))
                    queries.append(query_lower.replace(short, full))
                    queries.append(query_lower.replace(full.lower(), short))
                    queries.append(query_lower.replace(full.lower(), full))

        # Year expansion (2026 -> '26)
        elif ph == "<YEAR>":
            if len(val) == 4 and val.isdigit():
                short_year = val[-2:]
                queries.append(query_lower.replace(val_lower, short_year))
                queries.append(query_lower.replace(short_year, val_lower))

        # Generic expansions (state, zone, branch, etc.)
        else:
            tokens = val_lower.split()
            for token in tokens:
                queries.append(query_lower.replace(token, val_lower))
                queries.append(query_lower.replace(val_lower, token))

    return list(set(queries))


# -----------------------------------
# Replace placeholders in template
# -----------------------------------
def replace_placeholders(question: str, entities: dict) -> str:
    if not question:
        return ""

    out = question
    for placeholder, value in entities.items():
        val = value
        if placeholder == "<MONTH>":
            # Normalize month
            month_key = value.lower()[:3]
            for short, full in MONTHS.items():
                if month_key.startswith(short):
                    val = full
                    break
        out = re.sub(re.escape(placeholder), val, out, flags=re.I)
    return out


def rewrite_with_entities(question: str, query: str) -> str:
    """Fallback rewrite if placeholders still exist."""
    return re.sub(r"<[A-Z_]+>", query, question)


# # -----------------------------------
# # Save into corpus
# # -----------------------------------
# def save_to_corpus(collection, qid, user_question, normalized_q):
#     entities = extract_entities(user_question)
#     display_q = replace_placeholders(normalized_q, entities)
#     if re.search(r'<[A-Z_]+>', display_q):
#         display_q = rewrite_with_entities(normalized_q, user_question)

#     collection.add(
#         ids=[qid],
#         documents=[json.dumps({
#             "normalized_question": normalized_q,   # template
#             "raw_question": user_question,         # user asked
#             "display_question": display_q,         # rewritten
#             "asked_question": user_question
#         })],
#         metadatas=[{"type": "qa"}]
#     )


# -----------------------------------
# Search corpus (dynamic)
# -----------------------------------
def search_corpus(request):
    query = request.GET.get("q", "").strip()
    boost_raw = request.GET.get("boost", "[]")

    try:
        boost_keywords = json.loads(boost_raw)
        if not isinstance(boost_keywords, list):
            boost_keywords = BOOST_KEYWORDS
    except:
        boost_keywords = BOOST_KEYWORDS

    keywords_to_boost = set([kw.lower() for kw in BOOST_KEYWORDS + boost_keywords])

    entities = extract_entities(query)
    query_lower = query.lower()

    chroma_client = chromadb.PersistentClient(path="corpus_db")
    collection = chroma_client.get_or_create_collection("corpus_liberty")

    results = []
    if query:
        expanded_queries = expand_query_with_entities(query, entities)
        main_items = collection.query(query_texts=expanded_queries, n_results=25)

        for doc_id, doc, meta in zip(
            main_items["ids"][0],
            main_items["documents"][0],
            main_items["metadatas"][0],
        ):
            try:
                parsed = json.loads(doc)
            except:
                parsed = {"question": doc}

            # Always start from normalized template
            display_q = (
                parsed.get("normalized_question")
                or parsed.get("question")
                or ""
            )

            # Rewrite placeholders dynamically
            display_q = replace_placeholders(display_q, entities)
            if re.search(r'<[A-Z_]+>', display_q):
                display_q = rewrite_with_entities(display_q, query)

            if not display_q.strip():
                continue

            # Scoring
            score = 5.0
            display_lower = display_q.lower()
            for kw in keywords_to_boost:
                if kw in query_lower and kw in display_lower:
                    score += 10
                elif kw in display_lower:
                    score += 3

            results.append({
                "id": doc_id,
                "content": {**parsed, "display_question": display_q},
                "metadata": meta,
                "score": score
            })

    # Deduplicate
    seen = {}
    for r in results:
        key = r["content"]["display_question"].lower().strip()
        if key not in seen or seen[key]["score"] < r["score"]:
            seen[key] = r

    results = sorted(seen.values(), key=lambda x: x["score"], reverse=True)

    return JsonResponse(
        {"resolved_query": query, "results": results[:10]},
        safe=False,
    )


# --- Helper: extract dynamic tokens from question ---
def extract_dynamic_tokens(text: str):
    import re
    t = text.lower()
    tokens = set()

    # Months
    months = [
        "jan", "january", "feb", "february", "mar", "march", "apr", "april",
        "may", "jun", "june", "jul", "july", "aug", "august", "sep", "september",
        "oct", "october", "nov", "november", "dec", "december"
    ]
    for m in months:
        if m in t:
            tokens.add(m)

    # Years
    years = re.findall(r"\b(20\d{2}|19\d{2})\b", t)
    tokens.update(years)

    # Time refs
    time_refs = [
        "today", "yesterday", "last year", "this year", "last month",
        "this month", "quarter", "q1", "q2", "q3", "q4"
    ]
    for tr in time_refs:
        if tr in t:
            tokens.add(tr)

    # Geography refs
    geo_refs = ["state", "branch", "zone", "region", "district", "city"]
    for g in geo_refs:
        if g in t:
            tokens.add(g)

    # Words > 3 chars
    for word in re.findall(r"\b[a-z]{3,}\b", t):
        tokens.add(word)

    return list(tokens)






from .normalize_question_entities import normalize_question_entities
from .entity_cache import get_entities
import re

def resolve_followup_question(q: str, history: list) -> str:
    """
    Expands fragmentary follow-up queries like:
      'in kerala', 'for 2025', 'top 5 branches', 'premium trend'
    into a full question by merging with the last asked question.
    """
    q = q.strip()
    if not history:
        return q

    last_full = history[-1].get("asked_question") or history[-1].get("raw", "")
    if not last_full:
        return q

    norm_last = normalize_question_entities(last_full.lower())
    norm_q = normalize_question_entities(q.lower())

    entities = get_entities()

    # --- CASE 1: Pure location follow-up (state/zone/branch)
    for state in entities["state"]:
        if re.search(rf"\b{re.escape(state)}\b", q.lower()):
            return re.sub(r"\bstate\b.*", f"in {state}", last_full, flags=re.I)

    for zone in entities["zone"]:
        if re.search(rf"\b{re.escape(zone)}\b", q.lower()):
            return re.sub(r"\bzone\b.*", f"in {zone}", last_full, flags=re.I)

    for branch in entities["branch"]:
        if re.search(rf"\b{re.escape(branch)}\b", q.lower()):
            return re.sub(r"\bbranch.*", f"in {branch}", last_full, flags=re.I)

    # --- CASE 2: Year or month follow-up
    if re.match(r"^\d{4}$", q):  # "2025"
        return last_full + f" in {q}"
    if q.lower().startswith("for "):
        return last_full + " " + q

    months = ["jan", "feb", "mar", "apr", "may", "jun", 
              "jul", "aug", "sep", "oct", "nov", "dec",
              "january","february","march","april","may","june",
              "july","august","september","october","november","december"]
    if any(m in q.lower() for m in months):
        return last_full + f" in {q}"

    # --- CASE 3: Metric shift follow-up (churn, premium, policies, vehicles)
    metrics = ["churn", "premium", "policies", "vehicles"]
    if any(m in q.lower() for m in metrics):
        return f"Show {q}"  # "churn in 2025" → "Show churn in 2025"

    # --- CASE 4: Ranking follow-ups (top N)
    if re.match(r"top \d+", q.lower()):
        return f"Show {q} {last_full}"

    # --- Default: append context if it's too short
    if len(q.split()) <= 3:  
        return last_full + " " + q

    return q






# ============== STREAMING WITH ENHANCED CONVERSATION ==============
def user_wants_chart(question: str) -> bool:
    q = question.lower()
    keywords = [
        "chart", "graph", "plot", "visualize", "visualization",
        "trend", "over time", "distribution", "comparison", "compare",
        "versus", "vs", "ranking", "top", "bottom", "proportion", "percentage"
    ]
    return any(k in q for k in keywords)

from .normalize_question_entities import normalize_question_entities


import json

def _ev(event_type, **kwargs):
    """
    Emit a single NDJSON event with a consistent schema:
    {"type": <event_type>, ...}
    """
    payload = {"type": event_type}
    payload.update(kwargs)
    return json.dumps(payload) + "\n"






# -----------------------------------
def save_to_corpus_with_entities(collection, qid, user_question, normalized_q):
    entities = extract_entities(user_question)
    display_q = replace_placeholders(normalized_q, entities)
    if re.search(r'<[A-Z_]+>', display_q):
        display_q = rewrite_with_entities(normalized_q, user_question)

    collection.add(
        ids=[qid],
        documents=[json.dumps({
            "normalized_question": normalized_q,   # template
            "raw_question": user_question,         # user asked
            "display_question": display_q,         # rewritten
            "asked_question": user_question
        }, ensure_ascii=False)],
        metadatas=[{"type": "qa"}]
    )
    print(f"✅ Saved entity-aware QA into corpus: {qid}")
    return {"id": qid}


# ================= SAVE TO CORPUS =================
def save_to_corpus(
    question,
    summary=None,
    recommendations=None,
    sql=None,
    chart_config=None,
    row_count=0,
    db_id="liberty",
    existing_id=None,
    narrative=None,
    raw_examples=None,
    asked_question=None,
    normalized_q=None
):
    """
    Main entrypoint for saving structured answers into the corpus.
    Supports both structured (summary/sql/etc.) and entity-aware questions.
    """
    try:
        chroma_client = chromadb.PersistentClient(path="corpus_db")
        collection = chroma_client.get_or_create_collection(f"corpus_{db_id}")

        # If normalized_q is passed, also store entity-aware version
        if normalized_q:
            qid = existing_id or f"qa_{hash(question) % (10**8)}"
            save_to_corpus_with_entities(collection, qid, question, normalized_q)

        payload = {
            "question": question,
            "asked_question": asked_question or question,
            "summary": summary,
            "recommendations": recommendations,
            "sql": sql,
            "narrative": narrative,
            "chart_config": chart_config,
            "row_count": row_count,
            "timestamp": datetime.datetime.utcnow().isoformat()
        }

        metadata = {
            "question": question,
            "type": "qa",
            "asked_question": asked_question or question
        }

        if raw_examples:
            metadata["raw_examples"] = json.dumps(raw_examples, ensure_ascii=False)
            # ✅ also store full question for suggestions
            if "raw_question" in raw_examples:
                metadata["asked_question"] = raw_examples["raw_question"]

        # --- Update existing if provided ---
        if existing_id:
            collection.update(
                ids=[existing_id],
                documents=[json.dumps(payload, ensure_ascii=False)],
                metadatas=[metadata]
            )
            print(f"🔄 Updated structured answer in corpus: {existing_id}")
            return {"id": existing_id}

        # --- Check if similar already exists ---
        results = collection.query(query_texts=[question], n_results=1)
        if results and results.get("ids") and len(results["ids"][0]) > 0:
            found_id = results["ids"][0][0]
            collection.update(
                ids=[found_id],
                documents=[json.dumps(payload, ensure_ascii=False)],
                metadatas=[metadata]
            )
            print(f"🔄 Updated structured answer in corpus (found existing): {found_id}")
            return {"id": found_id}

        # --- Else add new ---
        doc_id = f"qa_{hash(question) % (10**8)}"
        collection.add(
            documents=[json.dumps(payload, ensure_ascii=False)],
            metadatas=[metadata],
            ids=[doc_id]
        )
        print(f"✅ Stored structured answer into corpus: {doc_id}")
        return {"id": doc_id}

    except Exception as e:
        print(f"⚠️ Failed to store into corpus: {e}")
        return {"error": str(e)}






@csrf_exempt
def save_to_corpus_endpoint(request):
    if request.method != "POST":
        return JsonResponse({"error": "POST only"}, status=405)

    try:
        body = json.loads(request.body or "{}")

        # Raw question from frontend (what user typed)
        raw_question = body.get("question", "").strip()
        if not raw_question:
            return JsonResponse({"error": "Missing question"}, status=400)

        # Resolve follow-ups against conversation history
        user_id = body.get("user_id", "admin")
        history = conversation_memory_store.get(user_id, [])
        resolved_q = resolve_followup_question(raw_question, history)

        # Normalize entities (replace Jan → <MONTH>, etc.)
        norm_question = normalize_question_entities(resolved_q)

        # Collect structured fields
        summary = body.get("summary", "")
        recommendations = body.get("recommendations", [])
        sql = body.get("sql", "")
        chart_config = body.get("chart_config")
        row_count = body.get("row_count", 0)
        db_id = body.get("db_id", "liberty")
        narrative = body.get("narrative")
        existing_id = body.get("existing_id")

        # ✅ Call main save
        result = save_to_corpus(
            question=resolved_q,                  # store resolved as main question
            asked_question=raw_question,          # what user typed
            normalized_q=norm_question,           # with placeholders
            summary=summary,
            recommendations=recommendations,
            sql=sql,
            chart_config=chart_config,
            row_count=row_count,
            db_id=db_id,
            narrative=narrative,
            existing_id=existing_id,
            raw_examples={
                "raw_question": raw_question,
                "resolved_question": resolved_q,
                "normalized_question": norm_question
            }
        )

        return JsonResponse({
            "success": True,
            "message": "Response saved to corpus memory",
            "corpus_id": result.get("id") if isinstance(result, dict) else None,
            "raw_question": raw_question,
            "resolved_question": resolved_q,
            "normalized_question": norm_question
        })

    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON"}, status=400)
    except Exception as e:
        print(f"⚠️ save_to_corpus_endpoint error: {e}")
        return JsonResponse({"error": str(e)}, status=500)







def generate_corpus_summary_and_chart(question: str, corpus_context: str):
    """Generate summary, table preview, chart config, and recommendations from corpus docs"""
    try:
        llm = get_llama_maverick_llm()

        system_prompt = (
             "You are a skilled data analyst. Based ONLY on the provided corpus context, "
            "generate a factual summary, recommendations, a simple table preview, and a chart config. "
            "If the corpus context is empty or contains no usable values, return an EMPTY summary and recommendations. "
            "Do NOT generate filler text like 'no data' or 'we don’t have data'. "
            "Keep the output concise, factual, and professional."
        )

        user_prompt = f"""
        Question: {question}
        Corpus Context:
        {corpus_context}

        Please respond in JSON with keys:
        {{
            "summary": "2-3 sentence summary",
            "recommendations": ["list of 2-3 recommendations"],
            "table_preview": [
                {{"metric": "example_metric", "value": "example_value"}},
                ...
            ],
            "chart_config": {{"type": "bar" | "line" | "pie", "x": [...], "y": [...], "title": "..."}}
        }}
        """

        raw = llm._call(
            prompt=user_prompt,
            system_prompt=system_prompt,
            temperature=0.5,
            max_tokens=700,
        )
        return _clean_and_extract_json(raw)

    except Exception as e:
        print(f"⚠️ Corpus summary/chart generation failed: {e}")
        return {
            "summary": "Analysis generated from corpus context.",
            "recommendations": ["Explore churn drivers further.", "Segment customers for deeper insights."],
            "table_preview": [],
            "chart_config": None,
        }





from .normalize_question_entities import normalize_question_entities

def retrieve_context_from_corpus(question: str, db_id="liberty", n_results=3, min_score=0.80):
    """
    Retrieve knowledge-grounded context from Chroma corpus.
    Always normalize question entities first, then check for exact match or embeddings.
    Returns docs with both normalized and full asked_question.
    """
    try:
        # ✅ Normalize first (months, years, states, zones, branches)
        norm_question = normalize_question_entities(question)

        chroma_client = chromadb.PersistentClient(path="corpus_db")
        collection = chroma_client.get_or_create_collection(f"corpus_{db_id}")

        # ✅ Exact text match (fast path)
        results = collection.get(where={"question": norm_question})
        if results and results.get("ids"):
            try:
                doc = json.loads(results["documents"][0])
            except Exception:
                doc = {"summary": results["documents"][0]}

            # merge metadata so we always include asked_question
            meta = results["metadatas"][0] or {}
            doc.update(meta)

            stored_q = meta.get("asked_question") or meta.get("question", "UNKNOWN")
            print(f"⚡ Exact match forced: served from corpus → {results['ids'][0]} | stored_q='{stored_q}'")
            return [doc]

        # ✅ Otherwise fall back to semantic similarity
        results = collection.query(
            query_texts=[question],
            n_results=n_results,
            include=["documents", "metadatas", "distances"]
        )
        print(results, "collectionresult")
        docs = []
        distances = results.get("distances", [[]])[0]
        for i, doc in enumerate(results.get("documents", [[]])[0]):
            score = 1 - distances[i]
            meta = results["metadatas"][0][i]

            stored_q = meta.get("asked_question") or meta.get("question", "UNKNOWN")

            if score >= min_score:
                try:
                    parsed_doc = json.loads(doc)
                    parsed_doc["asked_question"] = results["metadatas"][0][i].get("asked_question") \
                               or parsed_doc.get("asked_question") \
                               or results["metadatas"][0][i].get("question")

                except Exception:
                    parsed_doc = {"summary": doc}
                
                # attach both
                parsed_doc["asked_question"] = (
                        meta.get("asked_question")
                        or parsed_doc.get("asked_question")
                        or meta.get("question")
                    )


                docs.append(parsed_doc)
                print(f"✅ Corpus match accepted (score={score:.3f}) → {results['ids'][0][i]} | stored_q='{stored_q}'")

                
            else:
                print(f"⚡ Corpus skipped: similarity too low (score={score:.3f}, threshold={min_score}) | stored_q='{stored_q}'")

        return docs

    except Exception as e:
        print(f"⚠️ Corpus retrieval failed: {e}")
    return []



# PRONOUN_RE = re.compile(r"\b(this|that|it|same|these|those|above|same query|segmentation of this|breakdown of this)\b", re.I)

# def columns_from_schema(schema_text: str) -> list[str]:
#     """
#     Fallback extractor: pull quoted identifiers from FULL_SCHEMA.
#     We filter out obvious non-column names.
#     """
#     if not schema_text:
#         return []
#     raw = re.findall(r'"([A-Za-z0-9_]+)"', schema_text)
#     blacklist = {"bi_dwh","main_cai_lib","stage"}
#     return [c for c in raw if c.lower() not in blacklist]



# def parse_sql_context(sql: str, valid_cols: list[str]) -> dict:
#     """
#     Inspect SQL and extract:
#       - group_by columns
#       - metric aggregate (SUM/COUNT/AVG/MIN/MAX)
#       - filters (any valid column used with =, ILIKE, IN (...))
#       - time (policy_end_date_year/month if present)
#     Works for any columns that exist in your schema.
#     """
#     ctx = {"group_by": [], "metric": None, "filters": {}, "time": {}}
#     if not sql:
#         return ctx

#     # GROUP BY
#     m = re.search(r'GROUP\s+BY\s+(.+?)(?:\s+ORDER\s+BY|\s+LIMIT|$)', sql, re.I|re.S)
#     if m:
#         raw = m.group(1)
#         cols = [c.strip().strip('"') for c in raw.split(",")]
#         ctx["group_by"] = [c for c in cols if c]

#     # METRIC
#     m = re.search(r'(sum|count|avg|min|max)\s*\(\s*("?[\w\.\*]+"?)\s*\)', sql, re.I)
#     if m:
#         col = m.group(2).replace('"', '')
#         ctx["metric"] = {"agg": m.group(1).upper(), "col": col}

#     # TIME
#     y = re.search(r'policy_end_date_year\s*=\s*(\d{4})', sql, re.I)
#     mo = re.search(r'policy_end_date_month\s*=\s*(\d{1,2})', sql, re.I)
#     if y:  ctx["time"]["year"]  = int(y.group(1))
#     if mo: ctx["time"]["month"] = int(mo.group(1))

#     # WHERE → filters (only schema columns)
#     where_m = re.search(r'\bWHERE\b(.+?)(?:\bGROUP\b|\bORDER\b|\bLIMIT\b|$)', sql, re.I|re.S)
#     where_txt = where_m.group(1) if where_m else ""
#     for col in (valid_cols or []):
#         col_re = re.compile(rf'(?<!\.)\b{re.escape(col)}\b', re.I)
#         if not col_re.search(where_txt):
#             continue

#         # ILIKE
#         m_ilike = re.search(rf'\b{re.escape(col)}\b\s+ILIKE\s+\'([^\']+)\'', where_txt, re.I)
#         if m_ilike:
#             ctx["filters"][col] = m_ilike.group(1)
#             continue

#         # =
#         m_eq = re.search(rf'\b{re.escape(col)}\b\s*=\s*(\'[^\']+\'|\d+)', where_txt, re.I)
#         if m_eq:
#             v = m_eq.group(1).strip("'")
#             ctx["filters"][col] = v
#             continue

#         # IN (...)
#         m_in = re.search(rf'\b{re.escape(col)}\b\s+IN\s*\(([^)]+)\)', where_txt, re.I)
#         if m_in:
#             vals = [v.strip().strip("'") for v in m_in.group(1).split(",")]
#             if len(vals) == 1:
#                 ctx["filters"][col] = vals[0]

#     return ctx

# def latest_context(history: list[dict]) -> dict | None:
#     for entry in reversed(history or []):
#         if entry.get("context"):
#             return entry["context"]
#     return None

# def _humanize_context(ctx: dict) -> str:
#     """Produce a short, LLM-friendly reuse hint."""
#     parts = []
#     if ctx.get("group_by"):
#         parts.append("• Keep GROUP BY: " + ", ".join(ctx["group_by"]))
#     if ctx.get("metric"):
#         m = ctx["metric"]
#         parts.append(f'• Keep METRIC: {m["agg"]}({m["col"]})')
#     if ctx.get("filters"):
#         flt = "; ".join([f"{k}={v}" for k, v in ctx["filters"].items()])
#         parts.append(f"• Reuse FILTERS: {flt}")
#     if ctx.get("time"):
#         t = []
#         if "year" in ctx["time"]:  t.append(f'year={ctx["time"]["year"]}')
#         if "month" in ctx["time"]: t.append(f'month={ctx["time"]["month"]}')
#         if t: parts.append("• Previous TIME: " + ", ".join(t))
#     if not parts:
#         return ""
#     parts.append("• Unless I say otherwise, keep the same grouping/metric; only adjust time/filters mentioned in this question.")
#     return "\n".join(parts)

# def relax_location_filters(sql: str) -> str:
#     """
#     Make state/branch/zone text filters robust to spaces and punctuation.
#     Transforms:
#       state ILIKE 'tamil nadu'      -> state ILIKE '%' || REPLACE('tamil nadu',' ','%') || '%'
#       state = 'TamilNadu'           -> state ILIKE '%' || REPLACE('TamilNadu',' ','%') || '%'
#     Also works for branch_name and zone.
#     """
#     def repl(m):
#         col = m.group(1)
#         val = m.group(2)
#         # If the value already contains % we leave it alone
#         if "%" in val:
#             return f"{col} ILIKE '{val}'"
#         # Turn "tamil nadu" into "%tamil%nadu%"
#         return f"{col} ILIKE '%' || REPLACE('{val}', ' ', '%') || '%'"

#     # ILIKE 'value'
#     sql = re.sub(r"\b(state|branch_name|zone)\s+ILIKE\s+'([^']+)'", repl, sql, flags=re.IGNORECASE)
#     # = 'value'
#     sql = re.sub(r"\b(state|branch_name|zone)\s*=\s*'([^']+)'", repl, sql, flags=re.IGNORECASE)
#     return sql



# def augment_question_with_context(q: str, history: list) -> str:
#     if not history:
#         return q
#     last = next((h for h in reversed(history) if h.get("context")), None)
#     if not last:
#         return q

#     ctx = last["context"]
#     lines = [
#         "You are continuing a multi-turn SQL task. Reuse prior context where it does not conflict.",
#         f"Previous table(s): {', '.join(ctx.get('tables', [])) or 'unknown'}",
#         f"Previous filters: {ctx.get('filters_sql', 'none')}",
#         f"Previous time window: {ctx.get('time_range', 'none')}",
#         f"Previous grouping: {ctx.get('groupby', 'none')}",
#         f"Previous ordering: {ctx.get('order_by', 'none')}",
#     ]
#     return q + "\n\n" + "\n".join(lines)

# # ─────────────────────────────────────────────────────────────────────────────



# def _clean_model_text(text: str) -> str:
#     if not text:
#         return ""
#     # Strip SQL/code blocks and inline code
#     text = re.sub(r"```sql.*?```", "", text, flags=re.DOTALL | re.IGNORECASE)
#     text = re.sub(r"```.*?```", "", text, flags=re.DOTALL)
#     text = re.sub(r"`[^`]+`", "", text)
#     # Collapse whitespace
#     text = re.sub(r"\s+", " ", text).strip()
#     return text




# import re
# from difflib import SequenceMatcher

# def normalize_sql(sql: str) -> str:
#     sql = extract_sql_block(sql or "")
#     # Fix LIMIT1 → LIMIT 1 etc.
#     return re.sub(r'\bLIMIT(\d+)', r'LIMIT \1', sql, flags=re.IGNORECASE)

# def best_prior_sql(question: str, history: list, threshold: float = 0.86) -> str | None:
#     """Return the most similar prior SQL by fuzzy matching the question text."""
#     q = (question or "").strip().lower()
#     best_sql, best_score = None, 0.0
#     for h in reversed(history):
#         hq = (h.get("question") or "").strip().lower()
#         if not hq or not h.get("sql"): 
#             continue
#         score = SequenceMatcher(None, q, hq).ratio()
#         if score > best_score:
#             best_sql, best_score = h["sql"], score
#     return best_sql if best_score >= threshold else None

# import re
# from typing import Dict, Any, List, Optional, Tuple

# # If you already have parse_sql_context(), keep it. We add row-derived carry.

# GEO_COLS_IN_ROWS = [
#     "branch_name",                # "bi_dwh"."main_cai_lib"
#     "branch_name",      # alt
#     "state", "state",
#     "zone",  "zone",
# ]
# def _is_segmentation_question(q: str) -> bool:
#     ql = (q or "").strip().lower()
#     return "segment" in ql or "segmentation" in ql or "breakdown" in ql

# def derive_geo_carry_from_rows(sql: str, rows: list) -> dict:
#     """
#     Infer chosen branch/state/zone from rows (useful for 'top-N ... LIMIT 1' queries).
#     Returns e.g. {"branch_name": "Hyderabad"} or {"state": "Telangana"} or {"zone": "South"}.
#     """
#     if not rows:
#         return {}
#     r0 = rows[0]
#     # prioritize branch → state → zone (tweak as needed)
#     if isinstance(r0.get("branch_name"), str) and r0["branch_name"].strip():
#         return {"branch_name": r0["branch_name"].strip()}
#     if isinstance(r0.get("cleaned_branch_name_2"), str) and r0["cleaned_branch_name_2"].strip():
#         return {"branch_name": r0["cleaned_branch_name_2"].strip()}
#     if isinstance(r0.get("state"), str) and r0["state"].strip():
#         return {"state": r0["state"].strip()}
#     if isinstance(r0.get("state"), str) and r0["state"].strip():
#         return {"state": r0["state"].strip()}
#     if isinstance(r0.get("zone"), str) and r0["zone"].strip():
#         return {"zone": r0["zone"].strip()}
#     if isinstance(r0.get("zone"), str) and r0["zone"].strip():
#         return {"zone": r0["zone"].strip()}
#     return {}

# def get_last_carry(history: list) -> dict:
#     """
#     Fetch most recent context.carry from history.
#     """
#     for h in reversed(history or []):
#         ctx = (h.get("context") or {})
#         carry = ctx.get("carry")
#         if carry:
#             return carry
#     return {}

# def pick_level_and_value_from_carry(carry: dict):
#     """
#     Decide which geo to use and return (level, column_name, value).
#     Priority: branch → state → zone.
#     """
#     if "branch_name" in carry:
#         return "branch", "branch_name", carry["branch_name"]
#     if "state" in carry:
#         return "state", "state", carry["state"]
#     if "zone" in carry:
#         return "zone", "zone", carry["zone"]
#     return None, None, None



PRONOUN_RE = re.compile(r"\b(this|that|it|same|these|those|above|same query|segmentation of this|breakdown of this)\b", re.I)

def columns_from_schema(schema_text: str) -> list[str]:
    """
    Fallback extractor: pull quoted identifiers from FULL_SCHEMA.
    We filter out obvious non-column names.
    """
    if not schema_text:
        return []
    raw = re.findall(r'"([A-Za-z0-9_]+)"', schema_text)
    blacklist = {"bi_dwh","main_cai_lib","stage"}
    return [c for c in raw if c.lower() not in blacklist]

def parse_sql_context(sql: str, valid_cols: list[str]) -> dict:
    """
    Inspect SQL and extract:
      - group_by columns
      - metric aggregate (SUM/COUNT/AVG/MIN/MAX)
      - filters (any valid column used with =, ILIKE, IN (...))
      - time (policy_end_date_year/month if present)
    Works for any columns that exist in your schema.
    """
    ctx = {"group_by": [], "metric": None, "filters": {}, "time": {}}
    if not sql:
        return ctx

    # GROUP BY
    m = re.search(r'GROUP\s+BY\s+(.+?)(?:\s+ORDER\s+BY|\s+LIMIT|$)', sql, re.I|re.S)
    if m:
        raw = m.group(1)
        cols = [c.strip().strip('"') for c in raw.split(",")]
        ctx["group_by"] = [c for c in cols if c]

    # METRIC
    m = re.search(r'(sum|count|avg|min|max)\s*\(\s*("?[\w\.\*]+"?)\s*\)', sql, re.I)
    if m:
        col = m.group(2).replace('"', '')
        ctx["metric"] = {"agg": m.group(1).upper(), "col": col}

    # TIME
    y = re.search(r'policy_end_date_year\s*=\s*(\d{4})', sql, re.I)
    mo = re.search(r'policy_end_date_month\s*=\s*(\d{1,2})', sql, re.I)
    if y:  ctx["time"]["year"]  = int(y.group(1))
    if mo: ctx["time"]["month"] = int(mo.group(1))

    # WHERE → filters (only schema columns)
    where_m = re.search(r'\bWHERE\b(.+?)(?:\bGROUP\b|\bORDER\b|\bLIMIT\b|$)', sql, re.I|re.S)
    where_txt = where_m.group(1) if where_m else ""
    for col in (valid_cols or []):
        col_re = re.compile(rf'(?<!\.)\b{re.escape(col)}\b', re.I)
        if not col_re.search(where_txt):
            continue

        # ILIKE
        m_ilike = re.search(rf'\b{re.escape(col)}\b\s+ILIKE\s+\'([^\']+)\'', where_txt, re.I)
        if m_ilike:
            ctx["filters"][col] = m_ilike.group(1)
            continue

        # =
        m_eq = re.search(rf'\b{re.escape(col)}\b\s*=\s*(\'[^\']+\'|\d+)', where_txt, re.I)
        if m_eq:
            v = m_eq.group(1).strip("'")
            ctx["filters"][col] = v
            continue

        # IN (...)
        m_in = re.search(rf'\b{re.escape(col)}\b\s+IN\s*\(([^)]+)\)', where_txt, re.I)
        if m_in:
            vals = [v.strip().strip("'") for v in m_in.group(1).split(",")]
            if len(vals) == 1:
                ctx["filters"][col] = vals[0]

    return ctx

def latest_context(history: list[dict]) -> dict | None:
    for entry in reversed(history or []):
        if entry.get("context"):
            return entry["context"]
    return None

def _humanize_context(ctx: dict) -> str:
    """Produce a short, LLM-friendly reuse hint."""
    parts = []
    if ctx.get("group_by"):
        parts.append("• Keep GROUP BY: " + ", ".join(ctx["group_by"]))
    if ctx.get("metric"):
        m = ctx["metric"]
        parts.append(f'• Keep METRIC: {m["agg"]}({m["col"]})')
    if ctx.get("filters"):
        flt = "; ".join([f"{k}={v}" for k, v in ctx["filters"].items()])
        parts.append(f"• Reuse FILTERS: {flt}")
    if ctx.get("time"):
        t = []
        if "year" in ctx["time"]:  t.append(f'year={ctx["time"]["year"]}')
        if "month" in ctx["time"]: t.append(f'month={ctx["time"]["month"]}')
        if t: parts.append("• Previous TIME: " + ", ".join(t))
    if not parts:
        return ""
    parts.append("• Unless I say otherwise, keep the same grouping/metric; only adjust time/filters mentioned in this question.")
    return "\n".join(parts)

def augment_question_with_context(q: str, history: list) -> str:
    if not history:
        return q
    last = next((h for h in reversed(history) if h.get("context")), None)
    if not last:
        return q

    ctx = last["context"]
    lines = [
        "You are continuing a multi-turn SQL task. Reuse prior context where it does not conflict.",
        f"Previous table(s): {', '.join(ctx.get('tables', [])) or 'unknown'}",
        f"Previous filters: {ctx.get('filters_sql', 'none')}",
        f"Previous time window: {ctx.get('time_range', 'none')}",
        f"Previous grouping: {ctx.get('groupby', 'none')}",
        f"Previous ordering: {ctx.get('order_by', 'none')}",
    ]
    return q + "\n\n" + "\n".join(lines)

# ─────────────────────────────────────────────────────────────────────────────



def _clean_model_text(text: str) -> str:
    if not text:
        return ""
    # Strip SQL/code blocks and inline code
    text = re.sub(r"```sql.*?```", "", text, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r"```.*?```", "", text, flags=re.DOTALL)
    text = re.sub(r"`[^`]+`", "", text)
    # Collapse whitespace
    text = re.sub(r"\s+", " ", text).strip()
    return text


def llm_generate_recommendation(question: str, rows: List[Dict[str, Any]]) -> str:
    """
    Generate business recommendation with accurate row count handling.
    Prevents LLM from counting sample rows instead of total rows.
    """
    
    # *** FIX: Calculate actual total count FIRST ***
    actual_row_count = len(rows) if isinstance(rows, list) else 0
    row_count_formatted = f"{actual_row_count:,}"
    
    print(f"📊 Generating recommendation for {row_count_formatted} total rows")
    
    # *** FIX: Reduced from 10 to 5 sample rows to reduce confusion ***
    preview_rows = rows[:5] if rows else []
    
    if isinstance(preview_rows, list) and preview_rows and isinstance(preview_rows[0], dict):
        table_str = "\n".join(
            ", ".join(f"{k}: {v}" for k, v in row.items()) 
            for row in preview_rows
        )
    else:
        table_str = "No rows available."

    # *** FIX: Enhanced system prompt with explicit count warning ***
    system_prompt = (
        f"You are a data-driven business analyst assistant.\n\n"
        
        f"⚠️ CRITICAL: The dataset contains {row_count_formatted} total records.\n"
        f"You will see only 5 sample records below - DO NOT count these samples.\n"
        f"YOU MUST reference the total count ({row_count_formatted}) in your recommendation.\n\n"
        
        "Based on the user's question and the query output data, "
        "generate a concise, actionable, and professional business recommendation. "
        "Focus on explaining the key trends, risks, or opportunities in simple business terms, "
        "helping the user understand what actions or decisions they can take next.\n\n"
        
        "IMPORTANT RULES:\n"
        f"- Always mention that you're analyzing {row_count_formatted} records\n"
        "- The sample rows shown are for CONTEXT only - do not count them\n"
        "- Strictly avoid including any SQL code, technical jargon, or step-by-step query explanations\n"
        "- Only provide a high-level insight that would help a manager or decision-maker\n"
        "- Do not use phrases like 'no data', 'no records', 'unfortunately', or negative language\n"
        "- Do not mention SQL, filters, column names, or technical details"
    )

    # *** FIX: Enhanced user prompt with count emphasis ***
    user_prompt = f"""User Question: {question}

⚠️ DATASET INFORMATION:
- Total Records in Analysis: {row_count_formatted}
- Sample Records Shown Below: {len(preview_rows)} (for context only)

Sample Data (first {len(preview_rows)} rows out of {row_count_formatted} total):
{table_str}

CRITICAL INSTRUCTION:
Your recommendation MUST acknowledge that you are analyzing {row_count_formatted} total records.
Do NOT say "10 records" or count the samples above.
Start your recommendation by mentioning the total dataset size ({row_count_formatted}).

Example opening: "Based on analysis of {row_count_formatted} records, we observe..."

Please provide only a one-paragraph business recommendation. Do not include SQL queries or explanations.
"""

    # Size guard (rough)
    full_prompt_words = len((system_prompt + user_prompt).split())
    if full_prompt_words > 5000:
        logger.warning("Prompt too large, truncating...")
        # Hard truncation safeguards
        user_prompt = user_prompt[:8000]
        system_prompt = system_prompt[:2000]

    messages = [
        SystemMessage(content=system_prompt),
        UserMessage(content=user_prompt),
    ]

    max_retries = 3
    base_delay = 1.0

    for attempt in range(max_retries):
        try:
            resp = _get_client().complete(
                messages=messages,
                model=AZURE_MODEL,
                temperature=0.1,
                top_p=0.9,
                max_tokens=1000,
                presence_penalty=0.0,
                frequency_penalty=0.0,
            )

            if not resp.choices:
                logger.warning("Empty choices from Azure Inference.")
                raise RuntimeError("Empty response")

            raw_text = (resp.choices[0].message.content or "").strip()
            cleaned = _clean_model_text(raw_text)
            
            # *** FIX: Validate that the response contains the correct count ***
            if cleaned:
                import re
                
                # Check if response mentions the actual row count
                has_correct_count = (
                    row_count_formatted in cleaned or 
                    str(actual_row_count) in cleaned
                )
                
                # Check if response mentions wrong count (like "5", "10", etc.)
                wrong_count_pattern = r'\b(5|10|fifteen|five|ten)\s+(records?|rows?|policies?|vehicles?|customers?)\b'
                has_wrong_count = re.search(wrong_count_pattern, cleaned, re.IGNORECASE)
                
                if has_wrong_count and not has_correct_count:
                    logger.warning(
                        f"⚠️ LLM used wrong count in recommendation. "
                        f"Expected {row_count_formatted}, attempting correction..."
                    )
                    
                    # Force correction by prepending correct context
                    cleaned = _correct_recommendation_count(
                        cleaned, actual_row_count, row_count_formatted, question
                    )
                
                print(f"✅ Recommendation generated: {cleaned[:150]}...")
                return cleaned
            
            return "No clear recommendation could be generated from the data provided."

        except HttpResponseError as e:
            status = getattr(e, "status_code", None)
            logger.warning(f"Azure HttpResponseError (status={status}): {e}")
            if status in (429, 502, 503, 504) and attempt < max_retries - 1:
                time.sleep(base_delay * (2 ** attempt))
                continue
            break
        except (ServiceRequestError, ServiceResponseError, TimeoutError, ConnectionError) as e:
            logger.warning(f"Transient Azure error: {e}")
            if attempt < max_retries - 1:
                time.sleep(base_delay * (2 ** attempt))
                continue
            break
        except Exception as e:
            logger.warning(f"Attempt {attempt + 1}/{max_retries} failed: {e}")
            if attempt < max_retries - 1:
                time.sleep(base_delay * (2 ** attempt))
                continue
            break

    return "I apologize, but I'm having trouble processing your request right now."


def _correct_recommendation_count(
    recommendation: str, 
    actual_count: int, 
    formatted_count: str,
    question: str
) -> str:
    """
    Corrects a recommendation that used the wrong row count.
    Replaces sample counts (5, 10) with the actual total count.
    """
    import re
    
    # Pattern to find wrong counts
    patterns = [
        (r'\b(5|five)\s+(records?|rows?|policies?|vehicles?|customers?)\b', f'{formatted_count} \\2'),
        (r'\b(10|ten)\s+(records?|rows?|policies?|vehicles?|customers?)\b', f'{formatted_count} \\2'),
        (r'\bBased on analysis of\s+(5|10|five|ten)\s+', f'Based on analysis of {formatted_count} '),
        (r'\bAnalyzing\s+(5|10|five|ten)\s+', f'Analyzing {formatted_count} '),
        (r'\bWith\s+(5|10|five|ten)\s+', f'With {formatted_count} '),
    ]
    
    corrected = recommendation
    made_changes = False
    
    for pattern, replacement in patterns:
        if re.search(pattern, corrected, re.IGNORECASE):
            corrected = re.sub(pattern, replacement, corrected, flags=re.IGNORECASE)
            made_changes = True
    
    # If we made changes, prepend a clarification
    if made_changes:
        # Determine subject from question
        subject = "records"
        if "vehicle" in question.lower() or "car" in question.lower():
            subject = "vehicle insurance policies"
        elif "polic" in question.lower():
            subject = "policies"
        elif "customer" in question.lower():
            subject = "customers"
        
        prefix = f"Based on analysis of {formatted_count} {subject}: "
        
        # Remove any existing "Based on" prefix to avoid duplication
        corrected = re.sub(r'^Based on (analysis of|examining|reviewing)\s+\d[\d,]*\s+\w+:\s*', '', corrected, flags=re.IGNORECASE)
        
        corrected = prefix + corrected
        print(f"✅ Corrected recommendation count to {formatted_count}")
    
    return corrected

def llm_generate_recommendation2412(question: str, rows: List[Dict[str, Any]]) -> str:
    preview_rows = rows[:10] if rows else []
    if isinstance(preview_rows, list) and preview_rows and isinstance(preview_rows[0], dict):
        table_str = "\n".join(", ".join(f"{k}: {v}" for k, v in row.items()) for row in preview_rows)
    else:
        table_str = "No rows available."

    system_prompt = (
        "You are a data-driven business analyst assistant. Based on the user's question and the query output data, "
        "generate a concise, actionable, and professional business recommendation. "
        "Focus on explaining the key trends, risks, or opportunities in simple business terms, "
        "helping the user understand what actions or decisions they can take next. "
        "Strictly avoid including any SQL code, technical jargon, or step-by-step query explanations. "
        "Only provide a high-level insight that would help a manager or decision-maker."
    )

    user_prompt = f"""User Question: {question}

Data Preview (first 10 rows):
{table_str}

Please provide only a one-paragraph business recommendation. Do not include SQL queries or explanations.
"""

    # Size guard (rough)
    full_prompt_words = len((system_prompt + user_prompt).split())
    if full_prompt_words > 5000:
        logger.warning("Prompt too large, truncating...")
        # Hard truncation safeguards
        user_prompt = user_prompt[:8000]
        system_prompt = system_prompt[:2000]

    messages = [
        SystemMessage(content=system_prompt),
        UserMessage(content=user_prompt),
    ]

    max_retries = 3
    base_delay = 1.0

    for attempt in range(max_retries):
        try:
            resp = _get_client().complete(
                messages=messages,
                model=AZURE_MODEL,
                temperature=0.1,
                top_p=0.9,
                max_tokens=1000,
                presence_penalty=0.0,
                frequency_penalty=0.0,
            )

            if not resp.choices:
                logger.warning("Empty choices from Azure Inference.")
                raise RuntimeError("Empty response")

            raw_text = (resp.choices[0].message.content or "").strip()
            cleaned = _clean_model_text(raw_text)
            return cleaned or "No clear recommendation could be generated from the data provided."

        except HttpResponseError as e:
            status = getattr(e, "status_code", None)
            logger.warning(f"Azure HttpResponseError (status={status}): {e}")
            if status in (429, 502, 503, 504) and attempt < max_retries - 1:
                time.sleep(base_delay * (2 ** attempt))
                continue
            break
        except (ServiceRequestError, ServiceResponseError, TimeoutError, ConnectionError) as e:
            logger.warning(f"Transient Azure error: {e}")
            if attempt < max_retries - 1:
                time.sleep(base_delay * (2 ** attempt))
                continue
            break
        except Exception as e:
            logger.warning(f"Attempt {attempt + 1}/{max_retries} failed: {e}")
            if attempt < max_retries - 1:
                time.sleep(base_delay * (2 ** attempt))
                continue
            break

    return "I apologize, but I'm having trouble processing your request right now."


import re
from difflib import SequenceMatcher

def normalize_sql(sql: str) -> str:
    sql = extract_sql_block(sql or "")
    # Fix LIMIT1 → LIMIT 1 etc.
    return re.sub(r'\bLIMIT(\d+)', r'LIMIT \1', sql, flags=re.IGNORECASE)

# def best_prior_sql(question: str, history: list, threshold: float = 0.86) -> str | None:
#     """Return the most similar prior SQL by fuzzy matching the question text."""
#     q = (question or "").strip().lower()
#     best_sql, best_score = None, 0.0
#     for h in reversed(history):
#         hq = (h.get("question") or "").strip().lower()
#         if not hq or not h.get("sql"): 
#             continue
#         score = SequenceMatcher(None, q, hq).ratio()
#         if score > best_score:
#             best_sql, best_score = h["sql"], score
#     return best_sql if best_score >= threshold else None

def best_prior_sql(question: str, history: list, threshold: float = 0.85) -> str | None:
    """
    Enhanced version - uses existing normalize_question_entities
    """
    # Use existing normalization if available
    try:
        from .normalize_question_entities import normalize_question_entities
        q_normalized = normalize_question_entities(question.lower())
    except:
        # Fallback: simple normalization
        q_normalized = re.sub(r'[^\w\s]', '', question.lower()).strip()
    
    if not q_normalized:
        return None
    
    best_sql, best_score = None, 0.0
    
    for h in reversed(history[-20:]):  # Last 20 questions
        hq = h.get("question") or ""
        sql = h.get("sql") or ""
        
        if not hq or not sql:
            continue
        
        # Normalize history question
        try:
            hq_normalized = normalize_question_entities(hq.lower())
        except:
            hq_normalized = re.sub(r'[^\w\s]', '', hq.lower()).strip()
        
        # Exact match → immediate reuse
        if q_normalized == hq_normalized:
            print(f"♻️ EXACT MATCH - reusing SQL")
            return sql
        
        # Fuzzy match
        score = SequenceMatcher(None, q_normalized, hq_normalized).ratio()
        if score > best_score:
            best_sql, best_score = sql, score
    
    if best_score >= threshold:
        print(f"♻️ SIMILAR ({best_score:.2%}) - reusing SQL")
        return best_sql
    
    return None

import re
from typing import Dict, Any, List, Optional, Tuple

# If you already have parse_sql_context(), keep it. We add row-derived carry.

GEO_COLS_IN_ROWS = [
    "branch_name",                # "bi_dwh"."main_cai_lib"
    "branch_name",      # alt
    "state", "state",
    "zone",  "zone",
]
def _is_segmentation_question(q: str) -> bool:
    ql = (q or "").strip().lower()
    return "segment" in ql or "segmentation" in ql or "breakdown" in ql

def derive_geo_carry_from_rows(sql: str, rows: list) -> dict:
    """
    Infer chosen branch/state/zone from rows (useful for 'top-N ... LIMIT 1' queries).
    Returns e.g. {"branch_name": "Hyderabad"} or {"state": "Telangana"} or {"zone": "South"}.
    """
    if not rows:
        return {}
    r0 = rows[0]
    # prioritize branch → state → zone (tweak as needed)
    if isinstance(r0.get("branch_name"), str) and r0["branch_name"].strip():
        return {"branch_name": r0["branch_name"].strip()}
    if isinstance(r0.get("cleaned_branch_name_2"), str) and r0["cleaned_branch_name_2"].strip():
        return {"branch_name": r0["cleaned_branch_name_2"].strip()}
    if isinstance(r0.get("state"), str) and r0["state"].strip():
        return {"state": r0["state"].strip()}
    if isinstance(r0.get("state"), str) and r0["state"].strip():
        return {"state": r0["state"].strip()}
    if isinstance(r0.get("zone"), str) and r0["zone"].strip():
        return {"zone": r0["zone"].strip()}
    if isinstance(r0.get("zone"), str) and r0["zone"].strip():
        return {"zone": r0["zone"].strip()}
    return {}

def get_last_carry(history: list) -> dict:
    """
    Fetch most recent context.carry from history.
    """
    for h in reversed(history or []):
        ctx = (h.get("context") or {})
        carry = ctx.get("carry")
        if carry:
            return carry
    return {}

def pick_level_and_value_from_carry(carry: dict):
    """
    Decide which geo to use and return (level, column_name, value).
    Priority: branch → state → zone.
    """
    if "branch_name" in carry:
        return "branch", "branch_name", carry["branch_name"]
    if "state" in carry:
        return "state", "state", carry["state"]
    if "zone" in carry:
        return "zone", "zone", carry["zone"]
    return None, None, None


import re
# -----------------------------
# Imports
# -----------------------------
import re
from textwrap import dedent

# -----------------------------
# Your original extract_sql_block (kept; renamed to avoid overwrite)
# -----------------------------
def extract_sql_block_legacy(text):
    if isinstance(text, tuple):
        text = text[0]
    if not isinstance(text, str):
        # print("" extract_sql_block received non-string:", type(text))
        return ""

    match = re.search(r"```sql\s+(.*?)```", text, re.DOTALL | re.IGNORECASE)
    if match:
        return match.group(1).strip()

    match = re.search(r"```(.*?)```", text, re.DOTALL)
    if match:
        return match.group(1).strip()

    match = re.search(r"(SELECT\s+.+?)(;|\Z)", text, re.IGNORECASE | re.DOTALL)
    if match:
        return match.group(1).strip()

    print("⚠️ No valid SQL block found. Returning original text.")
    return text.strip()

# -----------------------------
# Your second extract_sql_block (kept; renamed)
# -----------------------------
def extract_sql_block_v2(s: str) -> str:
    s = str(s or "")
    m = re.search(r"```(?:sql)?\s*(.*?)```", s, flags=re.IGNORECASE | re.DOTALL)
    if m:
        return m.group(1).strip().rstrip(";").strip()
    # Fallback: return whole text trimmed (caller will slice from WITH/SELECT)
    return s.strip()

# -----------------------------
# Unified dispatcher that preserves both logics
# (safe for your other helpers that call extract_sql_block)
# -----------------------------
def extract_sql_block(s):
    """
    Try v2 first (better fenced handling); if that yields text without SQL,
    fall back to legacy which also searches for SELECT ... patterns.
    """
    candidate = extract_sql_block_v2(s)
    if candidate and re.search(r'(?is)\b(with|select)\b', candidate):
        return candidate
    # Fallback to legacy behavior
    return extract_sql_block_legacy(s)

# -----------------------------
# Your existing helpers (kept exactly as given)
# -----------------------------
def safe_extract_sql_text(maybe_sql: object) -> str:
    """
    Make sure we return a SQL string beginning with WITH/SELECT.
    If fenced, try extract_sql_block(); otherwise slice from the first WITH/SELECT.
    """
    text_ = str(maybe_sql or "").lstrip("\ufeff")
    if "```" in text_:
        try:
            unwrapped = extract_sql_block(text_)
            if isinstance(unwrapped, str) and unwrapped.strip():
                return unwrapped.strip()
        except Exception:
            pass
    m = re.search(r'(?is)\b(with|select)\b', text_)
    return text_[m.start():].strip() if m else text_.strip()


def repair_split_time_context(sql: str) -> str:
    """
    Fix pattern produced by memory merge:
      [leading SELECT ... ) SELECT ... CROSS JOIN time_context]
    by wrapping the leading SELECT into WITH time_context AS (...).
    """
    s = str(sql or "")
    low = s.lower()

    # already fine or not referencing time_context
    if low.startswith("with") or "time_context" not in low or re.search(r'\bwith\s+time_context\b', low):
        return s

    # find: [ ... ) SELECT ... ]  (first ') SELECT' boundary)
    m = re.search(r'\)\s*select\b', s, flags=re.IGNORECASE | re.DOTALL)
    if not (s.strip().lower().startswith("select") and m):
        return s

    paren_end = m.start()              # index of ')'
    # find the 'SELECT' right after that ')'
    sel_after = re.search(r'\bselect\b', s[paren_end:], flags=re.IGNORECASE)
    if not sel_after:
        return s
    sel_idx = paren_end + sel_after.start()

    # prefix becomes body of time_context; rest is the top-level SELECT
    prefix = s[:paren_end].rstrip()    # content before that ')'
    rest   = s[sel_idx:].lstrip()

    # Build repaired statement
    repaired = f"WITH time_context AS ({prefix})\n{rest}"
    return repaired

# -----------------------------
# New strict guards (additive; do not replace your code)
# -----------------------------
_BAD_MARKERS_RE = re.compile(
    r"(?i)\b("
    r"i['’]m|i am|sorry|apolog|error|cannot|can't|trouble processing|"
    r"stack trace|traceback|django|http|service unavailable|timeout"
    r")\b"
)

_SQL_START_RE = re.compile(r"(?is)^\s*(with|select)\b")
_DML_DDL_RE = re.compile(
    r"(?i)\b(insert|update|delete|drop|alter|create|truncate|grant|revoke|merge)\b"
)

def is_probably_sql_strict(s: str) -> bool:
    """
    Tight sanity checks to keep prose and broken SQL out of the DB.
    """
    if not s:
        return False
    if _BAD_MARKERS_RE.search(s):
        return False
    if not _SQL_START_RE.search(s):
        return False
    if s.count("'") % 2 == 1:  # odd single-quote count → likely unterminated string
        return False
    return True

def enforce_limit(sql: str, default_limit: int = 5000) -> str:
    if re.search(r"\blimit\s+\d+\b", sql, flags=re.IGNORECASE):
        return sql
    return f"{sql}\nLIMIT {default_limit}"

def prepare_sql_for_exec(raw_sql: str, *, require_readonly: bool = True) -> str:
    """
    End-to-end preparation called immediately before execution.
    Uses your helpers plus strict guards. Read-only by default.
    """
    # 1) Extract canonical SQL from LLM or reused text
    s = safe_extract_sql_text(raw_sql)

    # 2) Reject apology/error prose and malformed text
    if not is_probably_sql_strict(s):
        raise ValueError("Non-SQL or apology/error text detected.")

    # 3) Enforce read-only if desired
    if require_readonly and _DML_DDL_RE.search(s):
        raise ValueError("Only read-only SELECT queries are allowed.")

    # 4) Repair known time_context split
    try:
        s = repair_split_time_context(s)
    except Exception:
        pass

    # 5) Normalize minor issues (e.g., LIMIT1 → LIMIT 1)
    s = re.sub(r"\blimit\s*(\d+)\b", r"LIMIT \1", s, flags=re.IGNORECASE)

    # 6) Must start with WITH/SELECT
    if not _SQL_START_RE.search(s):
        raise ValueError("Query does not start with WITH or SELECT.")

    # 7) Safety: enforce a LIMIT unless present
    s = enforce_limit(s, default_limit=5000)

    return dedent(s).strip()

def repair_split_time_context(sql: str) -> str:
    """
    Fix pattern produced by memory merge:
      [leading SELECT ... ) SELECT ... CROSS JOIN time_context]
    by wrapping the leading SELECT into WITH time_context AS (...).
    """
    s = str(sql or "")
    low = s.lower()

    # already fine or not referencing time_context
    if low.startswith("with") or "time_context" not in low or re.search(r'\bwith\s+time_context\b', low):
        return s

    # find: [ ... ) SELECT ... ]  (first ') SELECT' boundary)
    m = re.search(r'\)\s*select\b', s, flags=re.IGNORECASE | re.DOTALL)
    if not (s.strip().lower().startswith("select") and m):
        return s

    paren_end = m.start()              # index of ')'
    # find the 'SELECT' right after that ')'
    sel_after = re.search(r'\bselect\b', s[paren_end:], flags=re.IGNORECASE)
    if not sel_after:
        return s
    sel_idx = paren_end + sel_after.start()

    # prefix becomes body of time_context; rest is the top-level SELECT
    prefix = s[:paren_end].rstrip()    # content before that ')'
    rest   = s[sel_idx:].lstrip()

    # Build repaired statement
    repaired = f"WITH time_context AS ({prefix})\n{rest}"
    return repaired




AZURE_ENDPOINT = os.getenv("AZURE_INFERENCE_ENDPOINT")  # https://genaiprochurn.services.ai.azure.com/models
AZURE_API_KEY = os.getenv("AZURE_INFERENCE_API_KEY")
AZURE_MODEL = os.getenv("AZURE_INFERENCE_MODEL", "Llama-4-Maverick-17B-128E-Instruct-FP8-prochurn-demo")
AZURE_API_VERSION = "2024-05-01-preview"
MAX_PROMPT_TOKENS = int(os.getenv("AZURE_MAX_PROMPT_TOKENS", "120000"))

_SYSTEM_PROMPT = os.getenv("AZURE_SYSTEM_PROMPT", "You are a helpful SQL assistant.")

_client: Optional[ChatCompletionsClient] = None

def _get_client() -> ChatCompletionsClient:
    global _client
    if _client is None:
        print(f"DEBUG - AZURE_ENDPOINT from Django: {os.getenv('AZURE_INFERENCE_ENDPOINT')}")
        print(f"DEBUG - AZURE_API_KEY from Django: {os.getenv('AZURE_INFERENCE_API_KEY', 'NOT SET')}")
        if not AZURE_ENDPOINT or not AZURE_API_KEY:
            logger.error("Azure credentials not configured")
            raise RuntimeError("Set AZURE_INFERENCE_ENDPOINT and AZURE_INFERENCE_API_KEY.")
        
        # Ensure endpoint has the correct format
        endpoint = AZURE_ENDPOINT.rstrip('/')
        if not endpoint.endswith('/models'):
            endpoint = f"{endpoint}/models"
        
        logger.info(f"Initializing Azure client with endpoint: {endpoint}")
        
        _client = ChatCompletionsClient(
            endpoint=endpoint,
            credential=AzureKeyCredential(AZURE_API_KEY),
            api_version=AZURE_API_VERSION,
        )
    return _client

def _yes_no(raw_response: str) -> str:
    """Extract YES/NO from response"""
    if not raw_response:
        return "NO"
    
    response = raw_response.strip().upper()
    if "YES" in response:
        return "YES"
    return "NO"


def _classify_intent(raw_response: str) -> str:
    """Extract YES/NO/UNCERTAIN from response"""
    if not raw_response:
        return "NO"
    
    response = raw_response.strip().upper()
    
    # Check for explicit responses first
    if "UNCERTAIN" in response:
        return "UNCERTAIN"
    elif "YES" in response:
        return "YES"
    elif "NO" in response:
        return "NO"
    
    # Default fallback
    return "NO"



@csrf_exempt
def check_intent(request):
    if request.method != "POST":
        return JsonResponse({"answer": "NO", "error": "Only POST allowed"}, status=405)

    try:
        data = json.loads(request.body or "{}")
    except Exception as e:
        logger.error(f"JSON parsing error: {e}")
        return JsonResponse({"answer": "NO", "error": "Invalid JSON body"}, status=400)

    question = (data.get("question") or "").strip()
    if not question:
        return JsonResponse({"answer": "NO"})

    # First-level check for incomplete questions
    if is_incomplete_question(question):
        return JsonResponse({"answer": "UNCERTAIN", "message": "Incomplete question. Please rephrase into a full question."})

    prompt = f"""
Classify the user's intent as YES, NO, UNCERTAIN, or PDF.

RULES:
1. **NO (General Knowledge/Greeting):**
   - Greetings: "hi", "hello", "how are you"
   - General world knowledge: "who is the PM of India?", "what is machine learning?", "what is AI?"
   - Casual explanations unrelated to insurance/business context

2. **YES (Clear Database Query - PRIORITY):**
   - ANY question with metrics/numbers/aggregations: "average", "total", "sum", "count", "how many", "top", "bottom"
   - Statistical queries: "average customer lifetime value", "top 5 branches by churn", "how many policies in Delhi?"
   - Data analysis: "total premium in 2024", "branch performance summary", "churn rate by state"
   - Comparisons: "compare X vs Y", "which branch has highest", "lowest retention rate"
   - Time-based queries: "last month", "in 2024", "this quarter"
   - Complete questions asking for metrics/counts/summaries from database
   - **IMPORTANT**: Questions like "what is the average/total/sum/count of X?" are ALWAYS database queries (YES), NOT definitions (PDF)

3. **PDF (Business Definitions/Conceptual Knowledge ONLY):**
   - Pure definitions without data requests: "what is a branch?" (definition), "what does customer segment mean?"
   - Conceptual explanations: "explain retention strategy", "what is the purpose of claims?"
   - Policy/insurance concepts: "what is surrender value?", "define claim ratio"
   - **NOT for queries asking for actual numbers/metrics even if they use business terms**

4. **UNCERTAIN (Vague/Ambiguous Fragments):**
   - Incomplete fragments: "top 5?", "Delhi?", "premium?"
   - Vague context: "me branches", "what about?"
   - Missing key info (metric, time period, entity)

KEY DISTINCTION:
- "What is customer lifetime value?" → PDF (asking for definition)
- "What is the average customer lifetime value?" → YES (asking for data/metric)
- "Show me customer lifetime value by segment" → YES (data query)
- "What is retention?" → PDF (definition)
- "What is the retention rate?" → YES (asking for metric)

Examples:
- "What is the average customer lifetime value?" → YES (data query)
- "What is the total revenue in 2024?" → YES (data query)
- "How many customers churned last month?" → YES (data query)
- "What is a branch?" → PDF (definition)
- "What is customer segment?" → PDF (definition)
- "Top 5 branches by churn in 2024" → YES (data query)
- "top 5?" → UNCERTAIN (incomplete)
- "Delhi?" → UNCERTAIN (incomplete)
- "Who is the PM of India?" → NO (general knowledge)
- "What is machine learning?" → NO (general knowledge)

Question: "{question}"

Respond with exactly one word: YES, NO, UNCERTAIN, or PDF
""".strip()

    try:
        client = _get_client()
        logger.info(f"Making request to Azure AI with model: {AZURE_MODEL}")

        response = client.complete(
            messages=[
                SystemMessage(content="You are an intent classifier. Reply only with YES, NO, UNCERTAIN, or PDF. Prioritize YES for any question asking for metrics, numbers, averages, totals, or data analysis."),
                UserMessage(content=prompt),
            ],
            model=AZURE_MODEL,
            temperature=0.0,
            top_p=1.0,
            max_tokens=15,
        )

        raw = response.choices[0].message.content if response.choices else ""
        answer = _classify_intent(raw)

        logger.info(f"Intent classification result: {answer} (raw: {raw})")

        if answer == "UNCERTAIN":
            history = conversation_memory_store.get("admin", [])
            tokens = extract_dynamic_tokens(question)

            history_suggestions = []
            for h in history[-10:]:
                asked = h.get("asked_question", "")
                if any(t in asked.lower() for t in tokens):
                    history_suggestions.append(asked)

            related = retrieve_context_from_corpus(question, db_id="liberty", min_score=0.30)
            semantic_suggestions = []
            if related:
                for r in related[:10]:
                    qtxt = (r.get("asked_question") or r.get("raw_question") or r.get("question") or "").strip()
                    if qtxt:
                        semantic_suggestions.append(qtxt)

            keyword_suggestions = [q for q in semantic_suggestions if any(t in q.lower() for t in tokens)]

            suggestions = (
                history_suggestions
                + [q for q in keyword_suggestions if q not in history_suggestions]
                + [q for q in semantic_suggestions if q not in keyword_suggestions and q not in history_suggestions]
            )

            return JsonResponse({
                "answer": "UNCERTAIN",
                "asked_question": question,
                "tokens_detected": tokens,
                "previous_questions": history_suggestions[:5],
                "related_questions": [
                    (r.get("asked_question") or r.get("question"))
                    for r in related if (r.get("asked_question") or r.get("question"))
                ]
            })

        return JsonResponse({"answer": answer})

    except HttpResponseError as e:
        logger.error(f"Azure HTTP error: {e.status_code} - {e.message}")
        return JsonResponse({"answer": "NO", "error": f"Azure service error: {e.message}"}, status=502)
        
    except (ServiceRequestError, ServiceResponseError) as e:
        logger.error(f"Azure service error: {e}")
        return JsonResponse({"answer": "NO", "error": f"Azure service error: {str(e)}"}, status=502)
        
    except (requests.exceptions.ConnectionError, requests.exceptions.Timeout) as e:
        logger.error(f"Connection error: {e}")
        return JsonResponse({"answer": "NO", "error": "Connection timeout or error"}, status=502)
        
    except Exception as e:
        logger.exception("check_intent failed with unexpected error")
        return JsonResponse({"answer": "NO", "error": f"Unexpected error: {str(e)}"}, status=500)


@csrf_exempt
def check_intent512(request):
    if request.method != "POST":
        return JsonResponse({"answer": "NO", "error": "Only POST allowed"}, status=405)

    try:
        data = json.loads(request.body or "{}")
    except Exception as e:
        logger.error(f"JSON parsing error: {e}")
        return JsonResponse({"answer": "NO", "error": "Invalid JSON body"}, status=400)

    question = (data.get("question") or "").strip()
    if not question:
        return JsonResponse({"answer": "NO"})

    prompt = f"""
Classify the user's intent as YES, NO, or UNCERTAIN.

RULES:
1. **NO (General Knowledge/Greeting):**
   - Greetings: "hi", "hello", "how are you"
   - Definitions: "what is a branch?", "explain churn", "what does premium mean?"
   - General knowledge: "who is the PM of India?", "what is machine learning?"
   - Explanatory questions that don't need database data

2. **YES (Clear Database Query):**
   - Statistical queries: "top 5 branches by churn", "how many policies in Delhi?"
   - Data analysis: "total premium in 2024", "branch performance summary"
   - Complete questions asking for database metrics/counts/summaries
   - Questions with clear entities and metrics

3. **UNCERTAIN (Vague/Ambiguous Fragments):**
   - Incomplete fragments: "top 5 branch?", "in Delhi?", "premium?"
   - Vague questions without clear context: "show me branches", "what about churn?"
   - Questions that could be either definition OR data query
   - Missing key information (metric, time period, or entity)

Examples:
- "What is a branch?" → NO (definition)
- "Top 5 branches by churn in 2024" → YES (clear DB query)
- "top 5 branch?" → UNCERTAIN (missing metric)
- "in Delhi?" → UNCERTAIN (fragment)
- "premium?" → UNCERTAIN (could be definition or data)
- "branches" → UNCERTAIN (too vague)

Question: "{question}"

Respond with exactly one word: YES, NO, or UNCERTAIN
""".strip()

    try:
        client = _get_client()
        logger.info(f"Making request to Azure AI with model: {AZURE_MODEL}")
        
        response = client.complete(
            messages=[
                SystemMessage(content="You are an intent classifier. Reply only with YES, NO, or UNCERTAIN."),
                UserMessage(content=prompt),
            ],
            model=AZURE_MODEL,
            temperature=0.0,
            top_p=1.0,
            max_tokens=15,  # Allow for UNCERTAIN response
        )
        
        # raw = response.choices[0].message.content if response.choices else ""
        # answer = _classify_intent(raw)
        raw = response.choices[0].message.content if response.choices else ""
        answer = _classify_intent(raw)
        
        logger.info(f"Intent classification result: {answer} (raw: {raw})")

        # 🆕 Handle UNCERTAIN → return suggestions with full asked_question
        if answer == "UNCERTAIN":
            history = conversation_memory_store.get("admin", [])
            tokens = extract_dynamic_tokens(question)

            # From history (full asked_question)
            history_suggestions = []
            for h in history[-10:]:
                asked = h.get("asked_question", "")
                if any(t in asked.lower() for t in tokens):
                    history_suggestions.append(asked)

            # From corpus (prefer asked_question if available)
            related = retrieve_context_from_corpus(question, db_id="liberty", min_score=0.30)
            semantic_suggestions = []
            if related:
                for r in related[:10]:
                    qtxt = (r.get("asked_question") or r.get("raw_question") or r.get("question") or "").strip()

                    if qtxt:
                        semantic_suggestions.append(qtxt)

            keyword_suggestions = [q for q in semantic_suggestions if any(t in q.lower() for t in tokens)]

            suggestions = (
                history_suggestions
                + [q for q in keyword_suggestions if q not in history_suggestions]
                + [q for q in semantic_suggestions if q not in keyword_suggestions and q not in history_suggestions]
            )

            return JsonResponse({
                "answer": "UNCERTAIN",
                "asked_question": question,
                "tokens_detected": tokens,
                "previous_questions": history_suggestions[:5],
                "related_questions": [
                (r.get("asked_question") or r.get("question"))
                for r in related if (r.get("asked_question") or r.get("question"))
            ]
  # ✅ FULL questions only
            })

        return JsonResponse({"answer": answer})


# ✅ Default return for YES/NO
        
        # logger.info(f"Intent classification result: {answer} (raw: {raw})")
        # return JsonResponse({"answer": answer})
        
    except HttpResponseError as e:
        logger.error(f"Azure HTTP error: {e.status_code} - {e.message}")
        return JsonResponse({"answer": "NO", "error": f"Azure service error: {e.message}"}, status=502)
        
    except (ServiceRequestError, ServiceResponseError) as e:
        logger.error(f"Azure service error: {e}")
        return JsonResponse({"answer": "NO", "error": f"Azure service error: {str(e)}"}, status=502)
        
    except (requests.exceptions.ConnectionError, requests.exceptions.Timeout) as e:
        logger.error(f"Connection error: {e}")
        return JsonResponse({"answer": "NO", "error": "Connection timeout or error"}, status=502)
        
    except Exception as e:
        logger.exception("check_intent failed with unexpected error")
        return JsonResponse({"answer": "NO", "error": f"Unexpected error: {str(e)}"}, status=500)


@csrf_exempt
def test_azure_endpoint(request):
    """Test endpoint to validate Azure AI connection"""
    if request.method != "GET":
        return JsonResponse({"error": "Only GET allowed"}, status=405)
    
    try:
        # Validate environment first
        is_valid, errors = validate_environment()
        if not is_valid:
            return JsonResponse({
                "status": "error",
                "message": "Environment validation failed",
                "errors": errors
            }, status=500)
        
        # Test connection
        success, result = test_azure_connection()
        if success:
            return JsonResponse({
                "status": "success",
                "message": "Azure AI connection successful",
                "response": result,
                "endpoint": AZURE_ENDPOINT,
                "model": AZURE_MODEL
            })
        else:
            return JsonResponse({
                "status": "error",
                "message": "Azure AI connection failed",
                "error": result
            }, status=500)
            
    except Exception as e:
        logger.exception("Test endpoint failed")
        return JsonResponse({
            "status": "error",
            "message": "Test failed with unexpected error",
            "error": str(e)
        }, status=500)

# settings.py additions for logging
LOGGING = {
    'version': 1,
    'disable_existing_loggers': False,
    'formatters': {
        'verbose': {
            'format': '{levelname} {asctime} {module} {process:d} {thread:d} {message}',
            'style': '{',
        },
    },
    'handlers': {
        'console': {
            'class': 'logging.StreamHandler',
            'formatter': 'verbose',
        },
        'file': {
            'class': 'logging.FileHandler',
            'filename': 'azure_ai.log',
            'formatter': 'verbose',
        },
    },
    'loggers': {
        'azure.ai.inference': {
            'handlers': ['console', 'file'],
            'level': 'DEBUG',
            'propagate': False,
        },
        '__main__': {  # or your app name
            'handlers': ['console', 'file'],
            'level': 'INFO',
            'propagate': False,
        },
    },
}

# Test function to validate configuration
def test_azure_connection():
    """Test function to validate Azure AI configuration"""
    try:
        client = _get_client()
        
        # Simple test request
        response = client.complete(
            messages=[
                SystemMessage(content="You are a helpful assistant."),
                UserMessage(content="Say 'Hello' if you can read this."),
            ],
            model=AZURE_MODEL,
            temperature=0.0,
            max_tokens=50,
        )
        
        result = response.choices[0].message.content if response.choices else ""
        logger.info(f"Azure connection test successful: {result}")
        return True, result
        
    except Exception as e:
        logger.error(f"Azure connection test failed: {e}")
        return False, str(e)

# Environment validation
def validate_environment():
    """Validate all required environment variables"""
    missing = []
    
    if not AZURE_ENDPOINT:
        missing.append("AZURE_INFERENCE_ENDPOINT")
    if not AZURE_API_KEY:
        missing.append("AZURE_INFERENCE_API_KEY")
    
    if missing:
        logger.error(f"Missing environment variables: {missing}")
        return False, missing
    
    # Validate endpoint format
    if not AZURE_ENDPOINT.startswith("https://"):
        logger.error("AZURE_INFERENCE_ENDPOINT must start with https://")
        return False, ["Invalid endpoint format"]
    
    logger.info("Environment validation passed")
    return True, []


# ✅ Register User View
session_store = {}
conversation_memory_store = {}

dataframe_map = {}
vectorstore_map = {}
conversation_memory = {}
memory_cache = {}
GROQ_API_KEY = os.getenv("GROQ_API_KEY")
GROQ_API_KEY = "REMOVEDeRuX7eFDYy4aOy7YrdkEWGdyb3FY6IjRCc9izTIPBvJaOMwdIQvJ"
OPENROUTER_API_KEY = os.getenv("OPENROUTER_API_KEY", "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")
os.environ["OPENAI_API_KEY"] = OPENROUTER_API_KEY 
OPENROUTER_API_BASE = os.getenv("OPENROUTER_API_BASE", "https://openrouter.ai/api/v1")

BASE_DIR = Path(__file__).resolve().parents[2]

DB_SSBI_Excel = settings.EXTERNAL_DATABASES["DB_SSBI_Excel"]
DB_SSBI = settings.EXTERNAL_DATABASES["DB_SSBI"]
DB_Auto = settings.EXTERNAL_DATABASES["DB_Auto"]
DB_PAST = settings.EXTERNAL_DATABASES["DB_PAST"]
DB_New_Pred = settings.EXTERNAL_DATABASES["DB_New_Pred"]
DB_New_Past = settings.EXTERNAL_DATABASES["DB_New_Past"]
# DB_Conv_AI = settings.EXTERNAL_DATABASES["DB_Conv_AI"]
# DB_Azure_Airflow = settings.EXTERNAL_DATABASES["DB_Azure_Airflow"]

# print("🌐 Loaded DB_Auto config:", DB_Auto) 
# print("🌐 Loaded DB_Auto config:", DB_SSBI)
# print("🌐 Loaded DB_Auto config:", DB_SSBI_Excel)
 # ✅ Debug print to confirm loaded DB config

logger = logging.getLogger(__name__)  # Add this at the top

session_store = {}
schema_context_store = {}  # store db schema context per session
query_preview_store = {}




FILE_MAPPING = {
    "base": "base",
    "pr": "pr",
    "claims": "claims",
    "feedback": "feedback",
    "additional": "additional_files"
}

MODEL_NAME = "distilgpt2"
DEVICE = torch.device("cpu")  # Change to "cuda" if using GPU
tokenizer = AutoTokenizer.from_pretrained(MODEL_NAME)
model = AutoModelForCausalLM.from_pretrained(MODEL_NAME, torch_dtype=torch.float32, device_map={"": DEVICE})
qa_pipeline = pipeline("question-answering", model="distilbert-base-uncased-distilled-squad")



PROCESS_STAGES = [
    ("Data Cleansing", ["Removed Duplicates", "Handled Missing Values", "Feature Engineering", "Normalization", "Encoding Categorical Values"]),
    ("Model Deployment", ["Pretrained Models", "Batch/API Prediction", "Confidence Score", "Segmented Customers", "Retention Strategies"]),
    ("Completed", [])
]

UPLOAD_DIR_PDF = os.path.join(settings.MEDIA_ROOT, "pdf_files")
VECTORSTORE_DIR = os.path.join(settings.MEDIA_ROOT, "vectorstorenew")
ADMIN_VECTORSTORE_NAME = "admin_base"
ADMIN_VECTORSTORE_PATH = os.path.join(VECTORSTORE_DIR, ADMIN_VECTORSTORE_NAME)
embedding_model = HuggingFaceEmbeddings(model_name="sentence-transformers/all-MiniLM-L6-v2")


UPLOAD_DIR = os.path.join(settings.MEDIA_ROOT, "uploaded_files") if settings.MEDIA_ROOT else None # ✅ Absolute Path
AIRFLOW_API_URL = "http://139.59.12.79:8080//api/v1/dags"

AIRFLOW_AUTH = ("airflow", "airflow")

DAG_ID = "data_processing_pipeline"

def ensure_dirs():
    """
    Create MEDIA_ROOT subfolders if they don't exist.
    This does NOT create a FAISS index; it only ensures directories exist.
    """
    for d in [settings.MEDIA_ROOT, UPLOAD_DIR_PDF, VECTORSTORE_DIR, ADMIN_VECTORSTORE_PATH]:
        os.makedirs(d, exist_ok=True)

def faiss_files_exist(path: str) -> bool:
    """
    Check for FAISS artifacts expected by LangChain's FAISS.load_local
    (filenames vary by lc version, but these are the common ones).
    """
    candidates = [
        os.path.join(path, "index.faiss"),
        os.path.join(path, "index.pkl"),      # metadata/docstore
        os.path.join(path, "docstore.pkl"),   # older variants
    ]
    return any(os.path.exists(p) for p in candidates)

from rest_framework.views import APIView
from rest_framework.response import Response
import psycopg2

import uuid
import json
import traceback
from urllib.parse import quote_plus
import time
import threading
from sqlalchemy import create_engine, text, event, Engine
from sqlalchemy.pool import QueuePool, NullPool
from contextlib import contextmanager
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt


POSTGRES_USER = "appadmin"
POSTGRES_PASSWORD = "prowesstics"
POSTGRES_HOST = "20.244.31.74"  # e.g. "localhost" or IP
POSTGRES_PORT = 5432
POSTGRES_DB = "liberty_updated"


ENCODED_PASSWORD = quote_plus(POSTGRES_PASSWORD)
DATABASE_URL = (
    f"postgresql+psycopg2://{POSTGRES_USER}:{ENCODED_PASSWORD}"
    f"@{POSTGRES_HOST}:{POSTGRES_PORT}/{POSTGRES_DB}"
)

print(f"🔗 Database URL: postgresql+psycopg2://{POSTGRES_USER}:****@{POSTGRES_HOST}:{POSTGRES_PORT}/{POSTGRES_DB}")

# ============================================================
#FIXED: AGGRESSIVE CONNECTION POOL SETTINGS
# ============================================================

ENGINE = create_engine(
    DATABASE_URL,
    
    # FIXED: More aggressive pool settings
    poolclass=QueuePool,
    pool_size=30,                    # 20 persistent connections
    max_overflow=50,                 # 30 additional connections during spikes
    pool_timeout=60,                 # Wait 45s for available connection
    pool_recycle=900,               # Recycle connections every 30 min
    pool_pre_ping=True,              # ✅ Test connections before use (prevents stale connections)
    
    # Execution Settings
    echo=False,                      # Set True for SQL debugging
    echo_pool=False,                 # Set True for pool debugging
    future=True,                     # Use SQLAlchemy 2.0 style
    
    # Connection Arguments (PostgreSQL-specific)
    connect_args={
         "connect_timeout": 20,                           # Increased from 10s
        "options": "-c statement_timeout=120000",        # 2 minutes (increased from 10 min default)
        "keepalives": 1,                                 # Enable TCP keepalives
        "keepalives_idle": 30,                           # Start keepalive after 30s idle
        "keepalives_interval": 10,                       # Send keepalive every 10s
        "keepalives_count": 5,                           # 5 failed keepalives = dead connection
        "application_name": "ProChurn_AI"                # Shows in pg_stat_activity
    }
)

print(f"✅ Database engine initialized (PRODUCTION MODE):")
print(f"   - Pool: 30 base + 50 overflow = 80 max connections")
print(f"   - Pool timeout: 60 seconds")
print(f"   - Connection timeout: 20 seconds")
print(f"   - Query timeout: 2 minutes")


# ============================================================

# FIXED: CONNECTION LIFECYCLE - SHORTER TIMEOUTS PER CONNECTION
# ============================================================

@event.listens_for(Engine, "connect")
def receive_connect(dbapi_conn, connection_record):
    """Set aggressive timeouts on each new connection"""
    print(f"🔌 New database connection created")
    
    with dbapi_conn.cursor() as cursor:
        # FIXED: Set reasonable per-query limits
        cursor.execute("SET statement_timeout = '120s'")     # 2 minutes max per query
        cursor.execute("SET lock_timeout = '10s'")           # Don't wait long for locks
        cursor.execute("SET idle_in_transaction_session_timeout = '30s'")  # Kill idle transactions


@event.listens_for(Engine, "checkout")
def receive_checkout(dbapi_conn, connection_record, connection_proxy):
    """Monitor pool utilization"""
    try:
        pool_obj = ENGINE.pool
        checked_out = pool_obj.checkedout()
        total_size = pool_obj.size() + pool_obj.overflow()
        
        # Warn if pool is filling up
        if checked_out > total_size * 0.9:
            print(f"🚨 CRITICAL: Pool at {checked_out}/{total_size} connections ({(checked_out/total_size)*100:.0f}%)")
        elif checked_out > total_size * 0.7:
            print(f"⚠️ WARNING: Pool at {checked_out}/{total_size} connections ({(checked_out/total_size)*100:.0f}%)")
        
    except Exception as e:
        print(f"⚠️ Pool status check failed: {e}")





# ============================================================
# SESSION MANAGEMENT
# ============================================================

# Store sessions (maps session_id -> user metadata)
# NOTE: We DON'T store engines here anymore - use the global ENGINE
session_store = {}
conversation_memory_store = {}



from urllib.parse import quote_plus
from sqlalchemy import create_engine


def ensure_session(session_id: str, user_id: str = "admin"):
    """Initialize session metadata (uses global ENGINE)"""
    
    if session_id not in session_store:
        session_store[session_id] = {
            "user_id": user_id,
            "created_at": time.time()
        }
        conversation_memory_store.setdefault(session_id, [])
        print(f"✅ New session: {session_id} (user: {user_id})")
    
    return ENGINE




# ============================================================
# FIXED: SAFE CONNECTION WITH PROPER ERROR HANDLING
# ============================================================

class DatabaseTimeoutError(Exception):
    """Raised when database operation exceeds timeout"""
    pass


class DatabaseBusyError(Exception):
    """Raised when database is under heavy load"""
    pass


@contextmanager
def get_db_connection(timeout_seconds=45):
    """
    FIXED: Proper connection handling with detailed error messages
    
    Args:
        timeout_seconds: How long to wait for a connection from pool
    
    Raises:
        DatabaseBusyError: When pool is exhausted
        DatabaseTimeoutError: When connection acquisition times out
    """
    conn = None
    start_time = time.time()
    
    try:
        # Get pool status BEFORE attempting connection
        pool_status = get_pool_status()
        utilization = pool_status.get('utilization_pct', 0)
        
        # Early warning if pool is stressed
        if utilization > 85:
            print(f"⚠️ High pool utilization ({utilization}%) - connection may be slow")
        
        # Try to acquire connection with timeout
        try:
            conn = ENGINE.connect()
            acquisition_time = time.time() - start_time
            
            if acquisition_time > 5:
                print(f"⚠️ Slow connection acquisition: {acquisition_time:.2f}s")
            
            yield conn
            
        except Exception as e:
            acquisition_time = time.time() - start_time
            error_msg = str(e).lower()
            
            # Classify the error
            if 'timeout' in error_msg or acquisition_time >= timeout_seconds:
                pool_status = get_pool_status()
                print(f"❌ Connection timeout after {acquisition_time:.1f}s")
                print(f"🔌 Pool: {pool_status['checked_out']}/{pool_status['max_connections']} in use")
                
                raise DatabaseBusyError(
                    "Database is busy. Please retry in 30 seconds."
                )
            else:
                # Other database error
                raise
    
    finally:
        if conn:
            try:
                # Explicitly rollback any pending transaction
                if conn.in_transaction():
                    conn.rollback()
                
                # Return to pool
                conn.close()
                
            except Exception as e:
                print(f"⚠️ Error during connection cleanup: {e}")
                try:
                    conn.close()
                except:
                    pass


# ============================================================
# FIXED: QUERY EXECUTION WITH BETTER TIMEOUT HANDLING
# ============================================================

def execute_query_with_timeout(query_text, timeout_seconds=110):
    """
    FIXED: Execute query with proper timeout and error handling
    
    Args:
        query_text: SQL query to execute
        timeout_seconds: Max time for query execution (default 110s = under 2min limit)
    
    Returns:
        List of result rows
    
    Raises:
        DatabaseTimeoutError: Query exceeded time limit
        DatabaseBusyError: Could not get connection
    """
    
    print(f"🔧 Executing SQL (timeout={timeout_seconds}s)")
    print(f"🔍 SQL Preview: {query_text[:200]}...")
    
    start_time = time.time()
    
    try:
        # Use the safe connection context manager
        with get_db_connection(timeout_seconds=45) as conn:
            
            # Execute with explicit timeout
            result = conn.execute(
                text(f"SET LOCAL statement_timeout = '{timeout_seconds}s'; {query_text}")
            )
            
            rows = result.fetchall()
            execution_time = time.time() - start_time
            
            print(f"✅ Query completed in {execution_time:.2f}s, returned {len(rows)} rows")
            
            return rows
    
    except DatabaseBusyError:
        # Re-raise as-is
        raise
    
    except Exception as e:
        execution_time = time.time() - start_time
        error_msg = str(e).lower()
        
        # Check if it's a query timeout
        if 'timeout' in error_msg or 'canceling statement' in error_msg:
            print(f"❌ Query timeout after {execution_time:.1f}s")
            raise DatabaseTimeoutError(
                f"Query exceeded {timeout_seconds}s time limit. "
                "Try simplifying your query or adding filters."
            )
        else:
            print(f"❌ Query failed after {execution_time:.1f}s: {str(e)}")
            raise

# ============================================================
# POOL MONITORING
# ============================================================

def get_pool_status():
    """Get current connection pool status"""
    try:
        pool_obj = ENGINE.pool
        size = pool_obj.size()
        checked_out = pool_obj.checkedout()
        overflow = pool_obj.overflow()
        max_connections = size + pool_obj._max_overflow
        
        return {
            "size": size,
            "checked_out": checked_out,
            "overflow": overflow,
            "available": size - checked_out + overflow,
            "max_connections": max_connections,
            "utilization_pct": round((checked_out / max_connections) * 100, 1) if max_connections > 0 else 0
        }
    except Exception as e:
        return {"error": str(e)}


def print_pool_status():
    """Print formatted pool status"""
    status = get_pool_status()
    if "error" in status:
        print(f"⚠️ Unable to get pool status: {status['error']}")
    else:
        print(f"🔌 Pool Status: {status['checked_out']}/{status['max_connections']} "
              f"({status['utilization_pct']}% utilization)")

# @csrf_exempt
# def ask_qwen(request):
#     try:
#         data = json.loads(request.body)
#         session_id = data.get("session_id", "").strip()
#         question = data.get("question", "").strip()

#         if not question:
#             return JsonResponse({'error': 'Question is required'}, status=400)
#         if not session_id:
#             return JsonResponse({'error': 'Session ID is required'}, status=400)

#         logger.info(f"Processing question for session {session_id}: {question[:100]}")

#         df = dataframe_map.get(session_id)
#         has_data = df is not None

#         if has_data:
#             TOKEN_BUDGETS = {
#                 'question': len(question.split()) * 1.5,
#                 'data_overview': 800,
#                 'memory_context': 1000,
#                 'semantic_context': 4000,
#                 'prompt_template': 1000,
#                 'response_buffer': 8000,
#                 'safety_margin': 2000
#             }

#             semantic_context = ""
#             memory_context = ""
#             data_overview = ""

#             prompt_parts = [
#                 "You are an expert data analyst AI that provides accurate answers based on uploaded datasets.",
#                 "",
#                 "IMPORTANT: Use ONLY the provided data. Never make assumptions or use external knowledge.",
#                 "",
#                 "### Dataset Overview:",
#                 data_overview,
#                 "",
#                 "### Previous Conversation Context:",
#                 memory_context or "No previous context.",
#                 "",
#                 "### Relevant Data Chunks:",
#                 semantic_context or "No relevant chunks found.",
#                 "",
#                 f"### User Question:",
#                 question,
#                 "",
#                 "### Instructions:",
#                 "1. Analyze the question carefully",
#                 "2. Use only the provided data chunks and dataset information",
#                 "3. If you need to perform calculations, show your work",
#                 "4. If the data doesn't contain enough information to answer, say so clearly",
#                 "5. Provide specific numbers, values, and examples from the actual data",
#                 "6. Be precise and factual - no guessing or assumptions",
#                 "7. Keep your answer concise but comprehensive",
#                 "",
#                 "Answer:"
#             ]

#             prompt = "\n".join(prompt_parts)
#             prompt_tokens = len(prompt.split()) * 1.5
#             logger.info(f"Final prompt: {prompt_tokens} tokens")

#             answer = call_llm_with_retry(prompt)
#             if not answer:
#                 answer = "I couldn't find an answer based on the uploaded data."

#             return JsonResponse({
#                 "question": question,
#                 "answer": answer,
#                 "session_id": session_id,
#                 "chunks_used": 0,
#                 "prompt_tokens": prompt_tokens,
#                 "timestamp": datetime.now().isoformat(),
#                 "success": True,
#                 "has_chart": False
#             })

#         fallback_prompt = f"""You are an intelligent assistant. Answer the following question as accurately and helpfully as possible.

# Question: {question}

# Answer:"""

#         answer = call_llm_with_retry(fallback_prompt)
#         if not answer:
#             answer = "Sorry, I couldn't generate a helpful answer. Please try rephrasing."

#         return JsonResponse({
#             "question": question,
#             "answer": answer,
#             "session_id": session_id,
#             "chunks_used": 0,
#             "timestamp": datetime.now().isoformat(),
#             "success": True,
#             "has_chart": False
#         })

#     except Exception as e:
#         logger.error(f"Exception in ask_qwen: {traceback.format_exc()}")
#         return JsonResponse({
#             'error': f'Processing error: {str(e)}',
#             'success': False,
#             'timestamp': datetime.now().isoformat(),
#             'has_chart': False
#         }, status=500)


@csrf_exempt
def ask_qwen(request):
    try:
        data = json.loads(request.body)
        # session_id = data.get("session_id", "").strip()
        question = data.get("question", "").strip()

        if not question:
            return JsonResponse({'error': 'Question is required'}, status=400)
        # if not session_id:
        #     return JsonResponse({'error': 'Session ID is required'}, status=400)

        # logger.info(f"Processing question for session {session_id}: {question[:100]}")

        # >>> NEW: sessionless shim (keep session_id field in responses but not used for lookup)
        session_id = "sessionless"

        df = dataframe_map.get(session_id)
        # >>> NEW: sessionless selection — if no df under the dummy key, use the most recently loaded df
        if df is None:
            try:
                df = list(dataframe_map.values())[-1] if dataframe_map else None
            except Exception:
                df = None

        has_data = df is not None

        if has_data:
            TOKEN_BUDGETS = {
                'question': len(question.split()) * 1.5,
                'data_overview': 800,
                'memory_context': 1000,
                'semantic_context': 4000,
                'prompt_template': 1000,
                'response_buffer': 8000,
                'safety_margin': 2000
            }

            semantic_context = ""
            memory_context = ""
            data_overview = ""

            prompt_parts = [
                "You are an expert data analyst AI that provides accurate answers based on uploaded datasets.",
                "",
                "IMPORTANT: Use ONLY the provided data. Never make assumptions or use external knowledge.",
                "",
                "### Dataset Overview:",
                data_overview,
                "",
                "### Previous Conversation Context:",
                memory_context or "No previous context.",
                "",
                "### Relevant Data Chunks:",
                semantic_context or "No relevant chunks found.",
                "",
                f"### User Question:",
                question,
                "",
                "### Instructions:",
                "1. Analyze the question carefully",
                "2. Use only the provided data chunks and dataset information",
                "3. If you need to perform calculations, show your work",
                "4. If the data doesn't contain enough information to answer, say so clearly",
                "5. Provide specific numbers, values, and examples from the actual data",
                "6. Be precise and factual - no guessing or assumptions",
                "7. Keep your answer concise but comprehensive",
                "",
                "Answer:"
            ]

            prompt = "\n".join(prompt_parts)
            prompt_tokens = len(prompt.split()) * 1.5
            logger.info(f"Final prompt: {prompt_tokens} tokens")

            answer = call_llm_with_retry(prompt)
            if not answer:
                answer = "I couldn't find an answer based on the uploaded data."

            return JsonResponse({
                "question": question,
                "answer": answer,
                "session_id": session_id,  # retained for compatibility; value is "sessionless"
                "chunks_used": 0,
                "prompt_tokens": prompt_tokens,
                "timestamp": datetime.now().isoformat(),
                "success": True,
                "has_chart": False
            })

        fallback_prompt = f"""You are an intelligent assistant. Answer the following question as accurately and helpfully as possible.

Question: {question}

Answer:"""

        answer = call_llm_with_retry(fallback_prompt)
        if not answer:
            answer = "Sorry, I couldn't generate a helpful answer. Please try rephrasing."

        return JsonResponse({
            "question": question,
            "answer": answer,
            "session_id": session_id,  # retained for compatibility; value is "sessionless"
            "chunks_used": 0,
            "timestamp": datetime.datetime.now().isoformat(),
            "success": True,
            "has_chart": False
        })

    except Exception as e:
        logger.error(f"Exception in ask_qwen: {traceback.format_exc()}")
        return JsonResponse({
            'error': f'Processing error: {str(e)}',
            'success': False,
            'timestamp': datetime.datetime.now().isoformat(),
            'has_chart': False
        }, status=500)


# def call_llm_with_retry(prompt: str, max_retries: int = 3, base_delay: float = 1.0) -> str:
#     # Rough token estimate; keep margin below model context
#     estimated_tokens = int(len(prompt.split()) * 1.5)
#     logger.info(f"Estimated prompt tokens: {estimated_tokens}")
#     if estimated_tokens > MAX_PROMPT_TOKENS:
#         logger.error(f"Prompt too long: {estimated_tokens} tokens (max: {MAX_PROMPT_TOKENS})")
#         return "The prompt is too long for the current model. Please try with a shorter question."

#     logger.info(f"Starting Azure Inference call with {max_retries} max retries")

#     for attempt in range(max_retries):
#         try:
#             logger.info(f"Attempt {attempt + 1}/{max_retries}")
#             resp = _get_client().complete(
#                 messages=[
#                     SystemMessage(content=_SYSTEM_PROMPT),
#                     UserMessage(content=prompt),
#                 ],
#                 model=AZURE_MODEL,
#                 temperature=0.1,
#                 top_p=0.9,
#                 frequency_penalty=0.0,
#                 presence_penalty=0.0,
#                 max_tokens=4000,
#             )

#             if resp.choices:
#                 content = (resp.choices[0].message.content or "").strip()
#                 if content:
#                     return content
#             logger.warning("Empty response from Azure Inference; retrying if attempts remain.")

#         except HttpResponseError as e:
#             status = getattr(e, "status_code", None)
#             logger.warning(f"Azure HttpResponseError (status={status}): {e}")
#             if status in (429, 502, 503, 504) and attempt < max_retries - 1:
#                 time.sleep(base_delay * (2 ** attempt))
#                 continue
#             break  # non-retryable or retries exhausted

#         except (ServiceRequestError, ServiceResponseError, TimeoutError, ConnectionError) as e:
#             logger.warning(f"Transient Azure error: {e}")
#             if attempt < max_retries - 1:
#                 time.sleep(base_delay * (2 ** attempt))
#                 continue
#             break

#         except Exception as e:
#             logger.warning(f"Retry {attempt + 1} failed: {e}")
#             if attempt < max_retries - 1:
#                 time.sleep(base_delay * (2 ** attempt))
#                 continue
#             break

#     logger.error("All retry attempts failed")
#     return "I'm currently experiencing issues reaching the AI service. Please try again later."

# def call_llm_with_retry(prompt: str, max_retries: int = 3, base_delay: float = 1.0) -> str:
#     if not GROQ_API_KEY:
#         logger.error("GROQ_API_KEY is not defined")
#         return "API key configuration error. Please check your settings."

#     estimated_tokens = len(prompt.split()) * 1.5
#     logger.info(f"Estimated prompt tokens: {estimated_tokens}")

#     if estimated_tokens > 220000:
#         logger.error(f"Prompt too long: {estimated_tokens} tokens (max: 220000)")
#         return "The prompt is too long for the current model. Please try with a shorter question."

#     headers = {
#         "Authorization": f"Bearer {GROQ_API_KEY}",
#         "Content-Type": "application/json",
#     }

#     model = "meta-llama/llama-4-maverick-17b-128e-instruct"

#     logger.info(f"Starting Groq LLM call with {max_retries} max retries")

#     for attempt in range(max_retries):
#         try:
#             logger.info(f"Attempt {attempt + 1}/{max_retries}")
#             payload = {
#                 "model": model,
#                 "messages": [{"role": "user", "content": prompt}],
#                 "temperature": 0.1,
#                 "top_p": 0.9,
#                 "frequency_penalty": 0.0,
#                 "presence_penalty": 0.0,
#                 "max_tokens": 4000
#             }
#             response = requests.post("https://api.groq.com/openai/v1/chat/completions", headers=headers, json=payload, timeout=60)
#             logger.info(f"Response status code: {response.status_code}")

#             if response.status_code == 200:
#                 result = response.json()
#                 if 'choices' in result and result['choices']:
#                     answer = result['choices'][0].get('message', {}).get('content', '').strip()
#                     if answer:
#                         return answer
#             elif response.status_code in [429, 502, 503, 504]:
#                 time.sleep(base_delay * (2 ** attempt))
#                 continue

#         except Exception as e:
#             logger.warning(f"Retry {attempt+1} failed: {e}")
#             time.sleep(base_delay)

#     logger.error("All retry attempts failed")
#     return "I'm currently experiencing issues reaching the AI service. Please try again later."



def call_llm_with_retry(prompt: str, max_retries: int = 3, base_delay: float = 1.0) -> str:
    # Estimate prompt tokens (rough)
    estimated_tokens = int(len(prompt.split()) * 1.5)
    logger.info(f"Estimated prompt tokens: {estimated_tokens}")
    MAX_PROMPT_TOKENS = 120_000  # adjust if your deployed model supports a different context length
    if estimated_tokens > MAX_PROMPT_TOKENS:
        logger.error(f"Prompt too long: {estimated_tokens} tokens (max: {MAX_PROMPT_TOKENS})")
        return "The prompt is too long for the current model. Please try with a shorter question."

    logger.info(f"Starting Azure Inference call with {max_retries} max retries")

    for attempt in range(max_retries):
        try:
            logger.info(f"Attempt {attempt + 1}/{max_retries}")
            resp = _get_client().complete(
                messages=[UserMessage(content=prompt)],
                model=AZURE_MODEL,
                temperature=0.1,
                top_p=0.9,
                frequency_penalty=0.0,
                presence_penalty=0.0,
                max_tokens=4000,
            )
            if resp.choices:
                content = getattr(resp.choices[0].message, "content", "") or ""
                content = content.strip()
                if content:
                    return content
            logger.warning("Empty response from Azure Inference; retrying if attempts remain.")

        except HttpResponseError as e:
            status = getattr(e, "status_code", None)
            logger.warning(f"Azure HttpResponseError (status={status}): {e}")
            if status in (429, 502, 503, 504) and attempt < max_retries - 1:
                time.sleep(base_delay * (2 ** attempt))
                continue
            break  # non-retryable or retries exhausted

        except (ServiceRequestError, ServiceResponseError, TimeoutError, ConnectionError) as e:
            logger.warning(f"Transient Azure error: {e}")
            if attempt < max_retries - 1:
                time.sleep(base_delay * (2 ** attempt))
                continue
            break

        except Exception as e:
            logger.warning(f"Retry {attempt + 1} failed: {e}")
            if attempt < max_retries - 1:
                time.sleep(base_delay * (2 ** attempt))
                continue
            break

    logger.error("All retry attempts failed")
    return "I'm currently experiencing issues reaching the AI service. Please try again later."

@csrf_exempt
def connect_database(request):
    """Database connection endpoint"""
    try:
        try:
            data = json.loads(request.body.decode("utf-8") or '{}')
        except json.JSONDecodeError:
            data = {}

        user_id = data.get("user_id", "admin")
        session_id = data.get("session_id") or str(uuid.uuid4())
        
        print(f"📞 connect_database: session={session_id}, user={user_id}")
        
        # Test connection
        with get_db_connection(timeout_seconds=30) as conn:
            conn.execute(text("SELECT 1"))
            print(f"✅ Database connectivity verified")
        
        # Create session metadata
        session_store[session_id] = {
            "user_id": user_id,
            "created_at": time.time()
        }
        conversation_memory_store[session_id] = []
        
        pool_status = get_pool_status()
        
        return JsonResponse({
            "message": "Connected successfully",
            "session_id": session_id,
            "user_id": user_id,
            "pool_status": pool_status
        })

    except DatabaseBusyError as e:
        return JsonResponse({
            "error": str(e),
            "retry_after": 30
        }, status=503)
    
    except Exception as e:
        print(f"❌ connect_database failed: {str(e)}")
        traceback.print_exc()
        return JsonResponse({
            "error": str(e),
            "details": "Database connection failed"
        }, status=500)



@csrf_exempt
def connect_database3012(request):
    try:
        # Optional user_id from frontend, fallback to 'admin'
        try:
            data = json.loads(request.body.decode("utf-8") or '{}')
        except json.JSONDecodeError:
            data = {}

        user_id = data.get("user_id", "admin")

        # Encode password and form connection string
        encoded_pwd = quote_plus(POSTGRES_PASSWORD)
        conn_str = f"postgresql://{POSTGRES_USER}:{encoded_pwd}@{POSTGRES_HOST}:{POSTGRES_PORT}/{POSTGRES_DB}"
        engine = create_engine(conn_str)

        # Test DB connection
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))

        # Extract schema and initialize session
        schema_text = extract_schema_from_sqlalchemy(engine)
        session_id = data.get("session_id", str(uuid.uuid4()))
        session_store[session_id] = {"engine": engine, "user_id": user_id}
        conversation_memory_store[session_id] = []

        # Embed schema for this session
        embed_schema_for_user(user_id=user_id, db_id=session_id, schema_text=schema_text)

        return JsonResponse({
            "message": "Connected and embedded successfully (static credentials)",
            "session_id": session_id
        })

    except Exception as e:
        print("❌ connect_database failed:", str(e))
        traceback.print_exc()
        return JsonResponse({"error": str(e)}, status=500)
    




# ============================================================
# HEALTH CHECK ENDPOINT (Optional but recommended)
# ============================================================

@csrf_exempt
def database_health_check(request):
    """Health check endpoint"""
    try:
        with get_db_connection(timeout_seconds=10) as conn:
            result = conn.execute(text("SELECT version()"))
            version = result.scalar()
        
        pool_status = get_pool_status()
        utilization = pool_status.get("utilization_pct", 0)
        
        if utilization > 90:
            health = "critical"
        elif utilization > 75:
            health = "warning"
        else:
            health = "healthy"
        
        return JsonResponse({
            "status": health,
            "database": "connected",
            "version": version,
            "pool": pool_status,
            "timestamp": time.time()
        })
        
    except Exception as e:
        return JsonResponse({
            "status": "unhealthy",
            "error": str(e),
            "timestamp": time.time()
        }, status=503)






from urllib.parse import quote_plus
from sqlalchemy import create_engine

# @csrf_exempt

# def check_intent(request):
#     if request.method == 'POST':
#         data = json.loads(request.body)
#         question = data.get("question", "")

#         prompt = f"""
# Classify the user's intent strictly as YES or NO.

# If the question is a general greeting, chit-chat, or general knowledge (e.g. "hi", "hello", "how are you", "who is the PM of India"), respond with NO.

# If it is about querying the connected database schema or fetching data from tables, respond with YES.
# If the question is trying to analyze, query, or summarize tabular or structured data (e.g. Excel, PDF tables, database), respond with YES.

# Question: "{question}"
# Only respond YES or NO.
# """

#         try:
#             response = requests.post(
#                 "https://api.groq.com/openai/v1/chat/completions",  # ✅ Groq Cloud endpoint
#                 headers={
#                     "Authorization": f"Bearer {GROQ_API_KEY}",  # ✅ Use your Groq API key
#                     "Content-Type": "application/json"
#                 },
#                 json={
#                     "model": "meta-llama/llama-4-maverick-17b-128e-instruct",  # ✅ Groq model
#                     "messages": [
#                         {"role": "system", "content": "You are an intent classifier."},
#                         {"role": "user", "content": prompt}
#                     ],
#                     "temperature": 0.0,
#                     "max_tokens": 10
#                 },
#                 timeout=30
#             )
#             result = response.json()
#             answer = result["choices"][0]["message"]["content"].strip()
#             return JsonResponse({"answer": answer})

#         except Exception as e:
#             return JsonResponse({"answer": "No", "error": str(e)}, status=500)



# @csrf_exempt
# def check_intent(request):
#     if request.method == 'POST':
#         data = json.loads(request.body)
#         question = data.get("question", "")

#         prompt = f"""
# Classify the user's intent strictly as YES or NO.

# If the question is a general greeting, chit-chat, or general knowledge (e.g. "hi", "hello", "how are you", "who is the PM of India"), respond with NO.

# If it is about querying the connected database schema or fetching data from tables, respond with YES.
# If the question is trying to analyze, query, or summarize tabular or structured data (e.g. Excel, PDF tables, database), respond with YES.

# Question: "{question}"
# Only respond YES or NO.
# """

#         try:
#             response = requests.post(
#                 "https://api.groq.com/openai/v1/chat/completions",  # ✅ Groq Cloud endpoint
#                 headers={
#                     "Authorization": f"Bearer {GROQ_API_KEY}",  # ✅ Use your Groq API key
#                     "Content-Type": "application/json"
#                 },
#                 json={
#                     "model": "meta-llama/llama-4-maverick-17b-128e-instruct",  # ✅ Groq model
#                     "messages": [
#                         {"role": "system", "content": "You are an intent classifier."},
#                         {"role": "user", "content": prompt}
#                     ],
#                     "temperature": 0.0,
#                     "max_tokens": 10
#                 },
#                 timeout=30
#             )
#             result = response.json()
#             answer = result["choices"][0]["message"]["content"].strip()
#             return JsonResponse({"answer": answer})

#         except Exception as e:
#             return JsonResponse({"answer": "No", "error": str(e)}, status=500)



# @csrf_exempt
# def check_intent(request):
#     if request.method != "POST":
#         return JsonResponse({"answer": "NO", "error": "Only POST allowed"}, status=405)

#     try:
#         data = json.loads(request.body or "{}")
#     except Exception:
#         return JsonResponse({"answer": "NO", "error": "Invalid JSON body"}, status=400)

#     question = (data.get("question") or "").strip()
#     if not question:
#         return JsonResponse({"answer": "NO"})

#     prompt = f"""
# Classify the user's intent strictly as YES or NO.

# If the question is a general greeting, chit-chat, or general knowledge (e.g., "hi", "hello", "how are you", "who is the PM of India"), respond with NO.

# If it is about querying the connected database schema or fetching data from tables, respond with YES.
# If the question is trying to analyze, query, or summarize tabular or structured data (e.g., Excel, PDF tables, database), respond with YES.

# Question: "{question}"
# Only respond YES or NO.
# """.strip()

#     try:
#         resp = _get_client().complete(
#             messages=[
#                 SystemMessage(content="You are an intent classifier. Reply only YES or NO."),
#                 UserMessage(content=prompt),
#             ],
#             model=AZURE_MODEL,
#             temperature=0.0,
#             top_p=1.0, 
#             max_tokens=3,
#         )
#         raw = resp.choices[0].message.content if resp.choices else ""
#         answer = _yes_no(raw)
#         return JsonResponse({"answer": answer})
#     except (HttpResponseError, ServiceRequestError, ServiceResponseError, TimeoutError, ConnectionError) as e:
#         logger.warning(f"Azure inference error: {e}")
#         return JsonResponse({"answer": "NO", "error": str(e)}, status=502)
#     except Exception as e:
#         logger.exception("check_intent failed")
#         return JsonResponse({"answer": "NO", "error": str(e)}, status=500)

FULL_SCHEMA = '''
Table: "bi_dwh"."main_cai_lib"
Columns:
- own_damage_premium
- vehicle_age
- third_party_premium
- total_premium_payable
- vehicle_idv
- total_revenue
- policy_tenure
- number_of_claims
- claims_approved
- claim_approval_rate
- customer_tenure
- customer_life_time_value
- customerid
- chassis_number
- engine_number
- vehicle_register_number
- state
- zone
- business_type
- car_manufacturer
- vehicle_model
- product_name
- policy_no
- tie_up
- vehicle_model_variant
- policy_start_date_year
- policy_end_date_year
- policy_start_date_month
- policy_end_date_month
- is_churn
- customer_segment
- branch_name
- main_churn_reason
- primary_recommendation
- insured_client_name
'''



def validate_sql_columns(sql: str, valid_columns: set) -> set:
    tokens = sqlparse.parse(sql)[0].tokens
    words = set()
    
    for token in tokens:
        for t in token.flatten():
            if t.ttype is None and t.value not in ('SELECT', 'FROM', 'WHERE', 'AND', 'OR', 'COUNT', '(', ')', '=', 'IS', 'NOT', 'NULL'):
                words.add(t.value)
    
    return words - valid_columns


def extract_sql_block22(text):
    if isinstance(text, tuple):
        text = text[0]
    if not isinstance(text, str):
        # print("" extract_sql_block received non-string:", type(text))
        return ""

    match = re.search(r"```sql\s+(.*?)```", text, re.DOTALL | re.IGNORECASE)
    if match:
        return match.group(1).strip()

    match = re.search(r"```(.*?)```", text, re.DOTALL)
    if match:
        return match.group(1).strip()

    match = re.search(r"(SELECT\s+.+?)(;|\Z)", text, re.IGNORECASE | re.DOTALL)
    if match:
        return match.group(1).strip()

    print("⚠️ No valid SQL block found. Returning original text.")
    return text.strip()



def extract_columns_from_schema(schema: str) -> set:
    return set(re.findall(r'- ([a-zA-Z0-9_]+)', schema))



def extract_column_names(schema: str) -> set:
    lines = schema.splitlines()
    columns = set()

    for line in lines:
        match = re.match(r'^\s*-\s+([a-zA-Z_][a-zA-Z0-9_]*)', line)
        if match:
            columns.add(match.group(1))
    
    return columns


VALID_COLUMNS = extract_columns_from_schema(FULL_SCHEMA)


print(VALID_COLUMNS)


def llm_generate_recommendationgroq(question, rows):
    
    preview_rows = rows[:10]
    table_str = "\n".join(", ".join(f"{k}: {v}" for k, v in row.items()) for row in preview_rows)

    system_prompt = (
    "You are a data-driven business analyst assistant. Based on the user's question and the query output data, "
    "generate a concise, actionable, and professional business recommendation. "
    "Focus on explaining the key trends, risks, or opportunities in simple business terms, "
    "helping the user understand what actions or decisions they can take next. "
    "Strictly avoid including any SQL code, technical jargon, or step-by-step query explanations. "
    "Only provide a high-level insight that would help a manager or decision-maker."
)


    user_prompt = f"""\
User Question: {question}

Data Preview (first 10 rows):
{table_str}

Please provide only a one-paragraph business recommendation. Do not include SQL queries or explanations.

"""

    # API headers
    GROQ_API_KEY = os.getenv("GROQ_API_KEY")  # Ensure this is set in your environment
    headers = {
        "Authorization": f"Bearer {GROQ_API_KEY}",
        "Content-Type": "application/json"
    }

    
    full_prompt = system_prompt + "\n\n" + user_prompt
    word_limit = 5000
    if len(full_prompt.split()) > word_limit:
        logger.warning("Prompt too large, truncating...")
        user_prompt = user_prompt[:8000]
        system_prompt = system_prompt[:2000]

    messages = []
    if system_prompt:
        messages.append({"role": "system", "content": system_prompt})
    messages.append({"role": "user", "content": user_prompt})

    payload = {
        "model": "meta-llama/llama-4-maverick-17b-128e-instruct",
        "messages": messages,
        "temperature": 0.1,
        "max_tokens": 1000
    }

    GROQ_BASE_URL = "https://api.groq.com/openai/v1"

    max_retries = 3
    for attempt in range(max_retries):
        try:
            response = requests.post(
                f"{GROQ_BASE_URL}/chat/completions",
                headers=headers,
                json=payload,
                timeout=30
            )

            if response.status_code == 200:
                raw_text = response.json()["choices"][0]["message"]["content"].strip()

              
                cleaned = re.sub(r"```sql.*?```", "", raw_text, flags=re.DOTALL | re.IGNORECASE)
                cleaned = re.sub(r"```.*?```", "", cleaned, flags=re.DOTALL)
                cleaned = re.sub(r"`[^`]+`", "", cleaned)  # Optional: inline `code` like `column_name`
                
                return cleaned.strip()

             

            elif response.status_code == 429:
                try:
                    error_data = response.json()
                    message = error_data.get("error", {}).get("message", "")
                    logger.warning(f"Rate limit hit: {message}")
                    match = re.search(r'try again in ([\\d\\.]+)s', message)
                    if match:
                        wait_time = float(match.group(1))
                        logger.info(f"Sleeping for {wait_time} seconds due to rate limit...")
                        time.sleep(wait_time)
                    else:
                        time.sleep(2 ** attempt)
                    continue
                except Exception as parse_error:
                    logger.error(f"Error parsing rate limit retry time: {parse_error}")
                    time.sleep(2 ** attempt)
                    continue

            else:
                logger.error(f"Groq Cloud API error: {response.status_code} - {response.text}")
                break

        except Exception as e:
            logger.error(f"Groq Cloud API call failed: {e}")
            time.sleep(2 ** attempt)

    return "I apologize, but I'm having trouble processing your request right now."



def _clean_model_text(text: str) -> str:
    if not text:
        return ""
    # Strip SQL/code blocks and inline code
    text = re.sub(r"sql.*?", "", text, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r".*?", "", text, flags=re.DOTALL)
    text = re.sub(r"[^]+`", "", text)
    # Collapse whitespace
    text = re.sub(r"\s+", " ", text).strip()
    return text


# def llm_generate_recommendation(question: str, rows: List[Dict[str, Any]]) -> str:
#     preview_rows = rows[:10] if rows else []
#     if isinstance(preview_rows, list) and preview_rows and isinstance(preview_rows[0], dict):
#         table_str = "\n".join(", ".join(f"{k}: {v}" for k, v in row.items()) for row in preview_rows)
#     else:
#         table_str = "No rows available."

#     system_prompt = (
#         "You are a data-driven business analyst assistant. Based on the user's question and the query output data, "
#         "generate a concise, actionable, and professional business recommendation. "
#         "Focus on explaining the key trends, risks, or opportunities in simple business terms, "
#         "helping the user understand what actions or decisions they can take next. "
#         "Strictly avoid including any SQL code, technical jargon, or step-by-step query explanations. "
#         "Only provide a high-level insight that would help a manager or decision-maker."
#     )

#     user_prompt = f"""User Question: {question}

# Data Preview (first 10 rows):
# {table_str}

# Please provide only a one-paragraph business recommendation. Do not include SQL queries or explanations.
# """

#     # Size guard (rough)
#     full_prompt_words = len((system_prompt + user_prompt).split())
#     if full_prompt_words > 5000:
#         logger.warning("Prompt too large, truncating...")
#         # Hard truncation safeguards
#         user_prompt = user_prompt[:8000]
#         system_prompt = system_prompt[:2000]

#     messages = [
#         SystemMessage(content=system_prompt),
#         UserMessage(content=user_prompt),
#     ]

#     max_retries = 3
#     base_delay = 1.0

#     for attempt in range(max_retries):
#         try:
#             resp = _get_client().complete(
#                 messages=messages,
#                 model=AZURE_MODEL,
#                 temperature=0.1,
#                 top_p=0.9,
#                 max_tokens=1000,
#                 presence_penalty=0.0,
#                 frequency_penalty=0.0,
#             )

#             if not resp.choices:
#                 logger.warning("Empty choices from Azure Inference.")
#                 raise RuntimeError("Empty response")

#             raw_text = (resp.choices[0].message.content or "").strip()
#             cleaned = _clean_model_text(raw_text)
#             return cleaned or "No clear recommendation could be generated from the data provided."

#         except HttpResponseError as e:
#             status = getattr(e, "status_code", None)
#             logger.warning(f"Azure HttpResponseError (status={status}): {e}")
#             if status in (429, 502, 503, 504) and attempt < max_retries - 1:
#                 time.sleep(base_delay * (2 ** attempt))
#                 continue
#             break
#         except (ServiceRequestError, ServiceResponseError, TimeoutError, ConnectionError) as e:
#             logger.warning(f"Transient Azure error: {e}")
#             if attempt < max_retries - 1:
#                 time.sleep(base_delay * (2 ** attempt))
#                 continue
#             break
#         except Exception as e:
#             logger.warning(f"Attempt {attempt + 1}/{max_retries} failed: {e}")
#             if attempt < max_retries - 1:
#                 time.sleep(base_delay * (2 ** attempt))
#                 continue
#             break

#     return "I apologize, but I'm having trouble processing your request right now."

# @csrf_exempt
# def ask_question(request):
#     try:
#         start_time = now()

     
#         if request.method == "GET" and request.GET.get("export") == "true":
#             session_id = request.GET.get("session_id")
#             question = request.GET.get("question")

#             print("🔍 Export request received")
#             print(f"🔎 Session ID requested: {session_id}")
#             print(f"🔎 Question: {question}")
#             print("🔍 Current session_store keys:", list(session_store.keys()))

#             if not session_id:
#                 return JsonResponse({"error": "Missing session_id parameter"}, status=400)
#             if not question:
#                 return JsonResponse({"error": "Missing question parameter"}, status=400)

            
#             # >>> ADD THIS BLOCK <<<
#             if session_id and session_id not in session_store:
#                 try:
#                     ensure_session(session_id, user_id="export_user")
#                     print(f"✅ ensure_session: export path recovered session {session_id}")
#                 except Exception as rec_err:
#                     print(f"❌ ensure_session (export) failed: {rec_err}")
#                     return JsonResponse({
#                         "error": f"Session ID {session_id} could not be initialized for export"
#                     }, status=404)
#             # <<< END ADD >>>

          
#             session_data = None
#             actual_session_id = None
       
#             if session_id in session_store:
#                 session_data = session_store[session_id]
#                 actual_session_id = session_id
#                 print(f"✅ Found exact session match: {session_id}")
#             else:
#                 # 🔧 FIX 2: Fallback - look for any session with the same user query in history
#                 print(f"❌ Session {session_id} not found, searching in conversation history...")
#                 for stored_session_id, stored_data in session_store.items():
#                     history = conversation_memory_store.get(stored_session_id, [])
#                     for entry in history:
#                         if entry.get("question", "").strip().lower() == question.strip().lower():
#                             session_data = stored_data
#                             actual_session_id = stored_session_id
#                             print(f"✅ Found session via question match: {stored_session_id}")
#                             break
#                     if session_data:
#                         break

#             if not session_data:
#                 print(f"❌ No session found for question: {question}")
#                 print(f"🔍 Available sessions: {list(session_store.keys())}")
#                 return JsonResponse({
#                     "error": f"Session ID {session_id} not found. Please run the query first via chat.",
#                     "available_sessions": list(session_store.keys()),
#                     "debug_info": f"Searched for question: '{question}'"
#                 }, status=404)

#             try:
#                 user_id = session_data.get("user_id", "export_user")
#                 engine = session_data.get("engine")

#                 if not engine:
#                     return JsonResponse({"error": "Database engine not found in session"}, status=500)

#                 print(f"✅ Session found. User ID: {user_id}, Actual Session: {actual_session_id}")

#                 history = conversation_memory_store.get(actual_session_id, [])
#                 sql = None

#                 # 🔧 FIX 3: More flexible question matching
#                 for entry in reversed(history):
#                     stored_question = entry.get("question", "").strip().lower()
#                     search_question = question.strip().lower()
                    
#                     # Try exact match first, then partial match
#                     if stored_question == search_question or search_question in stored_question:
#                         sql = entry.get("sql")
#                         print(f"📋 Found SQL in history: {sql}")
#                         break

#                 if not sql:
#                     print("🔄 Generating new SQL for export...")
#                     try:
#                         raw_response = run_sql_generation_graph(question, user_id=user_id, db_id=actual_session_id, history=history)
#                         sql, _ = raw_response if isinstance(raw_response, tuple) else (extract_sql_block(raw_response), None)
#                         print(f"🆕 Generated SQL: {sql}")
#                     except Exception as gen_error:
#                         print(f"❌ SQL generation failed: {str(gen_error)}")
#                         return JsonResponse({"error": f"Failed to generate SQL: {str(gen_error)}"}, status=500)

#                 if not sql or not sql.strip():
#                     return JsonResponse({"error": "No SQL query generated"}, status=400)

#                 sql_lower = sql.strip().lower()
#                 if not (sql_lower.startswith("select") or sql_lower.startswith("with")):
#                     return JsonResponse({"error": "Invalid SQL query type"}, status=400)

#                 print(f"🚀 Executing SQL query...")
#                 with engine.connect() as conn:
#                     result = conn.execute(text(sql))
#                     rows = [dict(row._mapping) for row in result]

#                 print(f"✅ Query executed successfully. Found {len(rows)} rows")

#                 if not rows:
#                     # 🔧 FIX 5: Return empty CSV instead of plain text
#                     response = HttpResponse(content_type='text/csv')
#                     response['Content-Disposition'] = 'attachment; filename="export_no_data.csv"'
#                     response['Access-Control-Allow-Origin'] = '*'
#                     response.write("No data found for this query")
#                     return response

              
#                 from urllib.parse import quote
#                 timestamp = now().strftime("%Y%m%d_%H%M%S")
#                 filename = f"export_{timestamp}.csv"
                
#                 response = HttpResponse(content_type='text/csv')
#                 response['Content-Disposition'] = f'attachment; filename="{filename}"'
#                 response['Access-Control-Allow-Origin'] = '*'
#                 response['Access-Control-Allow-Headers'] = 'Content-Type'
#                 response['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'

#                 writer = csv.DictWriter(response, fieldnames=rows[0].keys())
#                 writer.writeheader()
#                 writer.writerows(rows)

#                 print(f"📁 CSV file created successfully with {len(rows)} rows")
#                 return response

#             except Exception as e:
#                 print(f"❌ Export execution error: {str(e)}")
#                 traceback.print_exc()
#                 return JsonResponse({
#                     "error": f"Export failed: {str(e)}",
#                     "details": "Check server logs for more information",
#                     "session_used": actual_session_id
#                 }, status=500)

#         elif request.method == "POST":
#             data = json.loads(request.body)
#             session_id = data.get("session_id")
#             question = data.get("question")
#             chart_config = None
#             user_id = data.get("user_id", "admin")

#             print(f"📨 POST request received")
#             print(f"🔎 Session ID: {session_id}")
#             print(f"🔎 Question: {question}")
#             print(f"🔎 User ID: {user_id}")

#             if not all([session_id, question]):
#                 return JsonResponse({"error": "Missing session_id or question"}, status=400)

            
            
#             # 🔹 ADD: auto-recover session if missing (survives restarts / multi-workers)
#             try:
#                 engine = ensure_session(session_id, user_id=user_id)
#                 print(ensure_session, "post ensure_session")
#             except Exception as rec_err:
#                 print(f"❌ ensure_session failed: {rec_err}")
#                 return JsonResponse({
#                     "answer": "Session not found",
#                     "success": False,
#                     "error": f"Session ID {session_id} could not be initialized",
#                     "rows": [],
#                     "row_count": 0,
#                     "session_id": session_id,
#                     "response_time": "0.00s",
#                     "code": "SESSION_INIT_FAILED"
#                 }, status=404)

#             if session_id not in session_store:
#                 print(f"⚠️ Session {session_id} not found")

#             history = conversation_memory_store.get(session_id, [])
#             print(f"🧠 Running SQL generation with session_id={session_id}, user_id={user_id}")

#             try:
#                 # 🔧 FIX 7: Handle column validation gracefully
#                 try:
#                     VALID_COLUMNS = extract_columns_from_schema(FULL_SCHEMA)
#                     validation_enabled = True
#                 except NameError:
#                     print("⚠️ Column validation not available")
#                     validation_enabled = False

#                 raw_response = run_sql_generation_graph(question, user_id=user_id, db_id=session_id, history=history)
                
#                 if isinstance(raw_response, tuple) and len(raw_response) == 3:
#                     sql_raw, recommendation, summary = raw_response
#                 else:
#                     sql_raw, recommendation = raw_response if isinstance(raw_response, tuple) else (extract_sql_block(raw_response), None)
#                     summary = ""
#                 # if isinstance(raw_response, tuple):
#                 #     sql_raw, recommendation = raw_response
#                 # else:
#                 #     sql_raw, recommendation = raw_response, None

#                 sql = extract_sql_block(sql_raw)
#                 print("📝 Extracted SQL:", sql)
#                 # ✅ Fix spacing issues in LIMIT clauses (e.g., LIMIT1 → LIMIT 1)
#                 sql = re.sub(r'\bLIMIT(\d+)', r'LIMIT \1', sql, flags=re.IGNORECASE)


#                 if validation_enabled:
#                     invalid_cols = validate_sql_columns(sql, VALID_COLUMNS)
#                     if invalid_cols:
#                         return JsonResponse({
#                             "answer": f"Invalid columns in SQL: {', '.join(invalid_cols)}",
#                             "success": False,
#                             "query_used": sql,
#                             "rows": [],
#                             "row_count": 0,
#                             "session_id": session_id,
#                             "response_time": "0.00s"
#                         }, status=400)

#             except Exception as sql_gen_error:
#                 print(f"❌ SQL generation error: {str(sql_gen_error)}")
#                 return JsonResponse({
#                     "answer": "Failed to generate SQL query",
#                     "success": False,
#                     "error": str(sql_gen_error),
#                     "rows": [],
#                     "row_count": 0,
#                     "session_id": session_id,
#                     "response_time": "0.00s"
#                 }, status=500)

#             if not sql or not sql.strip().lower().startswith(("select", "with")):
#                 return JsonResponse({
#                     "answer": "Invalid or failed SQL generation",
#                     "success": False,
#                     "query_used": sql or "No SQL generated",
#                     "rows": [],
#                     "row_count": 0,
#                     "session_id": session_id,
#                     "response_time": "0.00s"
#                 }, status=500)

#             if session_id not in session_store:
#                 return JsonResponse({
#                     "answer": "Session not found",
#                     "success": False,
#                     "error": f"Session ID {session_id} not found in session_store",
#                     "rows": [],
#                     "row_count": 0,
#                     "session_id": session_id,
#                     "response_time": "0.00s"
#                 }, status=404)

#             engine = session_store[session_id]["engine"]

#             try:
#                 print(f"✅ Executing SQL on session: {session_id}")
#                 with engine.connect() as conn:
#                     result = conn.execute(text(sql))
#                     rows = [dict(row._mapping) for row in result]
#                 print(f"✅ SQL executed successfully. Row count: {len(rows)}")
#                 summary = generate_summary_from_rows(question, sql, rows)
#             except Exception as e:
#                 print("❌ SQL Execution Error:", str(e))
#                 traceback.print_exc()
#                 return JsonResponse({
#                     "answer": "SQL execution failed.",
#                     "success": False,
#                     "query_used": sql,
#                     "error": str(e),
#                     "rows": [],
#                     "row_count": 0,
#                     "response_time": "0.00s",
#                     "session_id": session_id
#                 }, status=500)

#             # chart_config = None 
#             answer = ""
#             if rows:
#                 if len(rows) > 50:
#                     answer = f"Found {len(rows)} results. Too many to display here - please download the full results using the download button."
#                 else:
#                     formatted_rows = [", ".join(str(v) for v in row.values()) for row in rows[:3]]
#                     answer = "\n".join(formatted_rows)
#                     if len(rows) > 3:
#                         answer += f"\n...and {len(rows) - 3} more rows."
#             else:
#                 answer = "No data found."


           
#             try:
#                     chart_config = llm_generate_chart_config(question, rows)
#                     print("📊 Generated chart config:", json.dumps(chart_config, indent=2))
                    
#                     # Validate chart config before sending
#                     if chart_config and isinstance(chart_config, dict):
#                         # Ensure required fields exist
#                         if 'series' not in chart_config or not chart_config['series']:
#                             print("⚠️ Invalid chart config - missing or empty series")
#                             chart_config = None
#                         else:
#                             # Validate each series has data
#                             valid_series = []
#                             for series in chart_config['series']:
#                                 if 'data' in series and series['data']:
#                                     valid_series.append(series)
                            
#                             if valid_series:
#                                 chart_config['series'] = valid_series
#                             else:
#                                 chart_config = None
                                
#             except Exception as chart_err:
#                     print("⚠️ Chart generation failed:", chart_err)
#                     chart_config = None


#             try:
#                 if not recommendation:
#                     recommendation = llm_generate_recommendation(question, rows)
#             except Exception as rec_err:
#                 print("⚠️ Recommendation generation failed:", rec_err)
#                 recommendation = "Could not generate recommendation at this time."

       
#             conversation_memory_store.setdefault(session_id, []).append({
#                 "question": question,
#                 "sql": sql,
#                 "row_count": len(rows),
#                 "timestamp": now().isoformat()
#             })

#             if session_id in session_store:
#                 session_store[session_id]["user_id"] = user_id

#             total_time = (now() - start_time).total_seconds()

#             return JsonResponse({
#                 "answer": answer,
#                 "success": True,
#                 "query_used": sql,
#                 "rows": rows,
#                 'summary': summary,
#                 "chart_config": chart_config,
#                 "row_count": len(rows),
#                 "recommendation": recommendation,
#                 "response_time": f"{total_time:.2f}s",
#                 "session_id": session_id,
#                 "history": conversation_memory_store[session_id]
#             })

#         elif request.method == "OPTIONS":
#             response = HttpResponse()
#             response['Access-Control-Allow-Origin'] = '*'
#             response['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
#             response['Access-Control-Allow-Headers'] = 'Content-Type, X-Requested-With'
#             return response

#         else:
#             return JsonResponse({"error": "Method not allowed. Use GET for export or POST for queries."}, status=405)

#     except Exception as e:
#             print("💥 Unexpected error in ask_question:")
#             traceback.print_exc()

#             # ✅ Don’t re-parse request.body here; it can raise again.
#             sid = None
#             try:
#                 if request.method == "GET":
#                     sid = request.GET.get("session_id")
#                 elif request.method == "POST":
#                     # use already-parsed 'data' if available
#                     if 'data' in locals() and isinstance(data, dict):
#                         sid = data.get("session_id")
#             except Exception:
#                 sid = None

#             return JsonResponse({
#                 "answer": "Something went wrong.",
#                 "success": False,
#                 "error": str(e),
#                 "rows": [],
#                 "row_count": 0,
#                 "response_time": "0.00s",
#                 "session_id": sid or "unknown"
#             }, status=500)

#     # except Exception as e:
#     #     print("💥 Unexpected error in ask_question:")
#     #     traceback.print_exc()
#     #     return JsonResponse({
#     #         "answer": "Something went wrong.",
#     #         "success": False,
#     #         "error": str(e),
#     #         "rows": [],
#     #         "row_count": 0,
#     #         "response_time": "0.00s",
#     #         "session_id": request.GET.get("session_id") if request.method == "GET" else json.loads(request.body).get("session_id", "unknown") if request.method == "POST" else "unknown"
#     #     }, status=500)

# @csrf_exempt
# def ask_question(request):
#     try:
#         start_time = now()

     
#         if request.method == "GET" and request.GET.get("export") == "true":
#             # session_id = request.GET.get("session_id")
#             question = request.GET.get("question")

#             print("🔍 Export request received")
#             # print(f"🔎 Query Session ID requested: {session_id}")
#             print(f"🔎 Question: {question}")
#             # print("🔍 Current session_store keys:", list(session_store.keys()))

#             # if not session_id:
#             #     return JsonResponse({"error": "Missing session_id parameter"}, status=400)
#             if not question:
#                 return JsonResponse({"error": "Missing question parameter"}, status=400)

             

#             try:
                

#                 user_id = request.GET.get("user_id", "export_user")
#                 memory_key = user_id
#                 engine = ENGINE

#                 history = conversation_memory_store.get(memory_key, [])
#                 sql = None

#                 # Try to reuse prior SQL from memory (exact or partial match)
#                 for entry in reversed(history):
#                     stored_question = (entry.get("question") or "").strip().lower()
#                     search_question = (question or "").strip().lower()
#                     if stored_question == search_question or search_question in stored_question:
#                         sql = entry.get("sql")
#                         print(f"📋 Found SQL in memory: {sql}")
#                         break

#                 if not sql:
#                     print("🔄 Generating new SQL for export (sessionless)...")

#                     # ✨ NEW: build a valid column list and augment the question using prior context
#                     try:
#                         VALID_COLUMNS = extract_columns_from_schema(FULL_SCHEMA)
#                     except NameError:
#                         VALID_COLUMNS = columns_from_schema(FULL_SCHEMA)

#                     aug_question = augment_question_with_context(question, history)

#                     raw_response = run_sql_generation_graph(
#                         aug_question, user_id=user_id, db_id="default", history=history
#                     )
#                     sql, _ = raw_response if isinstance(raw_response, tuple) else (extract_sql_block(raw_response), None)
#                     print(f"🆕 Generated SQL: {sql}")

#                 if not sql or not sql.strip():
#                     return JsonResponse({"error": "No SQL query generated"}, status=400)

#                 # Normalize common formatting issues (e.g., LIMIT1 -> LIMIT 1)
#                 sql = extract_sql_block(sql)
#                 sql = re.sub(r'\bLIMIT(\d+)', r'LIMIT \1', sql, flags=re.IGNORECASE)

#                 sql_lower = sql.strip().lower()
#                 if not (sql_lower.startswith("select") or sql_lower.startswith("with")):
#                     return JsonResponse({"error": "Invalid SQL query type"}, status=400)

#                 print(f"🚀 Executing SQL query (sessionless)...")
#                 with engine.connect() as conn:
#                     result = conn.execute(text(sql))
#                     rows = [dict(row._mapping) for row in result]

#                 print(f"✅ Query executed successfully. Found {len(rows)} rows")

#                 # ✅ Persist this interaction to conversation memory (per user)
#                 try:
#                     VALID_COLUMNS  # may exist already
#                 except NameError:
#                     try:
#                         VALID_COLUMNS = extract_columns_from_schema(FULL_SCHEMA)
#                     except NameError:
#                         VALID_COLUMNS = columns_from_schema(FULL_SCHEMA)

#                 ctx = parse_sql_context(sql, VALID_COLUMNS)

#                 conversation_memory_store.setdefault(memory_key, []).append({
#                     "question": question,
#                     "sql": sql,
#                     "row_count": len(rows),
#                     "timestamp": now().isoformat(),
#                     "context": ctx,   # ✨ NEW: store dynamic context
#                 })


#                 # ✅ Persist this interaction to conversation memory (per user)
#                 # conversation_memory_store.setdefault(memory_key, []).append({
#                 #     "question": question,
#                 #     "sql": sql,
#                 #     "row_count": len(rows),
#                 #     "timestamp": now().isoformat()
#                 # })

#                 if not rows:
#                     # Return empty CSV instead of plain text
#                     response = HttpResponse(content_type='text/csv')
#                     response['Content-Disposition'] = 'attachment; filename="export_no_data.csv"'
#                     response['Access-Control-Allow-Origin'] = '*'
#                     response.write("No data found for this query")
#                     return response

#                 timestamp = now().strftime("%Y%m%d_%H%M%S")
#                 filename = f"export_{timestamp}.csv"

#                 response = HttpResponse(content_type='text/csv')
#                 response['Content-Disposition'] = f'attachment; filename="{filename}"'
#                 response['Access-Control-Allow-Origin'] = '*'
#                 response['Access-Control-Allow-Headers'] = 'Content-Type'
#                 response['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'

#                 writer = csv.DictWriter(response, fieldnames=rows[0].keys())
#                 writer.writeheader()
#                 writer.writerows(rows)

#                 print(f"📁 CSV file created successfully with {len(rows)} rows")
#                 return response

#             except Exception as e:
#                 print(f"❌ Export execution error: {str(e)}")
#                 traceback.print_exc()
#                 return JsonResponse({
#                     "error": f"Export failed: {str(e)}",
#                     "details": "Check server logs for more information",
#                     "memory_key": memory_key
#                 }, status=500)


#         elif request.method == "POST":
#                 data = json.loads(request.body)
#                 # session_id = data.get("session_id")
#                 question = data.get("question")
#                 chart_config = None
#                 user_id = data.get("user_id", "admin")

#                 print(f"📨 POST request received")
#                 # print(f"🔎 Post request Session ID: {session_id}")
#                 print(f"🔎 Question: {question}")
#                 print(f"🔎 User ID: {user_id}")

#                 if not question:
#                     return JsonResponse({"error": "Missing question"}, status=400)

#                 # Use per-user conversation memory (sessionless)
#                 memory_key = user_id
#                 history = conversation_memory_store.get(memory_key, [])
#                 print(f"🧠 Running SQL generation (sessionless) with user_id={user_id}")

#                 # -------------------------------------------
#                 # ✅ EARLY BRANCH: generic "segmentation" follow-ups scoped by last carry
#                 # -------------------------------------------
#                 q = question
#                 if _is_segmentation_question(q):
#                     carry = get_last_carry(history)
#                     level, col, val = pick_level_and_value_from_carry(carry)

#                     base_table = '"bi_dwh"."main_cai_lib"'  # adjust if needed
#                     params = {}
#                     where_sql = ""
#                     if col and val:
#                         where_sql = f"WHERE {col} = :val"
#                         params["val"] = val

#                     seg_sql = f"""
#                         SELECT
#                         customer_segment,
#                         COUNT(policy_no) AS total_policies
#                         FROM {base_table}
#                         {where_sql}
#                         GROUP BY customer_segment
#                         ORDER BY total_policies DESC
#                     """

#                     try:
#                         with ENGINE.connect() as conn:
#                             result = conn.execute(text(seg_sql), params)
#                             rows = [dict(r._mapping) for r in result]

#                         summary = generate_summary_from_rows(q, seg_sql, rows)
#                         try:
#                             chart_config = llm_generate_chart_config(q, rows)
#                         except Exception:
#                             chart_config = None

#                         # Persist history, carrying forward the same geo scope
#                         ctx = {"filters": {}, "carry": carry}
#                         conversation_memory_store.setdefault(memory_key, []).append({
#                             "question": q,
#                             "sql": seg_sql,
#                             "row_count": len(rows),
#                             "timestamp": now().isoformat(),
#                             "context": ctx
#                         })

#                         total_time = (now() - start_time).total_seconds()
#                         return JsonResponse({
#                             "answer": "",
#                             "success": True,
#                             "query_used": seg_sql,
#                             "rows": rows,
#                             "summary": summary,
#                             "chart_config": chart_config,
#                             "row_count": len(rows),
#                             "recommendation": "",  # optionally call llm_generate_recommendation
#                             "response_time": f"{total_time:.2f}s",
#                             "user_id": user_id,
#                             "history": conversation_memory_store[memory_key]
#                         })
#                     except Exception as e:
#                         return JsonResponse({
#                             "answer": "Segmentation query failed.",
#                             "success": False,
#                             "query_used": seg_sql,
#                             "error": str(e),
#                             "rows": [],
#                             "row_count": 0,
#                             "response_time": "0.00s",
#                             "user_id": user_id
#                         }, status=500)

               

                
#                 validation_enabled = False            # FIX: always defined
#                 VALID_COLUMNS = []                    # FIX: safe default
#                 raw_response = None                   # FIX: so isinstance checks don't crash
#                 recommendation = None
#                 summary = ""
#                 sql = None
#                 used_reused_sql = False         

#                 try:
#                     # 🔧 Handle column list and augment question using prior context
#                     # Try to reuse a prior SQL if this looks like a repeat/near-repeat
#                     reused = best_prior_sql(question, history)
#                     if reused:
#                         sql = normalize_sql(reused)
#                         used_reused_sql = True
#                         print("♻️ Reusing prior SQL from memory for a near-duplicate question.")
#                         # FIX: define these so later code doesn't blow up
#                         raw_response = (sql, None)
#                         recommendation = None
#                         summary = ""
#                     else:
#                         # FIX: column validation with safe fallback even if FULL_SCHEMA isn't present
#                         try:
#                             VALID_COLUMNS = extract_columns_from_schema(FULL_SCHEMA)
#                             validation_enabled = True
#                         except Exception as e:
#                             print(f"⚠️ Column validation disabled (no FULL_SCHEMA or extractor failed): {e}")
#                             validation_enabled = False
#                             VALID_COLUMNS = []  # keep empty; validate step will be skipped

#                         aug_question = augment_question_with_context(question, history)
#                         raw_response = run_sql_generation_graph(
#                             aug_question, user_id=user_id, db_id="default", history=history
#                         )

#                         # FIX: only read sql_raw here; in reuse path we already have sql
#                         if isinstance(raw_response, tuple) and len(raw_response) == 3:
#                             sql_raw, recommendation, summary = raw_response
#                         else:
#                             sql_raw, recommendation = raw_response if isinstance(raw_response, tuple) else (extract_sql_block(raw_response), None)
#                             summary = ""

#                         # sql = extract_sql_block(sql_raw)
#                         # print("📝 Extracted SQL:", sql)
#                         # sql = re.sub(r'\bLIMIT(\d+)', r'LIMIT \1', sql, flags=re.IGNORECASE)
#                         sql = extract_sql_block(sql_raw)
#                         sql = re.sub(r'\bLIMIT(\d+)', r'LIMIT \1', sql, flags=re.IGNORECASE)
#                         sql = relax_location_filters(sql)     # 🔧 add this line
#                         print("🧰 Post-processed SQL:", sql)


#                     if validation_enabled:
#                         try:
#                             invalid_cols = validate_sql_columns(sql, VALID_COLUMNS)
#                         except Exception as ve:
#                             print(f"⚠️ validate_sql_columns crashed; skipping column validation: {ve}")
#                             invalid_cols = []
#                         if invalid_cols:
#                             return JsonResponse({
#                                 "answer": f"Invalid columns in SQL: {', '.join(invalid_cols)}",
#                                 "success": False,
#                                 "query_used": sql,
#                                 "rows": [],
#                                 "row_count": 0,
#                                 "user_id": user_id,
#                                 "response_time": "0.00s"
#                             }, status=400)

#                 except Exception as sql_gen_error:
#                     print(f"❌ SQL generation error: {str(sql_gen_error)}")
#                     return JsonResponse({
#                         "answer": "Failed to generate SQL query",
#                         "success": False,
#                         "error": str(sql_gen_error),
#                         "rows": [],
#                         "row_count": 0,
#                         "user_id": user_id,
#                         "response_time": "0.00s"
#                     }, status=500)

#                 if not sql or not sql.strip().lower().startswith(("select", "with")):
#                     return JsonResponse({
#                         "answer": "Invalid or failed SQL generation",
#                         "success": False,
#                         "query_used": sql or "No SQL generated",
#                         "rows": [],
#                         "row_count": 0,
#                         "user_id": user_id,
#                         "response_time": "0.00s"
#                     }, status=500)

               
#                 try:
#                     # print(f"✅ Executing SQL on session: {session_id}")
#                     print("✅ Executing SQL (sessionless)...")
#                     with ENGINE.connect() as conn:
#                         result = conn.execute(text(sql))
#                         rows = [dict(row._mapping) for row in result]
#                     print(f"✅ SQL executed successfully. Row count: {len(rows)}")

#                     summary = generate_summary_from_rows(question, sql, rows)
#                     try:
#                         ctx = parse_sql_context(sql, VALID_COLUMNS if validation_enabled else [])
#                     except Exception as ce:
#                         print(f"⚠️ parse_sql_context failed: {ce}")
#                         ctx = {"filters": {}}

#                     # Save carry so next turn like "segmentation of this" works
                    
#                     print(f"🧷 Saved carry from rows: {ctx}")
#                 except Exception as e:
#                     print("❌ SQL Execution Error:", str(e))
#                     traceback.print_exc()
#                     return JsonResponse({
#                         "answer": "SQL execution failed.",
#                         "success": False,
#                         "query_used": sql,
#                         "error": str(e),
#                         "rows": [],
#                         "row_count": 0,
#                         "response_time": "0.00s",
#                         "user_id": user_id,
#                     }, status=500)

#                 # chart_config = None
#                 answer = ""
#                 if rows:
#                     if len(rows) > 50:
#                         answer = f"Found {len(rows)} results. Too many to display here - please download the full results using the download button."
#                     else:
#                         formatted_rows = [", ".join(str(v) for v in row.values()) for row in rows[:3]]
#                         answer = "\n".join(formatted_rows)
#                         if len(rows) > 3:
#                             answer += f"\n...and {len(rows) - 3} more rows."
#                 else:
#                     answer = "No data found."

#                 try:
#                     chart_config = llm_generate_chart_config(question, rows)
#                     print("📊 Generated chart config:", json.dumps(chart_config, indent=2))
#                     # Validate chart config before sending
#                     if chart_config and isinstance(chart_config, dict):
#                         if 'series' not in chart_config or not chart_config['series']:
#                             print("⚠️ Invalid chart config - missing or empty series")
#                             chart_config = None
#                         else:
#                             valid_series = []
#                             for series in chart_config['series']:
#                                 if 'data' in series and series['data']:
#                                     valid_series.append(series)
#                             chart_config['series'] = valid_series or None
#                 except Exception as chart_err:
#                     print("⚠️ Chart generation failed:", chart_err)
#                     chart_config = None

#                 try:
#                     if not recommendation:
#                         recommendation = llm_generate_recommendation(question, rows)
#                 except Exception as rec_err:
#                     print("⚠️ Recommendation generation failed:", rec_err)
#                     recommendation = "Could not generate recommendation at this time."

#                 # ✅ Save carry derived from result rows (so follow-ups like "segmentation of this" work)
#                 try:
#                         carry = derive_geo_carry_from_rows(sql, rows)
#                         print(f"🧷 Saved carry from rows: {carry}")
#                 except Exception as de:
#                         print(f"⚠️ derive_geo_carry_from_rows failed: {de}")
#                         carry = {}

#                 if ctx is None:
#                         ctx = {}
#                 if "filters" not in ctx:
#                         ctx["filters"] = {}
#                 ctx["carry"] = carry
#                 print(f"🧪 is_segmentation_question={_is_segmentation_question(question)} raw='{question}'")


               

#                 conversation_memory_store.setdefault(memory_key, []).append({
#                     "question": question,
#                     "sql": sql,
#                     "row_count": len(rows),
#                     "timestamp": now().isoformat(),
#                     "context": ctx,
#                 })

#                 # if session_id in session_store:
#                 #     session_store[session_id]["user_id"] = user_id

#                 total_time = (now() - start_time).total_seconds()

#                 return JsonResponse({
#                     "answer": answer,
#                     "success": True,
#                     "query_used": sql,
#                     "rows": rows,
#                     "summary": summary,
#                     "chart_config": chart_config,
#                     "row_count": len(rows),
#                     "recommendation": recommendation,
#                     "response_time": f"{total_time:.2f}s",
#                     "user_id": user_id,
#                     "history": conversation_memory_store[memory_key]
#                     # "session_id": session_id,
#                     # "history": conversation_memory_store[session_id]
#                 })

#         elif request.method == "OPTIONS":
#             response = HttpResponse()
#             response['Access-Control-Allow-Origin'] = '*'
#             response['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
#             response['Access-Control-Allow-Headers'] = 'Content-Type, X-Requested-With'
#             return response

#         else:
#             return JsonResponse({"error": "Method not allowed. Use GET for export or POST for queries."}, status=405)

#     except Exception as e:
#             print("💥 Unexpected error in ask_question:")
#             traceback.print_exc()

#             # ✅ Don’t re-parse request.body here; it can raise again.
#             # sid = None
#             uid = None
#             try:
#                 if request.method == "GET":
#                     # sid = request.GET.get("session_id")
#                     uid = request.GET.get("user_id")
#                 elif request.method == "POST":
#                     # use already-parsed 'data' if available
#                     if 'data' in locals() and isinstance(data, dict):
#                         uid = data.get("user_id")
#                     # if 'data' in locals() and isinstance(data, dict):
#                     #     sid = data.get("session_id")
#             except Exception:
#                 # sid = None
#                 uid = None

#             return JsonResponse({
#                 "answer": "Something went wrong.",
#                 "success": False,
#                 "error": str(e),
#                 "rows": [],
#                 "row_count": 0,
#                 "response_time": "0.00s",
#                 "user_id": uid or "unknown"
#                 # "session_id": sid or "unknown"
#             }, status=500)



import json, time, traceback, re, datetime, uuid
from decimal import Decimal
from django.http import StreamingHttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.core.serializers.json import DjangoJSONEncoder
from sqlalchemy import text

# ---------- Enhanced SQL Processing Helpers ----------
def safe_extract_sql_text(maybe_sql: object) -> str:
    text_ = str(maybe_sql or "").lstrip("\ufeff")
    if "```" in text_:
        try:
            unwrapped = extract_sql_block(text_)
            if isinstance(unwrapped, str) and unwrapped.strip():
                return unwrapped.strip()
        except Exception:
            pass
    m = re.search(r'(?is)\b(with|select)\b', text_)
    return text_[m.start():].strip() if m else text_.strip()

def extract_time_context_from_question(question: str) -> dict:
    """Extract time-related context from user questions to ensure proper SQL generation"""
    context = {}
    question_lower = question.lower()
    
    # Extract year ranges
    year_patterns = [
        r'in (\d{4})',
        r'during (\d{4})',
        r'for (\d{4})',
        r'from (\d{4}) to (\d{4})',
        r'between (\d{4}) and (\d{4})',
        r'(\d{4})-(\d{4})',
    ]
    
    for pattern in year_patterns:
        match = re.search(pattern, question_lower)
        if match:
            groups = match.groups()
            if len(groups) == 1:
                context['single_year'] = groups[0]
            elif len(groups) == 2:
                context['year_range'] = {'start': groups[0], 'end': groups[1]}
            break
    
    # Extract month/quarter context
    if 'quarter' in question_lower or 'q1' in question_lower or 'q2' in question_lower:
        context['time_granularity'] = 'quarter'
    elif 'month' in question_lower or 'monthly' in question_lower:
        context['time_granularity'] = 'month'
    elif 'year' in question_lower or 'annual' in question_lower:
        context['time_granularity'] = 'year'
    
    return context

def fix_cte_structure(sql: str) -> str:
    """Fix Common Table Expression (CTE) structure issues"""
    if not sql or not sql.strip():
        return sql
    
    # Check if this is a WITH query
    if not re.match(r'\s*WITH\s+', sql, re.IGNORECASE):
        return sql
    
    try:
        # Pattern to find nested WITH clauses that should be at top level
        # Look for cases where a CTE is defined inside another CTE
        nested_with_pattern = r'(\s*WITH\s+\w+\s+AS\s*\([^)]*?WITH\s+(\w+)\s+AS\s*\([^)]+\)[^)]*?\))'
        
        if re.search(nested_with_pattern, sql, re.IGNORECASE | re.DOTALL):
            print("🔧 Fixing nested CTE structure...")
            
            # Extract all CTE definitions and reorganize them
            cte_definitions = []
            main_query = sql
            
            # Find all CTE patterns
            cte_pattern = r'WITH\s+(\w+)\s+AS\s*\(([^)]+(?:\([^)]*\)[^)]*)*)\)'
            matches = list(re.finditer(cte_pattern, sql, re.IGNORECASE | re.DOTALL))
            
            if len(matches) > 1:
                # Multiple CTEs found - need to restructure
                all_ctes = []
                for match in matches:
                    cte_name = match.group(1)
                    cte_body = match.group(2).strip()
                    all_ctes.append(f"{cte_name} AS (\n{cte_body}\n)")
                
                # Find the main SELECT query (after all CTEs)
                last_match = matches[-1]
                remaining_sql = sql[last_match.end():].strip()
                
                # Look for the main SELECT
                main_select_match = re.search(r'SELECT\s+.*', remaining_sql, re.IGNORECASE | re.DOTALL)
                if main_select_match:
                    main_query = main_select_match.group(0)
                    
                    # Reconstruct the query with proper CTE structure
                    fixed_sql = "WITH " + ",\n".join(all_ctes) + "\n" + main_query
                    return fixed_sql
        
        return sql
        
    except Exception as e:
        print(f"⚠️ CTE structure fix failed: {e}")
        return sql

def repair_split_time_context(sql: str) -> str:
    """Enhanced to better handle time-based CTEs and fix CTE structure"""
    s = str(sql or "")
    low = s.lower()
    
    # First fix any CTE structure issues
    s = fix_cte_structure(s)
    
    # If already has WITH clause, check if time context is properly formed
    if low.startswith("with") or "time_context" not in low:
        return s
    
    # Look for broken CTE patterns
    m = re.search(r'\)\s*select\b', s, flags=re.IGNORECASE | re.DOTALL)
    if not (s.strip().lower().startswith("select") and m):
        return s
    
    paren_end = m.start()
    sel_after = re.search(r'\bselect\b', s[paren_end:], flags=re.IGNORECASE)
    if not sel_after:
        return s
    
    sel_idx = paren_end + sel_after.start()
    prefix = s[:paren_end].rstrip()
    rest = s[sel_idx:].lstrip()
    
    # Enhanced: ensure time filtering is preserved
    time_filters = re.findall(r'policy_end_date_year\s*[>=<]+\s*\d{4}', prefix, flags=re.IGNORECASE)
    if time_filters:
        return f"WITH time_context AS ({prefix})\n{rest}"
    
    return s
def repair_time_context_alias(sql: str) -> str:
    """Enhanced alias repair with better time context handling"""
    pat = re.compile(r'WITH\s+time_context\s+AS\s*\(\s*(.*?)\s*\)',
                     flags=re.IGNORECASE | re.DOTALL)
    
    def _fix(m):
        body = m.group(1)
        
        # Check if there's a table alias being used but not defined
        # Look for patterns like "mcl.column_name" in WHERE clause
        alias_usage = re.findall(r'\b([a-zA-Z_][a-zA-Z0-9_]*)\s*\.\s*[a-zA-Z_]', body)
        
        if alias_usage:
            # Find the most common alias being used (usually 'mcl')
            from collections import Counter
            alias_counts = Counter(alias_usage)
            most_common_alias = alias_counts.most_common(1)[0][0] if alias_counts else 'mcl'
            
            # Check if the FROM clause already has an alias
            from_match = re.search(r'FROM\s+([^\s\)]+)(?:\s+(?:AS\s+)?([a-zA-Z_][a-zA-Z0-9_]*))?\s*(?:WHERE|GROUP|ORDER|LIMIT|\)|$)', 
                                 body, flags=re.IGNORECASE)
            
            if from_match:
                table_name = from_match.group(1)
                existing_alias = from_match.group(2)
                
                if not existing_alias:
                    # No alias defined, but alias is being used - add the alias
                    body = re.sub(
                        r'(FROM\s+' + re.escape(table_name) + r')(\s+)',
                        r'\1 ' + most_common_alias + r'\2',
                        body,
                        flags=re.IGNORECASE
                    )
                    print(f"🔧 Added missing alias '{most_common_alias}' to table '{table_name}' in time_context CTE")
                elif existing_alias != most_common_alias:
                    # Different alias defined than what's being used - replace all usages
                    body = re.sub(
                        r'\b' + re.escape(most_common_alias) + r'\s*\.',
                        existing_alias + '.',
                        body
                    )
                    print(f"🔧 Replaced alias '{most_common_alias}' with '{existing_alias}' in time_context CTE")
        
        return f'WITH time_context AS (\n{body}\n)'
    
    return pat.sub(_fix, sql)

def fix_all_cte_aliases(sql: str) -> str:
    """Fix aliases in all CTEs, not just time_context"""
    # Pattern to match any CTE
    cte_pattern = re.compile(r'WITH\s+(\w+)\s+AS\s*\(\s*(.*?)\s*\)(?=\s*,|\s+SELECT|\s*$)',
                           flags=re.IGNORECASE | re.DOTALL)
    
    def fix_cte_alias(match):
        cte_name = match.group(1)
        cte_body = match.group(2)
        
        # Find alias usage patterns
        alias_usage = re.findall(r'\b([a-zA-Z_][a-zA-Z0-9_]*)\s*\.\s*[a-zA-Z_]', cte_body)
        
        if alias_usage:
            from collections import Counter
            alias_counts = Counter(alias_usage)
            most_common_alias = alias_counts.most_common(1)[0][0] if alias_counts else None
            
            if most_common_alias:
                # Check FROM clause
                from_match = re.search(r'FROM\s+([^\s\)]+)(?:\s+(?:AS\s+)?([a-zA-Z_][a-zA-Z0-9_]*))?\s*(?:WHERE|GROUP|ORDER|LIMIT|\)|$)', 
                                     cte_body, flags=re.IGNORECASE)
                
                if from_match:
                    table_name = from_match.group(1)
                    existing_alias = from_match.group(2)
                    
                    if not existing_alias:
                        # Add the missing alias
                        cte_body = re.sub(
                            r'(FROM\s+' + re.escape(table_name) + r')(\s+)',
                            r'\1 ' + most_common_alias + r'\2',
                            cte_body,
                            flags=re.IGNORECASE
                        )
                        print(f"🔧 Fixed missing alias '{most_common_alias}' in CTE '{cte_name}'")
        
        return f'WITH {cte_name} AS (\n{cte_body}\n)'
    
    return cte_pattern.sub(fix_cte_alias, sql)

# def sanitize_sql_before_exec(sql: str, question: str = "") -> str:
#     """Enhanced sanitization with comprehensive CTE alias fixing"""
#     s = safe_extract_sql_text(sql)
    
#     # Fix CTE structure issues first
#     s = fix_cte_structure(s)
    
#     # Fix all CTE aliases (enhanced version)
#     s = fix_all_cte_aliases(s)
    
#     # Then apply other fixes
#     s = repair_split_time_context(s)
    
#     # Add time context if missing
#     if question:
#         s = enhance_sql_with_time_context(s, question)
    
#     # Fix LIMIT clause formatting
#     s = re.sub(r'\blimit\s*(\d+)\b', r'LIMIT \1', s, flags=re.IGNORECASE)
    
#     # Final validation - check for remaining alias issues
#     if re.search(r'\bmcl\s*\.', s) and not re.search(r'FROM\s+[^\s]+\s+(?:AS\s+)?mcl\b', s, re.IGNORECASE):
#         print("⚠️ Still found mcl alias usage without proper FROM clause definition")
#         # Emergency fix: replace all mcl. with nothing if no proper alias found
#         table_match = re.search(r'FROM\s+([^\s\)]+)', s, re.IGNORECASE)
#         if table_match:
#             table_name = table_match.group(1)
#             # Add mcl alias to the main table reference
#             s = re.sub(r'(FROM\s+' + re.escape(table_name) + r')(\s+)', r'\1 mcl\2', s, flags=re.IGNORECASE)
#             print(f"🚨 Emergency fix: Added mcl alias to {table_name}")
    
#     print(f"🔧 Final sanitized SQL:\n{s}")
#     return s

def enhance_sql_with_time_context(sql: str, question: str) -> str:
    """Add proper time context to SQL based on user question"""
    time_context = extract_time_context_from_question(question)
    
    if not time_context:
        return sql
    
    # If SQL already has time filters, don't modify
    if re.search(r'policy_end_date_year\s*[>=<]', sql, flags=re.IGNORECASE):
        return sql
    
    # Add time filtering based on extracted context
    if 'single_year' in time_context:
        year = time_context['single_year']
        if 'WHERE' in sql.upper():
            sql = re.sub(r'WHERE', f'WHERE policy_end_date_year = {year} AND', sql, count=1, flags=re.IGNORECASE)
        else:
            # Find a good place to insert WHERE clause
            from_match = re.search(r'FROM\s+[^\s]+(?:\s+[a-zA-Z_][a-zA-Z0-9_]*)?', sql, flags=re.IGNORECASE)
            if from_match:
                insert_pos = from_match.end()
                sql = sql[:insert_pos] + f' WHERE policy_end_date_year = {year}' + sql[insert_pos:]
    
    elif 'year_range' in time_context:
        start_year = time_context['year_range']['start']
        end_year = time_context['year_range']['end']
        if 'WHERE' in sql.upper():
            sql = re.sub(r'WHERE', f'WHERE policy_end_date_year BETWEEN {start_year} AND {end_year} AND', 
                        sql, count=1, flags=re.IGNORECASE)
        else:
            from_match = re.search(r'FROM\s+[^\s]+(?:\s+[a-zA-Z_][a-zA-Z0-9_]*)?', sql, flags=re.IGNORECASE)
            if from_match:
                insert_pos = from_match.end()
                sql = sql[:insert_pos] + f' WHERE policy_end_date_year BETWEEN {start_year} AND {end_year}' + sql[insert_pos:]
    
    return sql

# def sanitize_sql_before_exec(sql: str, question: str = "") -> str:
#     """Enhanced sanitization with time context awareness and CTE fixing"""
#     s = safe_extract_sql_text(sql)
    
#     # Fix CTE structure issues first
#     s = fix_cte_structure(s)
    
#     # Then apply other fixes
#     s = repair_split_time_context(s)
#     s = repair_time_context_alias(s)
    
#     # Add time context if missing
#     if question:
#         s = enhance_sql_with_time_context(s, question)
    
#     # Fix LIMIT clause formatting
#     s = re.sub(r'\blimit\s*(\d+)\b', r'LIMIT \1', s, flags=re.IGNORECASE)
    
#     print(f"🔧 Sanitized SQL:\n{s}")
#     return s

# ---------- Enhanced JSON Safety ----------
def _jsonable(x):
    """Recursively coerce DB types into JSON-safe values."""
    try:
        import numpy as _np
        _np_types = ( _np.integer, _np.floating, _np.bool_, _np.ndarray )
    except Exception:
        _np_types = tuple()

    if x is None or isinstance(x, (str, int, float, bool)):
        return x
    if isinstance(x, Decimal):
        return int(x) if x == int(x) else float(x)
    if isinstance(x, (datetime.date, datetime.datetime, datetime.time)):
        return x.isoformat()
    if isinstance(x, (uuid.UUID, bytes)):
        return str(x)
    if _np_types and isinstance(x, _np_types):
        try:
            return _jsonable(x.item())
        except Exception:
            try:
                return [_jsonable(v) for v in x.tolist()]
            except Exception:
                return str(x)
    if isinstance(x, dict):
        return {str(k): _jsonable(v) for k, v in x.items()}
    if isinstance(x, (list, tuple, set)):
        return [_jsonable(v) for v in x]
    return str(x)

def _line(event: str, **payload) -> str:
    obj = {"event": event}
    obj.update(payload)
    return json.dumps(_jsonable(obj), cls=DjangoJSONEncoder, ensure_ascii=False) + "\n"

def _ev(event: str, **payload):
    return _line(event, **payload)

# ============================================
# PRODUCTION FIX: Context-Aware Opener Generation
# ============================================


def has_metric_variance(rows):
    if not rows or len(rows) < 2:
        return False

    # Detect numeric metric column (e.g., churn_rate_percentage)
    numeric_keys = [
        k for k, v in rows[0].items()
        if isinstance(v, (int, float))
    ]

    if not numeric_keys:
        return False

    metric_key = numeric_keys[0]

    values = {r.get(metric_key) for r in rows}
    return len(values) > 1

def extract_exact_metrics(rows):
    """
    Extract percentage metrics directly from DB output.
    Only keys ending with '_percentage' are allowed.
    Values are formatted to 2 decimals (TRUNCATED, not guessed).
    """
    if not rows or not isinstance(rows, list):
        return {}

    metrics = {}
    first_row = rows[0]

    if not isinstance(first_row, dict):
        return metrics

    for key, value in first_row.items():
        # if key.endswith("_percentage") and isinstance(value, (int, float)):
        #     # metrics[key] = round(float(value), 2)  # 54.5677 → 54.56
        #     if ROUND_PERCENTAGES:  # 👈 ADD CONFIG CHECK
        #         metrics[key] = round(float(value))      # 54.56 → 55
        #     else:
        #         metrics[key] = round(float(value), 2)   # 54.56 → 54.56
        if key.endswith("_percentage") and isinstance(value, (int, float)):
            if ROUND_PERCENTAGES:
                metrics[key] = round(float(value))       # ✅ Whole number
            else:
                metrics[key] = round(float(value), 2)    # ✅ 2 decimals

    return metrics


def generate_conversational_opener_with_data(
    question: str, 
    row_count: int, 
    context: dict = None, 
    rows: list = None
) -> str:
    """
    Generate analytical opener that interprets what the data represents.
    Handles aggregated/grouped data intelligently.
    """
    llm = get_llama_maverick_llm()
    
    # Classify what type of result this is
    result_type = classify_result_type(rows) if rows else "UNKNOWN"
    
    # Extract business count (what to show to users)
    business_count = extract_business_count(rows, result_type)
    display_count = business_count if business_count is not None else row_count
    exact_metrics = extract_exact_metrics(rows)

    
    # Build data context for LLM
    data_context = {
        "result_type": result_type,
        "row_count": row_count,
        "business_count": business_count,
        "display_count": display_count,
    }

    

    
    # Get sample data structure
    sample_info = ""

    if rows and len(rows) > 0:
        first_row = rows[0]
        columns = list(first_row.keys()) if isinstance(first_row, dict) else []

        # ✅ Compute once, before branching
        metric_varies = (
            len(rows) > 1 and has_metric_variance(rows)
        )


        # Identify what the data represents
        if result_type == "SINGLE_METRIC":
            sample_info = f"Single aggregated metric: {first_row}"

        elif result_type == "GROUPED_METRIC" and len(rows) > 1 and not metric_varies:

            sample_info = (
                f"Grouped/segmented data with dimensions: {columns[:3]}"
                "\n⚠️ All segments show identical metric values"
            )

        elif result_type == "GROUPED_METRIC":
            sample_info = f"Grouped/segmented data with dimensions: {columns[:3]}"
            if len(rows) <= 5:
                sample_info += (
                    f"\nSegments: {[r.get(columns[0]) if columns else r for r in rows[:5]]}"
                )

        else:
            sample_info = f"Row-level data with columns: {columns[:5]}"

    
    # Enhanced prompt with result type awareness
    prompt = f"""
Write ONE business-facing opening sentence that explains the analytical result below.

Question:
"{question}"

ANALYSIS CONTEXT:
- Result type: {result_type}
- Business interpretation count: {display_count} {_get_business_entity_name(question, result_type, rows)}
{sample_info}

EXACT_METRICS:
{exact_metrics if exact_metrics else "None"}

IMPORTANT:
- You may ONLY mention a percentage if it exists in EXACT_METRICS
- Use the value EXACTLY as provided
- Format must be two decimal places (e.g., 54.56%)
- If EXACT_METRICS is None, DO NOT include numbers

🔒 RULE (must enforce globally)

If a column represents a rate shown to users, SQL must return a percentage (×100).


RULES:
1. Write the sentence itself — do NOT introduce or label it
2. Do NOT say "Here is", "Here's", "Below is", or similar
3. Do NOT mention SQL, queries, rows, records, fields, or tables
4. Focus on WHAT the data shows, not HOW it was produced
5. Sound like a human analyst explaining results to a stakeholder
6. Do NOT include any numeric values (percentages, counts, rates)
   unless an exact value is explicitly provided in the data context.
   If unsure, describe the trend WITHOUT numbers.

# 6. Use numbers only when they add business clarity
7. Keep it to ONE sentence (max two if absolutely necessary)
8. If all segments have the same metric value, DO NOT describe rankings, comparisons, 
   highest/lowest, or patterns across segments.
   If fewer than 2 segments are present, do NOT describe distribution,
    comparisons, uniformity, or differences.
    Instead, state that results are available only for the reported segment(s).

#    Instead, state that churn is uniformly distributed across segments.

ABSOLUTE RULES FOR OPENER:

- NEVER use placeholders like "X%", "Y%", or variables
- NEVER say "approximately", "around", or "roughly"
- If a percentage value is not confidently known, DO NOT mention a percentage at all
- If mentioning a percentage, it MUST be an exact value from the provided data
- Otherwise, describe the trend WITHOUT numbers


STYLE GUIDANCE:
- GROUPED_METRIC → explain patterns across segments
- SINGLE_METRIC → explain the overall outcome
- ROW_LEVEL → explain the scope of entities analyzed

GOOD OUTPUT EXAMPLES:
- "In the Bengaluru branch, policies are primarily held by renewed customers, with a smaller share coming from non-renewed customers."
- "Policy performance varies across customer segments, highlighting differences in renewal behavior."
- "The Bengaluru branch currently manages over 17,000 policies across its customer base."

BAD OUTPUT EXAMPLES:
- "Here’s a professional, analytical opening statement..."
- "We analyzed the SQL query and found..."
- "The dataset contains the following fields..."

Now write ONLY the opening sentence:
"""

    
    try:
        opener = llm._call(
            prompt=prompt,
            system_prompt="You are a skilled data analyst creating professional, context-aware opening statements. Focus on WHAT is being analyzed, not row counts.",
            temperature=0.7,
            max_tokens=150
        )
        
        opener = opener.strip().strip('"').strip("'")
        
        # Validation: reject if it still mentions "records" for grouped data
        if result_type in ["GROUPED_METRIC", "SINGLE_METRIC"]:
            if any(phrase in opener.lower() for phrase in [
                f"{row_count} records",
                f"{row_count} results", 
                f"{row_count} data points"
            ]):
                print(f"⚠️ Generated opener still mentions row count for grouped data, regenerating...")
                # Fallback to rule-based opener
                opener = _generate_smart_fallback_opener(question, result_type, display_count, rows)
        
        return opener
        
    except Exception as e:
        print(f"⚠️ Dynamic opener generation failed: {e}")
        return _generate_smart_fallback_opener(question, result_type, display_count, rows)


def _get_business_entity_name(question: str, result_type: str, rows: list = None) -> str:
    """Determine what business entity the data represents"""
    q_lower = question.lower()
    
    # For grouped data, interpret the grouping dimension
    if result_type in ["GROUPED_METRIC", "SINGLE_METRIC"]:
        if any(word in q_lower for word in ["by tenure", "tenure", "duration"]):
            return "tenure segments"
        elif any(word in q_lower for word in ["by state", "state", "geographic"]):
            return "states/regions"
        elif any(word in q_lower for word in ["by quarter", "quarter", "quarterly"]):
            return "quarters"
        elif any(word in q_lower for word in ["by month", "month", "monthly"]):
            return "months"
        elif any(word in q_lower for word in ["by product", "product type"]):
            return "product categories"
        elif any(word in q_lower for word in ["by branch", "branch"]):
            return "branches"
        elif any(word in q_lower for word in ["by customer type", "segment"]):
            return "customer segments"
        else:
            # Try to infer from data columns
            if rows and len(rows) > 0 and isinstance(rows[0], dict):
                cols = list(rows[0].keys())
                if any("state" in c.lower() for c in cols):
                    return "states"
                elif any("month" in c.lower() or "quarter" in c.lower() for c in cols):
                    return "time periods"
                elif any("type" in c.lower() or "category" in c.lower() for c in cols):
                    return "categories"
            return "segments"
    
    # For row-level data, identify the business entity
    if "customer" in q_lower:
        return "customers"
    elif "policy" in q_lower or "policies" in q_lower:
        return "policies"
    elif "claim" in q_lower:
        return "claims"
    elif "branch" in q_lower:
        return "branches"
    elif "transaction" in q_lower:
        return "transactions"
    else:
        return "records"


def _generate_smart_fallback_opener(
    question: str, 
    result_type: str, 
    display_count: int, 
    rows: list = None
) -> str:
    """Generate intelligent fallback opener based on result type"""
    entity_name = _get_business_entity_name(question, result_type, rows)
    
    if result_type == "SINGLE_METRIC":
        return f"We've calculated the key metric for your analysis of {question.lower()}."
    
    elif result_type == "GROUPED_METRIC":
        # For grouped data, focus on the analysis dimension
        if display_count <= 1:
            return f"We've analyzed your data for {question.lower()}."
        else:
            return f"We've analyzed patterns across {display_count} {entity_name} for your query."
    
    elif result_type == "ROW_LEVEL":
        if display_count == 0:
            return "We've completed the analysis, though no matching records were found."
        elif display_count == 1:
            return f"We've identified 1 {entity_name.rstrip('s')} matching your criteria."
        else:
            return f"We've analyzed {display_count:,} {entity_name} matching your criteria."
    
    else:
        return f"We've completed the analysis for your query about {question.lower()}."


# ============================================
# UPDATE: Ensure classify_result_type is robust
# ============================================

def classify_result_type(rows):
    """
    Determine how results should be interpreted and narrated
    """
    if not isinstance(rows, list) or not rows:
        return "EMPTY"
    
    if not rows or len(rows) == 0:
        return "EMPTY"

    # Single aggregated metric (e.g., total count, sum)
    if (
        len(rows) == 1
        and isinstance(rows[0], dict)
        and len(rows[0]) <= 2  # Usually 1-2 columns for single metrics
        and all(isinstance(v, (int, float, type(None))) for v in rows[0].values())
    ):
        return "SINGLE_METRIC"

    # Grouped / segmented metrics (small number of rows, contains aggregations)
    if all(isinstance(r, dict) for r in rows) and len(rows) <= 20:
        # Check if this looks like aggregated data
        first_row_values = list(rows[0].values())
        
        # Has mix of categorical and numeric columns (typical of GROUP BY results)
        has_categorical = any(isinstance(v, str) for v in first_row_values)
        has_numeric = any(isinstance(v, (int, float)) for v in first_row_values)
        
        # Common aggregation column names
        agg_keywords = ['count', 'sum', 'avg', 'total', 'rate', 'percentage', 'min', 'max']
        has_agg_cols = any(
            any(keyword in col.lower() for keyword in agg_keywords) 
            for col in rows[0].keys()
        )
        
        if (has_categorical and has_numeric) or has_agg_cols:
            return "GROUPED_METRIC"

    # Normal row-level data (individual records)
    return "ROW_LEVEL"


def extract_business_count(rows, result_type):
    """
    Extract the count that should be shown to users
    """
    if result_type == "EMPTY":
        return 0
    
    if result_type == "SINGLE_METRIC":
        # Return the actual metric value if it's a count
        first_row = rows[0]
        for key, value in first_row.items():
            if 'count' in key.lower() or 'total' in key.lower():
                return int(value) if value is not None else 0
        # If no count column, return the first numeric value
        return int(next(iter(first_row.values()))) if first_row.values() else 0

    if result_type == "GROUPED_METRIC":
        # For grouped data, the row count IS the business count (# of segments)
        return len(rows)

    if result_type == "ROW_LEVEL":
        # For row-level data, count the actual rows
        return len(rows)

    return len(rows)


# ---------- Enhanced Humanizer for Conversational Tone ----------
# def generate_conversational_opener(question: str, context: dict = None) -> str:
#     """Generate a completely dynamic, LLM-powered conversational opener"""
#     from .humanizer import generate_dynamic_conversational_opener, _derive_contextual_metrics
    
#     # Create rich metrics for dynamic generation
#     metrics = {
#         "row_count": 0,  # Will be updated when we have results
#         "data_size": "unknown",
#         "has_data": True,
#         "time_context": context.get('time_context') if context else {}
#     }
    
#     return generate_dynamic_conversational_opener(question, metrics)

# ---------- STREAMING WITH ENHANCED CONVERSATION ----------

import re
import time
import json
import traceback
import requests
from django.http import JsonResponse, StreamingHttpResponse
from django.views.decorators.csrf import csrf_exempt
from azure.core.exceptions import HttpResponseError, ServiceRequestError, ServiceResponseError

# ==================================================
# Helpers for classifying questions
# ==================================================

def is_incomplete_question(question: str) -> bool:
    """Detect incomplete starter questions like 'what is', 'how are', 'which'."""
    q = question.lower().strip()
    if len(q.split()) <= 2 and re.match(r"^(what|how|where|when|why|which)\b", q):
        return True
    return False

def is_fragment_followup(question: str) -> bool:
    """Detect short fragments like 'in Delhi', 'for 2025', 'by branch'."""
    q = question.lower().strip()
    if re.match(r"^(in|for|by|on|at)\b", q) and len(q.split()) <= 4:
        return True
    return False

def merge_followup_with_history(user_id: str, question: str) -> str:
    """Merge fragment with last full question if valid."""
    history = conversation_memory_store.get(user_id, [])
    if not history:
        return question
    if is_fragment_followup(question):
        last_q = history[-1].get("asked_question", "")
        return f"{last_q} {question}".strip()
    return question

# ==================================================
# Intent classification (your original + unchanged)
# ==================================================

def _classify_intent(raw_response: str) -> str:
    """Extract YES/NO/UNCERTAIN/PDF from response"""
    if not raw_response:
        return "NO"
    
    response = raw_response.strip().upper()

    if "PDF" in response:
        return "PDF"
    if "UNCERTAIN" in response:
        return "UNCERTAIN"
    elif "YES" in response:
        return "YES"
    elif "NO" in response:
        return "NO"
    
    return "NO"


# @csrf_exempt
# def check_intent(request):
#     if request.method != "POST":
#         return JsonResponse({"answer": "NO", "error": "Only POST allowed"}, status=405)

#     try:
#         data = json.loads(request.body or "{}")
#     except Exception as e:
#         logger.error(f"JSON parsing error: {e}")
#         return JsonResponse({"answer": "NO", "error": "Invalid JSON body"}, status=400)

#     question = (data.get("question") or "").strip()
#     if not question:
#         return JsonResponse({"answer": "NO"})

#     # 🆕 First-level check for incomplete questions
#     if is_incomplete_question(question):
#         return JsonResponse({"answer": "UNCERTAIN", "message": "Incomplete question. Please rephrase into a full question."})

#     prompt = f"""
# Classify the user's intent as YES, NO, UNCERTAIN, or PDF.

# RULES:
# 1. **NO (General Knowledge/Greeting):**
#    - Greetings: "hi", "hello", "how are you"
#    - General world knowledge: "who is the PM of India?", "what is machine learning?", "what is AI?"
#    - Casual explanations unrelated to insurance/business context

# 2. **YES (Clear Database Query):**
#    - Statistical queries: "top 5 branches by churn", "how many policies in Delhi?"
#    - Data analysis: "total premium in 2024", "branch performance summary"
#    - Complete questions asking for metrics/counts/summaries from DB

# 3. **PDF (Business Definitions/Policy Knowledge):**
#    - Definitions of domain-specific/business terms: "what is a branch?", "what is customer segment?", "what is retention?"
#    - Policy/insurance terms: "what is claim ratio?", "define surrender value"
#    - Explanatory business context that should be retrieved from PDF/vector knowledge base

# 4. **UNCERTAIN (Vague/Ambiguous Fragments):**
#    - Incomplete fragments: "top 5?", "Delhi?", "premium?"
#    - Vague context: "me branches", "what about?"
#    - Missing key info (metric, time period, entity)

# Examples:
# "What is a branch?" → PDF
# "What is customer segment?" → PDF
# "Top 5 branches by churn in 2024" → YES
# "top 5?" → UNCERTAIN
# "Delhi?" → UNCERTAIN
# "premium?" → UNCERTAIN
# "Who is the PM of India?" → NO
# "What is machine learning?" → NO

# Question: "{question}"

# Respond with exactly one word: YES, NO, UNCERTAIN, or PDF
# """.strip()

#     try:
#         client = _get_client()
#         logger.info(f"Making request to Azure AI with model: {AZURE_MODEL}")

#         response = client.complete(
#             messages=[
#                 SystemMessage(content="You are an intent classifier. Reply only with YES, NO, UNCERTAIN, or PDF."),
#                 UserMessage(content=prompt),
#             ],
#             model=AZURE_MODEL,
#             temperature=0.0,
#             top_p=1.0,
#             max_tokens=15,
#         )

#         raw = response.choices[0].message.content if response.choices else ""
#         answer = _classify_intent(raw)

#         logger.info(f"Intent classification result: {answer} (raw: {raw})")

#         if answer == "UNCERTAIN":
#             history = conversation_memory_store.get("admin", [])
#             tokens = extract_dynamic_tokens(question)

#             history_suggestions = []
#             for h in history[-10:]:
#                 asked = h.get("asked_question", "")
#                 if any(t in asked.lower() for t in tokens):
#                     history_suggestions.append(asked)

#             related = retrieve_context_from_corpus(question, db_id="liberty", min_score=0.30)
#             semantic_suggestions = []
#             if related:
#                 for r in related[:10]:
#                     qtxt = (r.get("asked_question") or r.get("raw_question") or r.get("question") or "").strip()
#                     if qtxt:
#                         semantic_suggestions.append(qtxt)

#             keyword_suggestions = [q for q in semantic_suggestions if any(t in q.lower() for t in tokens)]

#             suggestions = (
#                 history_suggestions
#                 + [q for q in keyword_suggestions if q not in history_suggestions]
#                 + [q for q in semantic_suggestions if q not in keyword_suggestions and q not in history_suggestions]
#             )

#             return JsonResponse({
#                 "answer": "UNCERTAIN",
#                 "asked_question": question,
#                 "tokens_detected": tokens,
#                 "previous_questions": history_suggestions[:5],
#                 "related_questions": [
#                     (r.get("asked_question") or r.get("question"))
#                     for r in related if (r.get("asked_question") or r.get("question"))
#                 ]
#             })

#         return JsonResponse({"answer": answer})

#     except HttpResponseError as e:
#         logger.error(f"Azure HTTP error: {e.status_code} - {e.message}")
#         return JsonResponse({"answer": "NO", "error": f"Azure service error: {e.message}"}, status=502)
        
#     except (ServiceRequestError, ServiceResponseError) as e:
#         logger.error(f"Azure service error: {e}")
#         return JsonResponse({"answer": "NO", "error": f"Azure service error: {str(e)}"}, status=502)
        
#     except (requests.exceptions.ConnectionError, requests.exceptions.Timeout) as e:
#         logger.error(f"Connection error: {e}")
#         return JsonResponse({"answer": "NO", "error": "Connection timeout or error"}, status=502)
        
#     except Exception as e:
#         logger.exception("check_intent failed with unexpected error")
#         return JsonResponse({"answer": "NO", "error": f"Unexpected error: {str(e)}"}, status=500)
    

import re
import time
import json
import traceback
import requests
from django.http import JsonResponse, StreamingHttpResponse
from django.views.decorators.csrf import csrf_exempt
from azure.core.exceptions import HttpResponseError, ServiceRequestError, ServiceResponseError

# ==================================================
# Helpers for classifying questions
# ==================================================

def is_incomplete_question(question: str) -> bool:
    """Detect incomplete starter questions like 'what is', 'how are', 'which'."""
    q = question.lower().strip()
    if len(q.split()) <= 2 and re.match(r"^(what|how|where|when|why|which)\b", q):
        return True
    return False

def is_fragment_followup(question: str) -> bool:
    """Detect short fragments like 'in Delhi', 'for 2025', 'by branch'."""
    q = question.lower().strip()
    if re.match(r"^(in|for|by|on|at)\b", q) and len(q.split()) <= 4:
        return True
    return False

def merge_followup_with_history(user_id: str, question: str) -> str:
    """Merge fragment with last full question if valid."""
    history = conversation_memory_store.get(user_id, [])
    if not history:
        return question
    if is_fragment_followup(question):
        last_q = history[-1].get("asked_question", "")
        return f"{last_q} {question}".strip()
    return question

# ==================================================
# Intent classification (your original + unchanged)
# ==================================================

def _classify_intent(raw_response: str) -> str:
    """Extract YES/NO/UNCERTAIN/PDF from response"""
    if not raw_response:
        return "NO"
    
    response = raw_response.strip().upper()

    if "PDF" in response:
        return "PDF"
    if "UNCERTAIN" in response:
        return "UNCERTAIN"
    elif "YES" in response:
        return "YES"
    elif "NO" in response:
        return "NO"
    
    return "NO"


@csrf_exempt
def check_intent(request):
    if request.method != "POST":
        return JsonResponse({"answer": "NO", "error": "Only POST allowed"}, status=405)

    try:
        data = json.loads(request.body or "{}")
    except Exception as e:
        logger.error(f"JSON parsing error: {e}")
        return JsonResponse({"answer": "NO", "error": "Invalid JSON body"}, status=400)

    question = (data.get("question") or "").strip()
    if not question:
        return JsonResponse({"answer": "NO"})

    # 🆕 First-level check for incomplete questions
    if is_incomplete_question(question):
        return JsonResponse({"answer": "UNCERTAIN", "message": "Incomplete question. Please rephrase into a full question."})

    prompt = f"""
Classify the user's intent as YES, NO, UNCERTAIN, or PDF.

RULES:
1. **NO (General Knowledge/Greeting):**
   - Greetings: "hi", "hello", "how are you"
   - General world knowledge: "who is the PM of India?", "what is machine learning?", "what is AI?"
   - Casual explanations unrelated to insurance/business context

2. **YES (Clear Database Query):**
   - Statistical queries: "top 5 branches by churn", "how many policies in Delhi?"
   - Data analysis: "total premium in 2024", "branch performance summary"
   - Complete questions asking for metrics/counts/summaries from DB

3. **PDF (Business Definitions/Policy Knowledge):**
   - Definitions of domain-specific/business terms: "what is a branch?", "what is customer segment?", "what is retention?"
   - Policy/insurance terms: "what is claim ratio?", "define surrender value"
   - Explanatory business context that should be retrieved from PDF/vector knowledge base

4. **UNCERTAIN (Vague/Ambiguous Fragments):**
   - Incomplete fragments: "top 5?", "Delhi?", "premium?"
   - Vague context: "me branches", "what about?"
   - Missing key info (metric, time period, entity)

Examples:
- "What is a branch?" → PDF
- "What is customer segment?" → PDF
- "Top 5 branches by churn in 2024" → YES
- "top 5?" → UNCERTAIN
- "Delhi?" → UNCERTAIN
- "premium?" → UNCERTAIN
- "Who is the PM of India?" → NO
- "What is machine learning?" → NO

Question: "{question}"

Respond with exactly one word: YES, NO, UNCERTAIN, or PDF
""".strip()

    try:
        client = _get_client()
        logger.info(f"Making request to Azure AI with model: {AZURE_MODEL}")

        response = client.complete(
            messages=[
                SystemMessage(content="You are an intent classifier. Reply only with YES, NO, UNCERTAIN, or PDF."),
                UserMessage(content=prompt),
            ],
            model=AZURE_MODEL,
            temperature=0.0,
            top_p=1.0,
            max_tokens=15,
        )

        raw = response.choices[0].message.content if response.choices else ""
        answer = _classify_intent(raw)

        logger.info(f"Intent classification result: {answer} (raw: {raw})")

        if answer == "UNCERTAIN":
            history = conversation_memory_store.get("admin", [])
            tokens = extract_dynamic_tokens(question)

            history_suggestions = []
            for h in history[-10:]:
                asked = h.get("asked_question", "")
                if any(t in asked.lower() for t in tokens):
                    history_suggestions.append(asked)

            related = retrieve_context_from_corpus(question, db_id="liberty", min_score=0.30)
            semantic_suggestions = []
            if related:
                for r in related[:10]:
                    qtxt = (r.get("asked_question") or r.get("raw_question") or r.get("question") or "").strip()
                    if qtxt:
                        semantic_suggestions.append(qtxt)

            keyword_suggestions = [q for q in semantic_suggestions if any(t in q.lower() for t in tokens)]

            suggestions = (
                history_suggestions
                + [q for q in keyword_suggestions if q not in history_suggestions]
                + [q for q in semantic_suggestions if q not in keyword_suggestions and q not in history_suggestions]
            )

            return JsonResponse({
                "answer": "UNCERTAIN",
                "asked_question": question,
                "tokens_detected": tokens,
                "previous_questions": history_suggestions[:5],
                "related_questions": [
                    (r.get("asked_question") or r.get("question"))
                    for r in related if (r.get("asked_question") or r.get("question"))
                ]
            })

        return JsonResponse({"answer": answer})

    except HttpResponseError as e:
        logger.error(f"Azure HTTP error: {e.status_code} - {e.message}")
        return JsonResponse({"answer": "NO", "error": f"Azure service error: {e.message}"}, status=502)
        
    except (ServiceRequestError, ServiceResponseError) as e:
        logger.error(f"Azure service error: {e}")
        return JsonResponse({"answer": "NO", "error": f"Azure service error: {str(e)}"}, status=502)
        
    except (requests.exceptions.ConnectionError, requests.exceptions.Timeout) as e:
        logger.error(f"Connection error: {e}")
        return JsonResponse({"answer": "NO", "error": "Connection timeout or error"}, status=502)
        
    except Exception as e:
        logger.exception("check_intent failed with unexpected error")
        return JsonResponse({"answer": "NO", "error": f"Unexpected error: {str(e)}"}, status=500)



# def query_corpus(question, n_results=5, min_score=0.80):
#     """Utility: search corpus and return top matches."""
#     chroma_client = chromadb.PersistentClient(path="corpus_db")
#     collection = chroma_client.get_or_create_collection("corpus_liberty")

#     entities = extract_entities(question)
#     expanded_queries = expand_query_with_entities(question, entities)

#     main_items = collection.query(query_texts=expanded_queries, n_results=n_results)
#     results = []
#     for doc_id, doc, meta in zip(
#         main_items["ids"][0],
#         main_items["documents"][0],
#         main_items["metadatas"][0],
#     ):
#         try:
#             parsed = json.loads(doc)
#         except:
#             parsed = {"question": doc}

#         # Fallback: always have a display question
#         display_q = (
#             parsed.get("normalized_question")
#             or parsed.get("question")
#             or ""
#         )
#         display_q = replace_placeholders(display_q, entities)
#         if re.search(r'<[A-Z_]+>', display_q):
#             display_q = rewrite_with_entities(display_q, question)

#         results.append({
#             "id": doc_id,
#             "content": parsed,
#             "metadata": meta,
#             "display_question": display_q,
#         })

#     return results


import re
import re


def sanitize_sql_before_exec(sql, question):
    """
    Sanitize SQL - ONLY remove truly orphaned WHERE clauses.
    """
    if not sql or not sql.strip():
        return "SELECT COUNT(*) as total_count FROM \"bi_dwh\".\"main_cai_lib\" WHERE policy_end_date_year IS NOT NULL"
    
    # First fix missing WHERE keywords
    sql = fix_missing_where_keyword(sql)
    
    # Now check for orphaned WHERE keywords
    lines = sql.split('\n')
    result_lines = []
    
    i = 0
    while i < len(lines):
        line = lines[i]
        stripped = line.strip()
        
        # Keep empty lines
        if not stripped:
            result_lines.append(line)
            i += 1
            continue
        
        # Only check lines that are EXACTLY "WHERE" and nothing else
        if stripped == 'WHERE':
            # Look ahead for a condition
            found_condition = False
            for j in range(i + 1, min(i + 5, len(lines))):
                next_stripped = lines[j].strip()
                if not next_stripped:
                    continue
                # If next non-empty line is NOT a SQL keyword, it's a condition
                keywords = ['GROUP', 'ORDER', 'LIMIT', 'HAVING', 'UNION', 
                           'SELECT', 'FROM', 'WHERE', 'WITH', ')']
                if not any(next_stripped.upper().startswith(kw) for kw in keywords):
                    found_condition = True
                break
            
            if not found_condition:
                print(f"🔧 Removing orphaned WHERE (no condition follows)")
                i += 1
                continue
        
        result_lines.append(line)
        i += 1
    
    final_sql = '\n'.join(result_lines)
    
    # Safety check
    if not final_sql.strip() or final_sql.strip().endswith(('WHERE', 'AND', 'OR')):
        print("🔧 SQL still invalid, using fallback")
        return "SELECT COUNT(*) as total_count FROM \"bi_dwh\".\"main_cai_lib\" WHERE policy_end_date_year IS NOT NULL"
    
    print("🔧 SQL sanitization complete")
    return final_sql

def semantic_question_match(user_question, corpus_question, similarity_score, min_similarity=0.50):
    """
    Validate that corpus question semantically matches user question.
    
    Args:
        user_question: User's normalized question
        corpus_question: Corpus normalized question
        similarity_score: Cosine similarity score
        min_similarity: Minimum similarity threshold (default 0.50 = 50%)
    
    Returns:
        (is_match: bool, reason: str)
    """
    # Check 1: Similarity threshold
    if similarity_score < min_similarity:
        return False, f"Similarity too low: {similarity_score:.3f} < {min_similarity}"
    
    # Check 2: Key intent words matching
    # Extract key business terms
    user_terms = set(re.findall(r'\b(?:total|count|number|sum|average|breakdown|analyze|compare|trend|churn|retention|customer|policy|premium|claim|revenue|profit|loss)\b', 
                                 user_question.lower()))
    corpus_terms = set(re.findall(r'\b(?:total|count|number|sum|average|breakdown|analyze|compare|trend|churn|retention|customer|policy|premium|claim|revenue|profit|loss)\b', 
                                   corpus_question.lower()))
    
    # At least 50% of user's key terms should be in corpus
    if user_terms:
        overlap = len(user_terms & corpus_terms) / len(user_terms)
        if overlap < 0.5:
            return False, f"Intent mismatch: User terms {user_terms} vs Corpus terms {corpus_terms} (overlap: {overlap:.1%})"
    
    # Check 3: Question type matching (what/how/when/which/analyze/compare)
    user_type = None
    corpus_type = None
    
    question_patterns = {
        'count': r'\b(?:total|count|number|how many)\b',
        'analyze': r'\b(?:analyze|analysis|pattern|trend|breakdown)\b',
        'compare': r'\b(?:compare|comparison|vs|versus|difference)\b',
        'average': r'\b(?:average|mean|typical)\b',
        'sum': r'\b(?:sum|total|aggregate)\b',
    }
    
    for qtype, pattern in question_patterns.items():
        if re.search(pattern, user_question.lower()):
            user_type = qtype
        if re.search(pattern, corpus_question.lower()):
            corpus_type = qtype
    
    if user_type and corpus_type and user_type != corpus_type:
        return False, f"Question type mismatch: User asks '{user_type}' vs Corpus answers '{corpus_type}'"
    
    return True, "Semantic match validated"


def query_corpus(question, user_entities=None, n_results=5, min_score=0.50):
    """
    Query corpus with STRICT entity matching AND semantic validation.
    
    Args:
        question: Normalized question template (with placeholders like <MONTH>, <YEAR>)
        user_entities: Dict of extracted entities from ORIGINAL question (NOT normalized)
        n_results: Number of results to return
        min_score: Minimum similarity score threshold (increased to 0.50 for better matches)
    
    Returns:
        List of matching corpus entries with exact entity alignment AND semantic match
    """
    print(f"🔍 query_corpus called with: question='{question}', min_score={min_score}")
    print(f"🔍 User entities passed: {user_entities}")
    
    try:
        chroma_client = chromadb.PersistentClient(path="corpus_db")
        collection = chroma_client.get_or_create_collection("corpus_liberty")
        
        # 🔥 CRITICAL: Use passed user_entities, don't extract again from normalized question
        if user_entities is None:
            user_entities = {}
            print("⚠️ WARNING: No user_entities passed, will extract from question")
            # Only extract if not provided (fallback)
            user_entities = extract_entities(question) if 'extract_entities' in globals() else {}
        
        print(f"🔍 Final user entities for matching: {user_entities}")
        
        # Expand queries for semantic search
        expanded_queries = expand_query_with_entities(question, user_entities) if 'expand_query_with_entities' in globals() else [question]
        print(f"🔍 Expanded queries: {expanded_queries}")
        
        # Semantic search from ChromaDB
        main_items = collection.query(query_texts=expanded_queries, n_results=n_results)
        print(f"🔍 ChromaDB returned: {len(main_items.get('ids', [[]])[0])} results")
        
        # Get distances for similarity calculation
        distances = main_items.get("distances", [[]])[0] if main_items.get("distances") else []
        print(f"🔍 Raw distances from ChromaDB: {distances}")
        
        results = []
        for i, (doc_id, doc, meta) in enumerate(zip(
            main_items["ids"][0],
            main_items["documents"][0],
            main_items["metadatas"][0],
        )):
            try:
                parsed = json.loads(doc)
            except:
                parsed = {"question": doc}
            
            # Calculate similarity score
            if distances and i < len(distances):
                cosine_distance = distances[i]
                similarity_score = max(0.0, min(1.0, 1.0 - (cosine_distance / 2.0)))
            else:
                similarity_score = 1.0
                
            print(f"🔍 Processing result {i}: distance={cosine_distance:.3f}, similarity={similarity_score:.3f}")
            
            # Skip if similarity too low
            if similarity_score < min_score:
                print(f"🔍 Skipping result {i}: similarity {similarity_score:.3f} < {min_score}")
                continue
            
            # 🔥 CRITICAL: Extract entities from corpus ORIGINAL question (not normalized)
            corpus_question = parsed.get("asked_question") or parsed.get("question") or ""
            print(f"🔍 Corpus original question: '{corpus_question}'")
            
            # Extract entities from corpus original question
            corpus_entities = extract_entities(corpus_question) if 'extract_entities' in globals() else {}
            print(f"🔍 Corpus question entities: {corpus_entities}")
            
            # 🔥 STRICT ENTITY MATCHING - ALL entities must match exactly
            entities_match, mismatch_reasons = entities_exact_match(user_entities, corpus_entities)
            
            if not entities_match:
                print(f"❌ REJECTED result {i}: Entity mismatch - {', '.join(mismatch_reasons)}")
                print(f"   User question entities: {user_entities}")
                print(f"   Corpus question: '{corpus_question}'")
                print(f"   Corpus entities: {corpus_entities}")
                continue
            
            print(f"✅ Entity check passed for result {i}")
            
            # 🔥 NEW: SEMANTIC VALIDATION - Questions must be semantically similar
            corpus_normalized = parsed.get("normalized_question") or corpus_question
            semantic_match, semantic_reason = semantic_question_match(
                question, corpus_normalized, similarity_score, min_similarity=min_score
            )
            
            if not semantic_match:
                print(f"❌ REJECTED result {i}: Semantic mismatch - {semantic_reason}")
                print(f"   User question: '{question}'")
                print(f"   Corpus question: '{corpus_normalized}'")
                continue
            
            print(f"✅ ACCEPTED result {i}: All validations passed")
            print(f"   ✓ Entities match: {user_entities} == {corpus_entities}")
            print(f"   ✓ Semantic match: {semantic_reason}")
            print(f"   ✓ Similarity: {similarity_score:.3f}")
            
            # Build display question
            display_q = (
                parsed.get("normalized_question")
                or parsed.get("question")
                or ""
            )
            
            # Replace placeholders with user's actual entities
            if 'replace_placeholders' in globals():
                display_q = replace_placeholders(display_q, user_entities)
            if re.search(r'<[A-Z_]+>', display_q) and 'rewrite_with_entities' in globals():
                display_q = rewrite_with_entities(display_q, corpus_question)
            
            result = {
                "id": doc_id,
                "content": parsed,
                "metadata": meta,
                "display_question": display_q,
                "similarity_score": similarity_score,
                "cosine_distance": cosine_distance,
                "user_entities": user_entities,
                "corpus_entities": corpus_entities,
                "entities_match": True,  # Only included if all match
                "semantic_match": semantic_match,
                "semantic_reason": semantic_reason
            }
            results.append(result)
            print(f"🔍 Added result {i}: similarity={similarity_score:.3f}, question='{corpus_question}'")
        
        print(f"🔍 query_corpus returning {len(results)} results (after strict entity + semantic filtering)")
        return results
        
    except Exception as e:
        print(f"❌ Error in query_corpus: {e}")
        import traceback
        print(traceback.format_exc())
        return []


import re

def fix_missing_where_keyword(sql):
    """Fix missing WHERE keywords in subqueries."""
    if not sql:
        return sql
    
    fixed = False
    
    # Pattern: FROM "schema"."table"\n    column_name IS
    pattern = r'(FROM\s+"[^"]+"\."[^"]+"\s*\n)(\s+)(\w+\s+(?:IS\s+(?:NOT\s+)?NULL|ILIKE|LIKE|=|!=|<>|IN\s*\())'
    
    def replacer(match):
        nonlocal fixed
        fixed = True
        return f"{match.group(1)}{match.group(2)}WHERE {match.group(3)}"
    
    result = re.sub(pattern, replacer, sql, flags=re.IGNORECASE)
    
    if fixed:
        print("🔧 Added missing WHERE keyword to subquery")
    
    return result


def sanitize_sql_before_exec(sql, question):
    """
    Sanitize SQL - ONLY remove truly orphaned WHERE clauses.
    """
    if not sql or not sql.strip():
        return "SELECT COUNT(*) as total_count FROM \"bi_dwh\".\"main_cai_lib\" WHERE policy_end_date_year IS NOT NULL"
    
    # First fix missing WHERE keywords
    sql = fix_missing_where_keyword(sql)
    
    # Now check for orphaned WHERE keywords
    lines = sql.split('\n')
    result_lines = []
    
    i = 0
    while i < len(lines):
        line = lines[i]
        stripped = line.strip()
        
        # Keep empty lines
        if not stripped:
            result_lines.append(line)
            i += 1
            continue
        
        # Only check lines that are EXACTLY "WHERE" and nothing else
        if stripped == 'WHERE':
            # Look ahead for a condition
            found_condition = False
            for j in range(i + 1, min(i + 5, len(lines))):
                next_stripped = lines[j].strip()
                if not next_stripped:
                    continue
                # If next non-empty line is NOT a SQL keyword, it's a condition
                keywords = ['GROUP', 'ORDER', 'LIMIT', 'HAVING', 'UNION', 
                           'SELECT', 'FROM', 'WHERE', 'WITH', ')']
                if not any(next_stripped.upper().startswith(kw) for kw in keywords):
                    found_condition = True
                break
            
            if not found_condition:
                print(f"🔧 Removing orphaned WHERE (no condition follows)")
                i += 1
                continue
        
        result_lines.append(line)
        i += 1
    
    final_sql = '\n'.join(result_lines)
    
    # Safety check
    if not final_sql.strip() or final_sql.strip().endswith(('WHERE', 'AND', 'OR')):
        print("🔧 SQL still invalid, using fallback")
        return "SELECT COUNT(*) as total_count FROM \"bi_dwh\".\"main_cai_lib\" WHERE policy_end_date_year IS NOT NULL"
    
    print("🔧 SQL sanitization complete")
    return final_sql


def fix_round_for_postgres(sql):
    """
    Fix PostgreSQL ROUND() type errors.
    Strategy: Only cast the numerator in division operations.
    """
    if not sql:
        return sql
    
    # Match: <anything>::numeric / NULLIF(...)
    # Just ensure the pattern is preserved
    pattern = r'ROUND\s*\(\s*(.+?)\s*,\s*(\d+)\s*\)'
    
    def replacer(match):
        inner_expr = match.group(1)
        digits = match.group(2)
        
        # Check if this is a division expression
        if '/' in inner_expr:
            parts = inner_expr.split('/', 1)
            numerator = parts[0].strip()
            denominator = parts[1].strip()
            
            # Only cast numerator if not already cast
            if '::numeric' not in numerator.lower():
                numerator = f'({numerator})::NUMERIC'
            
            return f'ROUND({numerator} / {denominator}, {digits})'
        else:
            # Not a division, cast the whole expression
            if '::numeric' not in inner_expr.lower():
                return f'ROUND(({inner_expr})::NUMERIC, {digits})'
            return f'ROUND({inner_expr}, {digits})'
    
    return re.sub(pattern, replacer, sql, flags=re.IGNORECASE)


def fix_nullif_for_postgres(sql):
    """
    Ensure NULLIF has both required arguments.
    If NULLIF is missing second argument, add default of 0.
    """
    if not sql:
        return sql
    
    # Find NULLIF with only one argument and add , 0
    pattern = r"NULLIF\s*\(\s*([^,)]+)\s*\)(?!\s*,)"
    
    def replacer(match):
        expr = match.group(1)
        return f"NULLIF({expr}, 0)"
    
    fixed_sql = re.sub(pattern, replacer, sql, flags=re.IGNORECASE)
    return fixed_sql



import time
import json
import traceback
from django.http import StreamingHttpResponse, JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from sqlalchemy import text, event
from sqlalchemy.engine import Engine
import threading

# ============== DATABASE TIMEOUT HANDLER ==============

class DatabaseTimeoutError(Exception):
    """Raised when database query exceeds timeout"""
    pass

# Add query timeout monitoring to all SQL executions
@event.listens_for(Engine, "before_cursor_execute")
def receive_before_cursor_execute(conn, cursor, statement, params, context, executemany):
    """Log query start time"""
    conn.info.setdefault('query_start_time', []).append(time.time())
    print(f"🔍 Executing query (timeout: 60s)")
    try:
        pool = ENGINE.pool
        print(f"🔌 Pool status - Size: {pool.size()}, Checked out: {pool.checkedout()}, Overflow: {pool.overflow()}")
    except:
        pass

@event.listens_for(Engine, "after_cursor_execute")
def receive_after_cursor_execute(conn, cursor, statement, params, context, executemany):
    """Log query completion time"""
    total = time.time() - conn.info['query_start_time'].pop()
    print(f"✅ Query completed in {total:.2f}s")



def execute_with_thread_timeout(func, args=(), kwargs=None, timeout_seconds=60):
    """
    Execute a function in a separate thread with hard timeout
    Works on all platforms (Windows/Linux/Mac)
    
    CRITICAL FIX: Properly unpack args tuple
    """
    if kwargs is None:
        kwargs = {}
    
    result_holder = {'result': None, 'error': None, 'completed': False}
    
    def worker():
        try:
            result_holder['result'] = func(*args, **kwargs)
            result_holder['completed'] = True
        except Exception as e:
            result_holder['error'] = e
            result_holder['completed'] = True
    
    thread = threading.Thread(target=worker, daemon=True)
    thread.start()
    thread.join(timeout=timeout_seconds)
    
    if not result_holder['completed']:
        raise DatabaseTimeoutError(f"Operation exceeded {timeout_seconds} second timeout")
    
    if result_holder['error']:
        raise result_holder['error']
    
    return result_holder['result']

# ============== SAFE SQL EXECUTION WRAPPER ==============
def execute_sql_with_timeout(
    sql,
    timeout_seconds=600,
    max_rows=None,
    question=None,
    user_entities=None,
    llm=None,
    corpus_results=None,
):
    """
    FIXED VERSION: Only check corpus for optimized SQL AFTER timeout.
    
    Flow:
    1. Always try original SQL first
    2. On timeout → Check corpus for cached optimization
    3. If not in corpus → Generate new optimization with LLM
    4. Store successful optimizations in corpus
    """
    rows = []
    sql_used = sql
    was_optimized = False
    conn = None
    
    # Cap timeout to PostgreSQL's max
    MAX_TIMEOUT_SECONDS = 600
    timeout_seconds = min(timeout_seconds, MAX_TIMEOUT_SECONDS)

    user_entities = user_entities or {}
    corpus_results = corpus_results or []

    # ------------------------------------------------------------------
    # STEP 1: Try ORIGINAL SQL first (NO corpus lookup yet)
    # ------------------------------------------------------------------
    try:
        print(f"🔧 Executing SQL (timeout={timeout_seconds}s)")
        print(f"🔍 SQL Preview: {sql[:300]}...")

        # Acquire connection with timeout protection
        def get_connection():
            return ENGINE.connect()

        try:
            conn = execute_with_thread_timeout(
                get_connection, 
                args=(),
                timeout_seconds=10
            )
            print("✅ Database connection acquired")
        except DatabaseTimeoutError:
            raise Exception(
                "Connection timeout: Database is busy. Please retry in 30 seconds."
            )

        # Apply PostgreSQL-level safety timeouts
        try:
            timeout_ms = min(timeout_seconds * 1000, 2147483647)
            conn.execute(text(f"SET statement_timeout = {timeout_ms}"))
            conn.execute(text("SET lock_timeout = '30s'"))
            conn.execute(text("SET idle_in_transaction_session_timeout = '60s'"))
            print(f"✅ Timeout set to {timeout_ms}ms ({timeout_seconds}s)")
        except Exception as e:
            print(f"⚠️ Timeout settings failed: {e}")

        start_time = time.time()

        # Execute query (protected)
        def run_query():
            return conn.execute(text(sql))

        try:
            result = execute_with_thread_timeout(
                run_query, 
                args=(),
                timeout_seconds=timeout_seconds
            )
        except DatabaseTimeoutError:
            try:
                conn.execute(text("SELECT pg_cancel_backend(pg_backend_pid())"))
            except:
                pass
            
            # ⚠️ TIMEOUT OCCURRED - NOW check corpus for optimization
            raise DatabaseTimeoutError(
                f"Query timeout: Took longer than {timeout_seconds}s."
            )

        # Fetch rows safely
        row_count = 0
        for row in result:
            elapsed = time.time() - start_time
            if elapsed > timeout_seconds * 0.9:
                print("⚠️ Near timeout during fetch, stopping early")
                break

            if max_rows and row_count >= max_rows:
                print(f"⚠️ Hit max row limit ({max_rows})")
                break

            rows.append({k: _jsonable(v) for k, v in dict(row._mapping).items()})
            row_count += 1

            if row_count % 10000 == 0:
                print(f"📊 Fetched {row_count} rows")

        execution_time = time.time() - start_time
        print(f"✅ SQL success: {len(rows)} rows in {execution_time:.2f}s")

        conn.commit()

        return (rows, sql, False) if question else rows

    # ==================================================================
    # TIMEOUT HANDLER: Check corpus → Generate optimization → Retry
    # ==================================================================
    except DatabaseTimeoutError as timeout_error:
        error_msg = str(timeout_error)
        print(f"⏱️ Timeout occurred: {error_msg}")

        if not question or not llm:
            raise Exception(error_msg + " Please add more specific filters.")

        # ------------------------------------------------------------------
        # STEP 2: NOW check corpus for cached optimization
        # ------------------------------------------------------------------
        optimized_sql = None
        optimization_source = None
        
        print("🔍 Checking corpus for cached optimization...")
        optimized_cache = retrieve_optimized_sql_from_corpus(
            question=question,
            current_sql=sql,
            user_entities=user_entities,
            corpus_results=corpus_results
        )
        
        if optimized_cache and optimized_cache.get("sql"):
            optimized_sql = optimized_cache["sql"]
            optimization_source = "corpus"
            print(f"✅ Found cached optimized SQL in corpus")
            print(f"   Previous execution: {optimized_cache.get('execution_time', 0):.2f}s")
        
        # ------------------------------------------------------------------
        # STEP 3: If not in corpus, generate new optimization with LLM
        # ------------------------------------------------------------------
        if not optimized_sql:
            print("🤖 No cached optimization found, generating with LLM...")
            try:
                optimized_sql = generate_optimized_sql_with_indexes(
                    original_sql=sql,
                    question=question,
                    timeout_error=error_msg,
                    llm=llm
                )
                optimization_source = "llm_generated"
                print(f"✅ Generated new optimized SQL")
            except Exception as opt_gen_error:
                print(f"❌ LLM optimization failed: {opt_gen_error}")
                raise Exception(
                    f"Query timeout and optimization failed. "
                    f"Original error: {error_msg}. Please add filters."
                )

        # ------------------------------------------------------------------
        # STEP 4: Retry with optimized SQL
        # ------------------------------------------------------------------
        print(f"🚀 Retrying with {optimization_source} SQL (timeout={timeout_seconds}s)...")
        
        try:
            optimized_rows, optimized_sql_used, _ = execute_sql_with_timeout_internal(
                optimized_sql,
                timeout_seconds=timeout_seconds,
                max_rows=max_rows,
                question=question,
                user_entities=user_entities,
                was_optimized=True
            )
            
            print(f"✅ Optimized SQL succeeded: {len(optimized_rows)} rows")
            
            # ------------------------------------------------------------------
            # STEP 5: Store successful optimization in corpus (if LLM-generated)
            # ------------------------------------------------------------------
            if optimization_source == "llm_generated":
                query_sig = extract_query_signature(optimized_sql, question, user_entities)
                store_optimized_sql_in_corpus(
                    question=question,
                    original_sql=sql,
                    optimized_sql=optimized_sql,
                    execution_time=time.time() - start_time,
                    row_count=len(optimized_rows),
                    user_entities=user_entities,
                    query_signature=query_sig
                )
                print("📝 Stored new optimization in corpus")
            
            return (optimized_rows, optimized_sql, True) if question else optimized_rows
            
        except Exception as retry_error:
            print(f"❌ Optimized retry also failed: {retry_error}")
            raise Exception(
                f"Query optimization failed. Original error: {error_msg}. "
                f"Optimization error: {str(retry_error)}"
            )

    # ==================================================================
    # OTHER ERRORS (syntax, connection, etc.)
    # ==================================================================
    except Exception as e:
        error_msg = str(e)
        print(f"❌ SQL failed: {error_msg}")

        # Friendly error mapping
        if "syntax error" in error_msg.lower():
            raise Exception(f"SQL syntax error: {error_msg}")
        elif "does not exist" in error_msg.lower():
            raise Exception("Database error: Table or column does not exist.")
        elif "out of memory" in error_msg.lower():
            raise Exception("Query returned too much data. Add filters.")
        elif "connection" in error_msg.lower() or "busy" in error_msg.lower():
            raise Exception("Database connection error. Please retry in 30 seconds.")
        else:
            raise Exception(f"Database error: {error_msg}")

    finally:
        if conn:
            try:
                conn.close()
                print("🔌 Connection released")
            except:
                pass

# Helper: Internal executor for retry (prevents infinite recursion)
# ------------------------------------------------------------------
def execute_sql_with_timeout_internal(
    sql,
    timeout_seconds,
    max_rows,
    question,
    user_entities,
    was_optimized
):
    """
    Internal executor for optimized SQL retry.
    Does NOT trigger recursive optimization.
    """
    rows = []
    conn = None
    
    try:
        print(f"🔧 Executing optimized SQL (timeout={timeout_seconds}s)")
        
        def get_connection():
            return ENGINE.connect()

        try:
            conn = execute_with_thread_timeout(
                get_connection, 
                args=(),
                timeout_seconds=10
            )
        except DatabaseTimeoutError:
            raise Exception("Connection timeout: Database is busy.")

        try:
            timeout_ms = min(timeout_seconds * 1000, 2147483647)
            conn.execute(text(f"SET statement_timeout = {timeout_ms}"))
            conn.execute(text("SET lock_timeout = '30s'"))
        except Exception as e:
            print(f"⚠️ Timeout settings failed: {e}")

        start_time = time.time()

        def run_query():
            return conn.execute(text(sql))

        try:
            result = execute_with_thread_timeout(
                run_query, 
                args=(),
                timeout_seconds=timeout_seconds
            )
        except DatabaseTimeoutError:
            raise Exception(f"Optimized query still timed out after {timeout_seconds}s")

        row_count = 0
        for row in result:
            if max_rows and row_count >= max_rows:
                break
            rows.append({k: _jsonable(v) for k, v in dict(row._mapping).items()})
            row_count += 1

        execution_time = time.time() - start_time
        print(f"✅ Optimized SQL success: {len(rows)} rows in {execution_time:.2f}s")

        conn.commit()

        return (rows, sql, True)

    finally:
        if conn:
            try:
                conn.close()
            except:
                pass

def execute_sql_with_timeout2912(
    sql,
    timeout_seconds=600,
    max_rows=None,
    question=None,
    user_entities=None,
    llm=None,
    corpus_results=None,
):
    """
    Robust SQL executor with proper timeout handling and optimization
    
    KEY FIXES:
    1. Fixed args passing to run_query (was passing int, needed tuple)
    2. Capped max timeout at PostgreSQL's limit (2147483647 ms = ~596 hours)
    3. Stop exponential backoff after first retry
    4. Better connection pool management
    """

    rows = []
    sql_used = sql
    was_optimized = False
    conn = None
    
    # Cap timeout to PostgreSQL's max (2147483647 ms / 1000 = 2147483 seconds)
    MAX_TIMEOUT_SECONDS = 600  # 10 minutes max per attempt
    timeout_seconds = min(timeout_seconds, MAX_TIMEOUT_SECONDS)

    user_entities = user_entities or {}
    corpus_results = corpus_results or []

    # ------------------------------------------------------------------
    # STEP 1: Try corpus-optimized SQL first
    # ------------------------------------------------------------------
    if question and llm:
        optimized_cache = retrieve_optimized_sql_from_corpus(
            question=question,
            current_sql=sql,
            user_entities=user_entities,
            corpus_results=corpus_results
        )
        if optimized_cache and optimized_cache.get("sql"):
            sql_used = optimized_cache["sql"]
            was_optimized = True
            print("🚀 Using cached optimized SQL from corpus")

    try:
        print(f"🔧 Executing SQL (timeout={timeout_seconds}s, optimized={was_optimized})")
        print(f"🔍 SQL Preview: {sql_used[:300]}...")

        # ------------------------------------------------------------------
        # STEP 2: Acquire connection with timeout protection
        # ------------------------------------------------------------------
        def get_connection():
            return ENGINE.connect()

        try:
            # FIX: Pass empty tuple for args, not an int
            conn = execute_with_thread_timeout(
                get_connection, 
                args=(),  # CRITICAL: Empty tuple, not an int
                timeout_seconds=10
            )
            print("✅ Database connection acquired")
        except DatabaseTimeoutError:
            raise Exception(
                "Connection timeout: Database is busy. Please retry in 30 seconds."
            )

        # ------------------------------------------------------------------
        # STEP 3: Apply PostgreSQL-level safety timeouts
        # ------------------------------------------------------------------
        try:
            timeout_ms = timeout_seconds * 1000
            # Ensure we don't exceed PostgreSQL's integer limit
            timeout_ms = min(timeout_ms, 2147483647)
            
            conn.execute(text(f"SET statement_timeout = {timeout_ms}"))
            conn.execute(text("SET lock_timeout = '30s'"))
            conn.execute(text("SET idle_in_transaction_session_timeout = '60s'"))
            print(f"✅ Timeout set to {timeout_ms}ms ({timeout_seconds}s)")
        except Exception as e:
            print(f"⚠️ Timeout settings failed: {e}")

        start_time = time.time()

        # ------------------------------------------------------------------
        # STEP 4: Execute query (protected)
        # ------------------------------------------------------------------
        def run_query():
            return conn.execute(text(sql_used))

        try:
            # FIX: Pass empty tuple for args
            result = execute_with_thread_timeout(
                run_query, 
                args=(),  # CRITICAL: Empty tuple, not an int
                timeout_seconds=timeout_seconds
            )
        except DatabaseTimeoutError:
            try:
                conn.execute(text("SELECT pg_cancel_backend(pg_backend_pid())"))
            except:
                pass
            raise Exception(
                f"Query timeout: Took longer than {timeout_seconds}s. "
                "Please add filters (date, state, product, etc.)."
            )

        # ------------------------------------------------------------------
        # STEP 5: Fetch rows safely
        # ------------------------------------------------------------------
        row_count = 0
        for row in result:
            elapsed = time.time() - start_time
            if elapsed > timeout_seconds * 0.9:
                print("⚠️ Near timeout during fetch, stopping early")
                break

            if max_rows and row_count >= max_rows:
                print(f"⚠️ Hit max row limit ({max_rows})")
                break

            rows.append({k: _jsonable(v) for k, v in dict(row._mapping).items()})
            row_count += 1

            if row_count % 10000 == 0:
                print(f"📊 Fetched {row_count} rows")

        execution_time = time.time() - start_time
        print(f"✅ SQL success: {len(rows)} rows in {execution_time:.2f}s")

        conn.commit()

        # ------------------------------------------------------------------
        # STEP 6: Store successful optimization
        # ------------------------------------------------------------------
        if question and was_optimized:
            query_sig = extract_query_signature(sql_used, question, user_entities)
            store_optimized_sql_in_corpus(
                question=question,
                original_sql=sql,
                optimized_sql=sql_used,
                execution_time=execution_time,
                row_count=len(rows),
                user_entities=user_entities,
                query_signature=query_sig
            )

        return (rows, sql_used, was_optimized) if question else rows

    # ==================================================================
    # TIMEOUT → TRY LLM OPTIMIZATION ONCE (NO EXPONENTIAL BACKOFF)
    # ==================================================================
    except Exception as e:
        error_msg = str(e)
        print(f"❌ SQL failed: {error_msg}")

        # Only attempt optimization once, not in a loop
        if (
            ("timeout" in error_msg.lower() or "busy" in error_msg.lower())
            and not was_optimized
            and question
            and llm
        ):
            print("🔄 Attempting LLM optimization retry (ONE TIME ONLY)...")

            try:
                optimized_sql = generate_optimized_sql_with_indexes(
                    original_sql=sql,
                    question=question,
                    timeout_error=error_msg,
                    llm=llm
                )

                # Retry with SAME timeout (no exponential increase)
                print(f"🚀 Retrying with optimized SQL (timeout={timeout_seconds}s)...")
                
                return execute_sql_with_timeout_internal(
                    optimized_sql,
                    timeout_seconds=timeout_seconds,  # SAME timeout
                    max_rows=max_rows,
                    question=question,
                    user_entities=user_entities,
                    was_optimized=True  # Mark as optimized to prevent infinite loop
                )
                
            except Exception as opt_error:
                print(f"❌ Optimized retry also failed: {opt_error}")
                raise Exception(
                    f"Query optimization failed. Original error: {error_msg}. "
                    f"Optimization error: {str(opt_error)}"
                )

        # Friendly error mapping
        if "syntax error" in error_msg.lower():
            raise Exception(f"SQL syntax error: {error_msg}")
        elif "does not exist" in error_msg.lower():
            raise Exception("Database error: Table or column does not exist.")
        elif "out of memory" in error_msg.lower():
            raise Exception("Query returned too much data. Add filters.")
        elif "connection" in error_msg.lower() or "busy" in error_msg.lower():
            raise Exception("Database connection error. Please retry in 30 seconds.")
        else:
            raise Exception(f"Database error: {error_msg}")

    finally:
        if conn:
            try:
                conn.close()
                print("🔌 Connection released")
            except:
                pass

def execute_sql_with_timeout_internal2912(
    sql,
    timeout_seconds,
    max_rows,
    question,
    user_entities,
    was_optimized
):
    """
    Internal executor that doesn't trigger recursive optimization.
    Used for the ONE retry after optimization.
    """
    rows = []
    conn = None
    
    try:
        print(f"🔧 Executing optimized SQL (timeout={timeout_seconds}s)")
        
        def get_connection():
            return ENGINE.connect()

        try:
            conn = execute_with_thread_timeout(
                get_connection, 
                args=(),
                timeout_seconds=10
            )
        except DatabaseTimeoutError:
            raise Exception("Connection timeout: Database is busy.")

        try:
            timeout_ms = min(timeout_seconds * 1000, 2147483647)
            conn.execute(text(f"SET statement_timeout = {timeout_ms}"))
            conn.execute(text("SET lock_timeout = '30s'"))
        except Exception as e:
            print(f"⚠️ Timeout settings failed: {e}")

        start_time = time.time()

        def run_query():
            return conn.execute(text(sql))

        try:
            result = execute_with_thread_timeout(
                run_query, 
                args=(),
                timeout_seconds=timeout_seconds
            )
        except DatabaseTimeoutError:
            raise Exception(f"Optimized query still timed out after {timeout_seconds}s")

        row_count = 0
        for row in result:
            if max_rows and row_count >= max_rows:
                break
            rows.append({k: _jsonable(v) for k, v in dict(row._mapping).items()})
            row_count += 1

        execution_time = time.time() - start_time
        print(f"✅ Optimized SQL success: {len(rows)} rows in {execution_time:.2f}s")

        conn.commit()

        if question:
            query_sig = extract_query_signature(sql, question, user_entities)
            store_optimized_sql_in_corpus(
                question=question,
                original_sql=sql,
                optimized_sql=sql,
                execution_time=execution_time,
                row_count=len(rows),
                user_entities=user_entities,
                query_signature=query_sig
            )

        return (rows, sql, True)

    finally:
        if conn:
            try:
                conn.close()
            except:
                pass

# ============== HELPER FUNCTIONS (placeholder implementations) ==============

def _jsonable(v):
    """Convert value to JSON-serializable format"""
    if hasattr(v, 'isoformat'):
        return v.isoformat()
    return v


def extract_query_signature(sql, question, user_entities):
    """Generate query signature for caching"""
    import hashlib
    sig_str = f"{sql[:500]}_{question}_{json.dumps(user_entities, sort_keys=True)}"
    return hashlib.md5(sig_str.encode()).hexdigest()


def store_optimized_sql_in_corpus(question, original_sql, optimized_sql, 
                                   execution_time, row_count, user_entities, 
                                   query_signature):
    """Store optimization in corpus"""
    print(f"📝 Storing optimization: {row_count} rows in {execution_time:.2f}s")
    # Implement your corpus storage logic here


def retrieve_optimized_sql_from_corpus(question, current_sql, user_entities, 
                                       corpus_results):
    """Retrieve cached optimization from corpus"""
    # Implement your corpus retrieval logic here
    return None


def generate_optimized_sql_with_indexes(original_sql, question, timeout_error, llm):
    """Generate optimized SQL using LLM"""
    print(f"🤖 Generating optimized SQL with LLM...")
    
    prompt = f"""Optimize this PostgreSQL query that timed out:

Original SQL:
{original_sql}

Error: {timeout_error}

Requirements:
1. Add index hints for policy_end_date_year
2. Simplify WHERE conditions
3. Add LIMIT if missing
4. Use CTEs for complex subqueries

Return ONLY the optimized SQL, no explanations."""

    try:
        response = llm._call(prompt=prompt, temperature=0.2, max_tokens=2000)
        optimized = response.strip().replace('```sql', '').replace('```', '').strip()
        print(f"✅ Generated optimized SQL ({len(optimized)} chars)")
        return optimized
    except Exception as e:
        print(f"⚠️ LLM optimization failed: {e}")
        return original_sql


# import time
# import json
# import traceback
# from django.http import StreamingHttpResponse, JsonResponse
# from django.views.decorators.csrf import csrf_exempt
# from sqlalchemy import text, event
# from sqlalchemy.engine import Engine

# # ============== DATABASE TIMEOUT HANDLER ==============

# class DatabaseTimeoutError(Exception):
#     """Raised when database query exceeds timeout"""
#     pass

# # Add query timeout monitoring to all SQL executions
# @event.listens_for(Engine, "before_cursor_execute")
# def receive_before_cursor_execute(conn, cursor, statement, params, context, executemany):
#     """Log query start time"""
#     conn.info.setdefault('query_start_time', []).append(time.time())
#     print(f"🔍 Executing query (timeout: 60s)")

# @event.listens_for(Engine, "after_cursor_execute")
# def receive_after_cursor_execute(conn, cursor, statement, params, context, executemany):
#     """Log query completion time"""
#     total = time.time() - conn.info['query_start_time'].pop()
#     print(f"✅ Query completed in {total:.2f}s")

# # ============== SAFE SQL EXECUTION WRAPPER ==============

# def execute_sql_with_timeout2412(sql, timeout_seconds=300, max_rows=None):
#     """
#     Execute SQL with comprehensive error handling and timeout protection
    
#     Args:
#         sql: SQL query string
#         timeout_seconds: Maximum execution time
#         max_rows: Maximum rows to fetch (prevent memory issues)
    
#     Returns:
#         List of row dictionaries
#     """
#     rows = []
    
#     try:
#         print(f"🔧 Preparing to execute SQL (timeout: {timeout_seconds}s, max_rows: {max_rows})")
#         print(f"🔍 SQL Preview: {sql[:200]}...")
        
#         # Set statement timeout at database level (PostgreSQL)
#         with ENGINE.connect() as conn:
#             # Set query timeout
#             conn.execute(text(f"SET statement_timeout = {timeout_seconds * 1000}"))  # milliseconds
            
#             start_time = time.time()
            
#             # Execute main query
#             result = conn.execute(text(sql))
            
#             # Fetch rows with limit
#             row_count = 0
#             for row in result:
#                 # if row_count >= max_rows:
#                 if max_rows is not None and row_count >= max_rows:
#                     print(f"⚠️ Hit max row limit ({max_rows}), stopping fetch")
#                     break
                
#                 rows.append({k: _jsonable(v) for k, v in dict(row._mapping).items()})
#                 row_count += 1
                
#                 # Yield control periodically for long fetches
#                 if row_count % 10000 == 0:
#                     elapsed = time.time() - start_time
#                     print(f"📊 Fetched {row_count} rows in {elapsed:.2f}s...")
            
#             execution_time = time.time() - start_time
#             print(f"✅ SQL executed successfully: {len(rows)} rows in {execution_time:.2f}s")
            
#             # Commit to clear any transaction state
#             conn.commit()
            
#     except DatabaseTimeoutError as e:
#         print(f"⏱️ Database timeout: {e}")
#         raise Exception(f"Query timeout: Your query took longer than {timeout_seconds} seconds. Please try a more specific filter.")
        
#     except Exception as e:
#         error_msg = str(e)
#         print(f"❌ SQL execution failed: {error_msg}")
#         print(f"❌ SQL that failed: {sql[:500]}")
        
#         # Provide user-friendly error messages
#         if "statement timeout" in error_msg.lower():
#             raise Exception(f"Query timeout: Your query is taking too long (>{timeout_seconds}s). Please add more specific filters.")
#         elif "out of memory" in error_msg.lower():
#             raise Exception("Query returned too much data. Please add filters to reduce result size.")
#         elif "syntax error" in error_msg.lower():
#             raise Exception(f"SQL syntax error: {error_msg}")
#         elif "does not exist" in error_msg.lower():
#             raise Exception(f"Database error: Referenced column or table does not exist.")
#         else:
#             raise Exception(f"Database error: {error_msg}")
    
#     return rows



import hashlib
import json
from typing import Dict, List, Any, Optional, Tuple

# ============== SQL OPTIMIZATION HELPER FUNCTIONS ==============

def extract_query_signature(sql: str, question: str, user_entities: Dict) -> str:
    """
    Create a unique signature for a query based on structure and filters.
    This ensures we only reuse optimized SQL for truly identical queries.
    """
    import re
    
    # Normalize SQL for comparison
    normalized = sql.upper()
    normalized = re.sub(r'\s+', ' ', normalized)  # Normalize whitespace
    
    # Extract key components
    components = {
        "tables": re.findall(r'FROM\s+"?(\w+)"?\.?"?(\w+)"?', normalized),
        "where_conditions": re.findall(r'WHERE\s+(.+?)(?:GROUP BY|ORDER BY|LIMIT|$)', normalized),
        "aggregations": re.findall(r'(COUNT|SUM|AVG|MAX|MIN)\s*\(', normalized),
        "year_filter": re.findall(r'policy_end_date_year\s*=\s*(\d{4})', normalized),
        "month_filter": re.findall(r'policy_end_date_month\s*=\s*(\d+)', normalized),
        "entities": user_entities,
        "question_keywords": sorted(set(question.lower().split()))[:10]  # Top 10 keywords
    }
    
    # Create hash from components
    signature_str = json.dumps(components, sort_keys=True)
    signature_hash = hashlib.md5(signature_str.encode()).hexdigest()
    
    print(f"🔑 Query signature: {signature_hash}")
    print(f"   Year: {components['year_filter']}, Month: {components['month_filter']}")
    
    return signature_hash


def generate_optimized_sql_with_indexes(
    original_sql: str,
    question: str,
    timeout_error: str,
    llm
) -> str:
    """
    Generate optimized SQL with index hints and query optimization.
    """
    
    optimization_prompt = f"""You are a PostgreSQL query optimization expert.

SITUATION:
This SQL query timed out during execution:

```sql
{original_sql}
```

Error: {timeout_error}
Original Question: {question}

TASK: Rewrite with these optimizations:

1. **Use Indexed Columns First in WHERE**:
   - policy_end_date_year (ALWAYS use this first if filtering by date)
   - customerid, state, branch_name, zone

# 2. **Add Aggressive LIMIT**:
#    - If no LIMIT: add "LIMIT 50000"
#    - If LIMIT > 50000: reduce to 50000

3. **Optimize JOINs**:
   - Use INNER JOIN instead of subqueries
   - Filter early in join chain

4. **Simplify Aggregations**:
   - Pre-filter before COUNT/SUM/AVG
   - Avoid DISTINCT in aggregations if possible

5. **PostgreSQL-Specific Hints**:
   - Use CTEs for complex subqueries
   - Add index hints as comments: /* INDEX: idx_name */

6. **Date Filtering**:
   - ALWAYS include policy_end_date_year in WHERE if date filtering
   - Use BETWEEN for date ranges

CRITICAL RULES:
- Return ONLY optimized SQL (no explanations)
- NO markdown code blocks
- Must answer the same question
- Valid PostgreSQL syntax only
- Add comment showing optimization strategy at top
- When asked "Which customers have highest churn", ALWAYS include `customerid` in SELECT and GROUP BY
- Customer identification requires `customerid` — never rely only on `insured_client_name`
- Format: GROUP BY customerid, insured_client_name, policy_end_date_year, policy_end_date_month

Generate optimized SQL:"""

    try:
        print(f"🔧 Generating optimized SQL with LLM...")
        
        response = llm._call(
            prompt=optimization_prompt,
            temperature=0.2,  # Lower for precision
            max_tokens=2000
        )
        
        # Clean response
        optimized = response.strip()
        optimized = optimized.replace('```sql', '').replace('```', '').strip()
        
        # Ensure LIMIT exists
        # if "LIMIT" not in optimized.upper():
        #     optimized = optimized.rstrip(';') + " LIMIT 50000;"
        #     print(f"⚠️ Added LIMIT 50000 to optimized SQL")
        
        print(f"✅ Optimized SQL generated ({len(optimized)} chars)")
        return optimized
        
    except Exception as e:
        print(f"⚠️ Optimization generation failed: {e}")
        # Fallback: add LIMIT to original
        fallback = original_sql.rstrip(';')
        # if "LIMIT" not in fallback.upper():
        #     fallback += " LIMIT 50000;"
        return fallback
    
def store_optimized_sql_in_corpus(
    question: str,
    original_sql: str,
    optimized_sql: str,
    execution_time: float,
    row_count: int,
    user_entities: Dict,
    query_signature: str
) -> bool:
    """
    Store optimized SQL in existing corpus system.
    """
    try:
        # Create corpus entry matching your existing format
        corpus_data = {
            "question": question,
            "sql": optimized_sql,
            "summary": f"Optimized query: {row_count:,} rows in {execution_time:.2f}s",
            "recommendations": [
                "This query has been optimized for better performance",
                f"Execution time: {execution_time:.2f} seconds"
            ],
            "chart_config": None,
            "narrative": {
                "insights": [f"Query optimized and executed in {execution_time:.2f}s"],
                "recommendations": ["Query performance has been optimized"],
                "next_steps": []
            },
            "metadata": {
                "optimized": True,
                "original_sql": original_sql,
                "execution_time": execution_time,
                "row_count": row_count,
                "optimization_timestamp": time.time(),
                "query_signature": query_signature,
                "entities": user_entities
            }
        }
        
        # Use your existing corpus storage function
        # Adjust this import path to match your project structure
        from app.corpus.corpus_utils import add_to_corpus
        
        add_to_corpus(
            question=question,
            answer_data=corpus_data,
            user_entities=user_entities
        )
        
        print(f"✅ Optimized SQL stored in corpus (signature: {query_signature[:8]}...)")
        return True
        
    except Exception as e:
        print(f"⚠️ Failed to store optimized SQL: {e}")
        print(f"⚠️ Traceback: {traceback.format_exc()}")
        return False


def retrieve_optimized_sql_from_corpus(
    question: str,
    current_sql: str,
    user_entities: Dict,
    corpus_results: List
) -> Optional[Dict]:
    """
    Check if optimized SQL exists in corpus results.
    Uses strict matching: year, month, CTEs, filters must match.
    """
    try:
        # Generate signature for current query
        current_signature = extract_query_signature(current_sql, question, user_entities)
        
        # Check corpus results for optimized version
        for result in corpus_results:
            content = result.get("content", {})
            metadata = content.get("metadata", {})
            
            # Check if this is an optimized query
            if not metadata.get("optimized"):
                continue
            
            # Check signature match
            stored_signature = metadata.get("query_signature")
            if not stored_signature:
                continue
            
            if stored_signature == current_signature:
                print(f"✅ Found matching optimized SQL in corpus")
                print(f"   Signature: {current_signature[:8]}...")
                print(f"   Previous execution: {metadata.get('execution_time', 0):.2f}s")
                
                return {
                    "sql": content.get("sql"),
                    "original_sql": metadata.get("original_sql"),
                    "execution_time": metadata.get("execution_time"),
                    "row_count": metadata.get("row_count"),
                    "similarity_score": result.get("similarity_score", 0)
                }
        
        print(f"ℹ️ No matching optimized SQL found in corpus")
        return None
        
    except Exception as e:
        print(f"⚠️ Failed to retrieve optimized SQL: {e}")
        return None
    

def execute_sql_with_timeoutold(
    sql, 
    timeout_seconds=300, 
    max_rows=None,
    question=None,          # ADD THIS
    user_entities=None,     # ADD THIS
    llm=None,               # ADD THIS
    corpus_results=None     # ADD THIS
):
    """
    Execute SQL with comprehensive error handling, timeout protection,
    and automatic optimization on timeout.
    
    Args:
        sql: SQL query string
        timeout_seconds: Maximum execution time
        max_rows: Maximum rows to fetch
        question: Original user question (for optimization)
        user_entities: Extracted entities (for signature matching)
        llm: LLM instance (for generating optimized SQL)
        corpus_results: Corpus lookup results (for checking cached optimizations)
    
    Returns:
        Tuple of (rows, sql_used, was_optimized)
    """
    rows = []
    sql_used = sql
    was_optimized = False
    
    # Normalize inputs
    if user_entities is None:
        user_entities = {}
    if corpus_results is None:
        corpus_results = []
    
    # STEP 1: Check if optimized version exists in corpus
    if question and llm:
        optimized_cache = retrieve_optimized_sql_from_corpus(
            question=question,
            current_sql=sql,
            user_entities=user_entities,
            corpus_results=corpus_results
        )
        
        if optimized_cache and optimized_cache.get("sql"):
            print(f"🚀 Using cached optimized SQL from corpus")
            sql_used = optimized_cache["sql"]
            was_optimized = True
    
    try:
        print(f"🔧 Executing SQL (timeout: {timeout_seconds}s, optimized: {was_optimized})")
        print(f"🔍 SQL Preview: {sql_used[:200]}...")
        
        with ENGINE.connect() as conn:
            conn.execute(text(f"SET statement_timeout = {timeout_seconds * 1000}"))
            
            start_time = time.time()
            result = conn.execute(text(sql_used))
            
            row_count = 0
            for row in result:
                if max_rows is not None and row_count >= max_rows:
                    print(f"⚠️ Hit max row limit ({max_rows}), stopping fetch")
                    break
                
                rows.append({k: _jsonable(v) for k, v in dict(row._mapping).items()})
                row_count += 1
                
                if row_count % 10000 == 0:
                    elapsed = time.time() - start_time
                    print(f"📊 Fetched {row_count} rows in {elapsed:.2f}s...")
            
            execution_time = time.time() - start_time
            print(f"✅ SQL executed: {len(rows)} rows in {execution_time:.2f}s")
            
            conn.commit()
            
            # STEP 2: If this was optimized and worked, update corpus
            if was_optimized and question:
                query_sig = extract_query_signature(sql_used, question, user_entities)
                store_optimized_sql_in_corpus(
                    question=question,
                    original_sql=sql,
                    optimized_sql=sql_used,
                    execution_time=execution_time,
                    row_count=len(rows),
                    user_entities=user_entities,
                    query_signature=query_sig
                )
            
            return (rows, sql_used, was_optimized) if question else rows
            
    except DatabaseTimeoutError as e:
        print(f"⏱️ Database timeout: {e}")
        error_msg = f"Query timeout: Your query took longer than {timeout_seconds} seconds."
        
        # STEP 3: Try optimization on timeout
        if not was_optimized and question and llm:
            print(f"🔄 Attempting query optimization...")
            
            try:
                optimized_sql = generate_optimized_sql_with_indexes(
                    original_sql=sql,
                    question=question,
                    timeout_error=str(e),
                    llm=llm
                )
                
                print(f"🚀 Retrying with optimized SQL...")
                
                # Retry with doubled timeout
                with ENGINE.connect() as conn:
                    conn.execute(text(f"SET statement_timeout = {timeout_seconds * 2 * 1000}"))
                    
                    start_time = time.time()
                    result = conn.execute(text(optimized_sql))
                    
                    row_count = 0
                    for row in result:
                        if max_rows is not None and row_count >= max_rows:
                            break
                        rows.append({k: _jsonable(v) for k, v in dict(row._mapping).items()})
                        row_count += 1
                    
                    execution_time = time.time() - start_time
                    print(f"✅ Optimized SQL succeeded: {len(rows)} rows in {execution_time:.2f}s")
                    
                    conn.commit()
                    
                    # STEP 4: Store successful optimization in corpus
                    query_sig = extract_query_signature(optimized_sql, question, user_entities)
                    store_optimized_sql_in_corpus(
                        question=question,
                        original_sql=sql,
                        optimized_sql=optimized_sql,
                        execution_time=execution_time,
                        row_count=len(rows),
                        user_entities=user_entities,
                        query_signature=query_sig
                    )
                    
                    return (rows, optimized_sql, True) if question else rows
                    
            except Exception as opt_error:
                print(f"❌ Optimization also failed: {opt_error}")
                raise Exception(f"Query optimization failed: {str(opt_error)}")
        
        raise Exception(error_msg + " Please add more specific filters.")
        
    except Exception as e:
        error_msg = str(e)
        print(f"❌ SQL execution failed: {error_msg}")
        print(f"❌ SQL that failed: {sql_used[:500]}")
        
        # Provide user-friendly error messages
        if "statement timeout" in error_msg.lower():
            
            # Try optimization if not already done
            if not was_optimized and question and llm:
                print(f"🔄 Statement timeout - attempting optimization...")
                try:
                    optimized_sql = generate_optimized_sql_with_indexes(
                        original_sql=sql,
                        question=question,
                        timeout_error=error_msg,
                        llm=llm
                    )
                    
                    # Retry once
                    with ENGINE.connect() as conn:
                        conn.execute(text(f"SET statement_timeout = {timeout_seconds * 2 * 1000}"))
                        start_time = time.time()
                        result = conn.execute(text(optimized_sql))
                        
                        row_count = 0
                        for row in result:
                            if max_rows is not None and row_count >= max_rows:
                                break
                            rows.append({k: _jsonable(v) for k, v in dict(row._mapping).items()})
                            row_count += 1
                        
                        execution_time = time.time() - start_time
                        print(f"✅ Optimized retry succeeded: {len(rows)} rows in {execution_time:.2f}s")
                        conn.commit()
                        
                        # Store optimization
                        query_sig = extract_query_signature(optimized_sql, question, user_entities)
                        store_optimized_sql_in_corpus(
                            question=question,
                            original_sql=sql,
                            optimized_sql=optimized_sql,
                            execution_time=execution_time,
                            row_count=len(rows),
                            user_entities=user_entities,
                            query_signature=query_sig
                        )
                        
                        return (rows, optimized_sql, True) if question else rows
                        
                except Exception as retry_error:
                    print(f"❌ Optimized retry also failed: {retry_error}")
            
            raise Exception(f"Query timeout: Your query is taking too long (>{timeout_seconds}s). Please add more specific filters.")
            
        elif "out of memory" in error_msg.lower():
            raise Exception("Query returned too much data. Please add filters to reduce result size.")
        elif "syntax error" in error_msg.lower():
            raise Exception(f"SQL syntax error: {error_msg}")
        elif "does not exist" in error_msg.lower():
            raise Exception(f"Database error: Referenced column or table does not exist.")
        else:
            raise Exception(f"Database error: {error_msg}")
    
    return (rows, sql_used, was_optimized) if question else rows



# @csrf_exempt
# def download_csv(request):
#     """
#     Fixed version - ensures consistent column ordering and no extra columns
#     """
#     try:
#         if request.method != "POST":
#             return JsonResponse({"error": "POST only"}, status=405)

#         body = json.loads(request.body or "{}")
#         sql = body.get("sql")

#         if not sql:
#             return JsonResponse({"error": "Missing SQL"}, status=400)

#         # Execute query with timeout
#         rows = execute_sql_with_timeout(
#             sql=sql,
#             timeout_seconds=600,
#             max_rows=None  # Get all rows for CSV download
#         )

#         if not rows or len(rows) == 0:
#             return JsonResponse({"error": "No data to download"}, status=404)

#         # ✅ FIX 1: Convert Decimals to float/int for consistent formatting
#         rows = _convert_decimals_to_float(rows)

#         # ✅ FIX 2: Get consistent column order from first row
#         first_row = rows[0]
#         ordered_columns = list(first_row.keys())
        
#         print(f"📥 CSV Download: {len(rows)} rows, columns: {ordered_columns}")

#         # ✅ FIX 3: Create DataFrame with explicit column order
#         import pandas as pd
        
#         # Ensure all rows have same columns in same order
#         normalized_rows = []
#         for row in rows:
#             normalized_row = {}
#             for col in ordered_columns:
#                 val = row.get(col, None)
#                 # Clean up None values
#                 normalized_row[col] = val if val is not None else ""
#             normalized_rows.append(normalized_row)
        
#         df = pd.DataFrame(normalized_rows, columns=ordered_columns)
        
#         # ✅ FIX 4: Clean column names (remove extra spaces, special chars)
#         df.columns = [str(col).strip() for col in df.columns]
        
#         # ✅ FIX 5: Format numbers consistently
#         for col in df.columns:
#             # If column has numeric data, format it properly
#             if df[col].dtype in ['float64', 'int64']:
#                 # Keep integers as integers, format floats to 2 decimals
#                 if df[col].dtype == 'int64':
#                     df[col] = df[col].astype(str)
#                 else:
#                     # Round floats to 2 decimals
#                     df[col] = df[col].apply(lambda x: f"{x:.2f}" if pd.notna(x) else "")

#         # ✅ FIX 6: Generate CSV with consistent formatting
#         csv_content = df.to_csv(index=False, lineterminator='\n')

#         response = HttpResponse(
#             csv_content,
#             content_type="text/csv;charset=utf-8"
#         )
#         response["Content-Disposition"] = 'attachment; filename="results.csv"'
        
#         print(f"✅ CSV generated: {len(df)} rows × {len(df.columns)} columns")
        
#         return response

#     except Exception as e:
#         print(f"❌ CSV download error: {e}")
#         return JsonResponse({
#             "error": str(e),
#             "type": type(e).__name__
#         }, status=500)

@csrf_exempt
def download_csv(request):
    """Simple, robust CSV download endpoint"""
    try:
        if request.method != "POST":
            return JsonResponse({"error": "POST only"}, status=405)

        body = json.loads(request.body or "{}")
        sql = body.get("sql")

        if not sql:
            return JsonResponse({"error": "Missing SQL"}, status=400)

        # Execute query
        rows = execute_sql_with_timeout(
            sql=sql,
            timeout_seconds=600,
            max_rows=None  # Get all rows
        )

        if not rows or len(rows) == 0:
            return JsonResponse({"error": "No data to download"}, status=404)

        print(f"📥 Generating CSV for {len(rows)} rows")

        # Convert to pandas DataFrame
        import pandas as pd
        df = pd.DataFrame(rows)
        
        # Generate CSV string
        csv_content = df.to_csv(index=False, lineterminator='\n')

        # Create response
        response = HttpResponse(
            csv_content,
            content_type="text/csv; charset=utf-8"
        )
        response["Content-Disposition"] = 'attachment; filename="results.csv"'
        
        print(f"✅ CSV ready: {len(df)} rows × {len(df.columns)} columns")
        
        return response

    except Exception as e:
        print(f"❌ CSV download error: {e}")
        import traceback
        traceback.print_exc()
        
        # CRITICAL: Always return JSON for errors, never HTML
        return JsonResponse({
            "error": str(e),
            "type": type(e).__name__
        }, status=500)

# ============================================================
# HELPER FUNCTION - Convert Decimals
# ============================================================

def _convert_decimals_to_float(rows):
    """Convert Decimal objects to float for consistent JSON/CSV serialization"""
    from decimal import Decimal
    
    if not isinstance(rows, list):
        return rows
    
    converted = []
    for row in rows:
        if not isinstance(row, dict):
            converted.append(row)
            continue
            
        new_row = {}
        for key, val in row.items():
            if isinstance(val, Decimal):
                # Convert to int if it's a whole number, else float
                if val % 1 == 0:
                    new_row[key] = int(val)
                else:
                    new_row[key] = float(val)
            else:
                new_row[key] = val
        converted.append(new_row)
    
    return converted

@csrf_exempt
def download_csv20_01(request):
    try:
        if request.method != "POST":
            return JsonResponse({"error": "POST only"}, status=405)

        body = json.loads(request.body or "{}")
        sql = body.get("sql")

        if not sql:
            return JsonResponse({"error": "Missing SQL"}, status=400)

        rows = execute_sql_with_timeout(
            sql=sql,
            timeout_seconds=600,
            max_rows=None
        )

        import pandas as pd
        df = pd.DataFrame(rows)

        response = HttpResponse(
            df.to_csv(index=False),
            content_type="text/csv"
        )
        response["Content-Disposition"] = "attachment; filename=results.csv"
        return response

    except Exception as e:
        # 🔴 CRITICAL: never return HTML
        return JsonResponse({
            "error": str(e),
            "type": type(e).__name__
        }, status=500)



from decimal import Decimal

# def round_table_values(rows: list) -> list:
#     """Config-based rounding for ALL table data"""
#     if not rows:
#         return rows
    
#     rounded = []
#     for row in rows:
#         new_row = {}
#         for key, val in row.items():
#             if isinstance(val, (float, Decimal)):
#                 val = float(val)
#                 key_lower = key.lower()
                
#                 # Check if percentage column
#                 is_pct_col = any(k in key_lower for k in [
#                     'rate', 'ratio', 'pct', 'percent',
#                     'churn', 'retention', 'conversion'
#                 ])
                
#                 if is_pct_col:
#                     # Percentage column
#                     if ROUND_PERCENTAGES:
#                         new_row[key] = round(val * 100) if 0 < val <= 1 else round(val)
#                     else:
#                         new_row[key] = round(val * 100, 2) if 0 < val <= 1 else round(val, 2)
#                 else:
#                     # Non-percentage column (e.g., policy_tenure: 1.0 → 1)
#                     rounded_val = round(val, 2)
#                     new_row[key] = int(rounded_val) if rounded_val == int(rounded_val) else rounded_val
#             else:
#                 new_row[key] = val
#         rounded.append(new_row)
    
#     return rounded

def round_table_values(rows: list) -> list:
    if not rows:
        return rows
    
    rounded = []
    for row in rows:
        new_row = {}
        for key, val in row.items():
            if isinstance(val, (float, Decimal)):
                val = float(val)
                key_lower = key.lower()
                
                # Check percentage column
                is_pct_col = any(k in key_lower for k in [
                    'rate', 'ratio', 'pct', 'percent',
                    'churn', 'retention', 'conversion'
                ])
                
                # ✅ ADD: Check currency column
                is_currency_col = any(k in key_lower for k in [
                    'premium', 'amount', 'revenue', 'value', 
                    'price', 'cost', 'total', 'idv'
                ])
                
                if is_pct_col:
                    # Percentage handling (existing code)
                    if ROUND_PERCENTAGES:
                        new_row[key] = round(val * 100) if 0 < val <= 1 else round(val)
                    else:
                        new_row[key] = round(val * 100, 2) if 0 < val <= 1 else round(val, 2)
                
                elif is_currency_col:  # ✅ NEW: Currency handling
                    if ROUND_CURRENCY:
                        new_row[key] = round(val)  # 128.04 → 128
                    else:
                        new_row[key] = round(val, 2)  # Keep 2 decimals
                
                else:
                    # Other numeric columns
                    rounded_val = round(val, 2)
                    new_row[key] = int(rounded_val) if rounded_val == int(rounded_val) else rounded_val
            else:
                new_row[key] = val
        rounded.append(new_row)
    
    return rounded
@csrf_exempt
def ask_question_stream(request):
    """
    PRODUCTION-READY VERSION with dynamic opener based on actual results
    """
    if request.method != "POST":
        return JsonResponse({"error": "POST only"}, status=405)

    try:
        body = json.loads(request.body or "{}")
        question = (body.get("question") or "").strip()
        user_id = body.get("user_id") or "admin"
        from_next_step_suggestion = body.get("from_next_step_suggestion", False)
        
        if not question:
            return JsonResponse({"error": "Missing question"}, status=400)

        if is_incomplete_question(question):
            return JsonResponse(
                {"error": "Incomplete question. Please rephrase into a full question."},
                status=400
            )

        question = merge_followup_with_history(user_id, question)
        start = time.time()

        def gen():
            print("🚨🚨🚨 GENERATOR STARTED 🚨🚨🚨")
            stream_completed = False
            final_payload = None
            last_heartbeat = time.time()
            HEARTBEAT_INTERVAL = 5
            
            def send_heartbeat_if_needed():
                nonlocal last_heartbeat
                current_time = time.time()
                if current_time - last_heartbeat > HEARTBEAT_INTERVAL:
                    last_heartbeat = current_time
                    print(f"💓 Sending heartbeat (elapsed: {current_time - start:.1f}s)")
                    return _ev("heartbeat", timestamp=current_time, elapsed=f"{current_time - start:.1f}s")
                return None
            
            try:
                yield _ev("heartbeat", timestamp=time.time())
                last_heartbeat = time.time()
                
                # ================= CONTEXT =================
                history = conversation_memory_store.get(user_id, [])
                time_context = extract_time_context_from_question(question)
                context = {"time_context": time_context}

                # DON'T generate opener yet - wait for actual data
                # opener = generate_conversational_opener(question, context)
                # yield _ev("conversational_opener", message=opener)
                
                opener = None  # Will be generated after we have results

                # ================= CORPUS LOOKUP =================
                corpus_results = []
                corpus_used = False
                norm_q = question
                user_entities = {}
                if from_next_step_suggestion:
                    print("🚀 Next-step suggestion - SKIPPING CORPUS LOOKUP")
                    yield _ev("phase", message="Generating fresh analysis…")
                    last_heartbeat = time.time()
                else:
                    yield _ev("phase", message="Checking knowledge corpus…")
                    last_heartbeat = time.time()
                    
                    try:
                        hb = send_heartbeat_if_needed()
                        if hb: yield hb
                        
                        # ✅ FIX 1: Extract entities from ORIGINAL question first
                        print(f"🔍 Original question: '{question}'")
                        user_entities = extract_entities(question)
                        print(f"🔍 Extracted user entities: {user_entities}")
                        
                        # ✅ FIX 2: Normalize the ORIGINAL question
                        norm_q = normalize_question_entities(question)
                        print(f"🔍 Normalized question: '{norm_q}'")
                        
                        # ✅ FIX 3: Then resolve for context (but use normalized for corpus)
                        resolved_q = resolve_followup_question(question, history)
                        print(f"🔍 Resolved question: '{resolved_q}'")

                        hb = send_heartbeat_if_needed()
                        if hb: yield hb

                        corpus_results = query_corpus(
                            question=norm_q,  # Use normalized question
                            user_entities=user_entities,  # Use entities from original
                            n_results=5,
                            min_score=0.70  # ✅ Higher threshold like Document 2
                        )
                        print(f"🔍 Corpus returned: {len(corpus_results) if corpus_results else 0} results")
        
        

                # ================= CORPUS LOOKUP =================
                # corpus_results = []
                # corpus_used = False
                # norm_q = question
                # user_entities = {}
                
                # if from_next_step_suggestion:
                #     print("🚀 Next-step suggestion - SKIPPING CORPUS LOOKUP")
                #     yield _ev("phase", message="Generating fresh analysis…")
                #     last_heartbeat = time.time()
                # else:
                #     yield _ev("phase", message="Checking knowledge corpus…")
                #     last_heartbeat = time.time()
                    
                #     try:
                #         hb = send_heartbeat_if_needed()
                #         if hb: yield hb
                        
                #         resolved_q = resolve_followup_question(question, history)
                #         print(f"🔍 Original question: '{question}'")
                #         print(f"🔍 Resolved question: '{resolved_q}'")
                        
                #         user_entities = extract_entities(resolved_q)
                #         print(f"🔍 Extracted user entities: {user_entities}")
                        
                #         norm_q = normalize_question_entities(resolved_q)
                #         print(f"🔍 Normalized question: '{norm_q}'")

                #         hb = send_heartbeat_if_needed()
                #         if hb: yield hb

                #         corpus_results = query_corpus(
                #             question=norm_q,
                #             user_entities=user_entities,
                #             n_results=5,
                #             min_score=0.50
                #         )
                #         print(f"🔍 Corpus returned: {len(corpus_results) if corpus_results else 0} results")
                        
                        yield _ev("phase", message="Corpus lookup completed")
                        last_heartbeat = time.time()
                        
                    except Exception as corpus_error:
                        print(f"⚠️ Corpus lookup error: {corpus_error}")
                        print(f"⚠️ Traceback: {traceback.format_exc()}")
                        corpus_results = []
                        yield _ev("phase", message="Corpus lookup skipped, proceeding with fresh analysis")
                        last_heartbeat = time.time()

                # ================= CORPUS PROCESSING =================
                if corpus_results and not from_next_step_suggestion:
                    print(f"✅ Found {len(corpus_results)} corpus matches")
                    
                    hb = send_heartbeat_if_needed()
                    if hb: yield hb
                    
                    first = corpus_results[0]
                    parsed = first["content"]
                    entities_match = first.get("entities_match", False)
                    semantic_match = first.get("semantic_match", False)
                    similarity_score = first.get("similarity_score", 0)
                    corpus_entities = first.get("corpus_entities", {})
                    semantic_reason = first.get("semantic_reason", "")
                    
                    corpus_question = parsed.get("question", "")
                    print(f"🔍 Top match: '{corpus_question}'")
                    print(f"🔍 Similarity: {similarity_score:.3f}")
                    
                    corpus_summary = parsed.get("summary", "")
                    corpus_recommendations = parsed.get("recommendations", [])
                    corpus_chart_config = parsed.get("chart_config")
                    corpus_sql = parsed.get("sql")
                    corpus_narrative = parsed.get("narrative", {})
                    
                    has_content = bool(corpus_summary or corpus_recommendations or corpus_narrative)
                    has_good_similarity = similarity_score >= 0.90
                    
                    if entities_match and semantic_match and has_content and has_good_similarity:
                        print(f"✅ ACCEPTING corpus")
                        
                        yield _ev("phase", message="Retrieving cached results…")
                        last_heartbeat = time.time()
                        
                        corpus_used = True
                        fresh_rows = []
                        fresh_summary = corpus_summary
                        fresh_chart_config = corpus_chart_config
                        
                        if corpus_sql:
                            print(f"🔄 Executing corpus SQL...")
                            try:
                                hb = send_heartbeat_if_needed()
                                if hb: yield hb
                                
                                is_valid, issues = validate_sql_entity_filters(corpus_sql, user_entities)
                                
                                if not is_valid:
                                    raise ValueError(f"SQL entity mismatch: {', '.join(issues)}")
                                
                                clean_sql = sanitize_sql_before_exec(corpus_sql, question)
                                
                                yield _ev("phase", message="Executing cached query…")
                                last_heartbeat = time.time()
                                
                                hb = send_heartbeat_if_needed()
                                if hb: yield hb
                                
                                fresh_rows = execute_sql_with_timeout(
                                    sql=clean_sql,
                                    timeout_seconds=600,
                                    max_rows=None
                                )
                                
                                print(f"✅ Corpus SQL executed: {len(fresh_rows)} rows")
                                
                                yield _ev("phase", message=f"Retrieved {len(fresh_rows)} records")
                                last_heartbeat = time.time()
                                
                                if fresh_rows:
                                    yield _ev("phase", message="Generating insights…")
                                    last_heartbeat = time.time()
                                    
                                    hb = send_heartbeat_if_needed()
                                    if hb: yield hb
                                    
                                    fresh_summary = generate_summary_from_rows(question, clean_sql, fresh_rows)
                                
                                if fresh_rows and not corpus_chart_config:
                                    yield _ev("phase", message="Creating visualization…")

                                    # ✅ Send heartbeat BEFORE starting slow operation
                                    hb = send_heartbeat_if_needed()
                                    if hb: yield hb
                                    
                                    llm = get_llama_maverick_llm()
                                    if llm.detect_chart_intent(question):
                                        hb = send_heartbeat_if_needed()
                                        if hb: yield hb

                                        fresh_chart_config = llm_generate_chart_config(question, fresh_rows)
                                
                            except (ValueError, Exception) as e:
                                print(f"❌ Corpus SQL failed: {e}")
                                corpus_used = False
                                corpus_results = []
                                yield _ev("phase", message="Corpus rejected, generating fresh SQL…")
                                last_heartbeat = time.time()
                        
                        if corpus_used and (fresh_rows or corpus_summary):
                            hb = send_heartbeat_if_needed()
                            if hb: yield hb
                            
                            # ✅ NOW generate opener with actual data AND rows for context
                            opener = generate_conversational_opener_with_data(
                                question=question,
                                row_count=len(fresh_rows),
                                context=context,
                                rows=fresh_rows  # Pass actual rows for interpretation
                            )
                            opener = re.sub(r'(\d+)\.0+%', lambda m: f"{m.group(1)}%", opener)
                            if ROUND_PERCENTAGES:
                                opener = re.sub(r'(\d+\.\d+)%', lambda m: f"{round(float(m.group(1)))}%", opener)
                            yield _ev("conversational_opener", message=opener)
                            last_heartbeat = time.time()
                            
                            narrative_insights = corpus_narrative.get("insights", [fresh_summary]) if corpus_narrative else [fresh_summary]
                            narrative_recommendations = corpus_narrative.get("recommendations", corpus_recommendations) if corpus_narrative else corpus_recommendations
                            narrative_next_steps = corpus_narrative.get("next_steps", []) if corpus_narrative else []
                            
                            if not narrative_next_steps:
                                try:
                                    metrics = _derive_contextual_metrics(question, fresh_rows)
                                    narrative_next_steps = _generate_dynamic_next_steps(question, metrics)
                                except:
                                    narrative_next_steps = [
                                        "Would you like to segment this analysis by customer type?",
                                        "Shall I compare with a different time period?"
                                    ]

                            narrative = {
                                "opener": opener,
                                "insights": narrative_insights,
                                "recommendations": narrative_recommendations,
                                "next_step": "Feel free to ask follow-up questions!",
                                "next_steps": narrative_next_steps,
                            }

                            final_payload = {
                                "answer": fresh_summary or corpus_summary,
                                "success": True,
                                "query_used": corpus_sql or "Knowledge base lookup",
                                "rows": fresh_rows,
                                "summary": fresh_summary or corpus_summary,
                                "chart_config": fresh_chart_config,
                                "row_count": len(fresh_rows) if fresh_rows else 0,
                                "recommendation": narrative_recommendations,
                                "narrative": narrative,
                                "conversational_opener": opener,
                                "response_time": f"{time.time() - start:.2f}s",
                                "user_id": user_id,
                                "history": conversation_memory_store.get(user_id, []),
                                "corpus_used": True,
                                "asked_question": question,
                                "normalized_question": norm_q,
                                "similarity_score": similarity_score,
                                "entities_matched": user_entities,
                                "semantic_match": semantic_match,
                                "semantic_reason": semantic_reason,
                            }

                            yield _ev("phase", message="Preparing results…")
                            last_heartbeat = time.time()
                            
                            if fresh_summary or corpus_summary:
                                yield _ev("summary", text=fresh_summary or corpus_summary)
                                last_heartbeat = time.time()
                            if narrative_recommendations:
                                yield _ev("recommendation", text="\n".join(narrative_recommendations))
                                last_heartbeat = time.time()
                            if fresh_chart_config:
                                yield _ev("chart", config=fresh_chart_config)
                                last_heartbeat = time.time()
                            if fresh_rows:
                                yield _ev("rows_preview", rows=fresh_rows[:8], row_count=len(fresh_rows))
                                last_heartbeat = time.time()
                            
                            yield _ev("narrative", obj=narrative)
                            yield _ev("final", payload=final_payload)
                            
                            stream_completed = True
                            print("✅ Corpus response served")
                            return
                    else:
                        print(f"⚠️ REJECTING corpus")
                        corpus_results = []
                        yield _ev("phase", message="Generating fresh analysis…")
                        last_heartbeat = time.time()
                        
                elif from_next_step_suggestion:
                    print("🚀 Next-step suggestion - generating fresh SQL")
                else:
                    print("❌ No corpus results found")
                    yield _ev("phase", message="Generating fresh analysis…")
                    last_heartbeat = time.time()

                # ================= SQL GENERATION =================
                print("⚡ Generating fresh SQL")

                yield _ev("phase", message="Understanding your question…")
                last_heartbeat = time.time()
                
                hb = send_heartbeat_if_needed()
                if hb: yield hb
                
                aug_q = augment_question_with_context(question, history)

                yield _ev("phase", message="Generating SQL query…")
                last_heartbeat = time.time()
                
                sql = None
                sql_recommendation = ""
                
                try:
                    hb = send_heartbeat_if_needed()
                    if hb: yield hb
                    
                    sql, sql_recommendation = run_sql_generation_graph(
                        aug_q, user_id=user_id, db_id="default", history=history
                    )
                    print("🔍 🚨 STREAM - BEFORE CLEANING SQL")
                    print(f"   Question: {question}")
                    print(f"   SQL length: {len(sql) if sql else 0}")

                    sql = validate_customer_query(sql, question)
                    sql = remove_unwanted_business_filters(sql, question)
                    print("=" * 50)
                    print("🔍 BEFORE fix_percentage_formatting:")
                    print(sql)
                    print("=" * 50)
                    sql = fix_percentage_formatting(sql)
                    print("=" * 50)
                    print("🔍 AFTER fix_percentage_formatting:")
                    print(sql)
                    print("=" * 50)
                    print("🔍 🚨 STREAM - AFTER CLEANING SQL")
                    print(f"   SQL length: {len(sql) if sql else 0}")
                    print(f"✅ SQL generated")
                except Exception as sql_gen_error:
                    print(f"❌ SQL generation failed: {sql_gen_error}")
                    sql = "SELECT COUNT(*) as total_count FROM \"bi_dwh\".\"main_cai_lib\" WHERE policy_end_date_year IS NOT NULL LIMIT 1000"
                    yield _ev("phase", message="Using fallback query…")
                    last_heartbeat = time.time()

                try:
                    sql = fix_missing_where_keyword(sql)
                    sql = fix_nullif_for_postgres(sql)
                    sql = fix_round_for_postgres(sql)
                    sql = sanitize_sql_before_exec(sql, question)
                    print(f"🔧 SQL sanitized")
                    
                        
                except Exception as e:
                    print(f"⚠️ SQL sanitization failed: {e}")
                    sql = "SELECT COUNT(*) as total_count FROM \"bi_dwh\".\"main_cai_lib\" WHERE policy_end_date_year IS NOT NULL LIMIT 1000"
                    yield _ev("phase", message="Using safe fallback query…")
                    last_heartbeat = time.time()

                yield _ev("sql", sql=sql)
                last_heartbeat = time.time()
                
                # ================= SAFE SQL EXECUTION =================
                yield _ev("phase", message="Executing database query…")
                last_heartbeat = time.time()

                hb = send_heartbeat_if_needed()
                if hb: yield hb

                rows = []
                sql_used = sql
                was_optimized = False
                try:
                    print(f"🚀 Starting SQL execution with optimization support")
                    
                    llm = get_llama_maverick_llm()
                    result = execute_sql_with_timeout(timeout_seconds=600,
                        sql=sql,
                        max_rows=None,
                        question=question,
                        user_entities=user_entities,
                        llm=llm,
                        corpus_results=corpus_results
                    )
                    
                    if isinstance(result, tuple):
                        rows, sql_used, was_optimized = result
                    else:
                        rows = result
                    
                    if was_optimized:
                        print(f"✅ Query executed with optimized SQL: {len(rows)} rows")
                        yield _ev("phase", message=f"Query optimized & executed: {len(rows)} records")
                    else:
                        print(f"✅ SQL executed successfully: {len(rows)} rows")
                        yield _ev("phase", message=f"Executing database query…")
                        # yield _ev("phase", message=f"Query returned {len(rows)} records")
                    last_heartbeat = time.time()
                    
                except Exception as e:
                    rows = round_table_values(rows)
                    error_msg = str(e)
                    print(f"❌ SQL execution failed: {error_msg}") 
                    
                    yield _ev("phase", message="Query execution failed, generating fallback response…")
                    last_heartbeat = time.time()
                    
                    # Generate opener for error case
                    opener = "I encountered an issue while processing your request."
                    yield _ev("conversational_opener", message=opener)
                    
                    error_summary = f"I encountered an issue executing your query: {error_msg}"
                    
                    error_narrative = {
                        "opener": opener,
                        "insights": [error_summary],
                        "recommendations": [
                            "Try adding more specific filters (date range, branch, state)",
                            "Simplify your question to focus on a smaller dataset",
                            "Contact support if this issue persists"
                        ],
                        "next_steps": ["Would you like to try a simpler query?"]
                    }
                    
                    yield _ev("summary", text=error_summary)
                    yield _ev("narrative", obj=error_narrative)
                    
                    error_payload = {
                        "answer": error_summary,
                        "success": False,
                        "error": error_msg,
                        "corpus_used": False,
                        "summary": error_summary,
                        "narrative": error_narrative,
                        "rows": [],
                        "query_used": sql,
                        "asked_question": question,
                        "response_time": f"{time.time() - start:.2f}s",
                    }
                    
                    yield _ev("final", payload=error_payload)
                    stream_completed = True
                    print("✅ Error response sent")
                    return
                rows = round_table_values(rows)  # 👈 ADD
                print(f"🔢 After rounding: {rows[:1]}")
                
                # ✅ NOW generate opener with actual data (after SQL execution)
                if opener is None:  # Only if not already generated from corpus path
                    opener = generate_conversational_opener_with_data(
                        question=question,
                        row_count=len(rows),
                        context=context,
                        rows=rows  # Pass actual rows for interpretation
                    )
                    # ✅ ADD THIS CLEANING HERE:
                    opener = re.sub(r'(\d+)\.0+%', lambda m: f"{m.group(1)}%", opener)
                    if ROUND_PERCENTAGES:
                        opener = re.sub(r'(\d+\.\d+)%', lambda m: f"{round(float(m.group(1)))}%", opener)
                    yield _ev("conversational_opener", message=opener)
                    last_heartbeat = time.time()
                    
                yield _ev("rows_preview", rows=rows[:8], row_count=len(rows))
                last_heartbeat = time.time()

                # Handle empty results for next-step
                if from_next_step_suggestion and len(rows) == 0:
                    yield _ev("phase", message="Generating conversational response…")
                    last_heartbeat = time.time()
                    
                    llm = get_llama_maverick_llm()
                    conv_answer = llm.invoke(question)
                    
                    yield _ev("summary", text=conv_answer)
                    
                    narrative = {
                        "opener": opener,
                        "insights": [conv_answer],
                        "recommendations": ["Try broadening your filters."],
                        "next_steps": ["Would you like to try a different analysis?"]
                    }
                    
                    yield _ev("narrative", obj=narrative)
                    
                    final_payload = {
                        "answer": conv_answer,
                        "success": True,
                        "summary": conv_answer,
                        "rows": [],
                        "narrative": narrative,
                        "corpus_used": False,
                        "query_used": sql,
                        "asked_question": question,
                        "response_time": f"{time.time() - start:.2f}s",
                    }
                    
                    yield _ev("final", payload=final_payload)
                    stream_completed = True
                    print("✅ Empty result response completed")
                    return

                yield _ev("phase", message="Analyzing results…")
                last_heartbeat = time.time()

                hb = send_heartbeat_if_needed()
                if hb: yield hb

                sql_summary = ""
                narrative = {}
                
                try:
                    sql_summary = generate_summary_from_rows(question, sql, rows)
                    yield _ev("summary", text=sql_summary)
                    last_heartbeat = time.time()
                    
                    yield _ev("phase", message="Creating narrative…")
                    last_heartbeat = time.time()
                    
                    hb = send_heartbeat_if_needed()
                    if hb: yield hb
                    
                    narrative = humanize_narrative(
                        question=question,
                        rows=rows,
                        summary=sql_summary,
                        recommendation="",
                        sql=sql,
                        opener=opener,
                        corpus_insights=None
                    )
                    
                    all_recommendations = narrative.get('recommendations', [])
                    if sql_recommendation and sql_recommendation not in all_recommendations:
                        all_recommendations.append(sql_recommendation)
                    
                    if not all_recommendations:
                        all_recommendations = ["Consider refining your query."]
                    
                    narrative['recommendations'] = all_recommendations
                    
                except Exception as e:
                    print(f"⚠️ Narrative generation failed: {e}")
                    sql_summary = "Analysis completed successfully." if rows else "No results found."
                    narrative = {
                        "opener": opener,
                        "insights": [sql_summary],
                        "recommendations": ["Consider adjusting filters."],
                        "next_steps": ["Would you like to explore other aspects?"]
                    }
                    yield _ev("phase", message="Narrative generated with fallback")
                    last_heartbeat = time.time()

                yield _ev("recommendation", text="\n".join(narrative.get('recommendations', [])))
                last_heartbeat = time.time()

                # Generate chart
                chart_config = None
                if len(rows) > 0:
                    try:
                        yield _ev("phase", message="Creating visualization…")
                        last_heartbeat = time.time()
                        
                        llm = get_llama_maverick_llm()
                        if llm.detect_chart_intent(question):
                            chart_config = llm_generate_chart_config(question, rows)
                            if chart_config:
                                yield _ev("phase", message="Visualization ready")
                                last_heartbeat = time.time()
                    except Exception as e:
                        print(f"⚠️ Chart generation failed: {e}")
                        yield _ev("phase", message="Skipping visualization")
                        last_heartbeat = time.time()

                yield _ev("chart", config=chart_config)
                yield _ev("narrative", obj=narrative)
                last_heartbeat = time.time()

                final_payload = {
                    "answer": narrative.get("opener") or sql_summary,
                    "success": True,
                    "query_used": sql,
                    "rows": rows,
                    "summary": sql_summary,
                    "chart_config": chart_config,
                    "row_count": len(rows),
                    "recommendation": narrative.get('recommendations', []),
                    "narrative": narrative,
                    "conversational_opener": opener,
                    "response_time": f"{time.time() - start:.2f}s",
                    "user_id": user_id,
                    "history": conversation_memory_store.get(user_id, []),
                    "corpus_used": False,
                    "asked_question": question,
                }

                yield _ev("final", payload=final_payload)
                stream_completed = True
                print("✅ SQL response completed")

            except Exception as e:
                error_msg = str(e)
                error_trace = traceback.format_exc()
                print(f"❌ CRITICAL Error: {error_msg}")
                print(f"❌ Traceback: {error_trace}")
                
                yield _ev("phase", message="Handling error…")
                yield _ev("error", message=error_msg, traceback=error_trace)
                
                # fallback_summary = f"I encountered an issue processing your question: {error_msg}"
                
                # fallback_narrative = {
                #     "opener": "I apologize for the inconvenience.",
                #     "insights": [fallback_summary],                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                             
                #     "recommendations": ["Please try rephrasing your question.", "Contact support if the issue persists."],
                #     "next_steps": ["Would you like to try a different query?"]
                # }
                fallback_narrative = {
                "opener": "No matching records found for your query.",
                "insights": ["The database returned no results matching your criteria."],
                "recommendations": [
                    "Try broadening your search criteria",
                    "Check if the filters are too restrictive"
                ],
                "next_steps": ["Would you like to try a different query?"]
            }
                yield _ev("summary", text=fallback_summary)
                yield _ev("narrative", obj=fallback_narrative)
                
                error_payload = {
                    "answer": fallback_summary,
                    "success": False,
                    "error": error_msg,
                    "corpus_used": False,
                    "summary": fallback_summary,
                    "narrative": fallback_narrative,
                    "rows": [],
                    "asked_question": question,
                    "response_time": f"{time.time() - start:.2f}s",
                }
                
                yield _ev("final", payload=error_payload)
                stream_completed = True
                print("✅ Error response completed")
            
            finally:
                if not stream_completed:
                    print("⚠️ Stream did not complete properly, sending emergency final event")
                    emergency_payload = {
                        "answer": "Response incomplete",
                        "success": False,
                        "error": "Stream terminated unexpectedly",
                        "corpus_used": False,
                        "summary": "The request did not complete successfully.",
                        "narrative": {
                            "opener": "I apologize.",
                            "insights": ["The analysis was interrupted."],
                            "recommendations": ["Please try again with more specific filters."],
                            "next_steps": []
                        },
                        "rows": [],
                        "asked_question": question,
                    }
                    yield _ev("final", payload=emergency_payload)
                    print("✅ Emergency final event sent")


        resp = StreamingHttpResponse(gen(), content_type="application/x-ndjson")
        resp["Cache-Control"] = "no-cache"
        resp["X-Accel-Buffering"] = "no"
        return resp

    except Exception as e:
        print(f"❌ Outer error: {str(e)}")
        print(f"❌ Traceback: {traceback.format_exc()}")
        return JsonResponse({"error": str(e), "traceback": traceback.format_exc()}, status=500)


     
def sanitize_sql_with_entity_validation(sql, question, user_entities):
    """
    Sanitize SQL and validate that entity filters match the user's question.
    This prevents using corpus SQL with wrong month/year/location filters.
    """
    import re
    
    print(f"🔧 Validating SQL entity filters against user question...")
    print(f"🔧 User entities: {user_entities}")
    
    # First, do standard sanitization
    sql = sanitize_sql_before_exec(sql, question)
    
    # Extract filters from SQL
    sql_lower = sql.lower()
    
    # Validate MONTH filter
    if "<MONTH>" in user_entities:
        user_month = user_entities["<MONTH>"]
        # Map month name to number
        month_map = {
            "January": 1, "February": 2, "March": 3, "April": 4,
            "May": 5, "June": 6, "July": 7, "August": 8,
            "September": 9, "October": 10, "November": 11, "December": 12
        }
        expected_month_num = month_map.get(user_month)
        
        # Check if SQL has correct month filter
        month_pattern = r"policy_end_date_month\s*=\s*(\d+)"
        month_match = re.search(month_pattern, sql_lower)
        
        if month_match:
            sql_month_num = int(month_match.group(1))
            if sql_month_num != expected_month_num:
                print(f"❌ ENTITY MISMATCH: SQL has month={sql_month_num} but user asked for {user_month} (month={expected_month_num})")
                raise ValueError(f"Entity mismatch: SQL month filter doesn't match user question")
        else:
            print(f"⚠️ WARNING: User asked for {user_month} but SQL has no month filter")
    
    # Validate YEAR filter
    if "<YEAR>" in user_entities:
        user_year = user_entities["<YEAR>"]
        year_pattern = r"policy_end_date_year\s*=\s*(\d{4})"
        year_match = re.search(year_pattern, sql_lower)
        
        if year_match:
            sql_year = year_match.group(1)
            if sql_year != user_year:
                print(f"❌ ENTITY MISMATCH: SQL has year={sql_year} but user asked for {user_year}")
                raise ValueError(f"Entity mismatch: SQL year filter doesn't match user question")
        else:
            print(f"⚠️ WARNING: User asked for {user_year} but SQL has no year filter")
    
    # Validate STATE filter (if you have state columns)
    if "<STATE>" in user_entities:
        user_state = user_entities["<STATE>"]
        # Check for state filter in SQL (adjust column name as needed)
        state_pattern = rf"state\s*=\s*['\"]({user_state})['\"]"
        if not re.search(state_pattern, sql, re.IGNORECASE):
            print(f"⚠️ WARNING: User asked for state={user_state} but SQL might not have correct state filter")
    
    # Validate ZONE filter
    if "<ZONE>" in user_entities:
        user_zone = user_entities["<ZONE>"]
        zone_pattern = rf"zone\s*=\s*['\"]({user_zone})['\"]"
        if not re.search(zone_pattern, sql, re.IGNORECASE):
            print(f"⚠️ WARNING: User asked for zone={user_zone} but SQL might not have correct zone filter")
    
    # Validate BRANCH filter
    if "<BRANCH>" in user_entities:
        user_branch = user_entities["<BRANCH>"]
        branch_pattern = rf"branch\s*=\s*['\"]({user_branch})['\"]"
        if not re.search(branch_pattern, sql, re.IGNORECASE):
            print(f"⚠️ WARNING: User asked for branch={user_branch} but SQL might not have correct branch filter")
    
    print(f"✅ SQL entity validation passed")
    return sql


# Additional helper function for better SQL sanitization
def sanitize_sql_before_exec1411(sql, question):
    """Enhanced SQL sanitization with better error handling."""
    if not sql or not sql.strip():
        return """
        SELECT COUNT(*) as total_count
        FROM "bi_dwh"."main_cai_lib" 
        WHERE policy_end_date_year IS NOT NULL
        """
    
    try:
        # Remove any problematic WHERE clauses that end abruptly
        lines = sql.split('\n')
        cleaned_lines = []
        
        for line in lines:
            stripped = line.strip()
            # Skip lines that end with WHERE without conditions
            if stripped.endswith('WHERE') and not any(word in stripped.lower() for word in ['and', 'or', '=']):
                print(f"🔧 Removing problematic WHERE line: {stripped}")
                continue
            # Skip lines with malformed WHERE ... AND patterns
            if 'WHERE  AND' in line or 'WHERE AND' in line:
                print(f"🔧 Removing malformed WHERE AND line: {stripped}")
                continue
            cleaned_lines.append(line)
        
        cleaned_sql = '\n'.join(cleaned_lines)
        
        # Final validation - if still problematic, use fallback
        if ('WHERE  AND' in cleaned_sql or 
            cleaned_sql.strip().endswith('WHERE') or 
            'column "customer_segment" does not exist' in str(cleaned_sql).lower()):
            
            print("🔧 Using fallback SQL due to validation issues")
            return """
            SELECT 
                policy_end_date_year,
                policy_end_date_month,
                COUNT(policy_no) as policy_count,
                SUM(total_premium_payable) as total_premium
            FROM "bi_dwh"."main_cai_lib" 
            WHERE policy_end_date_year IS NOT NULL 
                AND policy_end_date_month IS NOT NULL
            GROUP BY policy_end_date_year, policy_end_date_month
            ORDER BY policy_end_date_year, policy_end_date_month
            """
        
        return cleaned_sql
        
    except Exception as e:
        print(f"⚠️ Error in SQL sanitization: {e}")
        # Return a safe fallback query
        return """
        SELECT 
            policy_end_date_year,
            policy_end_date_month,
            COUNT(policy_no) as policy_count
        FROM "bi_dwh"."main_cai_lib" 
        WHERE policy_end_date_year IS NOT NULL 
            AND policy_end_date_month IS NOT NULL
        GROUP BY policy_end_date_year, policy_end_date_month
        ORDER BY policy_end_date_year, policy_end_date_month
        LIMIT 100
        """



# ---------- Enhanced Main Ask Question Endpoint ----------
@csrf_exempt
def ask_question(request):
    try:
        start_time = datetime.datetime.now()

        if request.method == "POST":
            data = json.loads(request.body)
            question = data.get("question")
            user_id = data.get("user_id", "admin")

            print(f"📨 POST request received")
            print(f"🔎 Question: {question}")
            print(f"🔎 User ID: {user_id}")

            if not question:
                return JsonResponse({"error": "Missing question"}, status=400)

            # Enhanced context extraction
            memory_key = user_id
            history = conversation_memory_store.get(memory_key, [])
            time_context = extract_time_context_from_question(question)
            context = {"time_context": time_context}
            
            # Generate conversational opener FIRST
            # conversational_opener = generate_conversational_opener(question, context)
            # print(f"🗣️ Generated opener: {conversational_opener}")
            conversational_opener = generate_conversational_opener(question, context)
            print(f"🗣️ Generated opener: {conversational_opener}")

            # Clean decimal percentages (always remove .00%)
            conversational_opener = re.sub(
                r'(\d+)\.0+%',  # 100.00% or 100.0% → 100%
                lambda m: f"{m.group(1)}%",
                conversational_opener
            )

            # If ROUND_PERCENTAGES, also remove other decimals
            if ROUND_PERCENTAGES:
                conversational_opener = re.sub(
                    r'(\d+\.\d+)%',  # 93.42% → 93%
                    lambda m: f"{round(float(m.group(1)))}%",
                    conversational_opener
                )
    
            print(f"✅ Cleaned opener: {conversational_opener}")

            # Enhanced SQL generation with proper context
            try:
                # Column validation setup
                validation_enabled = False
                VALID_COLUMNS = []
                
                try:
                    VALID_COLUMNS = extract_columns_from_schema(FULL_SCHEMA)
                    validation_enabled = True
                except Exception as e:
                    print(f"⚠️ Column validation disabled: {e}")
                    validation_enabled = False
                    VALID_COLUMNS = []

                # Check for SQL reuse
                reused = best_prior_sql(question, history) if 'best_prior_sql' in globals() else None
                if reused:
                    sql = sanitize_sql_before_exec(reused, question)
                    used_reused_sql = True
                    print("♻️ Reusing prior SQL with enhanced context.")
                    recommendation = None
                    summary = ""
                else:
                    # Generate new SQL with enhanced context
                    aug_question = augment_question_with_context(question, history) if 'augment_question_with_context' in globals() else question
                    
                    sql, recommendation = run_sql_generation_graph(
                        aug_question, user_id=user_id, db_id="default", history=history
                    )
                    
                    # Enhanced SQL processing with CTE fixes
                    sql = sanitize_sql_before_exec(sql, question)
                    print(f"🔧 Enhanced SQL with time context and CTE fixes:\n{sql}")

                # Column validation if enabled
                if validation_enabled:
                    try:
                        invalid_cols = validate_sql_columns(sql, VALID_COLUMNS)
                        if invalid_cols:
                            return JsonResponse({
                                "answer": f"Invalid columns in SQL: {', '.join(invalid_cols)}",
                                "success": False,
                                "query_used": sql,
                                "conversational_opener": "Hmm, I noticed some column issues in that query. Let me help you fix that.",
                                "rows": [],
                                "row_count": 0,
                                "user_id": user_id,
                                "response_time": "0.00s"
                            }, status=400)
                    except Exception as ve:
                        print(f"⚠️ Column validation failed: {ve}")

            except Exception as sql_gen_error:
                print(f"❌ SQL generation error: {str(sql_gen_error)}")
                return JsonResponse({
                    "answer": "I ran into some trouble generating that query. Could you try rephrasing your question?",
                    "success": False,
                    "conversational_opener": "Sorry about that! I had a bit of trouble understanding your question.",
                    "error": str(sql_gen_error),
                    "rows": [],
                    "row_count": 0,
                    "user_id": user_id,
                    "response_time": "0.00s"
                }, status=500)

            # SQL validation
            if not sql or not sql.strip().lower().startswith(("select", "with")):
                return JsonResponse({
                    "answer": "I couldn't generate a valid query for that question. Could you try asking it differently?",
                    "success": False,
                    "conversational_opener": "Hmm, I'm having trouble with that query. Let me try a different approach.",
                    "query_used": sql or "No SQL generated",
                    "rows": [],
                    "row_count": 0,
                    "user_id": user_id,
                    "response_time": "0.00s"
                }, status=500)

            # Execute SQL
            try:
                print("✅ Executing enhanced SQL...")
                with ENGINE.connect() as conn:
                    result = conn.execute(text(sql))
                    rows = [dict(row._mapping) for row in result]
                print(f"✅ SQL executed successfully. Row count: {len(rows)}")

                # Generate enhanced summary and analysis
                summary = generate_summary_from_rows(question, sql, rows)
                
                # Enhanced context parsing
                try:
                    ctx = parse_sql_context(sql, VALID_COLUMNS if validation_enabled else [])
                    if time_context:
                        ctx.setdefault('time_context', {}).update(time_context)
                except Exception as ce:
                    print(f"⚠️ Context parsing failed: {ce}")
                    ctx = {"filters": {}, "time_context": time_context}

            except Exception as e:
                print("❌ SQL Execution Error:", str(e))
                return JsonResponse({
                    "answer": "I encountered an issue running that query. The data might be temporarily unavailable.",
                    "success": False,
                    "conversational_opener": "Oops! I ran into a snag with the database.",
                    "query_used": sql,
                    "error": str(e),
                    "rows": [],
                    "row_count": 0,
                    "response_time": "0.00s",
                    "user_id": user_id,
                }, status=500)

            # Enhanced answer generation
            answer = ""
            if rows:
                if len(rows) > 50:
                    answer = f"Found {len(rows)} results. That's quite a bit! You can download the full dataset, or I can help you filter it down."
                else:
                    formatted_rows = [", ".join(str(v) for v in row.values()) for row in rows[:3]]
                    answer = "\n".join(formatted_rows)
                    if len(rows) > 3:
                        answer += f"\n...and {len(rows) - 3} more rows."
            else:
                answer = "No data found for those criteria. Want to try broadening the search or checking a different time period?"

            # Enhanced chart and recommendation generation
            chart_config = None
            try:
                chart_config = llm_generate_chart_config(question, rows)
                print("📊 Generated chart config")
            except Exception as chart_err:
                print("⚠️ Chart generation failed:", chart_err)
                chart_config = None

            try:
                if not recommendation:
                    recommendation = llm_generate_recommendation(question, rows)
            except Exception as rec_err:
                print("⚠️ Recommendation generation failed:", rec_err)
                recommendation = "Based on this data, you might want to explore the trends further or segment by different criteria."

            # Enhanced conversation memory storage
            conversation_memory_store.setdefault(memory_key, []).append({
                "question": question,
                "sql": sql,
                "row_count": len(rows),
                "timestamp": datetime.datetime.now().isoformat(),
                "context": ctx,
                "time_context": time_context,
            })

            # Enhanced narrative generation
            try:
                data_window = {"from": "", "to": ""}
                if time_context.get('single_year'):
                    data_window = {"from": time_context['single_year'], "to": time_context['single_year']}
                elif time_context.get('year_range'):
                    data_window = {"from": time_context['year_range']['start'], "to": time_context['year_range']['end']}

                narrative = humanize_narrative(
                    question=question,
                    rows=rows,
                    summary=summary,
                    recommendation=recommendation,
                    sql=sql,
                    data_window=data_window,
                    opener=conversational_opener
                )
            except Exception as e:
                print(f"⚠️ Narrative generation failed: {e}")
                narrative = {
                    "opener": conversational_opener,
                    "insights": [summary] if summary else ["I've analyzed your data successfully."],
                    "recommendations": [recommendation] if recommendation else [],
                    "next_step": "Want to dive deeper into any of these insights? Just ask!"
                }

            total_time = (datetime.datetime.now() - start_time).total_seconds()

            return JsonResponse({
                "answer": answer,
                "success": True,
                "conversational_opener": conversational_opener,
                "query_used": sql,
                "rows": rows,
                "summary": summary,
                "chart_config": chart_config,
                "row_count": len(rows),
                "recommendation": recommendation,
                "response_time": f"{total_time:.2f}s",
                "narrative": narrative,
                "user_id": user_id,
                "history": conversation_memory_store[memory_key],
                "time_context": time_context,
            })

        # Handle other HTTP methods (GET for export, OPTIONS)
        elif request.method == "GET" and request.GET.get("export") == "true":
            # Export functionality with conversational elements
            question = request.GET.get("question")
            user_id = request.GET.get("user_id", "admin")
            
            if not question:
                return JsonResponse({"error": "Missing question parameter"}, status=400)

            try:
                memory_key = user_id
                history = conversation_memory_store.get(memory_key, [])
                time_context = extract_time_context_from_question(question)
                
                # Try to reuse SQL from memory
                sql = None
                for entry in reversed(history):
                    stored_question = (entry.get("question") or "").strip().lower()
                    search_question = (question or "").strip().lower()
                    if stored_question == search_question or search_question in stored_question:
                        sql = entry.get("sql")
                        break

                if not sql:
                    print("🔄 Generating new SQL for export...")
                    aug_question = augment_question_with_context(question, history) if 'augment_question_with_context' in globals() else question
                    raw_response = run_sql_generation_graph(
                        aug_question, user_id=user_id, db_id="default", history=history
                    )
                    sql, _ = raw_response if isinstance(raw_response, tuple) else (extract_sql_block(raw_response), None)

                # Enhanced SQL processing for export with CTE fixes
                sql = sanitize_sql_before_exec(sql, question)

                if not sql or not sql.strip().lower().startswith(("select", "with")):
                    return JsonResponse({"error": "Invalid SQL query type"}, status=400)

                # Execute and export
                with ENGINE.connect() as conn:
                    result = conn.execute(text(sql))
                    rows = [dict(row._mapping) for row in result]

                print(f"✅ Export query executed successfully. Found {len(rows)} rows")

                if not rows:
                    from django.http import HttpResponse
                    response = HttpResponse(content_type='text/csv')
                    response['Content-Disposition'] = 'attachment; filename="export_no_data.csv"'
                    response['Access-Control-Allow-Origin'] = '*'
                    response.write("No data found for this query")
                    return response

                # Generate CSV export
                from django.http import HttpResponse
                import csv
                
                timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"export_{timestamp}.csv"

                response = HttpResponse(content_type='text/csv')
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                response['Access-Control-Allow-Origin'] = '*'

                writer = csv.DictWriter(response, fieldnames=rows[0].keys())
                writer.writeheader()
                writer.writerows(rows)

                
                print(f"📁 CSV export created successfully with {len(rows)} rows")
                return response

            except Exception as e:
                print(f"❌ Export error: {str(e)}")
                return JsonResponse({
                    "error": f"Export failed: {str(e)}",
                    "details": "Check server logs for more information"
                }, status=500)

        elif request.method == "OPTIONS":
            from django.http import HttpResponse
            response = HttpResponse()
            response['Access-Control-Allow-Origin'] = '*'
            response['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
            response['Access-Control-Allow-Headers'] = 'Content-Type, X-Requested-With'
            return response

        else:
            return JsonResponse({"error": "Method not allowed. Use GET for export or POST for queries."}, status=405)

    except Exception as e:
        print("💥 Unexpected error in ask_question:")
        traceback.print_exc()

        # Safe error handling without re-parsing request.body
        uid = None
        try:
            if request.method == "GET":
                uid = request.GET.get("user_id")
            elif request.method == "POST":
                # Use already-parsed 'data' if available
                if 'data' in locals() and isinstance(data, dict):
                    uid = data.get("user_id")
        except Exception:
            uid = None

        return JsonResponse({
            "answer": "Something went wrong on my end. Let me try that again.",
            "success": False,
            "conversational_opener": "Oops! I encountered an unexpected issue.",
            "error": str(e),
            "rows": [],
            "row_count": 0,
            "response_time": "0.00s",
            "user_id": uid or "unknown"
        }, status=500)
    # except Exception as e:
    #     print("💥 Unexpected error in ask_question:")
    #     traceback.print_exc()
    #     return JsonResponse({
    #         "answer": "Something went wrong.",
    #         "success": False,
    #         "error": str(e),
    #         "rows": [],
    #         "row_count": 0,
    #         "response_time": "0.00s",
    #         "session_id": request.GET.get("session_id") if request.method == "GET" else json.loads(request.body).get("session_id", "unknown") if request.method == "POST" else "unknown"
    #     }, status=500)

# @api_view(['GET'])
# def download_reason_data(request):
#     """
#     GET params:
#       - reason: str (required)
#       - year: int (optional)
#       - month: int (optional)
#       - type: "csv" | "xlsx" | "excel" (default "csv")
#       - columns: comma-separated column keys in the order to export (optional)
#       - display_names: "true" to use UI display names as headers (optional)
#     """
#     DISPLAY_NAME_MAP = {
#         "policy_no": "Policy No",
#         "corrected_name": "Customer Name",
#         "product_name": "Product",
#         "primary_recommendation": "Recommendation",
#         "number_of_claims": "Claim Count",
#         "make_clean": "Manufacturer",
#         "additional_offers": "Additional Offer",
#         "policy_end_date": "Policy End Date",
#         "cleaned_zone_2": "Zone",
#         "biztype": "Business Type",
#         "age": "Vehicle Age",
#         "customer_tenure": "Customer Tenure",
#         "cleaned_reg_no": "Registration No",
#         "total_premium_payable": "Premium (₹)",
#         "cleaned_branch_name_2": "Branch",
#         "cleaned_state2": "State",
#         "tie_up": "Tie Ups",
#         "policy_end_date_month": "Policy End Date Month",
#         "policy_end_date_year": "Policy End Date Year",
#         "predicted_status": "Predicted Status",
#         "policy_tenure": "Policy Tenure",
#         "main_reason": "Main Reason",
#         "vehicle_idv": "Vehicle Idv",
#         "churn_probability": "Churn Probability",
#         "top_3_reasons": "Top 3 Reasons"
#     }
#     ALLOWED_COLUMNS = set(DISPLAY_NAME_MAP.keys())

#     def _quote_ident(ident: str) -> str:
#         return '"' + ident.replace('"', '""') + '"'

#     reason = request.GET.get("reason")
#     year = request.GET.get("year")
#     month = request.GET.get("month")
#     filetype = (request.GET.get("type") or "csv").lower()
#     display_names = (request.GET.get("display_names") or "").lower() == "true"

#     if not reason:
#         return Response({"error": "Missing 'reason' parameter"}, status=400)

#     raw_cols = (request.GET.get("columns") or "").strip()
#     requested_cols = [c.strip() for c in raw_cols.split(",") if c.strip()] if raw_cols else []
#     if not requested_cols:
#         requested_cols = [
#             "policy_no",
#             "corrected_name",
#             "product_name",
#             "primary_recommendation",
#             "additional_offers",
#             "policy_end_date",
#         ]

#     seen = set()
#     requested_cols = [c for c in requested_cols if c in ALLOWED_COLUMNS and (c not in seen and not seen.add(c))]
#     if not requested_cols:
#         return Response({"error": "No valid columns requested."}, status=400)

#     try:
#         conn = psycopg2.connect(
#             dbname=DB_New_Pred["dbname"],
#             user=DB_New_Pred["user"],
#             password=DB_New_Pred["password"],
#             host=DB_New_Pred["host"],
#             port=DB_New_Pred["port"]
#         )
#         cursor = conn.cursor()

#         filters = ['"main_reason" = %s', '"predicted_status" = %s']
#         params = [reason, 'Not Renewed']

#         if year:
#             filters.append('"policy_end_date_year" = %s')
#             params.append(int(year))
#         if month:
#             filters.append('"policy_end_date_month" = %s')
#             params.append(int(month))

#         where_clause = " WHERE " + " AND ".join(filters)

#         # Build SELECT list with formatted policy_end_date when requested
#         select_exprs = []
#         for c in requested_cols:
#             if c == "policy_end_date":
#                 select_exprs.append('"policy_end_date"')

#         select_list = ", ".join(select_exprs)

#         order_clause = ' ORDER BY "total_premium_payable" DESC' if "total_premium_payable" in requested_cols else ""

#         query = f'''
#             SELECT {select_list}
#             FROM "Prediction"."2025_prediction_final_data"
#             {where_clause}
#             {order_clause}
#         '''

#         cursor.execute(query, params)
#         rows = cursor.fetchall()
#         cursor.close()
#         conn.close()

#         headers = [DISPLAY_NAME_MAP.get(c, c) for c in requested_cols] if display_names else requested_cols
#         filename = f"Recommendation_data_{now().strftime('%Y%m%d_%H%M%S')}"

#         if filetype == "csv":
#             def csv_row_generator():
#                 yield '\ufeff'
#                 buf = io.StringIO()
#                 writer = csv.writer(buf, quoting=csv.QUOTE_MINIMAL)

#                 writer.writerow(headers)
#                 yield buf.getvalue(); buf.seek(0); buf.truncate(0)

#                 for row in rows:
#                     writer.writerow([val if val is not None else '' for val in row])
#                     yield buf.getvalue(); buf.seek(0); buf.truncate(0)

#             resp = StreamingHttpResponse(csv_row_generator(), content_type="text/csv")
#             resp['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
#             return resp

#         elif filetype in ("xlsx", "excel"):
#             wb = openpyxl.Workbook()
#             ws = wb.active
#             ws.title = "Policies"

#             ws.append(headers)
#             for row in rows:
#                 ws.append(row)

#             for i, col in enumerate(headers, 1):
#                 max_len = max([len(str(col))] + [len(str(r[i-1])) for r in rows if r[i-1] is not None])
#                 ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 50)

#             out = io.BytesIO()
#             wb.save(out); out.seek(0)

#             resp = StreamingHttpResponse(
#                 out,
#                 content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#             )
#             resp['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
#             return resp

#         else:
#             return Response({"error": "Unsupported file type"}, status=400)

#     except Exception as e:
#         print("❌ Error in download-reason-data:", e)
#         print(traceback.format_exc())
#         return Response({"error": str(e)}, status=500)


@api_view(['GET'])
def download_reason_data(request):
    DISPLAY_NAME_MAP = {
        "policy_no": "Policy No",
        "corrected_name": "Customer Name",
        "product_name": "Product",
        "primary_recommendation": "Recommendation",
        "number_of_claims": "Claim Count",
        "make_clean": "Manufacturer",
        "additional_offers": "Additional Offer",
        "policy_end_date": "Policy End Date",
        "cleaned_zone_2": "Zone",
        "biztype": "Business Type",
        "age": "Vehicle Age",
        "customer_tenure": "Customer Tenure",
        "cleaned_reg_no": "Registration No",
        "total_premium_payable": "Premium (₹)",
        "cleaned_branch_name_2": "Branch",
        "cleaned_state2": "State",
        "tie_up": "Tie Ups",
        "policy_end_date_month": "Policy End Date Month",
        "policy_end_date_year": "Policy End Date Year",
        "predicted_status": "Predicted Status",
        "policy_tenure": "Policy Tenure",
        "main_reason": "Main Reason",
        "vehicle_idv": "Vehicle Idv",
        "churn_probability": "Churn Probability",
        "top_3_reasons": "Top 3 Reasons"
    }
    ALLOWED_COLUMNS = set(DISPLAY_NAME_MAP.keys())

    def as_headers(cols, use_display):
        return [DISPLAY_NAME_MAP.get(c, c) for c in cols] if use_display else cols

    reason = request.GET.get("reason")
    year = request.GET.get("year")
    month = request.GET.get("month")
    filetype = (request.GET.get("type") or "csv").lower()
    display_names = (request.GET.get("display_names") or "").lower() == "true"

    if not reason:
        return Response({"error": "Missing 'reason' parameter"}, status=400)
    if reason not in REASON_VIEW_MAP:
        return Response({"error": f"Unknown reason: {reason}"}, status=400)

    raw_cols = (request.GET.get("columns") or "").strip()
    requested_cols = [c.strip() for c in raw_cols.split(",") if c.strip()] or [
        # default set if none requested
        "policy_no","corrected_name","product_name",
        "primary_recommendation","additional_offers","policy_end_date"
    ]
    # validate + preserve order
    seen = set()
    requested_cols = [c for c in requested_cols if c in ALLOWED_COLUMNS and (c not in seen and not seen.add(c))]
    if not requested_cols:
        return Response({"error": "No valid columns requested."}, status=400)

    try:
        conn = psycopg2.connect(
            dbname=DB_New_Pred["dbname"],
            user=DB_New_Pred["user"],
            password=DB_New_Pred["password"],
            host=DB_New_Pred["host"],
            port=DB_New_Pred["port"]
        )
        cursor = conn.cursor()

        schema, view = REASON_VIEW_MAP[reason]

        # WHERE filters
        where_parts = [sql.SQL('"predicted_status" = %s')]
        params = ['Not Renewed']

        if year:
            where_parts.append(sql.SQL('"policy_end_date_year" = %s'))
            params.append(int(year))
        if month:
            where_parts.append(sql.SQL('"policy_end_date_month" = %s'))
            params.append(int(month))

        if where_parts:
            where_sql = sql.SQL(" WHERE ") + sql.SQL(" AND ").join(where_parts)
        else:
            where_sql = sql.SQL("")

        q = sql.SQL("""
    SELECT {cols} 
    FROM {src}
    {where}
    ORDER BY "policy_no"
""").format(
    cols=sql.SQL(', ').join(sql.Identifier(c) for c in requested_cols),
    src=_ident(schema, view),
    where=where_sql if where_sql else sql.SQL('')
)


        cursor.execute(q, params)
        rows = cursor.fetchall()
        cursor.close(); conn.close()

        headers = as_headers(requested_cols, display_names)
        fname = f"Recommendation_data_{now().strftime('%Y%m%d_%H%M%S')}"

        if filetype == "csv":
            def csv_stream():
                yield '\ufeff'
                buf = io.StringIO()
                w = csv.writer(buf, quoting=csv.QUOTE_MINIMAL)
                w.writerow(headers); yield buf.getvalue(); buf.seek(0); buf.truncate(0)
                for r in rows:
                    w.writerow([("" if v is None else v) for v in r])
                    yield buf.getvalue(); buf.seek(0); buf.truncate(0)

            resp = StreamingHttpResponse(csv_stream(), content_type="text/csv")
            resp['Content-Disposition'] = f'attachment; filename="{fname}.csv"'
            return resp

        elif filetype in ("xlsx", "excel"):
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Policies"
            ws.append(headers)
            for r in rows: ws.append(r)
            for i, col in enumerate(headers, 1):
                max_len = max([len(str(col))] + [len(str(r[i-1])) for r in rows if r[i-1] is not None])
                ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 50)
            out = io.BytesIO(); wb.save(out); out.seek(0)
            resp = StreamingHttpResponse(out, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            resp['Content-Disposition'] = f'attachment; filename="{fname}.xlsx"'
            return resp

        return Response({"error": "Unsupported file type"}, status=400)

    except Exception as e:
        print("Error in download_reason_data:", e)
        print(traceback.format_exc())
        return Response({"error": str(e)}, status=500)






def download_template(request, filename):
    file_path = os.path.join(settings.MEDIA_ROOT, 'excel_templates', filename)
    if os.path.exists(file_path):
        return FileResponse(open(file_path, 'rb'), as_attachment=True, filename=filename)
    else:
        raise Http404("Template not found.")

@api_view(['DELETE'])
def remove_profile_image(request, user_id):
    user = get_object_or_404(CustomUser, pk=user_id)
    if user.profile_image:
        user.profile_image.delete(save=False)  # delete file
        user.profile_image = None
        user.save()
    return Response({'message': 'Image removed'}, status=200)

# Update with your actual config import

class DownloadFullDataForSegment(APIView):
    # Map used to rename headers to exactly what the table shows.
    # Keep this in sync with your React `columnDisplayNames`.
    DISPLAY_NAME_MAP = {
        "policy_no": "Policy No",
        "corrected_name": "Customer Name",
        "product_name": "Product",
        "primary_recommendation": "Recommendation",
        "number_of_claims": "Claim Count",
        "make_clean": "Manufacturer",
        "additional_offers": "Additional Offer",
        "policy_end_date": "Policy End Date",
        "cleaned_zone_2": "Zone",
        "biztype": "Business Type",
        "age": "Vehicle Age",
        "new_customers": "New Customer",
        "predicted_status": "Predicted Status",
        "customer_tenure": "Customer Tenure",
        # "rto_location": "RTO Location",
        "cleaned_reg_no": "Registration No",
        "total_premium_payable": "Premium (₹)",
        "cleaned_branch_name_2": "Branch",
        "cleaned_state2": "State",
        "tie_up": "Tie Ups",
        "churn_probability": "Churn Probability",
        "main_reason": "Churn Reason",
        "model_clean":"Model",
        "vehicle_idv":"Vehicle Idv",
        "ncb_amount" : "Ncb Amount",
        "policy_tenure" : "Policy Tenure",
        "clv":"Clv",
        "customer_segment" : "Customer Segment",
        "top_3_reasons" : "Top 3 Reasons"
    }

    # If you want an allowlist to prevent arbitrary column requests:
    ALLOWED_COLUMNS = set(DISPLAY_NAME_MAP.keys()) | {
        # include any other columns you might expose in the UI
        "policy_end_date_month", "policy_end_date_year", "customer_segment"
    }

    def _quote_ident(self, ident: str) -> str:
        # Quote SQL identifiers safely (handles special chars like /)
        return '"' + ident.replace('"', '""') + '"'

    def get(self, request):
        try:
            file_type = (request.GET.get("type") or "excel").lower()
            segment = request.GET.get("segment")
            month = request.GET.get("month")
            year = request.GET.get("year")
            display_names = (request.GET.get("display_names") or "").lower() == "true"

            # Ordered list of requested columns (from the table)
            raw_cols = (request.GET.get("columns") or "").strip()
            requested_cols = [c.strip() for c in raw_cols.split(",") if c.strip()] if raw_cols else []

            # If nothing provided, fall back to a sane default (your mandatory keys)
            if not requested_cols:
                requested_cols = [
                    "policy_no", "product_name", "biztype", "cleaned_state2",
                    "cleaned_branch_name_2", "total_premium_payable",
                    "churn_probability", "main_reason", "policy_end_date",
                ]

            # Allowlist filter (optional but recommended)
            requested_cols = [c for c in requested_cols if c in self.ALLOWED_COLUMNS]

            if not requested_cols:
                return HttpResponse("No valid columns requested.", status=400)

            # Build filters
            filters = ['"predicted_status" = %s']
            values = ['Not Renewed']

            if segment:
                filters.append('"customer_segment" = %s')
                values.append(segment)

            if month:
                filters.append('EXTRACT(MONTH FROM "policy_end_date") = %s')
                values.append(month)

            if year:
                filters.append('EXTRACT(YEAR FROM "policy_end_date") = %s')
                values.append(year)

            where_clause = "WHERE " + " AND ".join(filters) if filters else ""

            # Build SELECT list in the same order as requested
            select_list = ", ".join(self._quote_ident(c) for c in requested_cols)

            query = f'''
                SELECT {select_list}
                FROM "Prediction"."2025_prediction_final_data"
                {where_clause}
                ORDER BY "policy_no"
            '''

            conn = psycopg2.connect(
                dbname=DB_New_Pred["dbname"],
                user=DB_New_Pred["user"],
                password=DB_New_Pred["password"],
                host=DB_New_Pred["host"],
                port=DB_New_Pred["port"]
            )
            cursor = conn.cursor()
            cursor.execute(query, values)
            rows = cursor.fetchall()
            cursor.close()
            conn.close()

            filename = f"segmentation_data_{now().strftime('%Y%m%d_%H%M%S')}"

            # Prepare headers: either DB column names or the display names
            if display_names:
                headers = [self.DISPLAY_NAME_MAP.get(c, c) for c in requested_cols]
            else:
                headers = requested_cols

            if file_type == "csv":
                def generate_csv():
                    yield '\ufeff'  # UTF-8 BOM
                    buffer = io.StringIO()
                    writer = csv.writer(buffer, quoting=csv.QUOTE_MINIMAL)

                    writer.writerow(headers)
                    yield buffer.getvalue()
                    buffer.seek(0)
                    buffer.truncate(0)

                    for row in rows:
                        writer.writerow([val if val is not None else '' for val in row])
                        yield buffer.getvalue()
                        buffer.seek(0)
                        buffer.truncate(0)

                response = StreamingHttpResponse(generate_csv(), content_type="text/csv")
                response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
                return response

            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "SegmentedData"

                ws.append(headers)
                for row in rows:
                    ws.append(row)

                # Optional: basic autofit
                for i, col in enumerate(headers, 1):
                    max_len = max([len(str(col))] + [len(str(r[i - 1])) for r in rows if r[i - 1] is not None])
                    ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 50)

                output = io.BytesIO()
                wb.save(output)
                output.seek(0)

                response = StreamingHttpResponse(
                    output,
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
                return response

        except Exception as e:
            print("❌ Error in download-full-data:", e)
            print(traceback.format_exc())
            return HttpResponse(f"Error: {str(e)}", status=500)



class UpdateUserView(APIView):
    def put(self, request, user_id):
        user = get_object_or_404(CustomUser, id=user_id)
        serializer = UpdateUserSerializer(user, data=request.data, partial=True)

        if serializer.is_valid():
            serializer.save()
            return Response({"message": "User updated successfully"}, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
def delete_user(request, user_id):
    user = get_object_or_404(CustomUser, id=user_id)
    user.delete()
    return Response({"message": "User deleted successfully."}, status=status.HTTP_204_NO_CONTENT)

class CurrentUserView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        serializer = UserSerializer(user)
        return Response(serializer.data)

class CreateUserView(generics.CreateAPIView):
    queryset = CustomUser.objects.all()
    serializer_class = CreateUserSerializer

class ListUsersView(generics.ListAPIView):
    queryset = CustomUser.objects.all()
    serializer_class = UserSerializer
    def get_serializer_context(self):
        return {'request': self.request}
class ListPagesView(generics.ListAPIView):
    queryset = Page.objects.all()
    serializer_class = PageSerializer

class AssignPageView(APIView):
    def post(self, request, *args, **kwargs):
        print("📥 AssignPageView called")
        print("🔍 Incoming request data:", request.data)

        user_id = request.data.get('user')
        page_name = request.data.get('page')

        if not user_id or not page_name:
            print("❌ Missing user or page")
            return Response(
                {'status': 'failed', 'reason': 'Missing user or page in request'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            user = CustomUser.objects.get(id=user_id)
            print(f"✅ Found user: {user.username}")
        except CustomUser.DoesNotExist:
            return Response(
                {'status': 'failed', 'reason': f'User with ID {user_id} not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        try:
            page = Page.objects.get(name=page_name)
            print(f"✅ Found page: {page.name}")
        except Page.DoesNotExist:
            return Response(
                {'status': 'failed', 'reason': f'Page with name {page_name} not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        access, created = UserPageAccess.objects.get_or_create(user=user, page=page)
        if created:
            print(f"✅ Access granted: user {user.username} -> page {page.name}")
            return Response({
                "id": access.id,
                "page": access.page.name
            }, status=status.HTTP_201_CREATED)
        else:
            print(f"ℹ️ Access already exists")
            return Response({
                "id": access.id,
                "page": access.page.name
            }, status=status.HTTP_200_OK)

# @csrf_exempt
# def connect_database(request):
#     """Connect to PostgreSQL via SQLAlchemy and prepare PostgreSQLChatAnalyzer"""
#     logger.info("=== STATIC DB CONNECTION TO Conversational_AI_Datafeed START ===")

#     try:
#         # Build SQLAlchemy connection string
#         db_url = f"postgresql+psycopg2://{DB_Conv_AI['user']}:{DB_Conv_AI['password']}@{DB_Conv_AI['host']}:{DB_Conv_AI['port']}/{DB_Conv_AI['dbname']}"
#         engine = create_engine(db_url)
        
#         # Test the connection and fetch schema info
#         inspector = inspect(engine)
#         columns = inspector.get_columns("GBM1_prediction_data_with_recommendations", schema="stage")

#         if not columns:
#             return JsonResponse({"error": "Table 'stage.GBM1_prediction_data_with_recommendations' not found."}, status=404)

#         column_names = [col["name"] for col in columns]
#         logger.info(f"Connected. Columns: {column_names}")

#         # Generate a session ID and create analyzer
#         session_id = str(uuid.uuid4())
#         analyzer = PostgreSQLChatAnalyzer(engine=engine, schema_info={})
#         session_store[session_id] = analyzer

#         return JsonResponse({
#             "message": "Connected successfully",
#             "session_id": session_id,
#             "columns": column_names,
#             "status": "Ready for querying"
#         })

#     except Exception as e:
#         logger.error(f"Connection error: {str(e)}")
#         logger.error(traceback.format_exc())
#         return JsonResponse({"error": f"Connection failed: {str(e)}"}, status=500)


class RemovePageAccessView(generics.DestroyAPIView):
    queryset = UserPageAccess.objects.all()
    lookup_field = 'id'

    def delete(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return Response({'status': 'deleted'})


@api_view(["GET"])
def get_roles(request):
    """Fetch distinct roles from the User model."""
    roles = User.objects.values_list('role', flat=True).distinct()
    return Response({"roles": list(roles)})

class MyTokenObtainPairView(TokenObtainPairView):

    role = serializers.CharField(write_only=True, required=True)

    def validate(self, attrs):
        username = attrs.get("username")
        password = attrs.get("password")
        requested_role = attrs.get("role")

        if not requested_role:
            raise serializers.ValidationError({"role": "Role is required."})

        # ✅ Authenticate using role-based backend
        user = authenticate(username=username, password=password)

        if user is None or str(user.role).lower() != requested_role.lower():
            raise serializers.ValidationError({"detail": "Invalid username, password, or role."})

        refresh = self.get_token(user)

        return {
            'refresh': str(refresh),
            'access': str(refresh.access_token),
            'role': str(user.role),
            'username': user.username,
            'email': user.email
        }

        # def post(self, request, *args, **kwargs):
        #     print("🔍 CUSTOM VIEW CALLED")
        #     print("🔍 Request data:", request.data)
        #     return super().post(request, *args, **kwargs)



def ensure_reports_table():
    """Create the saved_reports table if it doesn't exist."""
    conn = get_connection()
    if not conn:
        return False

    try:
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS bi_dwh.saved_reports (
                id SERIAL PRIMARY KEY,
                report_name VARCHAR(255) NOT NULL UNIQUE,
                table_name VARCHAR(255) NOT NULL,
                selected_columns JSON,
                slicers JSON,
                dashboard_components JSON,
                charts JSON,
                description TEXT,
                tags JSON,
                created_by VARCHAR(255),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)
        conn.commit()
        return True
    except Exception as e:
        logging.error(f"Error creating reports table: {str(e)}")
        return False
    finally:
        conn.close()


class ReportAPIView(APIView):
    """Handle report CRUD operations"""
    
    def post(self, request):
        """Save a new report or update existing one"""
        try:
            data = request.data
            
            # Validate required fields
            if not data.get('report_name'):
                return Response({'error': 'Report name is required'}, status=status.HTTP_400_BAD_REQUEST)
            
            if not data.get('table_name'):
                return Response({'error': 'Table name is required'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Ensure reports table exists
            if not ensure_reports_table():
                return Response({'error': 'Failed to initialize reports table'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
            conn = get_connection()
            if not conn:
                return Response({'error': 'Database connection failed'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
            try:
                cursor = conn.cursor(cursor_factory=RealDictCursor)
                
                # Check if report with same name already exists
                cursor.execute("""
                    SELECT id, report_name, table_name, selected_columns, slicers, 
                           dashboard_components, charts, description, tags, created_by,
                           created_at, updated_at
                    FROM bi_dwh.saved_reports 
                    WHERE report_name = %s
                """, (data.get('report_name'),))
                
                existing_report = cursor.fetchone()
                
                if existing_report:
                    # Update existing report
                    cursor.execute("""
                        UPDATE bi_dwh.saved_reports 
                        SET table_name = %s, selected_columns = %s, slicers = %s,
                            dashboard_components = %s, charts = %s, description = %s,
                            tags = %s, created_by = %s, updated_at = CURRENT_TIMESTAMP
                        WHERE report_name = %s
                        RETURNING id, report_name, table_name, selected_columns, slicers,
                                  dashboard_components, charts, description, tags, created_by,
                                  created_at, updated_at
                    """, (
                        data.get('table_name'),
                        json.dumps(data.get('selected_columns', [])),
                        json.dumps(data.get('slicers', {})),
                        json.dumps(data.get('dashboard_components', [])),
                        json.dumps(data.get('charts', [])),
                        data.get('description'),
                        json.dumps(data.get('tags', [])),
                        data.get('created_by'),
                        data.get('report_name')
                    ))
                    
                    updated_report = cursor.fetchone()
                    conn.commit()
                    
                    return Response({
                        'message': 'Report updated successfully',
                        'report': dict(updated_report)
                    }, status=status.HTTP_200_OK)
                else:
                    # Create new report
                    cursor.execute("""
                        INSERT INTO bi_dwh.saved_reports 
                        (report_name, table_name, selected_columns, slicers, dashboard_components,
                         charts, description, tags, created_by)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                        RETURNING id, report_name, table_name, selected_columns, slicers,
                                  dashboard_components, charts, description, tags, created_by,
                                  created_at, updated_at
                    """, (
                        data.get('report_name'),
                        data.get('table_name'),
                        json.dumps(data.get('selected_columns', [])),
                        json.dumps(data.get('slicers', {})),
                        json.dumps(data.get('dashboard_components', [])),
                        json.dumps(data.get('charts', [])),
                        data.get('description'),
                        json.dumps(data.get('tags', [])),
                        data.get('created_by')
                    ))
                    
                    new_report = cursor.fetchone()
                    conn.commit()
                    
                    return Response({
                        'message': 'Report saved successfully',
                        'report': dict(new_report)
                    }, status=status.HTTP_201_CREATED)
                    
            except psycopg2.IntegrityError:
                conn.rollback()
                return Response({'error': 'Report name already exists'}, status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                conn.rollback()
                logging.error(f"Error saving report: {str(e)}")
                return Response({'error': 'Failed to save report'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            finally:
                conn.close()
                
        except Exception as e:
            logging.error(f"Error in save_report: {str(e)}")
            return Response({'error': 'Failed to save report'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def get(self, request):
        """Get all saved reports"""
        try:
            conn = get_connection()
            if not conn:
                return Response({'error': 'Database connection failed'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            cursor.execute("""
                SELECT id, report_name, table_name, selected_columns, slicers,
                       dashboard_components, charts, description, tags, created_by,
                       created_at, updated_at
                FROM bi_dwh.saved_reports 
                ORDER BY updated_at DESC
            """)
            
            reports = cursor.fetchall()
            print(f"Fetched {len(reports)} reports from bi_dwh.saved_reports table")
            return Response({
                'reports': [dict(report) for report in reports]
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            logging.error(f"Error fetching reports: {str(e)}")
            return Response({'error': 'Failed to fetch reports'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        finally:
            if conn:
                conn.close()


class ReportDetailAPIView(APIView):
    """Handle individual report operations"""
    
    def get(self, request, report_id):
        """Get a specific report by ID"""
        try:
            conn = get_connection()
            if not conn:
                return Response({'error': 'Database connection failed'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            cursor.execute("""
                SELECT id, report_name, table_name, selected_columns, slicers,
                       dashboard_components, charts, description, tags, created_by,
                       created_at, updated_at
                FROM bi_dwh.saved_reports 
                WHERE id = %s
            """, (report_id,))
            
            report = cursor.fetchone()
            
            if not report:
                return Response({'error': 'Report not found'}, status=status.HTTP_404_NOT_FOUND)
            
            return Response({
                'report': dict(report)
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            logging.error(f"Error fetching report: {str(e)}")
            return Response({'error': 'Failed to fetch report'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        finally:
            if conn:
                conn.close()

    def put(self, request, report_id):
        """Update a specific report"""
        try:
            conn = get_connection()
            if not conn:
                return Response({'error': 'Database connection failed'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
            data = request.data
            
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            
            # Check if report exists
            cursor.execute("SELECT id FROM bi_dwh.saved_reports WHERE id = %s", (report_id,))
            if not cursor.fetchone():
                return Response({'error': 'Report not found'}, status=status.HTTP_404_NOT_FOUND)
            
            # Build update query dynamically
            update_fields = []
            update_values = []
            
            if 'report_name' in data:
                update_fields.append('report_name = %s')
                update_values.append(data['report_name'])
            if 'table_name' in data:
                update_fields.append('table_name = %s')
                update_values.append(data['table_name'])
            if 'selected_columns' in data:
                update_fields.append('selected_columns = %s')
                update_values.append(json.dumps(data['selected_columns']))
            if 'slicers' in data:
                update_fields.append('slicers = %s')
                update_values.append(json.dumps(data['slicers']))
            if 'dashboard_components' in data:
                update_fields.append('dashboard_components = %s')
                update_values.append(json.dumps(data['dashboard_components']))
            if 'charts' in data:
                update_fields.append('charts = %s')
                update_values.append(json.dumps(data['charts']))
            if 'description' in data:
                update_fields.append('description = %s')
                update_values.append(data['description'])
            if 'tags' in data:
                update_fields.append('tags = %s')
                update_values.append(json.dumps(data['tags']))
            if 'created_by' in data:
                update_fields.append('created_by = %s')
                update_values.append(data['created_by'])
            
            # Always update the timestamp
            update_fields.append('updated_at = CURRENT_TIMESTAMP')
            update_values.append(report_id)
            
            if len(update_fields) == 1:  # Only timestamp update
                return Response({'error': 'No fields to update'}, status=status.HTTP_400_BAD_REQUEST)
            
            query = f"""
                UPDATE bi_dwh.saved_reports 
                SET {', '.join(update_fields[:-1])} 
                WHERE id = %s
                RETURNING id, report_name, table_name, selected_columns, slicers,
                          dashboard_components, charts, description, tags, created_by,
                          created_at, updated_at
            """
            
            cursor.execute(query, update_values)
            updated_report = cursor.fetchone()
            conn.commit()
            
            return Response({
                'message': 'Report updated successfully',
                'report': dict(updated_report)
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            conn.rollback()
            logging.error(f"Error updating report: {str(e)}")
            return Response({'error': 'Failed to update report'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        finally:
            if conn:
                conn.close()

    def delete(self, request, report_id):
        """Delete a specific report"""
        try:
            conn = get_connection()
            if not conn:
                return Response({'error': 'Database connection failed'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
            cursor = conn.cursor()
            
            # Check if report exists
            cursor.execute("SELECT id FROM bi_dwh.saved_reports WHERE id = %s", (report_id,))
            if not cursor.fetchone():
                return Response({'error': 'Report not found'}, status=status.HTTP_404_NOT_FOUND)
            
            # Delete the report
            cursor.execute("DELETE FROM bi_dwh.saved_reports WHERE id = %s", (report_id,))
            conn.commit()
            
            return Response({
                'message': 'Report deleted successfully'
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            conn.rollback()
            logging.error(f"Error deleting report: {str(e)}")
            return Response({'error': 'Failed to delete report'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        finally:
            if conn:
                conn.close()


class ReportSearchAPIView(APIView):
    """Search reports by name or table"""
    
    def get(self, request):
        try:
            query = request.GET.get('q', '')
            table_name = request.GET.get('table', '')
            
            conn = get_connection()
            if not conn:
                return Response({'error': 'Database connection failed'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            
            # Build search query
            where_conditions = []
            search_params = []
            
            if query:
                where_conditions.append("report_name ILIKE %s")
                search_params.append(f'%{query}%')
            
            if table_name:
                where_conditions.append("table_name = %s")
                search_params.append(table_name)
            
            where_clause = ""
            if where_conditions:
                where_clause = "WHERE " + " AND ".join(where_conditions)
            
            search_query = f"""
                SELECT id, report_name, table_name, selected_columns, slicers,
                       dashboard_components, charts, description, tags, created_by,
                       created_at, updated_at
                FROM bi_dwh.saved_reports 
                {where_clause}
                ORDER BY updated_at DESC
            """
            
            cursor.execute(search_query, search_params)
            reports = cursor.fetchall()
            
            return Response({
                'reports': [dict(report) for report in reports]
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            logging.error(f"Error searching reports: {str(e)}")
            return Response({'error': 'Failed to search reports'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        finally:
            if conn:
                conn.close()


@api_view(['GET'])
def churn_dashboard(request):   
    year = request.GET.get('year')
    state = request.GET.get('state')
    branch = request.GET.get('branch')

    conn = psycopg2.connect(
        host=DB_New_Pred["host"],
        database=DB_New_Pred["dbname"],
        user=DB_New_Pred["user"],
        password=DB_New_Pred["password"],
        port=5432,
    )
    cursor = conn.cursor(cursor_factory=RealDictCursor)

    # ==================== Base Filter Query ====================
    filter_conditions = []
    if year:
        filter_conditions.append(f"EXTRACT(YEAR FROM \"policy end date\") = {year}")
    if state:
        filter_conditions.append(f"\"Cleaned State2\" = '{state}'")
    if branch:
        filter_conditions.append(f"\"Cleaned Branch Name 2\" = '{branch}'")

    filter_query = " AND ".join(filter_conditions)
    if filter_query:
        filter_query = "WHERE " + filter_query

    table_name = '"public"."overall_cleaned_base_and_pr_ef_policyef_with_reasons_bucket"'

    # ==================== Metrics Queries ====================

    # Total policies current year onboard
    cursor.execute(f"""
        SELECT COUNT(DISTINCT "policy no") as total_policies
        FROM {table_name}
        {filter_query} limit 1000
    """)
    total_policies = cursor.fetchone()['total_policies']

    # New policies
    new_policies_query = f"""
        SELECT COUNT(DISTINCT "policy no") as new_policies
        FROM {table_name} 
    """
    if filter_query:
        new_policies_query += f" {filter_query} AND \"New Customers\" = 'Yes'"
    else:
        new_policies_query += f" WHERE \"New Customers\" = 'Yes'"

    cursor.execute(new_policies_query)
    new_policies = cursor.fetchone()['new_policies']

    # Churned policies (policy_end_date_year = selected year -1)
    churned_policies = 0
    first_year_churned = 0
    churn_year = None

    if year:
        churn_year = int(year) - 1

        # Churned policies query
        churn_condition = f"EXTRACT(YEAR FROM \"policy end date\") = {churn_year} AND \"Churn Label\" = 'Yes'"
        churn_filter = filter_conditions.copy()
        churn_filter.append(churn_condition)
        churn_query = " AND ".join(churn_filter)
        if churn_query:
            churn_query = "WHERE " + churn_query

        cursor.execute(f"""
            SELECT COUNT(DISTINCT "policy no") as churned_policies
            FROM {table_name}
            {churn_query} limit 1000
        """)
        churned_policies = cursor.fetchone()['churned_policies']

        # First year churned query
        first_year_churn_condition = f"EXTRACT(YEAR FROM \"policy end date\") = {churn_year} AND \"Churn Label\" = 'Yes' AND \"New Customers\" = 'Yes'"
        first_year_filter = filter_conditions.copy()
        first_year_filter.append(first_year_churn_condition)
        first_year_query = " AND ".join(first_year_filter)
        if first_year_query:
            first_year_query = "WHERE " + first_year_query

        cursor.execute(f"""
            SELECT COUNT(DISTINCT "policy no") as first_year_churned
            FROM {table_name}
            {first_year_query} limit 1000
        """)
        first_year_churned = cursor.fetchone()['first_year_churned']

    # ==================== Churn Rate % ====================
    churn_rate = (churned_policies / total_policies * 100) if total_policies else 0

    # ==================== Top Highlighted Reasons ====================
    cursor.execute(f"""
        SELECT "Not Renewed Reasons",
               COUNT(DISTINCT "policy no")::float / NULLIF({total_policies}, 0) * 100 as percentage
        FROM {table_name}
        {filter_query}
        GROUP BY "Not Renewed Reasons"
        ORDER BY percentage DESC
        LIMIT 5
    """)
    top_reasons = cursor.fetchall()

    cursor.execute(f"""
        SELECT DISTINCT EXTRACT(YEAR FROM "policy end date") AS year
        FROM {table_name} 
        ORDER BY year DESC
        limit 1000
    """)
    years = [int(row['year']) for row in cursor.fetchall() if row['year']]

    # States
    cursor.execute(f"""
        SELECT DISTINCT "Cleaned State2" AS state
        FROM {table_name} 
        ORDER BY state 
    """)
    states = [row['state'] for row in cursor.fetchall() if row['state']]

    # Branches
    cursor.execute(f"""
        SELECT DISTINCT "Cleaned Branch Name 2" AS branch
        FROM {table_name} 
        ORDER BY branch
    """)
    branches = [row['branch'] for row in cursor.fetchall() if row['branch']]

    conn.close()

    return Response({
        "metrics": {
            "total_policies": total_policies,
            "new_policies": new_policies,
            "churned_policies": churned_policies,
            "first_year_churned": first_year_churned,
            "churn_rate": round(churn_rate, 2)
        },
        "top_highlighted_reasons": top_reasons,
        "filter_options": {
            "years": years,
            "states": states,
            "branches": branches
        }
    })


def get_connection():
    """Establish a connection to the database."""
    
    try:
        return psycopg2.connect(
            host=DB_New_Pred["host"],
            database=DB_New_Pred["dbname"],
            user=DB_New_Pred["user"],
            password=DB_New_Pred["password"]
        )
    except Exception as e:
        logger.error(f"Database connection error: {str(e)}")
    return None

@api_view(['GET'])
def get_table_data(request):
    """Fetch all rows from a given table inside 'Prediction' schema."""
    table_name = request.GET.get('table')

    if not table_name:
        logger.error("Table name parameter missing in request")
        return Response(
            {
                "error": "Table name is required.",
                "detail": "Please provide a table name as a query parameter (?table=your_table_name)"
            }, 
            status=status.HTTP_400_BAD_REQUEST
        )

    # Validate table name format
    if not re.match(r'^[a-zA-Z0-9_ \-]+$', table_name):
        return Response(
            {"error": "Invalid table name format."},
            status=status.HTTP_400_BAD_REQUEST
        )

    conn = None
    try:
        conn = get_connection()
        if not conn:
            return Response(
                {"error": "Unable to connect to the database."}, 
                status=status.HTTP_503_SERVICE_UNAVAILABLE
            )

        with conn.cursor() as cursor:
            # Safely quote the table name
            safe_table_name = table_name.replace('"', '""')
            query = f'SELECT * FROM "Prediction"."{safe_table_name}"'
            
            cursor.execute(query)
            columns = [desc[0] for desc in cursor.description]
            rows = cursor.fetchall()

            data = [dict(zip(columns, row)) for row in rows]
            return Response({"columns": columns, "data": data})

    except psycopg2.Error as e:
        logger.error(f"Database error fetching from {table_name}: {str(e)}")
        return Response(
            {"error": f"Database operation failed: {str(e)}"}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
    except Exception as e:
        logger.error(f"Error fetching data from {table_name}: {str(e)}")
        return Response(
            {"error": f"Failed to fetch data: {str(e)}"}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
    finally:
        if conn:
            conn.close()

@api_view(['GET'])
def get_tables(request):
    """Fetch a list of all tables inside 'Prediction' schema."""
    conn = None
    try:
        conn = get_connection()
        if not conn:
            logger.error("Database connection failed")
            return Response(
                {"error": "Unable to connect to database server"}, 
                status=status.HTTP_503_SERVICE_UNAVAILABLE
            )

        cursor = conn.cursor()
        cursor.execute("""
            SELECT table_name 
            FROM information_schema.tables 
            WHERE table_schema = 'Prediction'
            ORDER BY table_name;
        """)
        
        tables = [row[0] for row in cursor.fetchall()]
        return Response({"tables": tables})

    except Exception as e:
        logger.error(f"Error fetching tables: {str(e)}")
        return Response(
            {"error": "Failed to retrieve tables list"}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
    finally:
        if conn:
            conn.close()


@api_view(['GET'])
def get_columns(request):
    """Fetch a list of all columns for a given table inside 'public' schema."""
    table_name = request.GET.get('table')

    if not table_name:
        return Response(
            {"error": "Table name is required."}, 
            status=status.HTTP_400_BAD_REQUEST
        )

    conn = get_connection()
    if not conn:
        return Response(
            {"error": "Unable to connect to the database."}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

    try:
        cursor = conn.cursor()
        table_name = table_name.strip()
        # ✅ Ensure schema is used
        cursor.execute(sql.SQL("""
           SELECT column_name 
           FROM information_schema.columns 
           WHERE table_schema = 'Prediction' 
           AND LOWER(table_name) = LOWER(%s);
"""), (table_name,))


        columns = [row[0] for row in cursor.fetchall()]
        return Response({"columns": columns})

    except Exception as e:
        logger.error(f"Error fetching columns: {str(e)}")
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    finally:
        conn.close()

def extract_text_from_pdf(file_path):
    print(f"Extracting text from PDF: {file_path}")
    reader = PdfReader(file_path)
    text = ""
    for page in reader.pages:
        text += page.extract_text() or ""
    print("Text extraction complete.")
    return text


def ingest_pdf_to_vectorstore(file_path, vectorstore_path):
    print(f"Ingesting PDF to vectorstore: {file_path} -> {vectorstore_path}")
    text = extract_text_from_pdf(file_path)
    splitter = CharacterTextSplitter(chunk_size=CHUNK_SIZE, chunk_overlap=CHUNK_OVERLAP)
    chunks = [Document(page_content=c) for c in splitter.split_text(text)]
    vectordb = FAISS.from_documents(chunks, embedding_model)
    vectordb.save_local(vectorstore_path)
    print("Vectorstore ingestion complete.")

def get_best_vectorstore(query):
    print(f"Loading permanent admin vectorstore for query: {query}")

    try:
        vectordb = FAISS.load_local(
            ADMIN_VECTORSTORE_PATH,
            embedding_model,
            allow_dangerous_deserialization=True
        )
        print("Permanent vectorstore loaded.")
        return vectordb
    except Exception as e:
        print(f"Failed to load admin vectorstore: {e}")
        return None

@csrf_exempt
def upload_pdfbot(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Only POST allowed'}, status=405)

    pdf_file = request.FILES.get('file')
    if not pdf_file:
        return JsonResponse({'error': 'No file uploaded'}, status=400)

    os.makedirs(UPLOAD_DIR, exist_ok=True)
    os.makedirs(VECTORSTORE_DIR, exist_ok=True)

    file_path = os.path.join(UPLOAD_DIR, pdf_file.name)
    with open(file_path, 'wb+') as destination:
        for chunk in pdf_file.chunks():
            destination.write(chunk)

    print(f"PDF uploaded: {pdf_file.name}")

    # Always store to the admin_base vectorstore
    ingest_pdf_to_vectorstore(file_path, ADMIN_VECTORSTORE_PATH)

    return JsonResponse({'message': 'PDF uploaded and permanently indexed for all users.'})

def build_system_prompt_for_policy_schema(self):
            """Builds a smart system prompt for policy-related questions using known columns and fuzzy matching."""
            return (
                "You are a PostgreSQL expert helping translate user questions into precise SQL queries.\n"
                "Always use the table: \"stage\".\"GBM1_prediction_data_with_recommendations\"\n\n"

                "Relevant columns include:\n"
                "- predicted_status: 'Renewed' / 'Not Renewed'\n"
                "- churn_category, churn_probability, main_reason, not_renewed_reasons\n"
                "- make_clean, model_clean, variant, tie_up, vehicle_idv, segment_risk_score\n"
                "- total_premium_payable, total_od_premium, total_tp_premium, gst, clv, clv_category\n"
                "- corrected_name, policy_no, customerid, policy_status, retention_channel\n"
                "- customer_segment, discount_category, fuel_type_risk_factor, approved, number_of_claims\n"
                "- policy_start_date_year/month/day and policy_end_date_year/month/day\n"
                "- retention_streak, age, cleaned_branch_name_2, biztype, product_name\n\n"

                "Handle the following mappings:\n"
                "- If the user asks about segment keywords like 'elite', 'potential', 'low value', 'risky', match them against 'customer_segment'.\n"
                "    Eg.Use: UPPER(customer_segment) = UPPER('Elite Retainers') or appropriate full label.\n"
                "- If the user asks about communication method (SMS, Email, WhatsApp), use 'retention_channel'\n"
                "- If the user mentions values like 0.68, match with churn_probability or manufacturer_risk_rate depending on intent\n\n"

                "Fuzzy mapping examples:\n"
                "- 'How many people renewed?' → predicted_status = 'Renewed'\n"
                "- 'Churn reasons' → use main_reason or not_renewed_reasons\n"
                "- 'Top car brands' or 'most common vehicle companies' → group by make_clean\n"
                "- 'Total premium paid' → total_premium_payable\n"
                "- 'How many claims?' → number_of_claims\n"
                "- 'Retention channel usage' → group by retention_channel\n"
                "- 'Revenue' → total_premium_payable (or total_revenue if available)\n\n"

                "Date Handling:\n"
                "- If the user asks for a single year (e.g., 2024), use: policy_end_date_year = 2024\n"
                "- If the user asks for a range or comparison (e.g., 'between 2024 and 2025', or '2024 vs 2025'), use:\n"
                "    policy_end_date_year IN (2024, 2025)\n"
                "  or\n"
                "    policy_end_date_year BETWEEN 2024 AND 2025\n"
                "- If the user asks about months, use policy_end_date_month accordingly\n\n"

                "You are a SQL expert assistant working with a PostgreSQL table called \"stage\".\"GBM1_prediction_data_with_recommendations\".\n\n"
                "When the user asks about reasons for non-renewal, generate a SQL query using the columns:\n"
                "- not_renewed_reasons\n"
                "- main_reason\n\n"
                "If the user asks for recommendations, use:\n"
                "- primary_recommendation\n\n"
                "If the user asks for additional offers, use:\n"
                "- additional_offers\n\n"
                "Only use these columns if relevant to the user’s query.\n\n"
                "Always include filtering logic when applicable, for example:\n"
                "- WHERE \"predicted_status\" = 'Not Renewed'\n"
                "- WHERE column IS NOT NULL\n\n"
                "Respond only with a valid SQL query without explanations.\n\n"
                "Schema reference:\n"
                "- predicted_status (values: Renewed / Not Renewed)\n"
                "- not_renewed_reasons\n"
                "- main_reason\n"
                "- primary_recommendation\n"
                "- additional_offers\n\n"
                "User query: {{user_question}}\n\n"

                "Query Guidelines:\n"
                "- Use ONLY real column names (do NOT invent).\n"
                "- Use UPPER() around both column and string value for case-insensitive matching.\n"
                "- Add year/month filters from question (e.g., policy_end_date_year).\n"
                "- ❗ Always return SQL only — do NOT include explanation, markdown, or commentary.\n"
                "- ❗ Output must start with SELECT or WITH and contain only raw SQL.\n\n"

                "📍 Location Handling (Mandatory):\n\n"
                "If the user question contains any **location keyword** (city, state, or zone),\n"
                "you MUST add a WHERE clause that filters by location.\n\n"
                "Available location columns:\n"
                "1. cleaned_branch_name_2  → (e.g., 'delhi1', 'surat', 'mumbai') ✅ Preferred\n"
                "2. cleaned_state2         → (e.g., 'delhi', 'gujarat', 'punjab') ✅ Fallback\n"
                "3. cleaned_zone_2         → (e.g., 'north', 'west', 'south') ✅ Fallback\n\n"
                "🧠 Mapping strategy:\n"
                "- If a city is mentioned (like 'Delhi', 'Mumbai', 'Surat') → Match with `cleaned_branch_name_2`\n"
                "- If a state is mentioned (like 'Gujarat', 'Punjab') → Match with `cleaned_state2`\n"
                "- If a zone is mentioned (like 'North', 'West') → Match with `cleaned_zone_2`\n\n"
                "Examples:\n"
                "- User: \"in Mumbai\" → WHERE UPPER(cleaned_branch_name_2) = UPPER('mumbai')\n"
                "- User: \"in Gujarat\" → WHERE UPPER(cleaned_state2) = UPPER('gujarat')\n"
                "- User: \"in North zone\" → WHERE UPPER(cleaned_zone_2) = UPPER('north')\n\n"
                "⚠️ This filter is mandatory if any location keyword is present.\n"
            )

    


    
def generate_database_summary(self):
        """Generate comprehensive database summary"""
        if not self.is_connected:
            return "No database is currently connected."
        
        try:
            summary_parts = []
            
            # Basic database info
            with self.engine.connect() as conn:
                db_version = conn.execute(text("SELECT version()")).scalar()
                summary_parts.append(f"Database: PostgreSQL")
                summary_parts.append(f"Version: {db_version.split(',')[0]}")
            
            # Tables overview
            tables = self.schema_info.get('tables', {})
            summary_parts.append(f"Total Tables: {len(tables)}")
            
            # Detailed table information
            for table_name, table_info in tables.items():
                columns = table_info.get('columns', {})
                summary_parts.append(f"\nTable: {table_name}")
                summary_parts.append(f"  Columns: {len(columns)}")
                
                # Get row count
                try:
                    with self.engine.connect() as conn:
                        count_result = conn.execute(text(f'SELECT COUNT(*) FROM "{table_name}"'))
                        row_count = count_result.scalar()
                        summary_parts.append(f"  Rows: {row_count:,}")
                except Exception as e:
                    summary_parts.append(f"  Rows: Unable to count ({str(e)})")
                
                # Column details
                summary_parts.append("  Column Details:")
                for col_name, col_info in columns.items():
                    col_type = col_info.get('type', 'unknown')
                    nullable = "NULL" if col_info.get('nullable', True) else "NOT NULL"
                    summary_parts.append(f"    - {col_name}: {col_type} ({nullable})")
            
            return "\n".join(summary_parts)
            
        except Exception as e:
            logger.error(f"Error generating database summary: {e}")
            return f"Error generating database summary: {str(e)}"

def get_schema_info_structured(self):
        """Return schema information as a structured dictionary"""
        with self.engine.connect() as conn:
            count_result = conn.execute(text("SELECT COUNT(*) FROM information_schema.schemata;"))
            schema_count = count_result.scalar()

            names_result = conn.execute(text("SELECT schema_name FROM information_schema.schemata;"))
            schema_names = [row[0] for row in names_result.fetchall()]

        return {
            "schema_count": schema_count,
            "schema_names": schema_names
        }




def detect_best_column_for_value(self, value: str):
        """
        Try to infer which column contains the given value by probing common ID/number-like columns.
        This helps when user inputs a value like '1303624' without mentioning the column name.
        """
        probe_columns = [
            "cleaned_engine_number", "cleaned_chassis_number", "cleaned_reg_no",
            "customerid", "policy_no", "corrected_name"
        ]
        
        table = '"stage"."GBM1_prediction_data_with_recommendations"'
        
        for column in probe_columns:
            try:
                query = f'''
                SELECT COUNT(*) FROM {table}
                WHERE "{column}"::text = '{value}'
                '''
                with self.engine.connect() as conn:
                    count = conn.execute(text(query)).scalar()
                    if count and int(count) > 0:
                        return column
            except Exception as e:
                continue  # Silent fail on error (e.g., invalid column)
        
        return None



         

def generate_erd_and_usecases(self):
        """Generate ERD diagram summary (Mermaid) and use-case mappings per table"""
        if not self.is_connected:
            return "No database connected. Please connect first."

        try:
            tables = self.schema_info.get('tables', {})
            if not tables:
                return "No tables found to generate ERD or use-case mapping."

            erd_parts = []
            erd_parts.append("```mermaid")
            erd_parts.append("erDiagram")

            usecase_parts = []
            usecase_parts.append("### 📝 **Table-wise Use-case Mappings**")

            for table_name, table_info in tables.items():
                table_label = table_name.replace('"', '').replace('.', '_')

                columns = table_info.get('columns', {})
                erd_parts.append(f"  {table_label} {{")
                for col_name, col_info in columns.items():
                    col_type = col_info.get('type', 'unknown').split('(')[0]
                    erd_parts.append(f"    {col_type} {col_name}")
                erd_parts.append("  }")

                # Generate use-case mappings based on keywords
                col_text = " ".join(columns.keys()).lower()
                usecases = []
                if 'policy' in col_text:
                    usecases.append("Manage insurance policies and renewals")
                if 'customer' in col_text:
                    usecases.append("Store customer profiles and segmentation")
                if 'churn' in col_text:
                    usecases.append("Track churn prediction and status")
                if 'claim' in col_text:
                    usecases.append("Handle insurance claims processing")
                if 'vehicle' in col_text or 'make' in col_text:
                    usecases.append("Maintain vehicle information database")
                if not usecases:
                    usecases.append("General business data storage")

                usecase_parts.append(f"**Table `{table_name}` Use-cases:**")
                for uc in usecases:
                    usecase_parts.append(f"- {uc}")

            erd_parts.append("```")

            # Combine ERD and use-case mappings
            final_output = []
            final_output.append("## 🗂️ **Entity Relationship Diagram (ERD)**")
            final_output += erd_parts
            final_output += usecase_parts

            return "\n".join(final_output)

        except Exception as e:
            import traceback
            logger.error(f"ERD generation error: {e}")
            logger.error(traceback.format_exc())
            return f"Failed to generate ERD and use-case mappings: {str(e)}"

# @csrf_exempt
# def check_intent(request):
#     if request.method == 'POST':
#         data = json.loads(request.body)
#         question = data.get("question", "")

#         prompt = f"""
# Classify the user's intent strictly as YES or NO.

# If the question is a general greeting, chit-chat, or general knowledge (e.g. "hi", "hello", "how are you", "who is the PM of India"), respond with NO.

# If it is about querying the connected database schema or fetching data from tables, respond with YES.
# If the question is trying to analyze, query, or summarize tabular or structured data (e.g. Excel, PDF tables, database), respond with YES.

# Question: "{question}"
# Only respond YES or NO.
# """

#         try:
#             response = requests.post(
#                 "https://api.groq.com/openai/v1/chat/completions",  # ✅ Groq Cloud endpoint
#                 headers={
#                     "Authorization": f"Bearer {GROQ_API_KEY}",  # ✅ Use your Groq API key
#                     "Content-Type": "application/json"
#                 },
#                 json={
#                     "model": "meta-llama/llama-4-maverick-17b-128e-instruct",  # ✅ Groq model
#                     "messages": [
#                         {"role": "system", "content": "You are an intent classifier."},
#                         {"role": "user", "content": prompt}
#                     ],
#                     "temperature": 0.0,
#                     "max_tokens": 10
#                 },
#                 timeout=30
#             )
#             result = response.json()
#             answer = result["choices"][0]["message"]["content"].strip()
#             return JsonResponse({"answer": answer})

#         except Exception as e:
#             return JsonResponse({"answer": "No", "error": str(e)}, status=500)

 


def get_schema_info(engine):
    """Get comprehensive schema information with exact table naming"""
    logger.info("=== SCHEMA INFO RETRIEVAL START ===")
    
    try:
        inspector = inspect(engine)
        schema_info = {'tables': {}}

        for schema_name in inspector.get_schema_names():
            if schema_name not in ['information_schema', 'pg_catalog', 'pg_toast']:
                for table_name in inspector.get_table_names(schema=schema_name):
                    columns = inspector.get_columns(table_name, schema=schema_name)
                    
                    # Use double quotes around schema and table for exact referencing
                    qualified_table_name = f'"{schema_name}"."{table_name}"'

                    schema_info['tables'][qualified_table_name] = {
                        'columns': {
                            col['name']: {
                                'type': str(col['type']),
                                'nullable': col['nullable'],
                                'default': col.get('default')
                            }
                            for col in columns
                        }
                    }
        
        logger.info(f"Schema info retrieval completed. Total tables processed: {len(schema_info['tables'])}")
        return schema_info

    except Exception as e:
        logger.error(f"Critical error in schema inspection: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return {'tables': {}}

class PostgreSQLChatAnalyzer:
    """Streamlined PostgreSQL chat analyzer with OpenRouter Llama4 Maverick"""
    
    def __init__(self, engine=None, schema_info=None):
        self.engine = engine
        self.schema_info = schema_info
        self.is_connected = engine is not None

       
    def format_query_results(self, data, question):
        """Format query results into natural language using OpenRouter"""
        if not data:
            return "No results found for your query."
        
        question_lower = question.lower()

       

        
        if "schema" in question_lower:
            schema_names = [list(row.values())[0] for row in data]

            if "how many" in question_lower or "count" in question_lower:
                # If result is COUNT(*), extract its numeric value
                if len(schema_names) == 1 and isinstance(schema_names[0], (int, float)):
                    schema_count = schema_names[0]
                else:
                    schema_count = len(schema_names)

                response = f"The database contains {schema_count} schema(s)."
            else:
                schema_list = "\n".join([f"- {name}" for name in schema_names]).strip()
                response = f"The database contains {len(schema_names)} schema(s):\n{schema_list}"

            return response.strip()



    # Prepare data summary for general queries
        # Prepare data summary
        data_summary = f"Query returned {len(data)} rows.\n\n"
        
        # Show sample data (first 5 rows)
        sample_data = data[:5]
        data_summary += "Sample results:\n"
        for i, row in enumerate(sample_data, 1):
            data_summary += f"Row {i}: {row}\n"
        
        if len(data) > 5:
            data_summary += f"... and {len(data) - 5} more rows"
        
        system_prompt = f"""You are a database analyst. Format the query results into a clear, natural language response.

GUIDELINES:
1. Provide a clear, conversational answer
2. Summarize key findings
3. Use specific numbers and data points
4. If there are many rows, provide meaningful summaries
5. Make it easy to understand for business users
6.If the result contains schema names, mention the total number of schemas and list their names in bullet points.
7. Summarize key findings using specific numbers and data points.
8. If there are many rows, provide a concise summary mentioning only the most relevant information.
9. Make the explanation easy to understand for business users without technical jargon.

Original question: {question}
Data retrieved: {data_summary}

Provide a natural language response:"""
        
        return self.get_openrouter_response(data_summary, system_prompt)

           
    def execute_query_safely(self, query):
        """Execute query with safety checks"""
        if not self.is_connected:
            return None, "No database connection"
        
        # Safety checks
        query_lower = query.lower().strip()
        dangerous_keywords = ['drop', 'delete', 'truncate', 'alter', 'create', 'insert', 'update']
        
        # for keyword in dangerous_keywords:
        #     if keyword in query_lower:
        #         return None, f"Query contains potentially dangerous keyword: {keyword}"
        for keyword in dangerous_keywords:
        # Check using word boundaries to avoid partial matches (e.g. 'updated')
            if re.search(rf'\b{keyword}\b', query_lower):
                return None, f"Query contains potentially dangerous keyword: {keyword}"
            
        try:
            with self.engine.connect() as conn:
                result = conn.execute(text(query))
                
                if result.returns_rows:
                    rows = result.fetchall()
                    columns = list(result.keys())
                    
                    # Convert to list of dictionaries
                    data = []
                    for row in rows:
                        row_dict = {}
                        for i, col in enumerate(columns):
                            row_dict[col] = row[i]
                        data.append(row_dict)
                    
                    return data, None
                else:
                    return [], None
                    
        except Exception as e:
            error_msg = str(e)
            logger.error(f"Query execution error: {error_msg}")
            return None, error_msg


    
    def generate_case_insensitive_query(self, original_query, question):
        """Transform query to handle case-insensitive searches"""
        try:
            # Extract search index
            search_index = self.schema_info.get('search_index', {})
            
            words = re.findall(r'\b[a-zA-Z]+\b', question.lower())
            
            # Build replacement map for case-insensitive conditions
            replacements = {}
            
            for word in words:
                if len(word) > 2:  # Only consider words longer than 2 characters
                    for value_key, value_info in search_index.get('values', {}).items():
                        if word in value_key:
                            table = value_info['table']
                            column = value_info['column']
                            actual_value = value_info['actual_value']
                            
                            # Create case-insensitive pattern
                            pattern = f'"{column}" = \'{word}\''
                            replacement = f'UPPER("{column}") = UPPER(\'{actual_value}\')'
                            
                            # Also handle LIKE patterns
                            like_pattern = f'"{column}" LIKE \'%{word}%\''
                            like_replacement = f'UPPER("{column}") LIKE UPPER(\'%{actual_value}%\')'
                            
                            replacements[pattern] = replacement
                            replacements[like_pattern] = like_replacement
            
            # Apply replacements
            modified_query = original_query
            for pattern, replacement in replacements.items():
                modified_query = re.sub(pattern, replacement, modified_query, flags=re.IGNORECASE)
            
            modified_query = re.sub(
                r'WHERE\s+("?\w+"?)\s*=\s*\'([^\']+)\'',
                lambda m: f'WHERE UPPER({m.group(1)}) = UPPER(\'{m.group(2)}\')',
                modified_query,
                flags=re.IGNORECASE
            )
            
            modified_query = re.sub(
                r'("?\w+"?)\s+LIKE\s+\'([^\']+)\'',
                lambda m: f'UPPER({m.group(1)}) LIKE UPPER(\'{m.group(2)}\')',
                modified_query,
                flags=re.IGNORECASE
            )
            
            return modified_query
            
        except Exception as e:
            logger.error(f"Error in case-insensitive query generation: {e}")
            return original_query


    
    def detect_dynamic_location_filter(self, question: str):
        """
        Dynamically detect location (branch/city, state, or zone) from the question
        using fuzzy matching against actual values in the schema's search index.

        Returns:
            (location_column: str, matched_values: List[str]) or (None, [])
        """
        question = question.lower()
        words = re.findall(r'\b[a-z]+\b', question)  # all tokens

        # Extract known location values from schema or preloaded search_index
        search_values = self.schema_info.get('search_index', {}).get('values', {})

        city_values = set()
        state_values = set()
        zone_values = set()

        for _, info in search_values.items():
            col = info.get('column')
            val = str(info.get('actual_value', '')).lower().strip()
            if not val:
                continue

            if col == "cleaned_branch_name_2":
                city_values.add(val)
            elif col == "cleaned_state2":
                state_values.add(val)
            elif col == "cleaned_zone_2":
                zone_values.add(val)

        location_hits = []

        # Match each word from question with known values
        for word in words:
            city_match = difflib.get_close_matches(word, city_values, n=1, cutoff=0.85)
            if city_match:
                location_hits.append(("cleaned_branch_name_2", city_match[0]))
                continue

            state_match = difflib.get_close_matches(word, state_values, n=1, cutoff=0.85)
            if state_match:
                location_hits.append(("cleaned_state2", state_match[0]))
                continue

            zone_match = difflib.get_close_matches(word, zone_values, n=1, cutoff=0.85)
            if zone_match:
                location_hits.append(("cleaned_zone_2", zone_match[0]))
                continue

        if not location_hits:
            return None, []

        # Prefer more specific match: branch > state > zone
        for col in ["cleaned_branch_name_2", "cleaned_state2", "cleaned_zone_2"]:
            values = list({val for c, val in location_hits if c == col})
            if values:
                return col, values

        return None, []


    
    def clean_llm_sql_response(self, llm_response):
        """Extract pure SQL from LLM response by removing explanations or markdown"""
        import re

        if not llm_response:
            return ""

        # Remove code block markers
        llm_response = re.sub(r"```sql", "", llm_response, flags=re.IGNORECASE)
        llm_response = re.sub(r"```", "", llm_response, flags=re.IGNORECASE)

        # Split lines and find first valid SQL line
        sql_keywords = ['SELECT', 'WITH', 'INSERT', 'UPDATE', 'DELETE', 'CREATE']
        lines = llm_response.strip().splitlines()

        start_idx = 0
        for i, line in enumerate(lines):
            if any(line.strip().upper().startswith(k) for k in sql_keywords):
                start_idx = i
                break

        sql_lines = lines[start_idx:]
        cleaned_sql = "\n".join(sql_lines).strip()

        if not cleaned_sql.endswith(";"):
            cleaned_sql += ";"

        return cleaned_sql


       
    def get_openrouter_response(self, prompt, system_prompt=""):
            """Get response from Groq Cloud using Llama4 Maverick with rate-limit retry and safe truncation."""
            import re
            import time
            import requests
            from loguru import logger  # assuming logger is set up

            try:
                headers = {
                    "Authorization": f"Bearer {GROQ_API_KEY}",
                    "Content-Type": "application/json"
                }

                # Combine prompts and check total token size
                full_prompt = system_prompt + "\n\n" + prompt
                word_limit = 5000
                if len(full_prompt.split()) > word_limit:
                    logger.warning("Prompt too large, truncating...")
                    prompt = prompt[:8000]  # Truncate prompt portion
                    system_prompt = system_prompt[:2000]  # Truncate system message

                messages = []
                if system_prompt:
                    messages.append({"role": "system", "content": system_prompt})
                messages.append({"role": "user", "content": prompt})

                payload = {
                    "model": "meta-llama/llama-4-maverick-17b-128e-instruct",
                    "messages": messages,
                    "temperature": 0.1,
                    "max_tokens": 1000
                }

                GROQ_BASE_URL = "https://api.groq.com/openai/v1"

                max_retries = 3
                for attempt in range(max_retries):
                    response = requests.post(
                        f"{GROQ_BASE_URL}/chat/completions",
                        headers=headers,
                        json=payload,
                        timeout=30
                    )

                    if response.status_code == 200:
                        return response.json()["choices"][0]["message"]["content"].strip()

                    elif response.status_code == 429:
                        try:
                            error_data = response.json()
                            message = error_data.get("error", {}).get("message", "")
                            logger.warning(f"Rate limit hit: {message}")
                            match = re.search(r'try again in ([\\d\\.]+)s', message)
                            if match:
                                wait_time = float(match.group(1))
                                logger.info(f"Sleeping for {wait_time} seconds due to rate limit...")
                                time.sleep(wait_time)
                            else:
                                time.sleep(2 ** attempt)
                            continue
                        except Exception as parse_error:
                            logger.error(f"Error parsing rate limit retry time: {parse_error}")
                            time.sleep(2 ** attempt)
                            continue

                    else:
                        logger.error(f"Groq Cloud API error: {response.status_code} - {response.text}")
                        break

                return "I apologize, but I'm having trouble processing your request right now."

            except Exception as e:
                logger.error(f"Groq Cloud API call failed: {e}")
                return "I apologize, but I'm having trouble processing your request right now."


    
    def build_system_prompt_for_policy_schema(self):
            """Builds a smart system prompt for policy-related questions using known columns and fuzzy matching."""
            return (
                "You are a PostgreSQL expert helping translate user questions into precise SQL queries.\n"
                "Always use the table: \"stage\".\"GBM1_prediction_data_with_recommendations\"\n\n"

                "Relevant columns include:\n"
                "- predicted_status: 'Renewed' / 'Not Renewed'\n"
                "- churn_category, churn_probability, main_reason, not_renewed_reasons\n"
                "- make_clean, model_clean, variant, tie_up, vehicle_idv, segment_risk_score\n"
                "- total_premium_payable, total_od_premium, total_tp_premium, gst, clv, clv_category\n"
                "- corrected_name, policy_no, customerid, policy_status, retention_channel\n"
                "- customer_segment, discount_category, fuel_type_risk_factor, approved, number_of_claims\n"
                "- policy_start_date_year/month/day and policy_end_date_year/month/day\n"
                "- retention_streak, age, cleaned_branch_name_2, biztype, product_name\n\n"

                "Handle the following mappings:\n"
                "- If the user asks about segment keywords like 'elite', 'potential', 'low value', 'risky', match them against 'customer_segment'.\n"
                "    Eg.Use: UPPER(customer_segment) = UPPER('Elite Retainers') or appropriate full label.\n"
                "- If the user asks about communication method (SMS, Email, WhatsApp), use 'retention_channel'\n"
                "- If the user mentions values like 0.68, match with churn_probability or manufacturer_risk_rate depending on intent\n\n"

                "Fuzzy mapping examples:\n"
                "- 'How many people renewed?' → predicted_status = 'Renewed'\n"
                "- 'Churn reasons' → use main_reason or not_renewed_reasons\n"
                "- 'Top car brands' or 'most common vehicle companies' → group by make_clean\n"
                "- 'Total premium paid' → total_premium_payable\n"
                "- 'How many claims?' → number_of_claims\n"
                "- 'Retention channel usage' → group by retention_channel\n"
                "- 'Revenue' → total_premium_payable (or total_revenue if available)\n\n"

                "Date Handling:\n"
                "- If the user asks for a single year (e.g., 2024), use: policy_end_date_year = 2024\n"
                "- If the user asks for a range or comparison (e.g., 'between 2024 and 2025', or '2024 vs 2025'), use:\n"
                "    policy_end_date_year IN (2024, 2025)\n"
                "  or\n"
                "    policy_end_date_year BETWEEN 2024 AND 2025\n"
                "- If the user asks about months, use policy_end_date_month accordingly\n\n"

                "You are a SQL expert assistant working with a PostgreSQL table called \"stage\".\"GBM1_prediction_data_with_recommendations\".\n\n"
                "When the user asks about reasons for non-renewal, generate a SQL query using the columns:\n"
                "- not_renewed_reasons\n"
                "- main_reason\n\n"
                "If the user asks for recommendations, use:\n"
                "- primary_recommendation\n\n"
                "If the user asks for additional offers, use:\n"
                "- additional_offers\n\n"
                "Only use these columns if relevant to the user’s query.\n\n"
                "Always include filtering logic when applicable, for example:\n"
                "- WHERE \"predicted_status\" = 'Not Renewed'\n"
                "- WHERE column IS NOT NULL\n\n"
                "Respond only with a valid SQL query without explanations.\n\n"
                "Schema reference:\n"
                "- predicted_status (values: Renewed / Not Renewed)\n"
                "- not_renewed_reasons\n"
                "- main_reason\n"
                "- primary_recommendation\n"
                "- additional_offers\n\n"
                "User query: {{user_question}}\n\n"

                "Query Guidelines:\n"
                "- Use ONLY real column names (do NOT invent).\n"
                "- Use UPPER() around both column and string value for case-insensitive matching.\n"
                "- Add year/month filters from question (e.g., policy_end_date_year).\n"
                "- ❗ Always return SQL only — do NOT include explanation, markdown, or commentary.\n"
                "- ❗ Output must start with SELECT or WITH and contain only raw SQL.\n\n"

                "📍 Location Handling (Mandatory):\n\n"
                "If the user question contains any **location keyword** (city, state, or zone),\n"
                "you MUST add a WHERE clause that filters by location.\n\n"
                "Available location columns:\n"
                "1. cleaned_branch_name_2  → (e.g., 'delhi1', 'surat', 'mumbai') ✅ Preferred\n"
                "2. cleaned_state2         → (e.g., 'delhi', 'gujarat', 'punjab') ✅ Fallback\n"
                "3. cleaned_zone_2         → (e.g., 'north', 'west', 'south') ✅ Fallback\n\n"
                "🧠 Mapping strategy:\n"
                "- If a city is mentioned (like 'Delhi', 'Mumbai', 'Surat') → Match with `cleaned_branch_name_2`\n"
                "- If a state is mentioned (like 'Gujarat', 'Punjab') → Match with `cleaned_state2`\n"
                "- If a zone is mentioned (like 'North', 'West') → Match with `cleaned_zone_2`\n\n"
                "Examples:\n"
                "- User: \"in Mumbai\" → WHERE UPPER(cleaned_branch_name_2) = UPPER('mumbai')\n"
                "- User: \"in Gujarat\" → WHERE UPPER(cleaned_state2) = UPPER('gujarat')\n"
                "- User: \"in North zone\" → WHERE UPPER(cleaned_zone_2) = UPPER('north')\n\n"
                "⚠️ This filter is mandatory if any location keyword is present.\n"
            )


            
    


    def process_question(self, question, conversation_history):
            """Main question processing logic"""

            if isinstance(question, tuple):
                question = question[0] if question else ""

            if not isinstance(question, str):
                question = str(question)

            question_lower = question.lower().strip()

            year_match = re.search(r'(20\d{2})', question_lower)
            year = int(year_match.group(1)) if year_match else None

            
            month = None
            for i in range(1, 13):
                if calendar.month_name[i].lower() in question_lower or calendar.month_abbr[i].lower() in question_lower:
                    month = i
                    break

            # 🔍 Direct numeric value query: detect column and run
            # 🔍 Direct numeric value query: detect column and run
            numeric_match = re.search(r'\b\d{5,}\b', question)
            if numeric_match:
                numeric_val = numeric_match.group(0)
                matched_col = self.detect_best_column_for_value(numeric_val)
                if matched_col:
                    sql_query = f'''
                    SELECT * FROM "stage"."GBM1_prediction_data_with_recommendations"
                    WHERE "{matched_col}"::text = '{numeric_val}';
                    '''
                    data, error = self.execute_query_safely(sql_query)

                    if error:
                        return {
                            'answer': f"Query execution failed: {error}",
                            'success': False,
                            'query': sql_query,
                            'row_count': 0
                        }

                    formatted_answer = self.format_query_results(data, question)

                    return {
                        'answer': formatted_answer,
                        'success': True,
                        'query': sql_query,
                        'row_count': len(data)
                    }


            month = None
            for i in range(1, 13):
                month_name = calendar.month_name[i].lower()
                month_abbr = calendar.month_abbr[i].lower()
                if month_name in question_lower or month_abbr in question_lower:
                    month = i
                    break

            month = None
            for i in range(1, 13):
                if calendar.month_name[i].lower() in question_lower or calendar.month_abbr[i].lower() in question_lower:
                    month = i
                    break
            

            # FIX 1: Consolidate location detection logic
            known_locations = {}
            location_columns = ['cleaned_branch_name_2', 'cleaned_state2', 'cleaned_zone_2']
                
                # Debug: Check what's in schema_info
            logger.info(f"[Debug] schema_info keys: {list(self.schema_info.keys())}")
                
            search_index = self.schema_info.get('search_index', {})
            logger.info(f"[Debug] search_index keys: {list(search_index.keys())}")
                
                # Try multiple ways to get location data
            if 'values' in search_index:
                    logger.info(f"[Debug] Found {len(search_index['values'])} values in search_index")
                    for value_key, value_info in search_index['values'].items():
                        col = value_info.get('column', '')
                        actual_val = value_info.get('actual_value', '')
                        if col in location_columns and actual_val:
                            known_locations.setdefault(col, set()).add(actual_val)
                            logger.info(f"[Debug] Added location: {col} = {actual_val}")
                
                # Alternative: Try to get locations from tables directly
            if not known_locations and 'tables' in self.schema_info:
                    for table_name, table_info in self.schema_info['tables'].items():
                        columns = table_info.get('columns', {})
                        for col_name in location_columns:
                            if col_name in columns:
                                # If you have sample data or distinct values stored somewhere
                                distinct_vals = columns[col_name].get('distinct_values', [])
                                if distinct_vals:
                                    known_locations[col_name] = set(distinct_vals)
                                    logger.info(f"[Debug] Added from table schema: {col_name} = {distinct_vals}")
                
                # Convert sets to sorted lists
            known_locations = {col: sorted(list(vals)) for col, vals in known_locations.items()}
                
                # If still empty, try hardcoded fallback based on your data screenshot
            if not known_locations:
                    logger.warning("[Debug] No locations found in schema, using fallback")
                    known_locations = {
                        'cleaned_branch_name_2': [
        "agartala", "ahmedabad", "ahmednagar", "ambala", "amravati", "amritsar", "andheri", "angul", "aurangabad",
        "ballari", "balsore", "bangalore", "basirhat", "begusarai", "belagavi", "belgaum", "bellary", "bengaluru",
        "berhampore", "berhampur", "bhagalpur", "bhopal", "bhubaneshwar", "bijapur", "bilaspur", "burdwan", "calicut",
        "chandigarh", "chennai", "chhatrapatisambhajinagar", "coimbatore", "corporateoffice", "cuttack", "davanagere",
        "dehradun", "delhi1", "delhi2", "delhinauranghouse", "deoghar", "dhanbad", "durgapur", "gandhidham", "gaya",
        "gulbarga", "guntur", "gurgaon", "guwahati", "hubballi", "hubli", "hyderabad", "imphal", "indore", "jaipur",
        "jajpur", "jalandhar", "jammu", "jamnagar", "jamshedpur", "jeypore", "jorhat", "kadapa", "kalaburagi", "kangra",
        "kanpur", "karimnagar", "khammam", "kharagpur", "kochi", "kolhapur", "kolkata1", "kollam", "kurnool", "lucknow",
        "ludhiana", "madurai", "mahbubnagar", "maldah", "mandi", "mangalore", "mangaluru", "margao", "mumbai", "mumbai1",
        "muzaffarpur", "mysore", "mysuru", "nagpur", "nashik", "ncr", "nellore", "noida", "patna", "puducherry1", "pune",
        "punetpa", "purnea", "raipur", "rajahmundry", "rajkot", "ranchi", "rourkela", "salem", "sambalpur", "satara",
        "shillong", "shimoga", "shivamogga", "siliguri", "solan", "solapur", "srinagar", "surat", "thane", "thrissur",
        "tirunelveli", "tirupati", "trichy", "trivandrum", "tumakuru", "tumkur", "udaipur", "vadodara", "varanasi",
        "vellore", "vijayapura", "vijayawada", "vishakapatnam", "warangal"
    ],
                        'cleaned_state2': [
        "andhrapradesh", "assam","bihar","chandigarh",
        "chhattisgarh","delhi","goa","gujarat","haryana","himachalpradesh","jammukashmir","jharkhand","karnataka","kerala","madhyapradesh","maharashtra","manipur","meghalaya","mizoram","odisha","puducherry","punjab","rajasthan","tamilnadu","telangana","tripura","uttarakhand","uttarpradesh","westbengal"
    ],

                        'cleaned_zone_2': ['south', 'west', 'north',' east','corporate'],
                    }

                    

        

            def detect_location_filter(question, known_locations):
                """
                Detects the best matching location value from the question based on known location lists.
                Supports exact, space-insensitive, fuzzy, and word-boundary matches.
                """
                question_raw = question
                question_lower = question.lower()
                question_no_space = question_lower.replace(" ", "")

                for col, values in known_locations.items():
                    for val in values:
                        val_lower = val.lower()
                        val_no_space = val_lower.replace(" ", "")

                        # 🔹 Exact match with space preserved
                        if val_lower in question_lower:
                            logger.info(f"[Location Match] Exact match found: '{val}' in column '{col}'")
                            return col, val

                        # 🔹 Match with spaces removed
                        if val_no_space in question_no_space:
                            logger.info(f"[Location Match] Space-insensitive match found: '{val}' in column '{col}'")
                            return col, val

                        # 🔹 Word boundary match (surat, tamil nadu, etc.)
                        if re.search(r'\b' + re.escape(val_lower) + r'\b', question_lower):
                            logger.info(f"[Location Match] Word-boundary match: '{val}' in column '{col}'")
                            return col, val

                        # 🔹 Fuzzy match (safe threshold)
                        ratio = SequenceMatcher(None, val_no_space, question_no_space).ratio()
                        if ratio > 0.88:
                            logger.info(f"[Location Match] Fuzzy match: '{val}' (ratio: {ratio:.2f}) in column '{col}'")
                            return col, val

                logger.info(f"[Location Detection] No matches found for question: '{question_raw}'")
                return None, None

            
            location_col, location_val = detect_location_filter(question, known_locations)
                
                # Log location detection for debugging
            if location_col and location_val:
                    logger.info(f"[Location Detected] Column: {location_col}, Value: {location_val}")
            else:
                    logger.info(f"[Location Detection] No location found in question: '{question}'")
                    logger.info(f"[Available Locations] {known_locations}")


            
            # segment_keyword_map = {
            #     "elite": "Elite Retainers",
            #     "potential": "Potential Customers",
            #     "low value": "Low Value Customers",
            #     "risky": "Risk Segment"
            # }

            # question_lower = question.lower()
            # for keyword, segment_label in segment_keyword_map.items():
            #     if keyword in question_lower:
            #         # Inject proper SQL manually if intent is "count", "how many", etc.
            #         if any(word in question_lower for word in ["how many", "count", "total"]):
            #             sql_query = f'''
            #             SELECT COUNT(*) 
            #             FROM "stage"."GBM1_prediction_data_with_recommendations"
            #             WHERE UPPER(customer_segment) = UPPER('{segment_label}');
            #             '''
            #             return {
            #                 "answer": f"Counting records where customer_segment is '{segment_label}'",
            #                 "query": sql_query,
            #                 "success": True,
            #                 "row_count": None
            #             }
            #         break  # avoid multiple matches

            # 🔍 Direct numeric value query: detect column and run
            numeric_match = re.search(r'\b\d{5,}\b', question)
            if numeric_match:
                numeric_val = numeric_match.group(0)
                matched_col = self.detect_best_column_for_value(numeric_val)
                if matched_col:
                    sql_query = f'''
                    SELECT * FROM "stage"."GBM1_prediction_data_with_recommendations"
                    WHERE "{matched_col}"::text = '{numeric_val}';
                    '''
                    data, error = self.execute_query_safely(sql_query)

                    if error:
                        return {
                            'answer': f"Query execution failed: {error}",
                            'success': False,
                            'query': sql_query,
                            'row_count': 0
                        }

                    formatted_answer = self.format_query_results(data, question)

                    return {
                        'answer': formatted_answer,
                        'success': True,
                        'query': sql_query,
                        'row_count': len(data)
                    }


            if "average" in question_lower and "segment" in question_lower:
                    sql_query = '''
                    SELECT 
                        "customer_segment",
                        AVG("total_premium_payable") AS avg_premium,
                        AVG("churn_probability") AS avg_churn_prob,
                        COUNT(*) AS customer_count
                    FROM "stage"."GBM1_prediction_data_with_recommendations"
                    WHERE "customer_segment" IS NOT NULL AND TRIM("customer_segment") <> ''
                    GROUP BY "customer_segment"
                    ORDER BY customer_count DESC;
                    '''
                    
                    data, error = self.execute_query_safely(sql_query)
                    if error:
                        return {
                            'answer': f"Query execution failed: {error}",
                            'success': False,
                            'query': sql_query,
                            'row_count': 0
                        }

                    # Natural language summary
                    formatted_answer = self.format_query_results(data, question)

                    return {
                        'answer': formatted_answer,
                        'success': True,
                        'query': sql_query,
                        'row_count': len(data)
                    }


            # 🔍 Customer Segment Processing (FIXED)
            segment_keyword_map = {
                "elite retainers": "Elite Retainers",
                "elite": "Elite Retainers",
                "potential customers": "Potential Customers", 
                "potential": "Potential Customers",
                "low value customers": "Low Value Customers",
                "low value": "Low Value Customers",
                "risky": "Risk Segment",
                "risk segment": "Risk Segment"
            }

            # Check for customer segment keywords
            matched_segment = None
            matched_keyword = None
            
            for keyword, segment_label in segment_keyword_map.items():
                if keyword in question_lower:
                    matched_segment = segment_label
                    matched_keyword = keyword
                    break

            # Process customer segment queries
            if matched_segment:
                schema_table = '"stage"."GBM1_prediction_data_with_recommendations"'
                
                # Check if it's a count query
                if any(word in question_lower for word in ["how many", "count", "total", "number of"]):
                    sql_query = f'''
                    SELECT COUNT(*) AS segment_count
                    FROM {schema_table}
                    WHERE UPPER("customer_segment") = UPPER('{matched_segment}')
                    '''
                    
                    # Add date filters if present
                    if year:
                        sql_query += f' AND "policy_end_date_year" = {year}'
                    if month:
                        sql_query += f' AND "policy_end_date_month" = {month}'
                    
                    sql_query += ';'
                    
                    data, error = self.execute_query_safely(sql_query)
                    
                    if error:
                        return {
                            'answer': f"Query execution failed: {error}",
                            'success': False,
                            'query': sql_query,
                            'row_count': 0
                        }
                    
                    count = data[0]['segment_count'] if data else 0
                    
                    # Format timeframe
                    month_name = calendar.month_name[month] if month else ""
                    timeframe = f"{month_name} {year}" if month and year else str(year) if year else month_name
                    timeframe_text = f" for {timeframe}" if timeframe else ""
                    
                    answer = f" Total count of **{matched_segment}** customers{timeframe_text} is {count:,}."
                    
                    return {
                        'answer': answer,
                        'success': True,
                        'query': sql_query,
                        'row_count': count
                    }
                
                # Check if it's a summary query
                elif "summary" in question_lower:
                    sql_query = f'''
                    SELECT 
                        "customer_segment",
                        COUNT(*) AS segment_count,
                        AVG("total_premium_payable") AS avg_premium,
                        SUM("total_premium_payable") AS total_premium
                    FROM {schema_table}
                    WHERE UPPER("customer_segment") = UPPER('{matched_segment}')
                    '''
                    
                    # Add date filters if present
                    if year:
                        sql_query += f' AND "policy_end_date_year" = {year}'
                    if month:
                        sql_query += f' AND "policy_end_date_month" = {month}'
                    
                    sql_query += ' GROUP BY "customer_segment";'
                    
                    data, error = self.execute_query_safely(sql_query)
                    
                    if error:
                        return {
                            'answer': f"Query execution failed: {error}",
                            'success': False,
                            'query': sql_query,
                            'row_count': 0
                        }
                    
                    if data:
                        row = data[0]
                        month_name = calendar.month_name[month] if month else ""
                        timeframe = f"{month_name} {year}" if month and year else str(year) if year else month_name
                        timeframe_text = f" for {timeframe}" if timeframe else ""
                        
                        answer = f"""
        **{matched_segment} Summary{timeframe_text}**

        - Total customers: {row.get('segment_count', 0):,}
        - Average premium: ₹{row.get('avg_premium', 0):,.2f}
        - Total premium collected: ₹{row.get('total_premium', 0):,.2f}
                        """.strip()
                        
                        return {
                            'answer': answer,
                            'success': True,
                            'query': sql_query,
                            'row_count': row.get('segment_count', 0)
                        }
                    else:
                        return {
                            'answer': f"No data found for {matched_segment} customers.",
                            'success': True,
                            'query': sql_query,
                            'row_count': 0
                        }
                
                # Default: return sample records for the segment
                else:
                    sql_query = f'''
                    SELECT * FROM {schema_table}
                    WHERE UPPER("customer_segment") = UPPER('{matched_segment}')
                    '''
                    
                    # Add date filters if present
                    if year:
                        sql_query += f' AND "policy_end_date_year" = {year}'
                    if month:
                        sql_query += f' AND "policy_end_date_month" = {month}'
                    
                    sql_query += ' LIMIT 10;'
                    
                    data, error = self.execute_query_safely(sql_query)
                    
                    if error:
                        return {
                            'answer': f"Query execution failed: {error}",
                            'success': False,
                            'query': sql_query,
                            'row_count': 0
                        }
                    
                    formatted_answer = self.format_query_results(data, question)
                    
                    return {
                        'answer': formatted_answer,
                        'success': True,
                        'query': sql_query,
                        'row_count': len(data) if data else 0
                    }
                

                    # 🔍 Non-renewal reason handling
            if "reason" in question_lower and "not renew" in question_lower:
                sql_query = '''
                SELECT main_reason, COUNT(*) AS count
                FROM "stage"."GBM1_prediction_data_with_recommendations"
                WHERE predicted_status = 'Not Renewed' AND main_reason IS NOT NULL
                GROUP BY main_reason
                ORDER BY count DESC
                LIMIT 5;
                '''
                data, error = self.execute_query_safely(sql_query)
                if error:
                    return {
                        'answer': f"Query execution failed: {error}",
                        'success': False,
                        'query': sql_query,
                        'row_count': 0
                    }

                formatted = self.format_query_results(data, question)
                return {
                    'answer': formatted,
                    'success': True,
                    'query': sql_query,
                    'row_count': len(data)
                }

            # 🔍 Recommendation intent
            if "recommendation" in question_lower or "what to do" in question_lower or "next step" in question_lower:
                sql_query = '''
                SELECT primary_recommendation, COUNT(*) AS count
                FROM "stage"."GBM1_prediction_data_with_recommendations"
                WHERE primary_recommendation IS NOT NULL
                GROUP BY primary_recommendation
                ORDER BY count DESC
                LIMIT 5;
                '''
                data, error = self.execute_query_safely(sql_query)
                if error:
                    return {
                        'answer': f"Query execution failed: {error}",
                        'success': False,
                        'query': sql_query,
                        'row_count': 0
                    }

                formatted = self.format_query_results(data, question)
                return {
                    'answer': formatted,
                    'success': True,
                    'query': sql_query,
                    'row_count': len(data)
                }

            # 🔍 Additional offers
            if "offer" in question_lower or "additional benefit" in question_lower or "incentive" in question_lower:
                sql_query = '''
                SELECT additional_offers, COUNT(*) AS count
                FROM "stage"."GBM1_prediction_data_with_recommendations"
                WHERE additional_offers IS NOT NULL
                GROUP BY additional_offers
                ORDER BY count DESC
                LIMIT 5;
                '''
                data, error = self.execute_query_safely(sql_query)
                if error:
                    return {
                        'answer': f"Query execution failed: {error}",
                        'success': False,
                        'query': sql_query,
                        'row_count': 0
                    }

                formatted = self.format_query_results(data, question)
                return {
                    'answer': formatted,
                    'success': True,
                    'query': sql_query,
                    'row_count': len(data)
                }



            # Determine entity dynamically
            entity = None
            if "policy" in question_lower:
                entity = "policy"
            elif "churn" in question_lower:
                entity = "churn"
            elif "claim" in question_lower:
                entity = "claim"
            elif "vehicle" in question_lower:
                entity = "vehicle"

            # ✅ Location detection for group-by (zone/state/branch)
            location_keywords_map = {
                "branch": "cleaned_branch_name_2",
                "state": "cleaned_state2",
                "zone": "cleaned_zone_2"
            }

            detected_keyword = ""  # fallback default
            location_group_by_col = None
            for keyword, col in location_keywords_map.items():
                if keyword in question_lower:
                    location_group_by_col = col
                    detected_keyword = keyword  # 👈 save this for title
                    break

            location_title_map = {
                "branch": "Branches",
                "state": "States",
                "zone": "Zones"
            }
            location_title = location_title_map.get(detected_keyword, "Locations")




            # Dynamic direct SQL summarization for known entities
            if entity and (year or month):
                try:
                    schema_table = '"stage"."GBM1_prediction_data_with_recommendations"'

                    # Detect locations before entity branching
                    # location_col, location_val = self.detect_location_filter(question, known_locations)
                    # if location_col and location_val:
                    #     logger.info(f"[Location Detected] Column: {location_col}, Value: {location_val}")

                    if entity == "policy":
                        top_policy_groupby_columns = {
                            "product": "product_name",
                            "segment": "customer_segment",
                            "make": "make_clean",
                            "vehicle": "model_clean",
                            "location": "cleaned_branch_name_2"
                        }

                        group_by_col = None
                        for keyword, col in top_policy_groupby_columns.items():
                            if keyword in question_lower:
                                group_by_col = col
                                break

                        # ✅ Top policies by location
                        if location_group_by_col and "policy" in question_lower and any(word in question_lower for word in ["top", "which", "most", "more", "highest", "high", "low", "mid"]):
                            sql_query = f"""
                            SELECT "{location_group_by_col}" AS location, COUNT(*) AS policy_count
                            FROM {schema_table}
                            WHERE 1=1
                            """
                            if location_col and location_val:
                                sql_query += f' AND UPPER("{location_col}") = UPPER(\'{location_val}\')'
                            if year:
                                sql_query += f' AND "policy_end_date_year" = {year}'
                            if month:
                                sql_query += f' AND "policy_end_date_month" = {month}'
                            sql_query += f"""
                            GROUP BY "{location_group_by_col}"
                            ORDER BY policy_count DESC
                            LIMIT 5;
                            """

                            data, error = self.execute_query_safely(sql_query)
                            if error:
                                return {
                                    'answer': f"Query execution failed: {error}",
                                    'success': False,
                                    'query': sql_query,
                                    'row_count': 0
                                }

                            timeframe = f"{calendar.month_name[month]} {year}" if month and year else str(year) if year else "overall"
                            location_title_map = {
                                "branch": "Branches",
                                "state": "States",
                                "zone": "Zones"
                            }
                            location_title = location_title_map.get(detected_keyword, "Locations")

                            answer = f"**Top 5 policy counts by {location_title} for {timeframe}**\n"
                            for row in data:
                                loc = row['location'] or "Unknown"
                                answer += f"- {loc.title()}: {row['policy_count']:,}\n"

                            return {
                                'answer': answer,
                                'success': True,
                                'query': sql_query,
                                'row_count': len(data)
                            }

                        # ✅ Top policies by dimension
                        if ("top" in question_lower or "most" in question_lower or "highest" in question_lower) and group_by_col:
                            sql_query = f"""
                            SELECT "{group_by_col}", COUNT(*) AS policy_count
                            FROM {schema_table}
                            WHERE 1=1
                            """
                            if location_col and location_val:
                                sql_query += f' AND UPPER("{location_col}") = UPPER(\'{location_val}\')'
                            if year:
                                sql_query += f' AND "policy_end_date_year" = {year}'
                            if month:
                                sql_query += f' AND "policy_end_date_month" = {month}'
                            sql_query += f"""
                            GROUP BY "{group_by_col}"
                            ORDER BY policy_count DESC
                            LIMIT 5;
                            """

                            data, error = self.execute_query_safely(sql_query)
                            if error:
                                return {
                                    'answer': f"Query execution failed: {error}",
                                    'success': False,
                                    'query': sql_query,
                                    'row_count': 0
                                }

                            timeframe = f"{calendar.month_name[month]} {year}" if month and year else str(year) if year else "overall"
                            title_map = {
                                "product_name": "Policy Products",
                                "customer_segment": "Customer Segments",
                                "make_clean": "Vehicle Makes",
                                "model_clean": "Vehicle Models",
                                "cleaned_branch_name_2": "Policy Locations"
                            }
                            title = title_map.get(group_by_col, group_by_col.replace("_", " ").title())

                            answer = f"**Top 5 {title} for {timeframe}**\n"
                            for row in data:
                                label = row[group_by_col] or "Unknown"
                                answer += f"- {label}: {row['policy_count']:,}\n"
                            return {
                                'answer': answer,
                                'success': True,
                                'query': sql_query,
                                'row_count': len(data)
                            }

                        # ✅ Month-wise summary
                        if "summary" in question_lower and ("month wise" in question_lower or "monthly" in question_lower):
                            sql_query = f"""
                            SELECT "policy_end_date_month" AS month,
                                COUNT(*) AS total_policies,
                                SUM("total_premium_payable") AS total_premium,
                                AVG("total_premium_payable") AS avg_premium
                            FROM {schema_table}
                            WHERE 1=1
                            """
                            if location_col and location_val:
                                sql_query += f' AND UPPER("{location_col}") = UPPER(\'{location_val}\')'
                            if year:
                                sql_query += f' AND "policy_end_date_year" = {year}'
                            sql_query += """
                            GROUP BY month
                            ORDER BY month;
                            """

                            data, error = self.execute_query_safely(sql_query)
                            if error:
                                return {
                                    'answer': f"Query execution failed: {error}",
                                    'success': False,
                                    'query': sql_query,
                                    'row_count': 0
                                }

                            answer = f"**Month-wise Policy Summary for {year}**\n"
                            for row in data:
                                month_number = int(row['month']) if row['month'] else 0
                                month_name = calendar.month_name[month_number] if month_number else "Unknown"
                                answer += f"- {month_name}:\n"
                                answer += f"  • Total policies: {row.get('total_policies', 0):,}\n"
                                answer += f"  • Total premium collected: ₹{row.get('total_premium', 0):,.2f}\n"
                                answer += f"  • Average premium per policy: ₹{row.get('avg_premium', 0):,.2f}\n"

                            return {
                                'answer': answer,
                                'success': True,
                                'query': sql_query,
                                'row_count': sum(row.get('total_policies', 0) for row in data)
                            }

                        # ✅ Year summary
                        if "summary" in question_lower:
                            sql_query = f"""
                            SELECT COUNT(*) AS total_policies,
                                SUM("total_premium_payable") AS total_premium,
                                AVG("total_premium_payable") AS avg_premium
                            FROM {schema_table}
                            WHERE 1=1
                            """
                            if location_col and location_val:
                                sql_query += f' AND UPPER("{location_col}") = UPPER(\'{location_val}\')'
                            if year:
                                sql_query += f' AND "policy_end_date_year" = {year}'
                            if month:
                                sql_query += f' AND "policy_end_date_month" = {month}'
                            sql_query += ";"

                            data, error = self.execute_query_safely(sql_query)
                            if error:
                                return {
                                    'answer': f"Query execution failed: {error}",
                                    'success': False,
                                    'query': sql_query,
                                    'row_count': 0
                                }

                            row = data[0] if data else {}
                            timeframe = f"{calendar.month_name[month]} {year}" if month and year else str(year) if year else "overall"
                            answer = f"""
                    **Policy Summary for {timeframe}**

                    - Total policies: {row.get('total_policies', 0):,}
                    - Total premium collected: ₹{row.get('total_premium', 0):,.2f}
                    - Average premium per policy: ₹{row.get('avg_premium', 0):,.2f}
                            """.strip()

                            return {
                                'answer': answer,
                                'success': True,
                                'query': sql_query,
                                'row_count': row.get('total_policies', 0)
                            }

                        # ✅ Default: only total policies
                        sql_query = f"""
                        SELECT COUNT(*) AS total_policies
                        FROM {schema_table}
                        WHERE 1=1
                        """
                        if location_col and location_val:
                            sql_query += f' AND UPPER("{location_col}") = UPPER(\'{location_val}\')'
                        if year:
                            sql_query += f' AND "policy_end_date_year" = {year}'
                        if month:
                            sql_query += f' AND "policy_end_date_month" = {month}'
                        sql_query += ";"

                        data, error = self.execute_query_safely(sql_query)
                        if error:
                            return {
                                'answer': f"Query execution failed: {error}",
                                'success': False,
                                'query': sql_query,
                                'row_count': 0
                            }

                        policy_count = data[0]['total_policies'] if data else 0
                        month_name = calendar.month_name[month] if month else ""
                        timeframe = f"{month_name} {year}" if month and year else str(year) if year else month_name
                        answer = f"Total policies count for {timeframe or 'overall'} is {policy_count:,}."

                        return {
                            'answer': answer,
                            'success': True,
                            'query': sql_query,
                            'row_count': policy_count
                        }

                    # Build base query per entity
                    # if entity == "policy":
                    #     # 🔍 Dynamically detect group-by column for top policy summary
                    #     top_policy_groupby_columns = {
                    #         "product": "product_name",
                    #         "segment": "customer_segment",
                    #         "make": "make_clean",
                    #         "vehicle": "model_clean",
                    #         "location": "cleaned_branch_name_2"
                    #     }

                    #     group_by_col = None
                    #     for keyword, col in top_policy_groupby_columns.items():
                    #         if keyword in question_lower:
                    #             group_by_col = col
                    #             break

                    #     if location_group_by_col and "policy" in question_lower and any(word in question_lower for word in ["top", "which", "most","more", "highest","high","low","mid"]):
                    #         sql_query = f'''
                    #         SELECT "{location_group_by_col}" AS location, COUNT(*) AS policy_count
                    #         FROM {schema_table}
                    #         WHERE 1=1
                    #         '''
                    #         if location_col and location_val:
                    #             sql_query += f' AND UPPER("{location_col}") = UPPER(\'{location_val}\')'
                    #         if year:
                    #             sql_query += f' AND "policy_end_date_year" = {year}'
                    #         if month:
                    #             sql_query += f' AND "policy_end_date_month" = {month}'
                    #         sql_query += f'''
                    #         GROUP BY "{location_group_by_col}"
                    #         ORDER BY policy_count DESC
                    #         LIMIT 5;
                    #         '''

                    #         data, error = self.execute_query_safely(sql_query)
                    #         if error:
                    #             return {
                    #                 'answer': f"Query execution failed: {error}",
                    #                 'success': False,
                    #                 'query': sql_query,
                    #                 'row_count': 0
                    #             }

                    #         timeframe = f"{calendar.month_name[month]} {year}" if year and month else str(year) if year else "overall"
                    #         location_title_map = {
                    #             "branch": "Branches",
                    #             "state": "States",
                    #             "zone": "Zones"
                    #         }

                    #         location_title = location_title_map.get(detected_keyword, "Locations")

                    #         answer = f"**Top 5 policy counts by {location_title} for {timeframe}**\n"
                    #         for row in data:
                    #             loc = row['location'] or "Unknown"
                    #             answer += f"- {loc.title()}: {row['policy_count']:,}\n"

                    #         return {
                    #             'answer': answer,
                    #             'success': True,
                    #             'query': sql_query,
                    #             'row_count': len(data)
                    #         }


                    #     if ("top" in question_lower or "most" in question_lower or "highest" in question_lower) and group_by_col:
                    #         sql_query = f"""
                    #         SELECT "{group_by_col}", COUNT(*) AS policy_count
                    #         FROM {schema_table}
                    #         WHERE 1=1
                    #         """
                    #         if location_col and location_val:
                    #             sql_query += f' AND UPPER("{location_col}") = UPPER(\'{location_val}\')'
                    #         if year:
                    #             sql_query += f' AND "policy_end_date_year" = {year}'
                    #         if month:
                    #             sql_query += f' AND "policy_end_date_month" = {month}'
                    #         sql_query += f"""
                    #         GROUP BY "{group_by_col}"
                    #         ORDER BY policy_count DESC
                    #         LIMIT 5;
                    #         """

                    #         print(sql_query, "Final SQL Query")

                    #         data, error = self.execute_query_safely(sql_query)
                    #         if error:
                    #             return {
                    #                 'answer': f"Query execution failed: {error}",
                    #                 'success': False,
                    #                 'query': sql_query,
                    #                 'row_count': 0
                    #             }

                    #         timeframe = f"{calendar.month_name[month]} {year}" if month and year else str(year) if year else "overall"
                    #         title_map = {
                    #             "product_name": "Policy Products",
                    #             "customer_segment": "Customer Segments",
                    #             "make_clean": "Vehicle Makes",
                    #             "model_clean": "Vehicle Models",
                    #             "cleaned_branch_name_2": "Policy Locations"
                    #         }
                    #         title = title_map.get(group_by_col, group_by_col.replace("_", " ").title())

                    #         answer = f" **Top 5 {title} for {timeframe}**\n"
                    #         for row in data:
                    #             label = row[group_by_col] or "Unknown"
                    #             answer += f"- {label}: {row['policy_count']:,}\n"
                    #         row_count = len(data)

                    #     elif "summary" in question_lower and ("month wise" in question_lower or "monthly" in question_lower):
                    #         # ✅ Month-wise policy summary
                    #         sql_query = f"""
                    #         SELECT "policy_end_date_month" AS month,
                    #             COUNT(*) AS total_policies,
                    #             SUM("total_premium_payable") AS total_premium,
                    #             AVG("total_premium_payable") AS avg_premium
                    #         FROM {schema_table}
                    #         WHERE 1=1
                    #         """
                    #         if location_col and location_val:
                    #             sql_query += f' AND UPPER("{location_col}") = UPPER(\'{location_val}\')'
                    #         if year:
                    #             sql_query += f" AND \"policy_end_date_year\" = {year}"
                    #         sql_query += """
                    #         GROUP BY month
                    #         ORDER BY month;
                    #         """
                    #         print(sql_query, "Final SQL Query")
                            
                    #         data, error = self.execute_query_safely(sql_query)
                    #         if error:
                    #             return {
                    #                 'answer': f"Query execution failed: {error}",
                    #                 'success': False,
                    #                 'query': sql_query,
                    #                 'row_count': 0
                    #             }

                    #         answer = f" **Month-wise Policy Summary for {year}**\n"
                    #         for row in data:
                    #             month_number = int(row['month']) if row['month'] else 0
                    #             month_name = calendar.month_name[month_number] if month_number else "Unknown"
                    #             answer += f"- {month_name}:\n"
                    #             answer += f"  • Total policies: {row.get('total_policies', 0):,}\n"
                    #             answer += f"  • Total premium collected: ₹{row.get('total_premium', 0):,.2f}\n"
                    #             answer += f"  • Average premium per policy: ₹{row.get('avg_premium', 0):,.2f}\n"
                    #         row_count = sum(row.get('total_policies', 0) for row in data)

                    #     elif "summary" in question_lower:
                    #         # ✅ Yearly policy summary
                    #         sql_query = f"""
                    #         SELECT COUNT(*) AS total_policies,
                    #             SUM("total_premium_payable") AS total_premium,
                    #             AVG("total_premium_payable") AS avg_premium
                    #         FROM {schema_table}
                    #         WHERE 1=1
                    #         """
                    #         if location_col and location_val:
                    #             sql_query += f' AND UPPER("{location_col}") = UPPER(\'{location_val}\')'
                            
                    #         if year:
                    #             sql_query += f" AND \"policy_end_date_year\" = {year}"
                    #         if month:
                    #             sql_query += f" AND \"policy_end_date_month\" = {month}"
                    #         sql_query += ";"

                    #         print(sql_query, "Final SQL Query")

                    #         data, error = self.execute_query_safely(sql_query)
                    #         if error:
                    #             return {
                    #                 'answer': f"Query execution failed: {error}",
                    #                 'success': False,
                    #                 'query': sql_query,
                    #                 'row_count': 0
                    #             }

                    #         row = data[0] if data else {}
                    #         timeframe = f"{calendar.month_name[month]} {year}" if month and year else str(year) if year else "overall"
                    #         answer = f"""
                    #  **Policy Summary for {timeframe}**

                    # - Total policies: {row.get('total_policies', 0):,}
                    # - Total premium collected: ₹{row.get('total_premium', 0):,.2f}
                    # - Average premium per policy: ₹{row.get('avg_premium', 0):,.2f}
                    #         """.strip()
                    #         row_count = row.get('total_policies', 0)

                    #     else:
                    #         # ✅ Default: total policy count only
                    #         sql_query = f"""
                    #         SELECT COUNT(*) AS total_policies
                    #         FROM {schema_table}
                    #         WHERE 1=1
                    #         """
                    #         if location_col and location_val:
                    #             sql_query += f' AND UPPER("{location_col}") = UPPER(\'{location_val}\')'
                    #         if year:
                    #             sql_query += f" AND \"policy_end_date_year\" = {year}"
                    #         if month:
                    #             sql_query += f" AND \"policy_end_date_month\" = {month}"
                    #         sql_query += ";"

                    #         print(sql_query, "Final SQL Query")

                    #         data, error = self.execute_query_safely(sql_query)
                    #         if error:
                    #             return {
                    #                 'answer': f"Query execution failed: {error}",
                    #                 'success': False,
                    #                 'query': sql_query,
                    #                 'row_count': 0
                    #             }

                    #         policy_count = data[0]['total_policies'] if data else 0
                    #         month_name = calendar.month_name[month] if month else ""
                    #         timeframe = f"{month_name} {year}" if month and year else str(year) if year else month_name
                    #         answer = f" Total policies count for {timeframe or 'overall'} is {policy_count:,}."
                    #         row_count = policy_count

                    elif entity == "churn":

                        # 🔍 Dynamically detect group-by column for top churn analysis
                        top_groupby_columns = {
                            "location": "cleaned_branch_name_2",
                            "segment": "customer_segment",
                            "product": "product_name",
                            "make": "make_clean",
                            "vehicle": "model_clean"
                        }

                        group_by_col = None
                        for keyword, col in top_groupby_columns.items():
                            if keyword in question_lower:
                                group_by_col = col
                                break

                        # 🔍 Detect specific churn category FIRST
                        specific_category = None
                        category_word = ""
                        
                        if "highly" in question_lower or "high" in question_lower:
                            specific_category = "High"
                            category_word = "highly" if "highly" in question_lower else "high"
                        elif "mid" in question_lower or "medium" in question_lower:
                            specific_category = "Mid"
                            category_word = "mid" if "mid" in question_lower else "medium"
                        elif "lowest" in question_lower or "lower" in question_lower or "low" in question_lower:
                            specific_category = "Low"
                            category_word = "lowest" if "lowest" in question_lower else "lower" if "lower" in question_lower else "low"

                        # 🎯 PRIORITY 1: Location + Specific Category (e.g., "which branch get the high churn?")
                        if location_group_by_col and specific_category and any(word in question_lower for word in ["which", "what", "top", "most", "highest"]):
                            sql_query = f'''
                            SELECT "{location_group_by_col}" AS location, COUNT(*) AS churn_count
                            FROM {schema_table}
                            WHERE UPPER("churn_category") = UPPER('{specific_category}')
                            '''
                            if year:
                                sql_query += f' AND "policy_end_date_year" = {year}'
                            if month:
                                sql_query += f' AND "policy_end_date_month" = {month}'
                            sql_query += f'''
                            GROUP BY "{location_group_by_col}"
                            ORDER BY churn_count DESC
                            LIMIT 10;
                            '''

                            data, error = self.execute_query_safely(sql_query)
                            if error:
                                return {
                                    'answer': f"Query execution failed: {error}",
                                    'success': False,
                                    'query': sql_query,
                                    'row_count': 0
                                }

                            timeframe = f"{calendar.month_name[month]} {year}" if year and month else str(year) if year else "overall"
                            location_title = location_title_map.get(detected_keyword, "Locations")
                            
                            if data:
                                answer = f"**Top {location_title} with {category_word} churn for {timeframe}**\n"
                                for row in data:
                                    loc = row['location'] or "Unknown"
                                    answer += f"- {loc.title()}: {row['churn_count']:,}\n"
                            else:
                                answer = f"No {category_word} churn data found for {timeframe}."
                            
                            return {
                                'answer': answer,
                                'success': True,
                                'query': sql_query,
                                'row_count': len(data)
                            }
                        

                        # 🎯 PRIORITY 0: Specific branch + high churn questions
                        elif any(k in question_lower for k in ["branch", "state", "zone"]) and specific_category:
                                print("🔍 EXECUTING PRIORITY - Location-based churn aggregation")

                                location_col = "cleaned_branch_name_2" if "branch" in question_lower else \
                                            "cleaned_state2" if "state" in question_lower else \
                                            "cleaned_zone_2"  # Replace these with actual column names if available


                                sql_query = f'''
                                SELECT "{location_col}" AS location, COUNT(*) AS churn_count
                                FROM {schema_table}
                                WHERE UPPER("churn_category") = UPPER('{specific_category}')
                                '''

                                if year:
                                    sql_query += f' AND "policy_end_date_year" = {year}'
                                if month:
                                    sql_query += f' AND "policy_end_date_month" = {month}'

                                sql_query += f'''
                                GROUP BY "{location_col}"
                                ORDER BY churn_count DESC
                                LIMIT 1;
                                '''

                                data, error = self.execute_query_safely(sql_query)
                                if error:
                                    return {
                                        'answer': f"Query execution failed: {error}",
                                        'success': False,
                                        'query': sql_query,
                                        'row_count': 0
                                    }

                                if data:
                                    top_loc = data[0]
                                    loc = top_loc['location'] or "Unknown"
                                    count = top_loc['churn_count']
                                    answer = f"The {location_col.replace('_', ' ')} with the highest {category_word} churn is **{loc.title()}** with {count:,} customers."
                                else:
                                    answer = f"No {category_word} churn data found at {location_col} level."

                                return {
                                    'answer': answer,
                                    'success': True,
                                    'query': sql_query,
                                    'row_count': len(data)
                                }
                        # 🎯 PRIORITY 2: Location-based churn (general churn by location)
                        elif location_group_by_col and "churn" in question_lower and any(word in question_lower for word in ["top", "which", "most","more", "highest"]):
                            sql_query = f'''
                            SELECT "{location_group_by_col}" AS location, COUNT(*) AS churn_count
                            FROM {schema_table}
                            WHERE "churn_category" IS NOT NULL AND TRIM("churn_category") <> ''
                            '''
                            if year:
                                sql_query += f' AND "policy_end_date_year" = {year}'
                            if month:
                                sql_query += f' AND "policy_end_date_month" = {month}'
                            sql_query += f'''
                            GROUP BY "{location_group_by_col}"
                            ORDER BY churn_count DESC
                            LIMIT 5;
                            '''

                            data, error = self.execute_query_safely(sql_query)
                            if error:
                                return {
                                    'answer': f"Query execution failed: {error}",
                                    'success': False,
                                    'query': sql_query,
                                    'row_count': 0
                                }

                            timeframe = f"{calendar.month_name[month]} {year}" if year and month else str(year) if year else "overall"
                            location_title = location_title_map.get(detected_keyword, "Locations")

                            answer = f"**Top 5 churned {location_title} for {timeframe}**\n"
                            for row in data:
                                loc = row['location'] or "Unknown"
                                answer += f"- {loc.title()}: {row['churn_count']:,}\n"

                            return {
                                'answer': answer,
                                'success': True,
                                'query': sql_query,
                                'row_count': len(data)
                            }
                        
                    
                        # 🎯 PRIORITY 5.5: General churn with specific location (e.g., "churn in Tamil Nadu in Jan 2025")
                                            # 🎯 PRIORITY 5.5: General churn with specific location
                        
    # 🎯 PRIORITY 5.5: General churn for detected location
                        elif "churn" in question_lower and location_col and location_val:
                            sql_query = f"""
                            SELECT COUNT(*) AS churn_count
                            FROM {schema_table}
                            WHERE "churn_category" IS NOT NULL AND TRIM("churn_category") <> ''
                            AND UPPER("{location_col}") = UPPER('{location_val}')
                            """
                            if year:
                                sql_query += f' AND "policy_end_date_year" = {year}'
                            if month:
                                sql_query += f' AND "policy_end_date_month" = {month}'
                            sql_query += ";"
                            data, error = self.execute_query_safely(sql_query)
                            if error:
                                return {"answer": f"Query execution failed: {error}", "success": False, "query": sql_query, "row_count": 0}
                            churn_count = data[0]['churn_count'] if data else 0
                            month_name = calendar.month_name[month] if month else ""
                            timeframe = f"{month_name} {year}" if month and year else str(year) if year else month_name
                            answer = f"Total churn count for **{location_val.title()}** in {timeframe or 'overall'} is {churn_count:,}."
                            return {"answer": answer, "success": True, "query": sql_query, "row_count": churn_count}


                        # 🎯 PRIORITY 3: Top churn by other categories (segment, product, etc.)
                        elif "top" in question_lower and group_by_col:
                            sql_query = f"""
                            SELECT "{group_by_col}", COUNT(*) AS churn_count
                            FROM {schema_table}
                            WHERE "churn_category" IS NOT NULL AND TRIM("churn_category") <> ''
                            """
                            if year:
                                sql_query += f' AND "policy_end_date_year" = {year}'
                            if month:
                                sql_query += f' AND "policy_end_date_month" = {month}'
                            sql_query += f"""
                            GROUP BY "{group_by_col}"
                            ORDER BY churn_count DESC
                            LIMIT 5;
                            """

                            print(sql_query, "Final SQL Query")

                            data, error = self.execute_query_safely(sql_query)
                            if error:
                                return {
                                    'answer': f"Query execution failed: {error}",
                                    'success': False,
                                    'query': sql_query,
                                    'row_count': 0
                                }

                            timeframe = f"{calendar.month_name[month]} {year}" if year and month else str(year) if year else "overall"
                            location_title = location_title_map.get(detected_keyword, "Locations")

                            answer = f"**Top 5 {location_title} for {timeframe}**\n"
                            for row in data:
                                label = row[group_by_col] or "Unknown"
                                answer += f"- {label}: {row['churn_count']:,}\n"
                            row_count = len(data)

                        # 🎯 PRIORITY 4: Monthly churn summaries
                        elif "month wise" in question_lower or "monthly" in question_lower or ("summary" in question_lower and "per month" in question_lower):
                            # ✅ Month-wise churn category summary
                            sql_query = f"""
                            SELECT "policy_end_date_month" AS month,
                                "churn_category",
                                COUNT(*) AS count
                            FROM {schema_table}
                            WHERE 1=1
                            """
                            if year:
                                sql_query += f" AND \"policy_end_date_year\" = {year}"
                            sql_query += """
                            GROUP BY month, "churn_category"
                            ORDER BY month;
                            """
                            print(sql_query, "Final SQL Query")

                            data, error = self.execute_query_safely(sql_query)
                            if error:
                                return {
                                    'answer': f"Query execution failed: {error}",
                                    'success': False,
                                    'query': sql_query,
                                    'row_count': 0
                                }

                            answer = f"**Month-wise Churn Summary for {year or 'overall'}**\n"
                            for row in data:
                                month_number = int(row['month']) if row['month'] else 0
                                month_name = calendar.month_name[month_number] if month_number else "Unknown"
                                answer += f"- {month_name}: {row['churn_category']} → {row['count']:,}\n"
                            row_count = sum(row['count'] for row in data)

                        # 🎯 PRIORITY 5: Specific churn category total count
                        elif specific_category:
                            sql_query = f"""
                            SELECT COUNT(*) AS count
                            FROM {schema_table}
                            WHERE UPPER("churn_category") = UPPER('{specific_category}')
                            """
                            if year:
                                sql_query += f" AND \"policy_end_date_year\" = {year}"
                            if month:
                                sql_query += f" AND \"policy_end_date_month\" = {month}"
                            sql_query += ";"

                            print(sql_query, "Final SQL Query")

                            data, error = self.execute_query_safely(sql_query)
                            if error:
                                return {
                                    'answer': f"Query execution failed: {error}",
                                    'success': False,
                                    'query': sql_query,
                                    'row_count': 0
                                }

                            count = data[0]['count'] if data else 0
                            month_name = calendar.month_name[month] if month else ""
                            timeframe = f"{month_name} {year}" if month and year else str(year) if year else month_name
                            answer = f"Total **{category_word}** churn count for {timeframe or 'overall'} is {count:,}."
                            row_count = count

                        # 🎯 PRIORITY 6: Default total churn count
                        else:
                            sql_query = f"""
                            SELECT COUNT(*) AS total_churn
                            FROM {schema_table}
                            WHERE "churn_category" IS NOT NULL AND TRIM("churn_category") <> ''
                            """
                            if year:
                                sql_query += f" AND \"policy_end_date_year\" = {year}"
                            if month:
                                sql_query += f" AND \"policy_end_date_month\" = {month}"
                            sql_query += ";"

                            print(sql_query, "Final SQL Query")

                            data, error = self.execute_query_safely(sql_query)
                            if error:
                                return {
                                    'answer': f"Query execution failed: {error}",
                                    'success': False,
                                    'query': sql_query,
                                    'row_count': 0
                                }

                            churn_count = data[0]['total_churn'] if data else 0
                            month_name = calendar.month_name[month] if month else ""
                            timeframe = f"{month_name} {year}" if month and year else str(year) if year else month_name
                            answer = f"The total number of customers who have churned is {churn_count:,}."
                            row_count = churn_count

                    # elif entity == "claim":
                    #     # 🔍 Dynamically detect group-by column for top claim summary
                    #     top_claim_groupby_columns = {
                    #         "product": "product_name",
                    #         "make": "make_clean",
                    #         "vehicle": "model_clean",
                    #         "segment": "customer_segment",
                    #         "location": "cleaned_branch_name_2"
                    #     }

                    #     group_by_col = None
                    #     for keyword, col in top_claim_groupby_columns.items():
                    #         if keyword in question_lower:
                    #             group_by_col = col
                    #             break

                    #     def build_claim_filters(year=None, month=None, location_col=None, location_val=None):
                    #         filters = ['"approved" IS NOT NULL']
                    #         if year:
                    #             filters.append(f'"policy_end_date_year" = {year}')
                    #         if month:
                    #             filters.append(f'"policy_end_date_month" = {month}')
                    #         if location_col and location_val:
                    #             filters.append(f'UPPER("{location_col}") = UPPER(\'{location_val}\')')
                    #         return " AND ".join(filters)


                        # if location_group_by_col and "claim" in question_lower and any(word in question_lower for word in ["top", "which", "most","more", "highest","high","low","mid"]):
                        #     def build_vehicle_filters(year=None, month=None):
                        #         filters = ['"make_clean" IS NOT NULL', 'TRIM("make_clean") <> \'\'']
                        #         if year:
                        #             filters.append(f'"policy_end_date_year" = {year}')
                        #         if month:
                        #             filters.append(f'"policy_end_date_month" = {month}')
                        #         return " AND ".join(filters)

                        #     sql_query = f'''
                        #     SELECT "{location_group_by_col}" AS location, COUNT(*) AS vehicle_count
                        #     FROM {schema_table}
                        #     WHERE {build_vehicle_filters(year, month)}
                        #     GROUP BY "{location_group_by_col}"
                        #     ORDER BY vehicle_count DESC
                        #     LIMIT 5;
                        #     '''

                        #     data, error = self.execute_query_safely(sql_query)
                        #     if error:
                        #         return {
                        #             'answer': f"Query execution failed: {error}",
                        #             'success': False,
                        #             'query': sql_query,
                        #             'row_count': 0
                        #         }

                        #     timeframe = f"{calendar.month_name[month]} {year}" if year and month else str(year) if year else "overall"
                        #     location_title_map = {
                        #         "branch": "Branches",
                        #         "state": "States",
                        #         "zone": "Zones"
                        #     }

                        #     location_title = location_title_map.get(detected_keyword, "Locations")

                        #     answer = f"**Top 5 claimed {location_title} for {timeframe}**\n"
                        #     for row in data:
                        #         loc = row['location'] or "Unknown"
                        #         answer += f"- {loc.title()}: {row['claim_count']:,}\n"

                        #     return {
                        #         'answer': answer,
                        #         'success': True,
                        #         'query': sql_query,
                        #         'row_count': len(data)
                        #     }

                    #     if ("top" in question_lower or "most" in question_lower or "highest" in question_lower) and group_by_col:
                    #         sql_query = f"""
                    #         SELECT "{group_by_col}", COUNT(*) AS claim_count
                    #         FROM {schema_table}
                    #         WHERE "approved" IS NOT NULL
                    #         """
                    #         if year:
                    #             sql_query += f' AND "policy_end_date_year" = {year}'
                    #         if month:
                    #             sql_query += f' AND "policy_end_date_month" = {month}'
                    #         sql_query += f"""
                    #         GROUP BY "{group_by_col}"
                    #         ORDER BY claim_count DESC
                    #         LIMIT 5;
                    #         """

                    #         print(sql_query, "Final SQL Query")

                    #         data, error = self.execute_query_safely(sql_query)
                    #         if error:
                    #             return {
                    #                 'answer': f"Query execution failed: {error}",
                    #                 'success': False,
                    #                 'query': sql_query,
                    #                 'row_count': 0
                    #             }

                    #         timeframe = f"{calendar.month_name[month]} {year}" if month and year else str(year) if year else "overall"
                    #         title_map = {
                    #             "product_name": "Claimed Products",
                    #             "make_clean": "Claimed Vehicle Makes",
                    #             "model_clean": "Claimed Vehicle Models",
                    #             "customer_segment": "Claimed Segments",
                    #             "cleaned_branch_name_2": "Claim Locations"
                    #         }
                    #         title = title_map.get(group_by_col, group_by_col.replace("_", " ").title())

                    #         answer = f" **Top 5 {title} for {timeframe}**\n"
                    #         for row in data:
                    #             label = row[group_by_col] or "Unknown"
                    #             answer += f"- {label}: {row['claim_count']:,}\n"
                    #         row_count = len(data)

                    #     elif "summary" in question_lower and ("month wise" in question_lower or "monthly" in question_lower):
                    #         # ✅ Month-wise claims summary
                    #         sql_query = f"""
                    #         SELECT "policy_end_date_month" AS month,
                    #             COUNT(*) AS total_claims,
                    #             SUM("approved") AS total_approved
                    #         FROM {schema_table}
                    #         WHERE 1=1
                    #         """
                    #         if year:
                    #             sql_query += f" AND \"policy_end_date_year\" = {year}"
                    #         sql_query += """
                    #         GROUP BY month
                    #         ORDER BY month;
                    #         """
                    #         print(sql_query, "Final SQL Query")

                    #         data, error = self.execute_query_safely(sql_query)
                    #         if error:
                    #             return {
                    #                 'answer': f"Query execution failed: {error}",
                    #                 'success': False,
                    #                 'query': sql_query,
                    #                 'row_count': 0
                    #             }

                    #         answer = f" **Month-wise Claims Summary for {year}**\n"
                    #         for row in data:
                    #             month_number = int(row['month']) if row['month'] else 0
                    #             month_name = calendar.month_name[month_number] if month_number else "Unknown"
                    #             answer += f"- {month_name}:\n"
                    #             answer += f"  • Total claims: {row.get('total_claims', 0):,}\n"
                    #             answer += f"  • Approved: {row.get('total_approved', 0):,}\n"
                    #         row_count = sum(row.get('total_claims', 0) for row in data)

                    #     elif "summary" in question_lower:
                    #         # ✅ Yearly or monthly summary
                    #         sql_query = f"""
                    #         SELECT COUNT(*) AS total_claims,
                    #             SUM("approved") AS total_approved
                    #         FROM {schema_table}
                    #         WHERE 1=1
                    #         """
                    #         if year:
                    #             sql_query += f" AND \"policy_end_date_year\" = {year}"
                    #         if month:
                    #             sql_query += f" AND \"policy_end_date_month\" = {month}"
                    #         sql_query += ";"

                    #         print(sql_query, "Final SQL Query")

                    #         data, error = self.execute_query_safely(sql_query)
                    #         if error:
                    #             return {
                    #                 'answer': f"Query execution failed: {error}",
                    #                 'success': False,
                    #                 'query': sql_query,
                    #                 'row_count': 0
                    #             }

                    #         row = data[0] if data else {}
                    #         timeframe = f"{calendar.month_name[month]} {year}" if year and month else str(year) if year else "overall"
                    #         answer = f"""
                    #  **Claims Summary for {timeframe}**

                    # - Total claims: {row.get('total_claims', 0):,}
                    # - Approved: {row.get('total_approved', 0):,}
                    #         """.strip()
                    #         row_count = row.get('total_claims', 0)

                    #     else:
                    #         # ✅ Default: just count of claims
                    #         sql_query = f"""
                    #         SELECT COUNT(*) AS total_claims
                    #         FROM {schema_table}
                    #         WHERE 1=1
                    #         """
                    #         if year:
                    #             sql_query += f" AND \"policy_end_date_year\" = {year}"
                    #         if month:
                    #             sql_query += f" AND \"policy_end_date_month\" = {month}"
                    #         sql_query += ";"

                    #         print(sql_query, "Final SQL Query")

                    #         data, error = self.execute_query_safely(sql_query)
                    #         if error:
                    #             return {
                    #                 'answer': f"Query execution failed: {error}",
                    #                 'success': False,
                    #                 'query': sql_query,
                    #                 'row_count': 0
                    #             }

                    #         claims_count = data[0]['total_claims'] if data else 0
                    #         month_name = calendar.month_name[month] if month else ""
                    #         timeframe = f"{month_name} {year}" if month and year else str(year) if year else month_name
                    #         answer = f" Total claims count for {timeframe or 'overall'} is {claims_count:,}."
                    #         row_count = claims_count

                    elif entity == "claim":
                        # 🔍 Dynamically detect group-by column for top claim summary
                        top_claim_groupby_columns = {
                            "product": "product_name",
                            "make": "make_clean",
                            "vehicle": "model_clean",
                            "segment": "customer_segment",
                            "location": "cleaned_branch_name_2"
                        }

                        group_by_col = None
                        for keyword, col in top_claim_groupby_columns.items():
                            if keyword in question_lower:
                                group_by_col = col
                                break

                        def build_claim_filters(year=None, month=None, location_col=None, location_val=None):
                            filters = ['"approved" IS NOT NULL']
                            if year:
                                filters.append(f'"policy_end_date_year" = {year}')
                            if month:
                                filters.append(f'"policy_end_date_month" = {month}')
                            if location_col and location_val:
                                filters.append(f'UPPER("{location_col}") = UPPER(\'{location_val}\')')
                            return " AND ".join(filters)

                        if location_group_by_col and "claim" in question_lower and any(word in question_lower for word in ["top", "which", "most", "more", "highest", "high", "low", "mid"]):
                            sql_query = f'''
                            SELECT "{location_group_by_col}" AS location, COUNT(*) AS claim_count
                            FROM {schema_table}
                            WHERE {build_claim_filters(year, month)}
                            GROUP BY "{location_group_by_col}"
                            ORDER BY claim_count DESC
                            LIMIT 5;
                            '''

                            data, error = self.execute_query_safely(sql_query)
                            if error:
                                return {
                                    'answer': f"Query execution failed: {error}",
                                    'success': False,
                                    'query': sql_query,
                                    'row_count': 0
                                }

                            timeframe = f"{calendar.month_name[month]} {year}" if year and month else str(year) if year else "overall"
                            location_title_map = {
                                "branch": "Branches",
                                "state": "States",
                                "zone": "Zones"
                            }
                            location_title = location_title_map.get(detected_keyword, "Locations")

                            answer = f"**Top 5 claimed {location_title} for {timeframe}**\n"
                            for row in data:
                                loc = row['location'] or "Unknown"
                                answer += f"- {loc.title()}: {row['claim_count']:,}\n"

                            return {
                                'answer': answer,
                                'success': True,
                                'query': sql_query,
                                'row_count': len(data)
                            }

                        elif ("top" in question_lower or "most" in question_lower or "highest" in question_lower) and group_by_col:
                            sql_query = f"""
                            SELECT "{group_by_col}", COUNT(*) AS claim_count
                            FROM {schema_table}
                            WHERE {build_claim_filters(year, month, location_col, location_val)}
                            GROUP BY "{group_by_col}"
                            ORDER BY claim_count DESC
                            LIMIT 5;
                            """

                            data, error = self.execute_query_safely(sql_query)
                            if error:
                                return {
                                    'answer': f"Query execution failed: {error}",
                                    'success': False,
                                    'query': sql_query,
                                    'row_count': 0
                                }

                            timeframe = f"{calendar.month_name[month]} {year}" if month and year else str(year) if year else "overall"
                            title_map = {
                                "product_name": "Claimed Products",
                                "make_clean": "Claimed Vehicle Makes",
                                "model_clean": "Claimed Vehicle Models",
                                "customer_segment": "Claimed Segments",
                                "cleaned_branch_name_2": "Claim Locations"
                            }
                            title = title_map.get(group_by_col, group_by_col.replace("_", " ").title())

                            answer = f" **Top 5 {title} for {timeframe}**\n"
                            for row in data:
                                label = row[group_by_col] or "Unknown"
                                answer += f"- {label}: {row['claim_count']:,}\n"
                            return {
                                'answer': answer,
                                'success': True,
                                'query': sql_query,
                                'row_count': len(data)
                            }

                        elif "summary" in question_lower and ("month wise" in question_lower or "monthly" in question_lower):
                            sql_query = f"""
                            SELECT "policy_end_date_month" AS month,
                                COUNT(*) AS total_claims,
                                SUM("approved") AS total_approved
                            FROM {schema_table}
                            WHERE {build_claim_filters(year, None, location_col, location_val)}
                            GROUP BY month
                            ORDER BY month;
                            """

                            data, error = self.execute_query_safely(sql_query)
                            if error:
                                return {
                                    'answer': f"Query execution failed: {error}",
                                    'success': False,
                                    'query': sql_query,
                                    'row_count': 0
                                }

                            answer = f" **Month-wise Claims Summary for {year}**\n"
                            for row in data:
                                month_number = int(row['month']) if row['month'] else 0
                                month_name = calendar.month_name[month_number] if month_number else "Unknown"
                                answer += f"- {month_name}:\n"
                                answer += f"  • Total claims: {row.get('total_claims', 0):,}\n"
                                answer += f"  • Approved: {row.get('total_approved', 0):,}\n"
                            return {
                                'answer': answer,
                                'success': True,
                                'query': sql_query,
                                'row_count': sum(row.get('total_claims', 0) for row in data)
                            }

                        elif "summary" in question_lower:
                            sql_query = f"""
                            SELECT COUNT(*) AS total_claims,
                                SUM("approved") AS total_approved
                            FROM {schema_table}
                            WHERE {build_claim_filters(year, month, location_col, location_val)};
                            """

                            data, error = self.execute_query_safely(sql_query)
                            if error:
                                return {
                                    'answer': f"Query execution failed: {error}",
                                    'success': False,
                                    'query': sql_query,
                                    'row_count': 0
                                }

                            row = data[0] if data else {}
                            timeframe = f"{calendar.month_name[month]} {year}" if year and month else str(year) if year else "overall"
                            answer = f"""
                    **Claims Summary for {timeframe}**

                    - Total claims: {row.get('total_claims', 0):,}
                    - Approved: {row.get('total_approved', 0):,}
                            """.strip()
                            return {
                                'answer': answer,
                                'success': True,
                                'query': sql_query,
                                'row_count': row.get('total_claims', 0)
                            }

                        else:
                            sql_query = f"""
                            SELECT COUNT(*) AS total_claims
                            FROM {schema_table}
                            WHERE {build_claim_filters(year, month, location_col, location_val)};
                            """

                            data, error = self.execute_query_safely(sql_query)
                            if error:
                                return {
                                    'answer': f"Query execution failed: {error}",
                                    'success': False,
                                    'query': sql_query,
                                    'row_count': 0
                                }

                            claims_count = data[0]['total_claims'] if data else 0
                            month_name = calendar.month_name[month] if month else ""
                            timeframe = f"{month_name} {year}" if month and year else str(year) if year else month_name
                            return {
                                'answer': f"Total claims count for {timeframe or 'overall'} is {claims_count:,}.",
                                'success': True,
                                'query': sql_query,
                                'row_count': claims_count
                            }


                    # elif entity == "vehicle":
                    #     schema_table = '"stage"."GBM1_prediction_data_with_recommendations"'

                    #     # Build dynamic filters for the query
                    #     def build_vehicle_filters(year=None, month=None):
                    #         filters = ['"make_clean" IS NOT NULL', 'TRIM("make_clean") <> \'\'']
                    #         if year:
                    #             filters.append(f'"policy_end_date_year" = {year}')
                    #         if month:
                    #             filters.append(f'"policy_end_date_month" = {month}')
                    #         return " AND ".join(filters)

                    #     # Dynamically detect group-by column for top vehicle summary
                    #     top_vehicle_groupby_columns = {
                    #         "segment": "customer_segment",
                    #         "product": "product_name",
                    #         "location": "cleaned_branch_name_2",
                    #         "vehicle": "model_clean",
                    #         "make": "make_clean"
                    #     }

                    #     group_by_col = None
                    #     for keyword, col in top_vehicle_groupby_columns.items():
                    #         if keyword in question_lower:
                    #             group_by_col = col
                    #             break

                    #     if ("top" in question_lower or "most" in question_lower or "highest" in question_lower) and group_by_col:
                    #         # Query for top vehicle summary
                    #         sql_query = f"""
                    #         SELECT "{group_by_col}", COUNT(*) AS vehicle_count
                    #         FROM {schema_table}
                    #         WHERE {build_vehicle_filters(year, month)}
                    #         GROUP BY "{group_by_col}"
                    #         ORDER BY vehicle_count DESC
                    #         LIMIT 5;
                    #         """

                    #         print(sql_query, "Final SQL Query")

                    #         data, error = self.execute_query_safely(sql_query)
                    #         if error:
                    #             return {
                    #                 'answer': f"Query execution failed: {error}",
                    #                 'success': False,
                    #                 'query': sql_query,
                    #                 'row_count': 0
                    #             }

                    #         timeframe = f"{calendar.month_name[month]} {year}" if year and month else str(year) if year else "overall"
                    #         title_map = {
                    #             "make_clean": "Vehicle Makes",
                    #             "model_clean": "Vehicle Models",
                    #             "product_name": "Product-wise Vehicles",
                    #             "customer_segment": "Customer Segments",
                    #             "cleaned_branch_name_2": "Vehicle Locations"
                    #         }
                    #             # 🧠 Smart title mapping
                    #         location_title_map = {
                    #             "cleaned_branch_name_2": "Branches",
                    #             "cleaned_state2": "States",
                    #             "cleaned_zone_2": "Zones"
                    #         }

                    #         if group_by_col in location_title_map:
                    #             title = location_title_map[group_by_col]
                    #             answer = f"**Top 5 vehicle origins by {title} for {timeframe}**\n"
                    #             for row in data:
                    #                 loc = row[group_by_col] or "Unknown"
                    #                 answer += f"- {loc.title()}: {row['vehicle_count']:,}\n"
                    #         else:
                    #             title = title_map.get(group_by_col, group_by_col.replace("_", " ").title())
                    #             answer = f"**Top 5 {title} for {timeframe}**\n"
                    #             for row in data:
                    #                 label = row[group_by_col] or "Unknown"
                    #                 answer += f"- {label}: {row['vehicle_count']:,}\n"


                            

                    #     elif "summary" in question_lower and ("month wise" in question_lower or "monthly" in question_lower):
                    #         # Month-wise vehicle make summary
                    #         sql_query = f"""
                    #         SELECT "policy_end_date_month" AS month,
                    #             "make_clean" AS vehicle_make,
                    #             COUNT(*) AS vehicle_count
                    #         FROM {schema_table}
                    #         WHERE {build_vehicle_filters(year, None)}
                    #         GROUP BY month, vehicle_make
                    #         ORDER BY month, vehicle_count DESC;
                    #         """

                    #         print(sql_query, "Final SQL Query")

                    #         data, error = self.execute_query_safely(sql_query)
                    #         if error:
                    #             return {
                    #                 'answer': f"Query execution failed: {error}",
                    #                 'success': False,
                    #                 'query': sql_query,
                    #                 'row_count': 0
                    #             }

                    #         answer = f" **Month-wise Top Vehicle Makes for {year}**\n"
                    #         for row in data:
                    #             month_number = int(row['month']) if row['month'] else 0
                    #             month_name = calendar.month_name[month_number] if month_number else "Unknown"
                    #             answer += f"- {month_name}: {row['vehicle_make']} → {row['vehicle_count']:,}\n"
                    #         row_count = len(data)

                    #     elif "summary" in question_lower:
                    #         # Yearly or monthly top vehicle make summary
                    #         sql_query = f"""
                    #         SELECT "make_clean" AS vehicle_make,
                    #             COUNT(*) AS vehicle_count
                    #         FROM {schema_table}
                    #         WHERE {build_vehicle_filters(year, month)}
                    #         GROUP BY vehicle_make
                    #         ORDER BY vehicle_count DESC
                    #         LIMIT 5;
                    #         """

                    #         print(sql_query, "Final SQL Query")

                    #         data, error = self.execute_query_safely(sql_query)
                    #         if error:
                    #             return {
                    #                 'answer': f"Query execution failed: {error}",
                    #                 'success': False,
                    #                 'query': sql_query,
                    #                 'row_count': 0
                    #             }

                    #         timeframe = f"{calendar.month_name[month]} {year}" if year and month else str(year) if year else "overall"
                    #         answer = f" **Top Vehicle Makes for {timeframe}**\n"
                    #         for row in data:
                    #             answer += f"- {row['vehicle_make']}: {row['vehicle_count']:,}\n"
                    #         row_count = len(data)

                    #     else:
                    #         # Default: Total vehicle count
                    #         sql_query = f"""
                    #         SELECT COUNT(*) AS total_vehicles
                    #         FROM {schema_table}
                    #         WHERE {build_vehicle_filters(year, month)};
                    #         """

                    #         print(sql_query, "Final SQL Query")

                    #         data, error = self.execute_query_safely(sql_query)
                    #         if error:
                    #             return {
                    #                 'answer': f"Query execution failed: {error}",
                    #                 'success': False,
                    #                 'query': sql_query,
                    #                 'row_count': 0
                    #             }

                    #         vehicle_count = data[0]['total_vehicles'] if data else 0
                    #         month_name = calendar.month_name[month] if month else ""
                    #         timeframe = f"{month_name} {year}" if month and year else str(year) if year else month_name
                    #         answer = f" Total vehicle count for {timeframe or 'overall'} is {vehicle_count:,}."
                    #         row_count = vehicle_count

                    # return {
                    #     'answer': answer,
                    #     'success': True,
                    #     'query': sql_query,
                    #     'row_count': row_count
                    # }

                    elif entity == "vehicle":
                        schema_table = '"stage"."GBM1_prediction_data_with_recommendations"'

                        # Build dynamic filters for the query
                        def build_vehicle_filters(year=None, month=None, location_col=None, location_val=None):
                            filters = ['"make_clean" IS NOT NULL', 'TRIM("make_clean") <> \'\'']
                            if year:
                                filters.append(f'"policy_end_date_year" = {year}')
                            if month:
                                filters.append(f'"policy_end_date_month" = {month}')
                            if location_col and location_val:
                                filters.append(f'UPPER("{location_col}") = UPPER(\'{location_val}\')')
                            return " AND ".join(filters)

                        # Dynamically detect group-by column for top vehicle summary
                        top_vehicle_groupby_columns = {
                            "segment": "customer_segment",
                            "product": "product_name",
                            "location": "cleaned_branch_name_2",
                            "vehicle": "model_clean",
                            "make": "make_clean"
                        }

                        group_by_col = None
                        for keyword, col in top_vehicle_groupby_columns.items():
                            if keyword in question_lower:
                                group_by_col = col
                                break

                        if ("top" in question_lower or "most" in question_lower or "highest" in question_lower) and group_by_col:
                            # Query for top vehicle summary
                            sql_query = f"""
                            SELECT "{group_by_col}", COUNT(*) AS vehicle_count
                            FROM {schema_table}
                            WHERE {build_vehicle_filters(year, month, location_col, location_val)}
                            GROUP BY "{group_by_col}"
                            ORDER BY vehicle_count DESC
                            LIMIT 5;
                            """

                            print(sql_query, "Final SQL Query")

                            data, error = self.execute_query_safely(sql_query)
                            if error:
                                return {
                                    'answer': f"Query execution failed: {error}",
                                    'success': False,
                                    'query': sql_query,
                                    'row_count': 0
                                }

                            timeframe = f"{calendar.month_name[month]} {year}" if year and month else str(year) if year else "overall"
                            title_map = {
                                "make_clean": "Vehicle Makes",
                                "model_clean": "Vehicle Models",
                                "product_name": "Product-wise Vehicles",
                                "customer_segment": "Customer Segments",
                                "cleaned_branch_name_2": "Vehicle Locations"
                            }

                            location_title_map = {
                                "cleaned_branch_name_2": "Branches",
                                "cleaned_state2": "States",
                                "cleaned_zone_2": "Zones"
                            }

                            if group_by_col in location_title_map:
                                title = location_title_map[group_by_col]
                                answer = f"**Top 5 vehicle origins by {title} for {timeframe}**\n"
                                for row in data:
                                    loc = row[group_by_col] or "Unknown"
                                    answer += f"- {loc.title()}: {row['vehicle_count']:,}\n"
                            else:
                                title = title_map.get(group_by_col, group_by_col.replace("_", " ").title())
                                answer = f"**Top 5 {title} for {timeframe}**\n"
                                for row in data:
                                    label = row[group_by_col] or "Unknown"
                                    answer += f"- {label}: {row['vehicle_count']:,}\n"
                            row_count = len(data)

                        elif "summary" in question_lower and ("month wise" in question_lower or "monthly" in question_lower):
                            # Month-wise vehicle make summary
                            sql_query = f"""
                            SELECT "policy_end_date_month" AS month,
                                "make_clean" AS vehicle_make,
                                COUNT(*) AS vehicle_count
                            FROM {schema_table}
                            WHERE {build_vehicle_filters(year, None, location_col, location_val)}
                            GROUP BY month, vehicle_make
                            ORDER BY month, vehicle_count DESC;
                            """

                            print(sql_query, "Final SQL Query")

                            data, error = self.execute_query_safely(sql_query)
                            if error:
                                return {
                                    'answer': f"Query execution failed: {error}",
                                    'success': False,
                                    'query': sql_query,
                                    'row_count': 0
                                }

                            answer = f" **Month-wise Top Vehicle Makes for {year}**\n"
                            for row in data:
                                month_number = int(row['month']) if row['month'] else 0
                                month_name = calendar.month_name[month_number] if month_number else "Unknown"
                                answer += f"- {month_name}: {row['vehicle_make']} → {row['vehicle_count']:,}\n"
                            row_count = len(data)

                        elif "summary" in question_lower:
                            # Yearly or monthly top vehicle make summary
                            sql_query = f"""
                            SELECT "make_clean" AS vehicle_make,
                                COUNT(*) AS vehicle_count
                            FROM {schema_table}
                            WHERE {build_vehicle_filters(year, month, location_col, location_val)}
                            GROUP BY vehicle_make
                            ORDER BY vehicle_count DESC
                            LIMIT 5;
                            """

                            print(sql_query, "Final SQL Query")

                            data, error = self.execute_query_safely(sql_query)
                            if error:
                                return {
                                    'answer': f"Query execution failed: {error}",
                                    'success': False,
                                    'query': sql_query,
                                    'row_count': 0
                                }

                            timeframe = f"{calendar.month_name[month]} {year}" if year and month else str(year) if year else "overall"
                            answer = f" **Top Vehicle Makes for {timeframe}**\n"
                            for row in data:
                                answer += f"- {row['vehicle_make']}: {row['vehicle_count']:,}\n"
                            row_count = len(data)

                        else:
                            # Default: Total vehicle count
                            sql_query = f"""
                            SELECT COUNT(*) AS total_vehicles
                            FROM {schema_table}
                            WHERE {build_vehicle_filters(year, month, location_col, location_val)};
                            """

                            print(sql_query, "Final SQL Query")

                            data, error = self.execute_query_safely(sql_query)
                            if error:
                                return {
                                    'answer': f"Query execution failed: {error}",
                                    'success': False,
                                    'query': sql_query,
                                    'row_count': 0
                                }

                            vehicle_count = data[0]['total_vehicles'] if data else 0
                            month_name = calendar.month_name[month] if month else ""
                            timeframe = f"{month_name} {year}" if month and year else str(year) if year else month_name
                            answer = f" Total vehicle count for {timeframe or 'overall'} is {vehicle_count:,}."
                            row_count = vehicle_count

                        return {
                            'answer': answer,
                            'success': True,
                            'query': sql_query,
                            'row_count': row_count
                        }


                except Exception as e:
                    return {
                        'answer': f"Error processing date-based {entity} summary: {str(e)}",
                        'success': False,
                        'query': None,
                        'row_count': 0
                    }


            # if "schema" in question_lower:
            #     if self.is_connected:
            #         schema_info = self.get_schema_info_structured()
            #         return {
            #             'answer': f"There are {schema_info['schema_count']} schemas: {', '.join(schema_info['schema_names'])}",
            #             'success': True,
            #             'query': None,
            #             'row_count': schema_info['schema_count']
            #         }
            #     else:
            #         return {
            #             'answer': "Please connect to a database first to view schema information.",
            #             'success': False,
            #             'query': None,
            #             'row_count': 0
            #         }


            if any(keyword in question_lower for keyword in ["erd", "entity relationship diagram", "use-case", "usecase", "use case mapping", "table mappings"]):
                if self.is_connected:
                    erd_usecase_output = self.generate_erd_and_usecases()
                    return {
                        'answer': erd_usecase_output,
                        'success': True,
                        'query': None,
                        'row_count': 0
                    }
                else:
                    return {
                        'answer': "Please connect to a database first to generate ERD or use-case mappings.",
                        'success': False,
                        'query': None,
                        'row_count': 0
                    }
            
            # Check for summary requests
            summary_keywords = [ 'overview', 'describe database', 'show tables', 'database info', 'summary of database', 'give me the summary of database','database summary', 'what does this database contain', 'purpose of database', 'db purpose']
            if any(keyword in question_lower for keyword in summary_keywords):
                if self.is_connected:
                    summary = self.generate_high_level_db_summary()
                    return {
                        'answer': summary,
                        'success': True,
                        'query': None,
                        'row_count': 0
                    }
                else:
                    return {
                        'answer': "Please connect to a database first to view its summary.",
                        'success': False,
                        'query': None,
                        'row_count': 0
                    }
            
            # If not connected, provide general response
            if not self.is_connected:
                general_response = self.get_openrouter_response(
                    question, 
                    "You are a helpful database assistant. The user hasn't connected to a database yet. Provide a general, helpful response and suggest connecting to a database for specific data queries."
                )
                return {
                    'answer': general_response,
                    'success': True,
                    'query': None,
                    'row_count': 0
                }
            
            # Generate and execute SQL query
            try:
                sql_query = self.generate_sql_query(question, conversation_history, year=year, month=month)
                if not sql_query:
                    return {
                        'answer': "I couldn't generate a SQL query for your question. Please try rephrasing it.",
                        'success': False,
                        'query': None,
                        'row_count': 0
                    }
                
                # Execute query
                data, error = self.execute_query_safely(sql_query)
                
                if error and not data:
                    return {
                        'answer': f"Query execution failed:{error}",
                        'success': False,
                        'query': sql_query,
                        'row_count': 0
                    }
                
                # Format results
                formatted_answer = self.format_query_results(data, question)
                
                return {
                    'answer': formatted_answer,
                    'success': True,
                    'query': sql_query,
                    'row_count': len(data) if data else 0
                }
                
            except Exception as e:
                logger.error(f"Question processing error: {e}")
                return {
                    'answer': f"An error occurred while processing your question: {str(e)}",
                    'success': False,
                    'query': None,
                    'row_count': 0
                }
        


    def generate_case_insensitive_query(self, original_query, question):
        """Transform query to handle case-insensitive searches"""
        try:
            # Extract search index
            search_index = self.schema_info.get('search_index', {})
            
            words = re.findall(r'\b[a-zA-Z]+\b', question.lower())
            
            # Build replacement map for case-insensitive conditions
            replacements = {}
            
            for word in words:
                if len(word) > 2:  # Only consider words longer than 2 characters
                    for value_key, value_info in search_index.get('values', {}).items():
                        if word in value_key:
                            table = value_info['table']
                            column = value_info['column']
                            actual_value = value_info['actual_value']
                            
                            # Create case-insensitive pattern
                            pattern = f'"{column}" = \'{word}\''
                            replacement = f'UPPER("{column}") = UPPER(\'{actual_value}\')'
                            
                            # Also handle LIKE patterns
                            like_pattern = f'"{column}" LIKE \'%{word}%\''
                            like_replacement = f'UPPER("{column}") LIKE UPPER(\'%{actual_value}%\')'
                            
                            replacements[pattern] = replacement
                            replacements[like_pattern] = like_replacement
            
            # Apply replacements
            modified_query = original_query
            for pattern, replacement in replacements.items():
                modified_query = re.sub(pattern, replacement, modified_query, flags=re.IGNORECASE)
            
            modified_query = re.sub(
                r'WHERE\s+("?\w+"?)\s*=\s*\'([^\']+)\'',
                lambda m: f'WHERE UPPER({m.group(1)}) = UPPER(\'{m.group(2)}\')',
                modified_query,
                flags=re.IGNORECASE
            )
            
            modified_query = re.sub(
                r'("?\w+"?)\s+LIKE\s+\'([^\']+)\'',
                lambda m: f'UPPER({m.group(1)}) LIKE UPPER(\'{m.group(2)}\')',
                modified_query,
                flags=re.IGNORECASE
            )
            
            return modified_query
            
        except Exception as e:
            logger.error(f"Error in case-insensitive query generation: {e}")
            return original_query

    def clean_llm_sql_response(self, llm_response):
        """Extract pure SQL from LLM response by removing explanations or markdown"""
        import re

        if not llm_response:
            return ""

        # Remove code block markers
        llm_response = re.sub(r"```sql", "", llm_response, flags=re.IGNORECASE)
        llm_response = re.sub(r"```", "", llm_response, flags=re.IGNORECASE)

        # Split lines and find first valid SQL line
        sql_keywords = ['SELECT', 'WITH', 'INSERT', 'UPDATE', 'DELETE', 'CREATE']
        lines = llm_response.strip().splitlines()

        start_idx = 0
        for i, line in enumerate(lines):
            if any(line.strip().upper().startswith(k) for k in sql_keywords):
                start_idx = i
                break

        sql_lines = lines[start_idx:]
        cleaned_sql = "\n".join(sql_lines).strip()

        if not cleaned_sql.endswith(";"):
            cleaned_sql += ";"

        return cleaned_sql
    
    
    import difflib
    import re

    def detect_dynamic_location_filter(self, question: str):
        """
        Dynamically detect location (branch/city, state, or zone) from the question
        using fuzzy matching against actual values in the schema's search index.

        Returns:
            (location_column: str, matched_values: List[str]) or (None, [])
        """
        question = question.lower()
        words = re.findall(r'\b[a-z]+\b', question)  # all tokens

        # Extract known location values from schema or preloaded search_index
        search_values = self.schema_info.get('search_index', {}).get('values', {})

        city_values = set()
        state_values = set()
        zone_values = set()

        for _, info in search_values.items():
            col = info.get('column')
            val = str(info.get('actual_value', '')).lower().strip()
            if not val:
                continue

            if col == "cleaned_branch_name_2":
                city_values.add(val)
            elif col == "cleaned_state2":
                state_values.add(val)
            elif col == "cleaned_zone_2":
                zone_values.add(val)

        location_hits = []

        # Match each word from question with known values
        for word in words:
            city_match = difflib.get_close_matches(word, city_values, n=1, cutoff=0.85)
            if city_match:
                location_hits.append(("cleaned_branch_name_2", city_match[0]))
                continue

            state_match = difflib.get_close_matches(word, state_values, n=1, cutoff=0.85)
            if state_match:
                location_hits.append(("cleaned_state2", state_match[0]))
                continue

            zone_match = difflib.get_close_matches(word, zone_values, n=1, cutoff=0.85)
            if zone_match:
                location_hits.append(("cleaned_zone_2", zone_match[0]))
                continue

        if not location_hits:
            return None, []

        # Prefer more specific match: branch > state > zone
        for col in ["cleaned_branch_name_2", "cleaned_state2", "cleaned_zone_2"]:
            values = list({val for c, val in location_hits if c == col})
            if values:
                return col, values

        return None, []



    def generate_sql_query(self, question, conversation_history,year=None, month=None):
            """Generate SQL query using OpenRouter Llama4 Maverick with case-insensitive enhancements"""
            if not self.is_connected:
                return None
            
            question_lower = question.lower()

# Direct override for renewed / not renewed questions using predicted_status column
            if 'not renewed' in question_lower:
                status_value = 'Not Renewed'
            elif 'renewed' in question_lower:
                status_value = 'Renewed'
            else:
                status_value = None


            # FIX 1: Consolidate location detection logic
            known_locations = {}
            location_columns = ['cleaned_branch_name_2', 'cleaned_state2', 'cleaned_zone_2']
            
            # Debug: Check what's in schema_info
            logger.info(f"[Debug] schema_info keys: {list(self.schema_info.keys())}")
            
            search_index = self.schema_info.get('search_index', {})
            logger.info(f"[Debug] search_index keys: {list(search_index.keys())}")
            
            # Try multiple ways to get location data
            if 'values' in search_index:
                logger.info(f"[Debug] Found {len(search_index['values'])} values in search_index")
                for value_key, value_info in search_index['values'].items():
                    col = value_info.get('column', '')
                    actual_val = value_info.get('actual_value', '')
                    if col in location_columns and actual_val:
                        known_locations.setdefault(col, set()).add(actual_val)
                        logger.info(f"[Debug] Added location: {col} = {actual_val}")
            
            # Alternative: Try to get locations from tables directly
            if not known_locations and 'tables' in self.schema_info:
                for table_name, table_info in self.schema_info['tables'].items():
                    columns = table_info.get('columns', {})
                    for col_name in location_columns:
                        if col_name in columns:
                            # If you have sample data or distinct values stored somewhere
                            distinct_vals = columns[col_name].get('distinct_values', [])
                            if distinct_vals:
                                known_locations[col_name] = set(distinct_vals)
                                logger.info(f"[Debug] Added from table schema: {col_name} = {distinct_vals}")
            
            # Convert sets to sorted lists
            known_locations = {col: sorted(list(vals)) for col, vals in known_locations.items()}
            
            # If still empty, try hardcoded fallback based on your data screenshot
            if not known_locations:
                logger.warning("[Debug] No locations found in schema, using fallback")
                known_locations = {
                    'cleaned_branch_name_2': [
    "agartala", "ahmedabad", "ahmednagar", "ambala", "amravati", "amritsar", "andheri", "angul", "aurangabad",
    "ballari", "balsore", "bangalore", "basirhat", "begusarai", "belagavi", "belgaum", "bellary", "bengaluru",
    "berhampore", "berhampur", "bhagalpur", "bhopal", "bhubaneshwar", "bijapur", "bilaspur", "burdwan", "calicut",
    "chandigarh", "chennai", "chhatrapatisambhajinagar", "coimbatore", "corporateoffice", "cuttack", "davanagere",
    "dehradun", "delhi1", "delhi2", "delhinauranghouse", "deoghar", "dhanbad", "durgapur", "gandhidham", "gaya",
    "gulbarga", "guntur", "gurgaon", "guwahati", "hubballi", "hubli", "hyderabad", "imphal", "indore", "jaipur",
    "jajpur", "jalandhar", "jammu", "jamnagar", "jamshedpur", "jeypore", "jorhat", "kadapa", "kalaburagi", "kangra",
    "kanpur", "karimnagar", "khammam", "kharagpur", "kochi", "kolhapur", "kolkata1", "kollam", "kurnool", "lucknow",
    "ludhiana", "madurai", "mahbubnagar", "maldah", "mandi", "mangalore", "mangaluru", "margao", "mumbai", "mumbai1",
    "muzaffarpur", "mysore", "mysuru", "nagpur", "nashik", "ncr", "nellore", "noida", "patna", "puducherry1", "pune",
    "punetpa", "purnea", "raipur", "rajahmundry", "rajkot", "ranchi", "rourkela", "salem", "sambalpur", "satara",
    "shillong", "shimoga", "shivamogga", "siliguri", "solan", "solapur", "srinagar", "surat", "thane", "thrissur",
    "tirunelveli", "tirupati", "trichy", "trivandrum", "tumakuru", "tumkur", "udaipur", "vadodara", "varanasi",
    "vellore", "vijayapura", "vijayawada", "vishakapatnam", "warangal"
],
                    'cleaned_state2': [
    "andhrapradesh", "assam","bihar","chandigarh",
    "chhattisgarh","delhi","goa","gujarat","haryana","himachalpradesh","jammukashmir","jharkhand","karnataka","kerala","madhyapradesh","maharashtra","manipur","meghalaya","mizoram","odisha","puducherry","punjab","rajasthan","tamilnadu","telangana","tripura","uttarakhand","uttarpradesh","westbengal"
],

                    'cleaned_zone_2': ['south', 'west', 'north',' east','corporate'],
                }

                

            # def detect_location_filter(question, known_locations):
            #     """Enhanced location detection with fuzzy matching"""
            #     question_lower = question.lower()
                
            #     for col, values in known_locations.items():
            #         for val in values:
            #             val_lower = val.lower()
            #             # Exact match
            #             if val_lower in question_lower:
            #                 logger.info(f"[Location Match] Exact match found: '{val}' in column '{col}'")
            #                 return col, val
                        
            #             # Fuzzy match
            #             ratio = SequenceMatcher(None, val_lower, question_lower).ratio()
            #             if ratio > 0.85:
            #                 logger.info(f"[Location Match] Fuzzy match found: '{val}' (ratio: {ratio:.2f}) in column '{col}'")
            #                 return col, val
                        
            #             # Word boundary match (e.g., "surat" should match even if part of larger text)
            #             import re
            #             if re.search(r'\b' + re.escape(val_lower) + r'\b', question_lower):
            #                 logger.info(f"[Location Match] Word boundary match found: '{val}' in column '{col}'")
            #                 return col, val
                
            #     logger.info(f"[Location Detection] No matches found for question: '{question}'")
            #     return None, None
            def detect_location_filter(question, known_locations):
                    """
                    Detects the best matching location value from the question based on known location lists.
                    Supports exact, space-insensitive, fuzzy, and word-boundary matches.
                    """
                    question_raw = question
                    question_lower = question.lower()
                    question_no_space = question_lower.replace(" ", "")

                    for col, values in known_locations.items():
                        for val in values:
                            val_lower = val.lower()
                            val_no_space = val_lower.replace(" ", "")

                            # 🔹 Exact match with space preserved
                            if val_lower in question_lower:
                                logger.info(f"[Location Match] Exact match found: '{val}' in column '{col}'")
                                return col, val

                            # 🔹 Match with spaces removed
                            if val_no_space in question_no_space:
                                logger.info(f"[Location Match] Space-insensitive match found: '{val}' in column '{col}'")
                                return col, val

                            # 🔹 Word boundary match (surat, tamil nadu, etc.)
                            if re.search(r'\b' + re.escape(val_lower) + r'\b', question_lower):
                                logger.info(f"[Location Match] Word-boundary match: '{val}' in column '{col}'")
                                return col, val

                            # 🔹 Fuzzy match (safe threshold)
                            ratio = SequenceMatcher(None, val_no_space, question_no_space).ratio()
                            if ratio > 0.88:
                                logger.info(f"[Location Match] Fuzzy match: '{val}' (ratio: {ratio:.2f}) in column '{col}'")
                                return col, val

                    logger.info(f"[Location Detection] No matches found for question: '{question_raw}'")
                    return None, None

            # FIX 2: Detect location BEFORE building the query
            location_col, location_val = detect_location_filter(question, known_locations)
            
            # Log location detection for debugging
            if location_col and location_val:
                logger.info(f"[Location Detected] Column: {location_col}, Value: {location_val}")
            else:
                logger.info(f"[Location Detection] No location found in question: '{question}'")
                logger.info(f"[Available Locations] {known_locations}")

            if status_value:
                table_name = '"stage"."GBM1_prediction_data_with_recommendations"'
                sql_query = f"""
                SELECT COUNT(*)
                FROM {table_name}
                WHERE UPPER("predicted_status") = UPPER('{status_value}')
                """

                if year:
                    sql_query += f" AND \"policy_end_date_year\" = {year}"
                if month:
                    sql_query += f" AND \"policy_end_date_month\" = {month}"

                # FIX 3: Apply location filter if detected
                if location_col and location_val:
                    sql_query += f" AND UPPER(\"{location_col}\") = UPPER('{location_val}')"

                sql_query += ";"

                logger.info(f"[Renewed status specific] Final SQL Query: {sql_query}")
                return sql_query


#             if status_value:
#                 table_name = '"stage"."GBM1_prediction_data_with_recommendations"'
#   # Adjust if multiple tables
#                 sql_query = f"""
#                 SELECT COUNT(*)
#                 FROM {table_name}
#                 WHERE UPPER("predicted_status") = UPPER('{status_value}')
#                 """
#                 if year:
#                     sql_query += f" AND \"policy_end_date_year\" = {year}"
#                 if month:
#                     sql_query += f" AND \"policy_end_date_month\" = {month}"
#                 sql_query += ";"

#                 logger.info(f"[Renewed status specific] Final SQL Query: {sql_query}")
#                 return sql_query

            segment_keywords = {
        "elite": "Elite Retainers",
        "potential": "Potential Customers",
        "low value": "Low Value Customers",
        "risky": "Risk Segment",
    }

            segment_value = None
            for keyword, value in segment_keywords.items():
                if keyword in question_lower:
                    segment_value = value
                    break

            # Build contextual metadata
            date_context = ""
            if year and month:
                month_name = calendar.month_name[month]
                date_context = f"The user is asking about data for {month_name} {year}.\n"
            elif year:
                date_context = f"The user is asking about data for the year {year}.\n"
            elif month:
                month_name = calendar.month_name[month]
                date_context = f"The user is asking about data for the month {month_name}.\n"

            # 🔍 Schema structure
            schema_context = "DATABASE SCHEMA:\n"
            for table_name, table_info in self.schema_info.get('tables', {}).items():
                schema_context += f"\nTable: {table_name}\n"
                columns = table_info.get('columns', {})
                for col_name, col_info in columns.items():
                    col_type = col_info.get('type', 'unknown')
                    schema_context += f"  - {col_name}: {col_type}\n"

            # 🧪 Sample values for grounding
            search_index = self.schema_info.get('search_index', {})
            if search_index.get('values'):
                schema_context += "\nSAMPLE VALUES (for reference):\n"
                current_table = None
                for value_key, value_info in list(search_index['values'].items())[:20]:  # Limit to 20 for clarity
                    if value_info['table'] != current_table:
                        current_table = value_info['table']
                        schema_context += f"\nTable {current_table}:\n"
                    schema_context += f"  {value_info['column']}: {value_info['actual_value']}\n"

            # # 🧠 Last 3 conversations (if any)
            # history_context = ""
            # if conversation_history:
            #     history_context = "\nRECENT CONVERSATION:\n"
            #     for entry in conversation_history[-3:]:
            #         history_context += f"- Q: {entry.get('question', '')}\n"
            #         history_context += f"  A: {entry.get('answer', '')}\n"


            # 🧠 Last 3 conversations (if any) + dynamic memory for follow-up question resolution
            history_context = ""
            if conversation_history:
                history_context = "\nRECENT CONVERSATION:\n"
                for entry in conversation_history[-3:]:
                    history_context += f"- Q: {entry.get('question', '')}\n"
                    history_context += f"  A: {entry.get('answer', '')}\n"

                # 🔄 Rewrite vague question if it's a follow-up like "in jan?", "what about churn?", etc.
                recent_q = conversation_history[-1]['question'].lower()
                recent_a = conversation_history[-1]['answer'].lower()
                user_q = question.lower()

                # Detect vague follow-up pattern
                followup_phrases = ['in jan', 'in feb', 'what about', 'same', 'again', 'then', 'and in', 'how about', 'next']
                is_followup = any(p in user_q for p in followup_phrases)

                # Reuse last question intent if follow-up is detected and new question is short
                if is_followup and len(user_q.split()) < 6:
                    months = {calendar.month_name[i].lower(): i for i in range(1, 13)}
                    found_month = next((m for m in months if m in user_q), None)

                    # Replace month if new one is given
                    if found_month:
                        enriched_q = re.sub(r'\b(' + '|'.join(months.keys()) + r')\b', found_month, recent_q, flags=re.IGNORECASE)
                    else:
                        # If no new month, just reuse previous question as context
                        enriched_q = recent_q

                    logger.info(f"[Memory-Resolved] Original: '{question}' → Enriched: '{enriched_q}'")
                    question = enriched_q


            policy_schema_prompt = self.build_system_prompt_for_policy_schema()
            sql_rules_prompt = f"""You are a PostgreSQL SQL expert. Generate ONLY the SQL query based on the user's question.

### 🔴 STRICT RULES:

        1. Understand the user's question carefully.

        2. Use double quotes for table and column names to handle case sensitivity.

        3. Always place SQL clauses in correct order:
        - WHERE
        - GROUP BY
        - HAVING (if needed)
        - ORDER BY
        - LIMIT

        4. For case-insensitive filters, use UPPER():
        - Example: UPPER(column_name) = UPPER('value')

        5. For LIKE searches, use:
        - Example: UPPER(column_name) LIKE UPPER('%value%')

        6. When generating summary or count queries:
        - Use COUNT(*) for totals.
        - Use SUM(), AVG(), or appropriate aggregates for financial or numeric summaries.
        - Include GROUP BY if multiple categories or month-wise summaries are requested.

        7. For month-wise summaries:
        - Use "policy_end_date_month" and group by it.

        8. For year-wise summaries:
        - Use "policy_end_date_year" and group by it.

        9. For vehicles summary, include:
        - Manufacturer/make
        - Model
        - Variant
        - Segment
        - Fuel type
        - Example: GROUP BY "manufacturer/make"

        10. For churn summaries:
            - Check distinct values in "churn_category".
            - If 'summary' is asked, aggregate counts by churn_category.

        11. For claims summaries:
            - Provide total claims, approved, and denied counts.
            - Example:
            SELECT COUNT(*) AS total_claims,
                    SUM("approved") AS total_approved,
                    SUM("denied") AS total_denied
            FROM schema.table
            WHERE ...

        12. Always use fully qualified table names with schema, e.g. "stage"."GBM1_prediction_data_with_recommendations".

        13. Do not place AND after LIMIT. Ensure WHERE clauses come before GROUP BY, and LIMIT is last.

        14. When the question is ambiguous (e.g. 'give me the summary'), prefer providing month-wise summaries if month or year is mentioned, otherwise provide an overall category-wise summary.

        15. Return only valid PostgreSQL SQL. No MySQL or T-SQL syntax.

        ### ⚠️ EXAMPLES:

        ❌ Wrong:
        SELECT * FROM table
        LIMIT 10 AND column = 'value';

        ✅ Correct:
        SELECT * FROM table
        WHERE column = 'value'
        LIMIT 10;

IMPORTANT RULES:
        Your goals:
           - Understand the user's question carefully.
           - If the question asks for a **summary or overview**, such as containing words like:
             summary, breakdown, report, overview, analysis, month wise, year wise, trend
             - Customer segments such as "Elite Retainers", "Potential Loyalist", "Low Value", etc., are always stored in the column "customer_segment".
            - Do NOT use "clv_category" or "retentional_channel" to filter elite or segment-related questions.
             "- If the user asks about segment names like 'elite', 'potential', etc., match against customer_segment\n"
            "- If the user asks about retention strategy or communication channel (SMS, Email), then use retention_channel\n\n"
            "- If user gives a float value like 0.68, match it with the closest numeric columns (e.g., manufacturer_risk_rate, churn_probability) based on context."
            "- manufacturer_risk_rate: vehicle-specific risk score used in pricing"
            "- churn_probability: likelihood of a customer leaving the service"

        👉 Then:
        - Generate an aggregate SQL query summarising the data.
        - Include GROUP BY if month-wise or year-wise summary is requested.

        - If the question asks for a **count or direct retrieval**:
        - Generate a simple SELECT COUNT(*) query with appropriate WHERE filters.

        - If the user requests **full data or details**, such as "show all policies in January", use:
        - SELECT * with WHERE filters.

        - Always follow these SQL rules:
        - Use double quotes for all table and column names for PostgreSQL.
        - Use UPPER() for case-insensitive comparisons.
        - Always include schema names (e.g., "stage"."GBM1_prediction_data_with_recommendations").
        - For LIKE searches use: UPPER(column) LIKE UPPER('%value%').

1. Return ONLY the SQL query, no explanations or formatting  
2. Use double quotes for table/column names to handle case sensitivity  
3. Always use proper PostgreSQL syntax  
4. For aggregations, use appropriate GROUP BY clauses  
5. Use LIMIT for large result sets when appropriate  
6. CRITICAL: For text comparisons, ALWAYS use UPPER() function for case-insensitive matching  
7. When comparing text values, use: UPPER(column_name) = UPPER('search_value')  
8. For LIKE searches, use: UPPER(column_name) LIKE UPPER('%search_value%')  
9. The database may contain values in different cases (e.g., 'HONDA', 'Honda', 'honda')  

EXAMPLE CASE-INSENSITIVE PATTERNS:
- Instead of: WHERE brand = 'honda'
- Use: WHERE UPPER(brand) = UPPER('honda')
- Instead of: WHERE brand LIKE '%honda%'  
- Use: WHERE UPPER(brand) LIKE UPPER('%honda%')

10. Always use fully qualified table names with schema (e.g., "stage"."GBM1_prediction_data_with_recommendations")  
11. When generating schema queries:
    - Exclude system schemas by adding WHERE schema_name NOT IN ('information_schema', 'pg_catalog', 'pg_toast')

12. For schema name queries, always provide:
    SELECT schema_name FROM information_schema.schemata WHERE schema_name NOT IN ('information_schema', 'pg_catalog', 'pg_toast');

13. For questions mentioning "churn", check the distinct values in churn_category column.
    - If 'churn' is NOT a value but the column has High, Mid, Low:
        - Aggregate and return counts for each churn_category.
        - Do NOT use WHERE churn_category = 'churn' if 'churn' does not exist as value.
    - Example:
        - Instead of:
            SELECT COUNT(*)
            FROM table
            WHERE churn_category = 'churn'
        - Use:
            SELECT churn_category, COUNT(*)
            FROM "stage"."GBM1_prediction_data_with_recommendations"
            WHERE "policy_end_date_year" = 2024
            GROUP BY churn_category;

14. For questions mentioning 'churn':
    - Check distinct values in churn_category column.
    - If the value 'churn' does not exist, aggregate counts for each churn_category instead of filtering WHERE churn_category = 'churn'.
    - Example:
        SELECT "churn_category", COUNT(*)
        FROM "stage"."GBM1_prediction_data_with_recommendations"
        WHERE "policy_end_date_year" = 2024
        GROUP BY "churn_category"

15. For questions mentioning 'policy' with 'year':
    - Provide counts or summaries grouped by year or filtered by the specific year mentioned.
    - Example:
        SELECT "policy_end_date_year" as policy_year, COUNT(*)
        FROM "stage"."GBM1_prediction_data_with_recommendations"
        GROUP BY policy_year
        ORDER BY policy_year;

16. If the table contains pre-computed date parts (e.g., "policy_end_date_year", "policy_end_date_month"), always use them instead of EXTRACT().

17. If the question mentions a year (e.g., 2024), filter using "policy_end_date_year".

18. If no year is mentioned, provide the overall summary.

19. For vehicle summaries, include: "manufacturer/make", model, variant, vehicle_segment, fuel_type.

20. For zone summaries, include: cleaned_zone_2, cleaned_state2, cleaned_branch_name_2.

21. For policy summaries, include: product_name, policy_tenure, policy_status, total_premium_payable.

22. For summary or aggregate queries, include meaningful GROUP BY clauses and counts.

23. For count queries, use COUNT(*) with appropriate filters.

24. If the question requests 'all details', 'full data', or 'complete record', use SELECT * to include all columns for that table.

25. If the user asks for column names without specifying a table, return all columns with their table and schema names using information_schema.columns.

26. If the user mentions a location (e.g. Tamil Nadu), filter using the relevant state or zone column.

27. If the user mentions "month wise", include GROUP BY "policy_end_date_month".

28. Always use UPPER() for case-insensitive filters.

29. If the user mentions 'month wise', include GROUP BY "policy_end_date_month".

30. If the user mentions a location (e.g. Tamil Nadu), filter using the relevant state or zone column with UPPER().

31. Generate appropriate aggregations and GROUP BY clauses based on the question.

32. If the table contains null or blank strings in categorical columns, use:
    WHERE column IS NOT NULL AND TRIM(column) <> ''

33. For vague churn questions like "churn summary" without a value, provide a grouped churn_category count summary.

34. For "how many types" or "list all X" questions, use SELECT DISTINCT on the relevant column.

35. Always use the table "stage"."GBM1_prediction_data_with_recommendations" unless specified otherwise.

36.To answer questions about car models, use the column "variant" (or "model_name" if present) from the table "stage"."GBM1_prediction_data_with_recommendations".

37.To answer questions about vehicles, use the column "variant" (or "model_name" or "if present) from the table "stage"."GBM1_prediction_data_with_recommendations".

38. For customer segment-based queries:
    - Use the column "customer_segment" in WHERE clause.
    - Segment values include 'Elite Retainers', 'Potential Customers', 'Low Value', etc.
    - Always use UPPER() for comparisons.
    - Example:
      SELECT COUNT(*)
      FROM "stage"."GBM1_prediction_data_with_recommendations"
      WHERE UPPER("customer_segment") = UPPER('Low Value');


        If the user asks for segment-based insights like:
        - "Show churn of elite customers"
        - "Policy summary for low value"
        - "Claims from potential customers"

        → THEN:
        - Use UPPER("customer_segment") = UPPER('...') appropriately in the WHERE clause.

39. If the question mentions "elite", "potential","low value", etc., apply UPPER("customer_segment") = UPPER('...') filter.

40.If the data is month-wise, highlight trends (e.g., peak month, average, low).
        - If the data contains categories, mention dominant categories.
        - Use natural, non-technical tone for business users.

41.If the question includes “between YEAR1 and YEAR2”, generate a query using:
WHERE policy_end_date_year BETWEEN YEAR1 AND YEAR2
or
WHERE policy_end_date_year IN (YEAR1, YEAR2)

42.When the user asks about reasons for non-renewal, generate a SQL query using the columns:
- not_renewed_reasons
- main_reason

If the user asks for recommendations, use:
- primary_recommendation

If the user asks for additional offers, use:
- additional_offers

Only use these columns if relevant to the user’s query.

Always include filtering logic when applicable, for example:
- WHERE "predicted_status" = 'Not Renewed'
- WHERE column IS NOT NULL

Respond only with a valid SQL query without explanations.

Schema reference:
- predicted_status (values: Renewed / Not Renewed)
- not_renewed_reasons
- main_reason
- primary_recommendation
- additional_offers

43.You are querying the table "stage"."GBM1_prediction_data_with_recommendations".

The location-related columns are:
- cleaned_branch_name_2 (e.g., 'delhi1', 'surat', 'mumbai')
- cleaned_state2 (e.g., 'delhi', 'gujarat', 'punjab')
- cleaned_zone_2 (e.g., 'north', 'west', 'south')

If a user asks a question involving a city/state/zone like 'delhi', try to match it using:
1. cleaned_branch_name_2 (most granular)
2. cleaned_state2 (fallback)
3. cleaned_zone_2 (if broader zone is mentioned)

Always include this location filter if found.




        
        {date_context}

        {schema_context}
        {history_context}

        Generate a PostgreSQL query for the following question:
        {question} 
        """
            system_prompt = f"{policy_schema_prompt}\n\n{sql_rules_prompt}\n\n{date_context}{schema_context}{history_context}\n\nGenerate a PostgreSQL query for the following question:\n{question}"
            
            
            # query = self.get_openrouter_response(question, system_prompt)
            llm_response = self.get_openrouter_response(question, system_prompt)
            query = self.clean_llm_sql_response(llm_response)
            
            # Clean the query
            # query = query.strip()
            # # Remove code block markers if present
            # query = re.sub(r'sql\s*', '', query)
            # # query = re.sub(r'\s*', '', query)
            # query = re.sub(r'^```sql', '', query, flags=re.IGNORECASE).strip()
            # query = re.sub(r'^```', '', query, flags=re.IGNORECASE).strip()
            # query = re.sub(r'```$', '', query, flags=re.IGNORECASE).strip()

             # Enhanced query cleaning to handle more edge cases
            query = query.strip()

             # Remove code block markers if present
            query = re.sub(r'sql\s*', '', query)
            # query = re.sub(r'\s*', '', query)
            query = re.sub(r'^```sql', '', query, flags=re.IGNORECASE).strip()
            query = re.sub(r'^```', '', query, flags=re.IGNORECASE).strip()
            query = re.sub(r'```$', '', query, flags=re.IGNORECASE).strip()
            
            # Remove markdown code blocks completely
            query = re.sub(r'^```(?:sql)?\s*', '', query, flags=re.IGNORECASE | re.MULTILINE)
            query = re.sub(r'\s*```$', '', query, flags=re.IGNORECASE | re.MULTILINE)
            
            # Remove any explanatory text before the actual SQL
            # Look for common SQL starting patterns
            sql_start_patterns = [
                r'^\s*SELECT\s+',
                r'^\s*WITH\s+',
                r'^\s*INSERT\s+',
                r'^\s*UPDATE\s+',
                r'^\s*DELETE\s+',
                r'^\s*CREATE\s+',
                r'^\s*ALTER\s+',
                r'^\s*DROP\s+'
            ]
            
            # Find the first line that starts with a SQL keyword
            lines = query.split('\n')
            sql_start_line = 0
            
            for i, line in enumerate(lines):
                line_stripped = line.strip()
                if any(re.match(pattern, line_stripped, re.IGNORECASE) for pattern in sql_start_patterns):
                    sql_start_line = i
                    break
            
            # Extract only the SQL part
            if sql_start_line > 0:
                query = '\n'.join(lines[sql_start_line:])
            
            # Remove any trailing explanatory text after the SQL
            # Look for common ending patterns
            sql_end_patterns = [
                r';\s*$',
                r';\s*\n\s*$'
            ]
            
            # Find the last semicolon and cut everything after it
            semicolon_match = re.search(r';(?:\s*\n.*)?$', query, re.DOTALL)
            if semicolon_match:
                query = query[:semicolon_match.start() + 1]
            
            # Final cleanup
            query = query.strip()
            
            # Ensure the query ends with a semicolon
            if not query.endswith(';'):
                query += ';'
            
            # Log the cleaned query for debugging
            logger.info(f"[Cleaned SQL Query]: {query}")
            
            # Validate that the query starts with a valid SQL keyword
            if not any(re.match(pattern, query, re.IGNORECASE) for pattern in sql_start_patterns):
                logger.error(f"[Invalid SQL Query]: Query doesn't start with valid SQL keyword: {query}")
                # Return a fallback query or None
                return None

            churn_category = None
            if 'low churn' in question.lower():
                churn_category = 'Low'
            elif 'mid churn' in question.lower():
                churn_category = 'Mid'
            elif 'high churn' in question.lower():
                churn_category = 'High'

            location_keywords = ['branch','location', 'state', 'zone']
            location_col = None
            for kw in location_keywords:
                if kw in question_lower:
                    if kw == 'branch' or "location" in question_lower:
                        location_col = 'cleaned_branch_name_2'
                    elif kw == 'state':
                        location_col = 'cleaned_state2'  # replace with actual column if exists
                    elif kw == 'zone':
                        location_col = 'cleaned_zone_2'   # replace with actual column if exists

            # PRIORITY: Specific churn category + location
            # ==== Location-based churn query ====
            if churn_category and location_col:
                table_name = '"stage"."GBM1_prediction_data_with_recommendations"'
                if location_col not in self.schema_info.get("columns", []):
                    return {
                        'answer': f"❌ Column `{location_col}` not found in current table. This question may only work on the churn prediction table.",
                        'success': False,
                        'query': "",
                        'row_count': 0
                    }

                sql_query = f'''
                SELECT "{location_col}" AS location, COUNT(*) AS churn_count
                FROM {table_name}
                WHERE UPPER("churn_category") = UPPER('{churn_category}')
                '''
                if year:
                    sql_query += f' AND "policy_end_date_year" = {year}'
                if month:
                    sql_query += f' AND "policy_end_date_month" = {month}'
                sql_query += f'''
                GROUP BY "{location_col}"
                ORDER BY churn_count DESC
                LIMIT 1;
                '''
                logger.info(f"[Churn by location] Final SQL Query: {sql_query}")
                return sql_query


            

            # If general churn count is requested
            # if 'churn' in question_lower and churn_category is None and 'how many' in question_lower:
            if 'churn' in question_lower and churn_category is None:

                table_name = '"stage"."GBM1_prediction_data_with_recommendations"'
                sql_query = f"""
                SELECT COUNT(*)
                FROM {table_name}
                WHERE "churn_category" IS NOT NULL AND TRIM("churn_category") <> ''
                """
                
                # if year:
                #     sql_query += f" AND \"policy_end_date_year\" = {year}"

                if year:
                    sql_query += f" AND \"policy_end_date_year\" = {year}"
                if month:
                    sql_query += f" AND \"policy_end_date_month\" = {month}"

                # ⬇️ Inject detected location filter
                location_col, location_val = detect_location_filter(question, known_locations)
                if location_col and location_val:
                    sql_query += f" AND UPPER(\"{location_col}\") = UPPER('{location_val}')"

                sql_query += ";"

                logger.info(f"[General churn count] Final SQL Query: {sql_query}")
                return sql_query

            # If a specific churn category is detected
            if churn_category:
                table_name = '"stage"."GBM1_prediction_data_with_recommendations"'
                sql_query = f"""
                SELECT COUNT(*)
                FROM {table_name}
                WHERE UPPER("churn_category") = UPPER('{churn_category}')
                """
                
                if year:
                    sql_query += f" AND \"policy_end_date_year\" = {year}"
                if month:
                    sql_query += f" AND \"policy_end_date_month\" = {month}"
                sql_query += ";"

                logger.info(f"[Specific churn count] Final SQL Query: {sql_query}")
                return sql_query
            
                        # === If only location with year/month is present without churn category ===
            if location_col and (year or month) and churn_category is None:
                table_name = '"stage"."GBM1_prediction_data_with_recommendations"'
                sql_query = f'''
                SELECT "{location_col}" AS location, COUNT(*) AS record_count
                FROM {table_name}
                WHERE "{location_col}" IS NOT NULL
                '''
                if year:
                    sql_query += f' AND "policy_end_date_year" = {year}'
                if month:
                    sql_query += f' AND "policy_end_date_month" = {month}'

                sql_query += f'''
                GROUP BY "{location_col}"
                ORDER BY record_count DESC
                LIMIT 1;
                '''
                logger.info(f"[Location with year/month] Final SQL Query: {sql_query}")
                return sql_query

            

            question_lower = question.lower()

            # Mapping from keyword → actual customer_segment value in DB
            # segment_keywords = {
            #     "elite": "Elite Retainers",
            #     "potential": "Potential Customers",
            #     "loyalist": "Potential Loyalist",
            #     "low value": "Low Value",
            #     "risky": "Risk Segment",
            # }

            # segment_clause = ""
            # for keyword, segment in segment_keywords.items():
            #     if keyword in question_lower:
            #         # Use UPPER() for both sides to satisfy your rule
            #         segment_clause = f'UPPER("customer_segment") = UPPER(\'{segment}\')'
            #         break

            # if segment_clause:
            #     sql_query = f"""
            #     SELECT COUNT(*)
            #     FROM "stage"."GBM1_prediction_data_with_recommendations"
            #     WHERE {segment_clause}
            #     """
            #     if year:
            #         sql_query += f" AND \"policy_end_date_year\" = {year}"
            #     if month:
            #         sql_query += f" AND \"policy_end_date_month\" = {month}"
            #     sql_query += ";"
            # 🔍 Inject dynamic location filter if missing
            location_col, location_values = self.detect_dynamic_location_filter(question)

            if location_col and location_col.lower() not in query.lower():
                if len(location_values) == 1:
                    clause = f'AND UPPER("{location_col}") = UPPER(\'{location_values[0].upper()}\')'
                else:
                    value_list = ", ".join([f"'{val.upper()}'" for val in location_values])
                    clause = f'AND UPPER("{location_col}") IN ({value_list})'

                # Inject into WHERE clause
                if "where" in query.lower():
                    query = query.rstrip(';') + f"\n  {clause};"
                else:
                    query = query.rstrip(';') + f"\nWHERE {clause};"

                logger.info(f"[Location Injected] Column: {location_col}, Values: {location_values}")

            # ✅ Final cleanup and return
            
            
            
            # Apply additional case-insensitive transformations
            query = self.generate_case_insensitive_query(query, question)

            logger.info(f"Final SQL Query: {query}")
            
            return query

def find_matching_column(self, value, value_type="float"):
        sample_row = self.schema_info.get("sample_row", {})
        for column, col_value in sample_row.items():
            try:
                if value_type == "float" and abs(float(col_value) - float(value)) < 0.001:
                    return column
                elif value_type == "str" and str(value).lower() in str(col_value).lower():
                    return column
            except:
                continue
        return None

def generate_high_level_db_summary(self):
        """Generate natural language purpose summary for the database"""
        if not self.is_connected:
            return "No database is currently connected."

        try:
            tables = self.schema_info.get('tables', {})
            if not tables:
                return "No tables found in the connected database."

            summary_parts = []
            summary_parts.append(" **Database High-Level Summary**")

            for table_name, table_info in tables.items():
                columns = table_info.get('columns', {})
                column_names = list(columns.keys())

                table_summary = f"🔹 **Table:** `{table_name}`\n"

                # Convert column list to lowercase string for keyword matching
                col_text = ", ".join(column_names).lower()

                # Assign purpose based on keywords in columns
                if 'policy' in col_text:
                    table_summary += "  - **Purpose:** Contains insurance policy data with churn, tenure, premium, and dates.\n"
                elif 'customer' in col_text:
                    table_summary += "  - **Purpose:** Contains customer details, segments, and retention info.\n"
                elif 'claim' in col_text:
                    table_summary += "  - **Purpose:** Contains insurance claims data with approvals and reasons.\n"
                elif 'vehicle' in col_text:
                    table_summary += "  - **Purpose:** Contains vehicle information such as make, model, variant.\n"
                else:
                    table_summary += "  - **Purpose:** Contains general business operational data.\n"

                summary_parts.append(table_summary)

            # Append overall database purpose
            summary_parts.append("\n✅ **This database is used for:**")
            summary_parts.append("- Insurance policy and claims management")
            summary_parts.append("- Churn prediction and customer retention analytics")
            summary_parts.append("- Customer profiling and segmentation")
            summary_parts.append("- Vehicle and premium data analysis")

            # Return final combined summary
            return "\n".join(summary_parts)

        except Exception as e:
            return f"An error occurred while generating the database summary: {str(e)}"


    
def get_openrouter_response(self, prompt, system_prompt=""):
            """Get response from Groq Cloud using Llama4 Maverick with rate-limit retry and safe truncation."""
            import re
            import time
            import requests
            from loguru import logger  # assuming logger is set up

            try:
                headers = {
                    "Authorization": f"Bearer {GROQ_API_KEY}",
                    "Content-Type": "application/json"
                }

                # Combine prompts and check total token size
                full_prompt = system_prompt + "\n\n" + prompt
                word_limit = 5000
                if len(full_prompt.split()) > word_limit:
                    logger.warning("Prompt too large, truncating...")
                    prompt = prompt[:8000]  # Truncate prompt portion
                    system_prompt = system_prompt[:2000]  # Truncate system message

                messages = []
                if system_prompt:
                    messages.append({"role": "system", "content": system_prompt})
                messages.append({"role": "user", "content": prompt})

                payload = {
                    "model": "meta-llama/llama-4-maverick-17b-128e-instruct",
                    "messages": messages,
                    "temperature": 0.1,
                    "max_tokens": 1000
                }

                GROQ_BASE_URL = "https://api.groq.com/openai/v1"

                max_retries = 3
                for attempt in range(max_retries):
                    response = requests.post(
                        f"{GROQ_BASE_URL}/chat/completions",
                        headers=headers,
                        json=payload,
                        timeout=30
                    )

                    if response.status_code == 200:
                        return response.json()["choices"][0]["message"]["content"].strip()

                    elif response.status_code == 429:
                        try:
                            error_data = response.json()
                            message = error_data.get("error", {}).get("message", "")
                            logger.warning(f"Rate limit hit: {message}")
                            match = re.search(r'try again in ([\\d\\.]+)s', message)
                            if match:
                                wait_time = float(match.group(1))
                                logger.info(f"Sleeping for {wait_time} seconds due to rate limit...")
                                time.sleep(wait_time)
                            else:
                                time.sleep(2 ** attempt)
                            continue
                        except Exception as parse_error:
                            logger.error(f"Error parsing rate limit retry time: {parse_error}")
                            time.sleep(2 ** attempt)
                            continue

                    else:
                        logger.error(f"Groq Cloud API error: {response.status_code} - {response.text}")
                        break

                return "I apologize, but I'm having trouble processing your request right now."

            except Exception as e:
                logger.error(f"Groq Cloud API call failed: {e}")
                return "I apologize, but I'm having trouble processing your request right now."



# import json
# import uuid
# import traceback
# from datetime import datetime
# from django.http import JsonResponse
# from django.views.decorators.csrf import csrf_exempt
# from sqlalchemy import create_engine, text
# from .analyzer import PostgreSQLChatAnalyzer  # ✅ adjust path as needed
# from .config import DB_Conv_AI  # ✅ your DB credentials dictionary
# from .stores import session_store, conversation_memory_store  # ✅ session memory globals

# def call_llm_with_retry(prompt: str, max_retries: int = 3, base_delay: float = 1.0) -> str:
#     if not GROQ_API_KEY:
#         logger.error("GROQ_API_KEY is not defined")
#         return "API key configuration error. Please check your settings."

#     estimated_tokens = len(prompt.split()) * 1.5
#     logger.info(f"Estimated prompt tokens: {estimated_tokens}")

#     if estimated_tokens > 220000:
#         logger.error(f"Prompt too long: {estimated_tokens} tokens (max: 220000)")
#         return "The prompt is too long for the current model. Please try with a shorter question."

#     headers = {
#         "Authorization": f"Bearer {GROQ_API_KEY}",
#         "Content-Type": "application/json",
#     }

#     model = "meta-llama/llama-4-maverick-17b-128e-instruct"

#     logger.info(f"Starting Groq LLM call with {max_retries} max retries")

#     for attempt in range(max_retries):
#         try:
#             logger.info(f"Attempt {attempt + 1}/{max_retries}")
#             payload = {
#                 "model": model,
#                 "messages": [{"role": "user", "content": prompt}],
#                 "temperature": 0.1,
#                 "top_p": 0.9,
#                 "frequency_penalty": 0.0,
#                 "presence_penalty": 0.0,
#                 "max_tokens": 4000
#             }
#             response = requests.post("https://api.groq.com/openai/v1/chat/completions", headers=headers, json=payload, timeout=60)
#             logger.info(f"Response status code: {response.status_code}")

#             if response.status_code == 200:
#                 result = response.json()
#                 if 'choices' in result and result['choices']:
#                     answer = result['choices'][0].get('message', {}).get('content', '').strip()
#                     if answer:
#                         return answer
#             elif response.status_code in [429, 502, 503, 504]:
#                 time.sleep(base_delay * (2 ** attempt))
#                 continue

#         except Exception as e:
#             logger.warning(f"Retry {attempt+1} failed: {e}")
#             time.sleep(base_delay)

#     logger.error("All retry attempts failed")
#     return "I'm currently experiencing issues reaching the AI service. Please try again later."

# @csrf_exempt
# def ask_qwen(request):
#     try:
#         data = json.loads(request.body)
#         session_id = data.get("session_id", "").strip()
#         question = data.get("question", "").strip()

#         if not question:
#             return JsonResponse({'error': 'Question is required'}, status=400)
#         if not session_id:
#             return JsonResponse({'error': 'Session ID is required'}, status=400)

#         logger.info(f"Processing question for session {session_id}: {question[:100]}")

#         df = dataframe_map.get(session_id)
#         has_data = df is not None

#         if has_data:
#             TOKEN_BUDGETS = {
#                 'question': len(question.split()) * 1.5,
#                 'data_overview': 800,
#                 'memory_context': 1000,
#                 'semantic_context': 4000,
#                 'prompt_template': 1000,
#                 'response_buffer': 8000,
#                 'safety_margin': 2000
#             }

#             semantic_context = ""
#             memory_context = ""
#             data_overview = ""

#             prompt_parts = [
#                 "You are an expert data analyst AI that provides accurate answers based on uploaded datasets.",
#                 "",
#                 "IMPORTANT: Use ONLY the provided data. Never make assumptions or use external knowledge.",
#                 "",
#                 "### Dataset Overview:",
#                 data_overview,
#                 "",
#                 "### Previous Conversation Context:",
#                 memory_context or "No previous context.",
#                 "",
#                 "### Relevant Data Chunks:",
#                 semantic_context or "No relevant chunks found.",
#                 "",
#                 f"### User Question:",
#                 question,
#                 "",
#                 "### Instructions:",
#                 "1. Analyze the question carefully",
#                 "2. Use only the provided data chunks and dataset information",
#                 "3. If you need to perform calculations, show your work",
#                 "4. If the data doesn't contain enough information to answer, say so clearly",
#                 "5. Provide specific numbers, values, and examples from the actual data",
#                 "6. Be precise and factual - no guessing or assumptions",
#                 "7. Keep your answer concise but comprehensive",
#                 "",
#                 "Answer:"
#             ]

#             prompt = "\n".join(prompt_parts)
#             prompt_tokens = len(prompt.split()) * 1.5
#             logger.info(f"Final prompt: {prompt_tokens} tokens")

#             answer = call_llm_with_retry(prompt)
#             if not answer:
#                 answer = "I couldn't find an answer based on the uploaded data."

#             return JsonResponse({
#                 "question": question,
#                 "answer": answer,
#                 "session_id": session_id,
#                 "chunks_used": 0,
#                 "prompt_tokens": prompt_tokens,
#                 "timestamp": datetime.now().isoformat(),
#                 "success": True,
#                 "has_chart": False
#             })

#         fallback_prompt = f"""You are an intelligent assistant. Answer the following question as accurately and helpfully as possible.

# Question: {question}

# Answer:"""

#         answer = call_llm_with_retry(fallback_prompt)
#         if not answer:
#             answer = "Sorry, I couldn't generate a helpful answer. Please try rephrasing."

#         return JsonResponse({
#             "question": question,
#             "answer": answer,
#             "session_id": session_id,
#             "chunks_used": 0,
#             "timestamp": datetime.now().isoformat(),
#             "success": True,
#             "has_chart": False
#         })

#     except Exception as e:
#         logger.error(f"Exception in ask_qwen: {traceback.format_exc()}")
#         return JsonResponse({
#             'error': f'Processing error: {str(e)}',
#             'success': False,
#             'timestamp': datetime.now().isoformat(),
#             'has_chart': False
#         }, status=500)

# @csrf_exempt
# def ask_question(request):
#     """Question processing endpoint"""
    
#     try:
#         data = json.loads(request.body)
#         session_id = data.get("session_id")
#         question = data.get("question", "").strip()
        
#         logger.info(f"Question received: {question}")
        
#         if not question:
#             return JsonResponse({"error": "Missing question"}, status=400)
        
#         # Get or create analyzer
#         if session_id and session_id in session_store:
#             analyzer = session_store[session_id]
#             logger.info(f"Retrieved analyzer for session {session_id}. Connected: {analyzer.is_connected}")
#             if session_id not in conversation_memory_store:
#                conversation_memory_store[session_id] = []
#         else:
#             logger.warning(f"Session {session_id} not found. Reconnecting to DB...")

#             # Create engine from DB config
#             db_url = f"postgresql+psycopg2://{DB_Conv_AI['user']}:{DB_Conv_AI['password']}@{DB_Conv_AI['host']}:{DB_Conv_AI['port']}/{DB_Conv_AI['dbname']}"
#             engine = create_engine(db_url)

#             # Create analyzer instance
#             analyzer = PostgreSQLChatAnalyzer(engine=engine, schema_info={})
#             session_id = str(uuid.uuid4())
#             session_store[session_id] = analyzer
#             conversation_memory_store[session_id] = []

#         # Get session history
#         conversation_history = conversation_memory_store.get(session_id, [])

#         # Process the question
#         start_time = datetime.now()
#         result = analyzer.process_question(question, conversation_history)
#         end_time = datetime.now()

#         total_time = (end_time - start_time).total_seconds()

#         sql_query = result.get("query")
#         rows_data = []

#         if analyzer.engine and sql_query:
#             try:
#                 with analyzer.engine.connect() as conn:
#                     db_result = conn.execute(text(sql_query))
#                     columns = db_result.keys()
#                     fetched_rows = db_result.fetchall()
#                     rows_data = [dict(zip(columns, row)) for row in fetched_rows]
#             except Exception as db_err:
#                 logger.error(f"DB execution error: {db_err}")
#                 logger.error(traceback.format_exc())

#         # Save memory
#         conversation_entry = {
#             "question": question,
#             "answer": result.get("answer", ""),
#             "query": sql_query,
#             "timestamp": datetime.now().isoformat(),
#             "success": result.get("success", False)
#         }
#         conversation_memory_store[session_id].append(conversation_entry)

#         # Keep only recent 5
#         if len(conversation_memory_store[session_id]) > 5:
#             conversation_memory_store[session_id] = conversation_memory_store[session_id][-5:]

#         return JsonResponse({
#             "answer": result.get("answer", ""),
#             "success": result.get("success", False),
#             "query_used": sql_query,
#             "rows": rows_data,
#             "row_count": len(rows_data),
#             "response_time": f"{total_time:.2f}s",
#             "session_id": session_id
#         })

#     except Exception as e:
#         logger.error(f"Question processing error: {e}")
#         logger.error(traceback.format_exc())  # 🔍 shows exact line of failure
#         return JsonResponse({"error": str(e)}, status=500)



# @csrf_exempt
# def ask_questionbot(request):
#     if request.method != 'POST':
#         return JsonResponse({'error': 'Only POST allowed'}, status=405)

#     data = json.loads(request.body)
#     question = data.get("query")
#     print(f"Received question: {question}")

#     vectordb = get_best_vectorstore(question)
#     if not vectordb:
#         return JsonResponse({'answer': "Sorry, the knowledge base is not available. Please contact admin."})

    

#     retriever = vectordb.as_retriever(search_type="similarity", k=5)

#     llm = ChatOpenAI(
#         model="google/gemma-3-27b-it:free",  
#         openai_api_key=OPENROUTER_API_KEY,
#         openai_api_base="https://openrouter.ai/api/v1",  
#         temperature=0,
#         max_tokens=1024
#     )

#     custom_prompt = PromptTemplate(
#         input_variables=["context", "question"],
#         template="""
#     You are an expert AI assistant. Use the below context to answer the user's question.

#     IMPORTANT:
#     - Do NOT start your answer with phrases like 'Based on the provided text' or 'According to the text'.
#     - Answer directly in a clear, confident, and natural tone.

#     Context:
#     {context}

#     Question:
#     {question}

#     Answer:
#     """
#     )

#     qa_chain = RetrievalQA.from_chain_type(
#         llm=llm,
#         retriever=retriever,
#         chain_type_kwargs={"prompt": custom_prompt}
#     )

#     print("Generating answer using LLM...")
#     result = qa_chain.invoke({"query": question})
#     answer = result if isinstance(result, str) else result.get("result", "")

#     print(f"Answer generated: {answer}")
#     return JsonResponse({'answer': answer})



import json, requests
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from langchain_openai import ChatOpenAI
from langchain.prompts import PromptTemplate
from langchain.chains import RetrievalQA  # NEW
from langchain.retrievers import ContextualCompressionRetriever  # NEW
from langchain.retrievers.document_compressors import CrossEncoderReranker  # NEW

# # === Endpoint for any user to ask questions ===
# @csrf_exempt
# def ask_questionbot(request):
#     if request.method != 'POST':
#         return JsonResponse({'error': 'Only POST allowed'}, status=405)

#     # robust JSON parsing
#     try:
#         data = json.loads((request.body or b"{}").decode("utf-8"))
#     except Exception:
#         return JsonResponse({'answer': "Invalid JSON body."}, status=400)

#     question = str(data.get("query") or "").strip()
#     print(f"Received question: {question}")
#     if not question:
#         return JsonResponse({'answer': "Please type a question."})

#     vectordb = get_best_vectorstore(question)
#     if not vectordb:
#         return JsonResponse({'answer': "Sorry, the knowledge base is not available. Please contact admin."})

#     # ---------- NEW: probe KB first with a threshold ----------
#     try:
#         probe = vectordb.as_retriever(
#             search_type="similarity_score_threshold",
#             search_kwargs={"k": 5, "score_threshold": 0.30}  # soft threshold
#         )
#         try:
#             probe_docs = probe.invoke(question)  # new LC API
#         except Exception:
#             probe_docs = probe.get_relevant_documents(question)  # backwards compat
#     except Exception:
#         probe_docs = []

#     # ---------- NEW: if nothing relevant in KB, answer generally ----------
#     if not probe_docs:
#         try:
#             general_llm = ChatOpenAI(
#                 model="google/gemma-3-27b-it:free",
#                 openai_api_key=OPENROUTER_API_KEY,
#                 openai_api_base="https://openrouter.ai/api/v1",
#                 temperature=0,
#                 max_tokens=256
#             )
#             general_prompt = (
#                 "You are a helpful assistant. Answer the user's question clearly in 2–4 sentences. "
#                 "Provide a complete, standalone answer (not a fragment). "
#                 "Do NOT mention Prochurn/ProSync or documentation unless the user asked about it.\n\n"
#                 f"Question:\n{question}\n\nAnswer:"
#             )
#             general_answer = general_llm.invoke(general_prompt).content
#             return JsonResponse({'answer': general_answer})
#         except Exception:
#             # if general generation fails, fall through to RAG as a safe default
#             pass

#     # ---------- YOUR ORIGINAL RAG CODE (UNCHANGED) ----------
#     # retriever = vectordb.as_retriever(search_type="similarity", k=5)

#     # llm = ChatOpenAI(
#     #     # model="meta-llama/llama-4-maverick:free",  
#     #     model="google/gemma-3-27b-it:free",  
#     #     openai_api_key=OPENROUTER_API_KEY,
#     #     openai_api_base="https://openrouter.ai/api/v1",  
#     #     temperature=0,
#     #     max_tokens=1024
#     # )

#     # qa_chain = RetrievalQA.from_chain_type(
#     #     llm=llm,
#     #     retriever=retriever,
#     # )

#     # print("Generating answer using LLM...")
#     # result = qa_chain.invoke({"query": question})
#     # answer = result if isinstance(result, str) else result.get("result", "")
#     # print(f"Answer generated: {answer}")
#     # return JsonResponse({'answer': answer})

#     retriever = vectordb.as_retriever(search_type="similarity", k=5)

#     # ===== NEW: Fix reranker wiring (no pydantic error) with safe fallbacks =====
#     try:
#         # HuggingFaceCrossEncoder instance is required (not a string)
#         from langchain_community.cross_encoders import HuggingFaceCrossEncoder
#         cross_encoder = HuggingFaceCrossEncoder("BAAI/bge-reranker-large")
#         reranker = CrossEncoderReranker(model=cross_encoder, top_n=3)
#         retriever = ContextualCompressionRetriever(
#             base_retriever=retriever,
#             compressor=reranker
#         )
#     except Exception as e:
#         print("Reranker unavailable, switching to embeddings filter:", e)
#         try:
#             # Lightweight fallback that removes off-topic text using your vectorstore's embedder
#             from langchain_community.document_transformers import EmbeddingsFilter
#             embedder = getattr(vectordb, "_embedding_function", None)
#             if embedder is not None:
#                 ef = EmbeddingsFilter(embeddings=embedder, similarity_threshold=0.60)
#                 retriever = ContextualCompressionRetriever(
#                     base_retriever=retriever,
#                     compressor=ef
#                 )
#             else:
#                 print("No embedder found on vectorstore; using base retriever.")
#         except Exception as e2:
#             print("EmbeddingsFilter unavailable; using base retriever:", e2)

#     llm = ChatOpenAI(
#         model="google/gemma-3-27b-it:free",  
#         openai_api_key=OPENROUTER_API_KEY,
#         openai_api_base="https://openrouter.ai/api/v1",  
#         temperature=0,
#         max_tokens=1024
#     )

#     custom_prompt = PromptTemplate(
#         input_variables=["context", "question"],
#         template="""
#     You are an expert AI assistant. Use the below context to answer the user's question.

#     IMPORTANT:
#     - Answer ONLY what was asked; ignore unrelated sections from the context.
#     - If the question asks "what is/define/explain X", reply with a concise 1–3 sentence definition that starts with "X is ...".
#     - Do NOT include headings or long lists unless explicitly requested.
#     - If context is insufficient, say so briefly and ask for specifics.
#     - Do NOT start with phrases like 'Based on the provided text' or 'According to the text'.

#     Context:
#     {context}

#     Question:
#     {question}

#     Answer (max 80 words):
#     """
#     )

#     qa_chain = RetrievalQA.from_chain_type(
#         llm=llm,
#         retriever=retriever,
#         chain_type_kwargs={"prompt": custom_prompt}
#     )

#     print("Generating answer using LLM...")
#     result = qa_chain.invoke({"query": question})
#     answer = result if isinstance(result, str) else result.get("result", "")

#     return JsonResponse({'answer': answer})


# STOPWORDS = {
#     "the","is","are","a","an","of","on","in","for","to","and","or","with",
#     "what","which","who","whom","when","why","how","explain","define","tell",
#     "about","give","me","us","you","i"
# }

# def should_use_pdf_context(question: str, vectordb, k: int = 3, min_overlap_ratio: float = 0.25) -> bool:
#     """
#     Decide whether to use the PDF/vectorstore context.

#     Method:
#     - Pull top-k chunks from the vector store.
#     - Compute overlap of meaningful query tokens with those chunks.
#     - If overlap ratio >= threshold, treat as document-grounded; else general answer.
#     """
#     if not question or not vectordb:
#         return False

#     try:
#         top_docs = vectordb.similarity_search(question, k=k)
#     except Exception as e:
#         # Fail open to your original PDF path if retrieval errors
#         print(f"[intent] similarity_search error -> default to PDF path: {e}")
#         return True

#     joined_text = " ".join([getattr(d, "page_content", "") for d in top_docs]).lower()

#     q_terms = {
#         t for t in re.findall(r"[a-zA-Z]+", (question or "").lower())
#         if len(t) >= 3 and t not in STOPWORDS
#     }
#     if not q_terms:
#         return False

#     matches = sum(1 for t in q_terms if re.search(rf"\b{re.escape(t)}\b", joined_text))
#     overlap_ratio = matches / max(1, len(q_terms))
#     print(f"[intent] terms={len(q_terms)} matches={matches} ratio={overlap_ratio:.2f}")

#     return overlap_ratio >= min_overlap_ratio
# # ---------------------------------------------------------------------------


# @csrf_exempt
# def ask_questionbot(request):
#     if request.method != 'POST':
#         return JsonResponse({'error': 'Only POST allowed'}, status=405)

#     data = json.loads(request.body)
#     question = data.get("query")
#     print(f"Received question: {question}")

#     if question.lower() in ["what is your name", "what's your name", "who are you"]:
#         return JsonResponse({'answer': "My name is Sara, your AI assistant. How can i help you today?"})

#     if question.lower() in ["what is your purpose", "what's your purpose"]:
#         return JsonResponse({'answer': "My purpose is to help you with your queries and provide you with the best possible answers."})

#     vectordb = get_best_vectorstore(question)
#     if not vectordb:
#         return JsonResponse({'answer': "Sorry, the knowledge base is not available. Please contact admin."})

    

#     retriever = vectordb.as_retriever(search_type="similarity", k=5)

#     llm = ChatOpenAI(
#         model="google/gemma-3-27b-it:free",  
#         openai_api_key=OPENROUTER_API_KEY,
#         openai_api_base="https://openrouter.ai/api/v1",  
#         temperature=0,
#         max_tokens=1024
#     )

#     custom_prompt = PromptTemplate(
#         input_variables=["context", "question"],
#         template="""
#     You are an expert AI assistant. Use the below context to answer the user's question.

#     IMPORTANT:
#     - Do NOT start your answer with phrases like 'Based on the provided text', 'The text describes' or 'According to the text'.
#     - Answer directly in a clear, confident, and natural tone.

#     Context:
#     {context}

#     Question:
#     {question}

#     Answer:
#     """
#     )

#     qa_chain = RetrievalQA.from_chain_type(
#         llm=llm,
#         retriever=retriever,
#         chain_type_kwargs={"prompt": custom_prompt}
#     )

#     # -------------------- NEW: intent gate (no deletions to your code) --------------------
#     try:
#         use_pdf = should_use_pdf_context(question, vectordb)
#     except Exception as e:
#         print(f"[intent] heuristic failed; default to PDF path: {e}")
#         use_pdf = True

#     if not use_pdf:
#         print("[intent] Route: general (no PDF context)")
#         general_prompt = PromptTemplate(
#             input_variables=["question"],
#             template=(
#                 "Answer clearly and concisely.\n\n"
#                 "If the user greets (e.g., hi/hello/hey), reply with a short friendly greeting only.\n"
#                 "Question:\n{question}\n\n"
#                 "Answer:"
#             ),
#         )
#         # Use same LLM; no retrieval
#         general_chain = general_prompt | llm
#         general_msg = general_chain.invoke({"question": question})
#         general_answer = getattr(general_msg, "content", str(general_msg))
#         print(f"General answer generated.")
#         return JsonResponse({'answer': general_answer})
#     else:
#         print("[intent] Route: PDF/vector context")
#     # --------------------------------------------------------------------------------------

#     print("Generating answer using LLM...")
#     result = qa_chain.invoke({"query": question})
#     answer = result if isinstance(result, str) else result.get("result", "")

#     print(f"Answer generated: {answer}")
#     return JsonResponse({'answer': answer})


STOPWORDS = {
    "the","is","are","a","an","of","on","in","for","to","and","or","with",
    "what","which","who","whom","when","why","how","explain","define","tell",
    "about","give","me","us","you","i"
}

def should_use_pdf_context(question: str, vectordb, k: int = 3, min_overlap_ratio: float = 0.25) -> bool:
    """
    Decide whether to use the PDF/vectorstore context.

    Method:
    - Pull top-k chunks from the vector store.
    - Compute overlap of meaningful query tokens with those chunks.
    - If overlap ratio >= threshold, treat as document-grounded; else general answer.
    """
    if not question or not vectordb:
        return False

    try:
        top_docs = vectordb.similarity_search(question, k=k)
    except Exception as e:
        # Fail open to your original PDF path if retrieval errors
        print(f"[intent] similarity_search error -> default to PDF path: {e}")
        return True

    joined_text = " ".join([getattr(d, "page_content", "") for d in top_docs]).lower()

    q_terms = {
        t for t in re.findall(r"[a-zA-Z]+", (question or "").lower())
        if len(t) >= 3 and t not in STOPWORDS
    }
    if not q_terms:
        return False

    matches = sum(1 for t in q_terms if re.search(rf"\b{re.escape(t)}\b", joined_text))
    overlap_ratio = matches / max(1, len(q_terms))
    print(f"[intent] terms={len(q_terms)} matches={matches} ratio={overlap_ratio:.2f}")

    return overlap_ratio >= min_overlap_ratio
# ---------------------------------------------------------------------------


# @csrf_exempt
# def ask_questionbot(request):
#     if request.method != 'POST':
#         return JsonResponse({'error': 'Only POST allowed'}, status=405)

#     data = json.loads(request.body)
#     question = data.get("query")
#     print(f"Received question: {question}")

#     if question.lower() in ["what is your name", "what's your name", "who are you"]:
#         return JsonResponse({'answer': "My name is Sara, your AI assistant. How can i help you today?"})

#     if question.lower() in ["what is your purpose", "what's your purpose"]:
#         return JsonResponse({'answer': "My purpose is to help you with your queries and provide you with the best possible answers."})

#     vectordb = get_best_vectorstore(question)
#     if not vectordb:
#         return JsonResponse({'answer': "Sorry, the knowledge base is not available. Please contact admin."})

    

#     retriever = vectordb.as_retriever(search_type="similarity", k=5)

#     llm = ChatOpenAI(
#         model="google/gemma-3-27b-it:free",  
#         openai_api_key=OPENROUTER_API_KEY,
#         openai_api_base="https://openrouter.ai/api/v1",  
#         temperature=0,
#         max_tokens=1024
#     )

#     custom_prompt = PromptTemplate(
#         input_variables=["context", "question"],
#         template="""
#     You are an expert AI assistant. Use the below context to answer the user's question.

#     IMPORTANT:
#     - Do NOT start your answer with phrases like 'Based on the provided text', 'The text describes' or 'According to the text'.
#     - Answer directly in a clear, confident, and natural tone.
#     - Respond using BULLET POINTS ONLY:
#             - No paragraphs, intros, or conclusions.
        

#     Context:
#     {context}

#     Question:
#     {question}

#     Answer:
#     """
#     )

#     qa_chain = RetrievalQA.from_chain_type(
#         llm=llm,
#         retriever=retriever,
#         chain_type_kwargs={"prompt": custom_prompt}
#     )

#     # -------------------- NEW: intent gate (no deletions to your code) --------------------
#     try:
#         use_pdf = should_use_pdf_context(question, vectordb)
#     except Exception as e:
#         print(f"[intent] heuristic failed; default to PDF path: {e}")
#         use_pdf = True

#     if not use_pdf:
#         print("[intent] Route: general (no PDF context)")
#         general_prompt = PromptTemplate(
#             input_variables=["question"],
#             template=(
#                 "Answer clearly and concisely.\n\n"
#                 "If the user greets (e.g., hi/hello/hey), reply with a short friendly greeting only.\n"
#                 "Question:\n{question}\n\n"
#                 "Answer:"
#             ),
#         )
#         # Use same LLM; no retrieval
#         general_chain = general_prompt | llm
#         general_msg = general_chain.invoke({"question": question})
#         general_answer = getattr(general_msg, "content", str(general_msg))
#         print(f"General answer generated.")
#         return JsonResponse({'answer': general_answer})
#     else:
#         print("[intent] Route: PDF/vector context")
#     # --------------------------------------------------------------------------------------

#     print("Generating answer using LLM...")
#     result = qa_chain.invoke({"query": question})
#     answer = result if isinstance(result, str) else result.get("result", "")

#     print(f"Answer generated: {answer}")
#     return JsonResponse({'answer': answer})


def build_azure_llm(temperature: float = 0, max_tokens: int = 1024):
    """
    Returns a LangChain chat model backed by Azure AI Model Inference.
    Endpoint must point to .../models (serverless API). Model name is passed in calls.
    """
    if not AZURE_ENDPOINT or not AZURE_API_KEY:
        raise RuntimeError("Missing AZURE_INFERENCE_ENDPOINT or AZURE_INFERENCE_API_KEY")

    # Cap max_tokens by optional env safety limit (generation tokens, not context)
    gen_tokens = min(max_tokens, MAX_PROMPT_TOKENS)

    llm = AzureAIChatCompletionsModel(
        endpoint=AZURE_ENDPOINT,            # e.g., https://<resource>.services.ai.azure.com/models
        credential=AZURE_API_KEY,           # key auth; Entra ID also supported if you pass a TokenCredential
        api_version=AZURE_API_VERSION,      # 2024-05-01-preview per your config
        model=AZURE_MODEL,                  # deployment/model id (your custom Llama-4 Maverick)
        temperature=temperature,
        max_tokens=gen_tokens,
    )
    return llm

# ------------------------------------------------------------------------------
# Your existing helpers are assumed available in your project:
# - get_best_vectorstore(question)  -> returns a vector store or None
# - should_use_pdf_context(question, vectordb) -> bool heuristic
# ------------------------------------------------------------------------------


@csrf_exempt
def ask_questionbot(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Only POST allowed'}, status=405)

    data = json.loads(request.body)
    question = data.get("query")
    print(f"Received question: {question}")

    # Short-circuit small-talk
    if question and question.lower() in ["what is your name", "what's your name", "who are you"]:
        return JsonResponse({'answer': "My name is Sara, your AI assistant. How can i help you today?"})

    if question and question.lower() in ["what is your purpose", "what's your purpose"]:
        return JsonResponse({'answer': "My purpose is to help you with your queries and provide you with the best possible answers."})

    # Retrieve best KB (kept as-is)
    vectordb = get_best_vectorstore(question)
    if not vectordb:
        return JsonResponse({'answer': "Sorry, the knowledge base is not available. Please contact admin."})

    retriever = vectordb.as_retriever(search_type="similarity", k=5)

    # ==================== LLM INIT (Azure AI Inference) ====================
    # Replaces previous OpenRouter/Gemma client; no other logic changed.
    llm = build_azure_llm(temperature=0, max_tokens=1024)
    # ======================================================================

    # Your custom RAG prompt (unchanged)
    custom_prompt = PromptTemplate(
        input_variables=["context", "question"],
        template="""
     You are an expert AI assistant. Use the below context to answer the user's question.

    IMPORTANT:
    - Do NOT start your answer with phrases like 'Based on the provided text', 'The text describes' or 'According to the text'.
    - Answer directly in a clear, confident, and natural tone.
    - Respond using BULLET POINTS ONLY:
        - No paragraphs, intros, or conclusions.

    Context:
    {context}

    Question:
    {question}

    Answer:
    """
    )

    qa_chain = RetrievalQA.from_chain_type(
        llm=llm,
        retriever=retriever,
        chain_type_kwargs={"prompt": custom_prompt}
    )

    # -------------------- NEW: intent gate (no deletions to your code) --------------------
    try:
        use_pdf = should_use_pdf_context(question, vectordb)
    except Exception as e:
        print(f"[intent] heuristic failed; default to PDF path: {e}")
        use_pdf = True

    if not use_pdf:
        print("[intent] Route: general (no PDF context)")
        general_prompt = PromptTemplate(
            input_variables=["question"],
            template=(
                "Answer clearly and concisely.\n\n"
                "If the user greets (e.g., hi/hello/hey), reply with a short friendly greeting only.\n"
                "Question:\n{question}\n\n"
                "Answer:"
            ),
        )
        # Use same Azure LLM; no retrieval
        general_chain = general_prompt | llm
        general_msg = general_chain.invoke({"question": question})
        general_answer = getattr(general_msg, "content", str(general_msg))
        print("General answer generated.")
        return JsonResponse({'answer': general_answer})
    else:
        print("[intent] Route: PDF/vector context]")
    # --------------------------------------------------------------------------------------

    print("Generating answer using LLM...")
    result = qa_chain.invoke({"query": question})
    answer = result if isinstance(result, str) else result.get("result", "")

    print(f"Answer generated: {answer}")
    return JsonResponse({'answer': answer})


@api_view(['POST'])
def trigger_azure_import(request):
    try:
        # Get parameters from request
        connection_string = request.data.get('connection_string')
        container_name = request.data.get('container_name')
        selected_files = request.data.get('selected_files', [])
        schema_name = request.data.get('schema_name', 'stage')
        sensitive_columns = request.data.get('sensitive_columns', [])

        if not connection_string or not container_name:
            return Response({"error": "connection_string and container_name are required"}, status=400)

        # Store parameters in Airflow Variables (temporary storage)
        Variable.set("azure_connection_string", connection_string)
        Variable.set("azure_container_name", container_name)
        Variable.set("selected_files", json.dumps(selected_files))
        Variable.set("schema_name", schema_name)
        Variable.set("sensitive_columns", json.dumps(sensitive_columns))

        # Trigger the DAG
        dag_id = "azure_blob_to_postgres_etl_pg_hook_v3"
        response = requests.post(
            f"{AIRFLOW_API_URL}/dags/{dag_id}/dagRuns",
            auth=AIRFLOW_AUTH,
            json={
                "conf": {
                    "connection_string": connection_string,
                    "container_name": container_name,
                    "selected_files": selected_files,
                    "schema_name": schema_name,
                    "sensitive_columns": sensitive_columns
                }
            }
        )

        if response.status_code not in [200, 201]:
            return Response({"error": f"Failed to trigger DAG: {response.text}"}, status=500)

        return Response({
            "message": "DAG triggered successfully",
            "dag_run_id": response.json().get("dag_run_id")
        })

    except Exception as e:
        return Response({"error": str(e)}, status=500)

@api_view(['GET'])
def check_dag_status(request):
    dag_run_id = request.GET.get('dag_run_id')
    if not dag_run_id:
        return Response({"error": "dag_run_id is required"}, status=400)

    try:
        dag_id = "azure_blob_to_postgres_etl_pg_hook_v3"
        response = requests.get(
            f"{AIRFLOW_API_URL}/dags/{dag_id}/dagRuns/{dag_run_id}",
            auth=AIRFLOW_AUTH
        )

        if response.status_code != 200:
            return Response({"error": f"Failed to get DAG status: {response.text}"}, status=500)

        return Response({
            "state": response.json().get("state"),
            "start_date": response.json().get("start_date"),
            "end_date": response.json().get("end_date")
        })

    except Exception as e:
        return Response({"error": str(e)}, status=500)

@api_view(['POST'])
def list_azure_files(request):
    connection_string = request.data.get('connection_string')
    container_name = request.data.get('container_name')

    if not connection_string or not container_name:
        return JsonResponse({"error": "connection_string and container_name are required"}, status=400)

    try:
        blob_service_client = BlobServiceClient.from_connection_string(connection_string)
        container_client = blob_service_client.get_container_client(container_name)
        blob_list = [blob.name for blob in container_client.list_blobs()]
        return JsonResponse({"files": blob_list})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

class BusinessMetricsView(APIView):
    def get(self, request):
        try:
            # === DB connection ===
            conn = psycopg2.connect(
                dbname=DB_New_Past["dbname"],
                user=DB_New_Past["user"],
                password=DB_New_Past["password"],
                host=DB_New_Past["host"],
                port=DB_New_Past["port"]
            )

            with conn.cursor() as cursor:
                cursor.execute("""SELECT "total_churned", "total_policies", "revenue_loss", "avg_tenure" FROM public.churn_summary_mv""")
                churn_summary = cursor.fetchone()

                if churn_summary is None:
                    raise Exception("churn_summary_mv returned no rows!")

                total_churned, total_policies, revenue_loss, avg_tenure = churn_summary
                churn_rate = round((total_churned / total_policies) * 100, 2) if total_policies else 0

                cursor.execute("""SELECT "year", "churned_policies" FROM public.churn_trend_mv ORDER BY "year" """)
                trend = [{"year": str(r[0]), "churnedPolicies": r[1]} for r in cursor.fetchall()]

                cursor.execute("""SELECT * FROM public.top_regions""")
                top_regions = [{"name": r[0], "value": r[1]} for r in cursor.fetchall()]

                cursor.execute("""SELECT * FROM public.top_biztype_mv""")
                top_biztypes = [{"name": r[0], "value": r[1]} for r in cursor.fetchall()]

                cursor.execute("""SELECT * FROM public.top_products""")
                top_products = [{"name": r[0], "value": r[1]} for r in cursor.fetchall()]

                cursor.execute("""SELECT * FROM public.top_vehicle_age_mv""")
                top_ages = [{"name": str(r[0]), "value": r[1]} for r in cursor.fetchall()]

            return Response({
                "totalChurnedPolicies": total_churned,
                "totalPolicies": total_policies,
                "churnRate": churn_rate,
                "revenueLoss": revenue_loss,
                "avgTenure": avg_tenure,
                "yearlyTrend": trend,
                "topRegions": top_regions,
                "topBusinessTypes": top_biztypes,
                "topProducts": top_products,
                "topVehicleAges": top_ages
            })

        except Exception as e:
    
            print("❌ Error in BusinessMetricsView:")
            print(traceback.format_exc())  # Add this for full traceback
            return Response({"error": str(e)}, status=500)








# @api_view(['GET'])
# def reason_summary_api(request):  
#     try:
#         conn = psycopg2.connect( dbname=DB_New_Pred["dbname"],
#             user=DB_New_Pred["user"],
#             password=DB_New_Pred["password"],
#             host=DB_New_Pred["host"],
#             port=DB_New_Pred["port"])  # your DB credentials
#         cursor = conn.cursor()

#         query = """
#             SELECT
#                 main_reason,
#                 COUNT(*) AS policy_count,
#                 SUM(total_premium_payable) AS total_premium
#             FROM (
#                 SELECT DISTINCT policy_no, main_reason, total_premium_payable
#                 FROM "Prediction"."2025_prediction_final_data"
#                 WHERE predicted_status = 'Not Renewed'
#             ) AS sub
#             GROUP BY main_reason
#             ORDER BY policy_count DESC
#         """
#         cursor.execute(query)
#         rows = cursor.fetchall()
#         columns = [desc[0] for desc in cursor.description]
#         reasons = [dict(zip(columns, row)) for row in rows]

#         return Response({'reasons': reasons})

#     except Exception as e:
#         return Response({'error': str(e)}, status=500)


REASON_VIEW_MAP = {
    "Claims Happened":                ("Prediction", "Claims Happened"),
    "High Own-Damage Premium":        ("Prediction", "High Own-Damage Premium"),
    "High Third-Party Premium":       ("Prediction", "High Third-Party Premium"),
    "Low Discount with NCB":          ("Prediction", "Low Discount with NCB"),
    "Low Vehicle IDV":                ("Prediction", "Low Vehicle IDV"),
    "Tie Up with Non-OEM":            ("Prediction", "Tie Up with Non-OEM"),
    "Young Vehicle Age":              ("Prediction", "Young Vehicle Age"),
}

def _ident(*parts: str) -> sql.SQL:
    """Safely quote schema/table/column identifiers as `schema"."table`."""
    return sql.SQL('.').join(sql.Identifier(p) for p in parts)



@api_view(['GET'])
def reason_summary_api(request):
    try:
        conn = psycopg2.connect(
            dbname=DB_New_Pred["dbname"],
            user=DB_New_Pred["user"],
            password=DB_New_Pred["password"],
            host=DB_New_Pred["host"],
            port=DB_New_Pred["port"]
        )
        cur = conn.cursor()

        query = """
            SELECT
                main_reason,
                COUNT(DISTINCT policy_no) AS policy_count,
                SUM(total_premium_payable) AS total_premium
            FROM "Prediction"."2025_prediction_final_data"
            WHERE predicted_status = 'Not Renewed'
            GROUP BY main_reason
            ORDER BY policy_count DESC
        """
        cur.execute(query)
        rows = cur.fetchall()
        columns = [desc[0] for desc in cur.description]

        reasons = [dict(zip(columns, row)) for row in rows]

        cur.close()
        conn.close()

        return Response({"reasons": reasons})

    except Exception as e:
        return Response({"error": str(e)}, status=500)








# @api_view(['GET'])
# def reason_detail_api(request, reason):
#     try:
#         # Connect to DB_New_Pred database
#         conn = psycopg2.connect(
#             dbname=DB_New_Pred["dbname"],
#             user=DB_New_Pred["user"],
#             password=DB_New_Pred["password"],
#             host=DB_New_Pred["host"],
#             port=DB_New_Pred["port"]
#         )
#         cursor = conn.cursor()

#         # Check if table exists
#         table_check_query = """
#             SELECT EXISTS (
#                 SELECT FROM information_schema.tables 
#                 WHERE table_schema = 'Prediction'
#                 AND table_name = '2025_prediction_final_data'
#             )
#         """
#         cursor.execute(table_check_query)
#         table_exists = cursor.fetchone()[0]

#         if not table_exists:
#             return Response({"error": "Table not found in prediction schema"}, status=404)

#         # Get detailed policies for the specific main_reason
#         policies_query = """
#       SELECT 
#         policy_no,
#         product_name,
#         biztype,
#         age,
#         make_clean,
#         vehicle_idv,
#         cleaned_zone_2,
#         tie_up,
#         number_of_claims,
#         customer_tenure,
#         churn_probability,
#         cleaned_branch_name_2,
#         cleaned_state2,
#         cleaned_reg_no,
#         total_premium_payable,
#         corrected_name,
#         policy_end_date_month,
#         predicted_status,
#         policy_end_date_year,
#         policy_tenure,
#         policy_end_date,
#         main_reason,
#         additional_offers,
#         primary_recommendation,
#         top_3_reasons
#        FROM "Prediction"."2025_prediction_final_data"
#        WHERE main_reason = %s AND predicted_status='Not Renewed'
#        ORDER BY total_premium_payable DESC
#     """

#         cursor.execute(policies_query, [reason])

#         columns = [desc[0] for desc in cursor.description]
#         policies = [dict(zip(columns, row)) for row in cursor.fetchall()]

#         cursor.close()
#         conn.close()

#         return Response({
#             'policies': policies,
#             'main_reason': reason
#         })

#     except psycopg2.Error as e:
#         logger.error(f"Database error in reason_detail_api: {str(e)}", exc_info=True)
#         return Response({"error": str(e), "type": "database_error"}, status=500)
#     except Exception as e:
#         logger.error(f"Error in reason_detail_api: {str(e)}", exc_info=True)
#         return Response({"error": str(e), "traceback": traceback.format_exc()}, status=500)


@api_view(['GET'])
def reason_detail_api(request, reason):
    try:
        if reason not in REASON_VIEW_MAP:
            return Response({"error": f"Unknown reason: {reason}"}, status=400)

        schema, view = REASON_VIEW_MAP[reason]

        conn = psycopg2.connect(
            dbname=DB_New_Pred["dbname"],
            user=DB_New_Pred["user"],
            password=DB_New_Pred["password"],
            host=DB_New_Pred["host"],
            port=DB_New_Pred["port"]
        )
        cursor = conn.cursor()

        # Choose the columns you want to expose (must exist in each view)
        cols = [
            "policy_no","product_name","biztype","age","make_clean","vehicle_idv",
            "cleaned_zone_2","tie_up","number_of_claims","customer_tenure",
            "churn_probability","cleaned_branch_name_2","cleaned_state2",
            "cleaned_reg_no","total_premium_payable","corrected_name",
            "policy_end_date_month","predicted_status","policy_end_date_year",
            "policy_tenure","policy_end_date","main_reason",
            "additional_offers","primary_recommendation","top_3_reasons"
        ]

        query = sql.SQL("""
            SELECT {cols}
            FROM {src}
            WHERE "predicted_status" = 'Not Renewed'
            ORDER BY "policy_no"
        """).format(
            cols=sql.SQL(', ').join(sql.Identifier(c) for c in cols),
            src=_ident(schema, view),
        )

        cursor.execute(query)
        rows = cursor.fetchall()
        columns = [d.name for d in cursor.description]

        cursor.close(); conn.close()

        return Response({
            "policies": [dict(zip(columns, r)) for r in rows],
            "main_reason": reason
        })

    except psycopg2.Error as e:
        logger.error(f"Database error in reason_detail_api: {str(e)}", exc_info=True)
        return Response({"error": str(e), "type": "database_error"}, status=500)
    except Exception as e:
        logger.error(f"Error in reason_detail_api: {str(e)}", exc_info=True)
        return Response({"error": str(e), "traceback": traceback.format_exc()}, status=500)




@api_view(['GET'])
def churned_policies_7days(request):
    try:
        # Calculate date range
        #today = datetime.now().date()
        #end_date = today + timedelta(days=7)
         
        today = date(2025, 4, 1)
        end_date = date(2025, 4, 7)
        
        # Connect to DB_New_Pred database
        conn = psycopg2.connect(
            dbname=DB_New_Pred["dbname"],
            user=DB_New_Pred["user"],
            password=DB_New_Pred["password"],
            host=DB_New_Pred["host"],
            port=DB_New_Pred["port"]
        )
        cursor = conn.cursor()
        
        # Get trend data (count by day)
        trend_query = """
            SELECT 
                policy_end_date as date,
                COUNT(policy_no) as count
            FROM "Prediction"."7days_M"
            WHERE policy_end_date BETWEEN %s AND %s AND predicted_status='Not Renewed'
            GROUP BY policy_end_date
            ORDER BY policy_end_date
        """
        cursor.execute(trend_query, [today, end_date])
        table_exists = cursor.fetchone()[0]
        print(f"🔍 Table exists: {table_exists}")  # Debug
        
        if not table_exists:
            return Response({"error": "Table not found in prediction schema"}, status=404)
        trend_data = [
            {"date": row[0], "count": row[1]} 
            for row in cursor.fetchall()
        ]
        
        # Get detailed policy data
        policies_query = """
            SELECT 
                policy_no,
                product_name,
                biztype,
                age,
                make_clean,
                vehicle_idv,
                cleaned_zone_2,
                tie_up,
                number_of_claims,
                customer_tenure,
                churn_probability,
                cleaned_branch_name_2,
                cleaned_state2,
                cleaned_reg_no,
                total_premium_payable,
                corrected_name,
                policy_tenure,
                policy_end_date,
                main_reason,
                top_3_reasons
            FROM "Prediction"."7days_M"
            WHERE policy_end_date BETWEEN %s AND %s AND predicted_status='Not Renewed'
            ORDER BY policy_no
        """
        cursor.execute(policies_query, [today, end_date])
        
        # Get column names from cursor description
        columns = [desc[0] for desc in cursor.description]
        policies = [
            dict(zip(columns, row))
            for row in cursor.fetchall()
        ]
        
        cursor.close()
        conn.close()
        
        # Format trend data for chart
        trend_labels = []
        trend_counts = []
        current_date = today
        
        while current_date <= end_date:
            date_str = current_date.strftime('%Y-%m-%d')
            count = next(
                (item['count'] for item in trend_data 
                 if item['date'] == current_date), 
                0
            )
            trend_labels.append(date_str)
            trend_counts.append(count)
            current_date += timedelta(days=1)
        
        return Response({
            'trend_data': {
                'labels': trend_labels,
                'counts': trend_counts
            },
            'policies': policies
        })
        
    except Exception as e:
        logger.error(f"Error in churned_policies_7days: {str(e)}", exc_info=True)
        return Response({
            'error': str(e),
            'traceback': traceback.format_exc()
        }, status=500)
    except psycopg2.Error as e:
        print(f"❌ Database error: {e}")  # Debug
        return Response({
            "error": str(e),
            "type": "database_error"
        }, status=500)
    except Exception as e:
        print(f"❌ Unexpected error: {e}")  # Debug
        return Response({
            "error": str(e),
            "traceback": traceback.format_exc()
        }, status=500)  
    

@api_view(['GET'])
def churned_policies_30days(request):
    try:
        # Calculate date range
        #today = datetime.now().date()
        #end_date = today + timedelta(days=30)
        today = date(2025, 4, 1)
        end_date = date(2025, 4, 30)
       
        
        # Connect to DB_New_Pred database
        conn = psycopg2.connect(
            dbname=DB_New_Pred["dbname"],
            user=DB_New_Pred["user"],
            password=DB_New_Pred["password"],
            host=DB_New_Pred["host"],
            port=DB_New_Pred["port"]
        )
        cursor = conn.cursor()
        
        # Get trend data (count by day)
        trend_query = """
            SELECT 
                policy_end_date as date,
                COUNT(policy_no) as count
            FROM "Prediction"."30days_M"
            WHERE policy_end_date BETWEEN %s AND %s AND predicted_status='Not Renewed'
            GROUP BY policy_end_date
            ORDER BY policy_end_date
        """
        cursor.execute(trend_query, [today, end_date])
        table_exists = cursor.fetchone()[0]
        print(f"🔍 Table exists: {table_exists}")  # Debug
        
        if not table_exists:
            return Response({"error": "Table not found in prediction schema"}, status=404)
        trend_data = [
            {"date": row[0], "count": row[1]} 
            for row in cursor.fetchall()
        ]
        
        # Get detailed policy data
        policies_query = """
            SELECT 
                policy_no,
                product_name,
                biztype,
                age,
                make_clean,
                vehicle_idv,
                cleaned_zone_2,
                tie_up,
                number_of_claims,
                customer_tenure,
                churn_probability,
                cleaned_branch_name_2,
                cleaned_state2,
                cleaned_reg_no,
                total_premium_payable,
                corrected_name,
                policy_tenure,
                policy_end_date,
                main_reason,
                top_3_reasons
            FROM "Prediction"."30days_M"
            WHERE policy_end_date BETWEEN %s AND %s AND predicted_status='Not Renewed'
            ORDER BY policy_no
        """
        cursor.execute(policies_query, [today, end_date])
        
        # Get column names from cursor description
        columns = [desc[0] for desc in cursor.description]
        policies = [
            dict(zip(columns, row))
            for row in cursor.fetchall()
        ]
        
        cursor.close()
        conn.close()
        
        # Format trend data for chart
        trend_labels = []
        trend_counts = []
        current_date = today
        
        while current_date <= end_date:
            date_str = current_date.strftime('%Y-%m-%d')
            count = next(
                (item['count'] for item in trend_data 
                 if item['date'] == current_date), 
                0
            )
            trend_labels.append(date_str)
            trend_counts.append(count)
            current_date += timedelta(days=1)
        
        return Response({
            'trend_data': {
                'labels': trend_labels,
                'counts': trend_counts
            },
            'policies': policies
        })
        
    except Exception as e:
        logger.error(f"Error in churned_policies_7days: {str(e)}", exc_info=True)
        return Response({
            'error': str(e),
            'traceback': traceback.format_exc()
        }, status=500)
    except psycopg2.Error as e:
        print(f"❌ Database error: {e}")  # Debug
        return Response({
            "error": str(e),
            "type": "database_error"
        }, status=500)
    except Exception as e:
        print(f"❌ Unexpected error: {e}")  # Debug
        return Response({
            "error": str(e),
            "traceback": traceback.format_exc()
        }, status=500)

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def user_details(request):
    print("🧠 Authenticated user:", request.user)
    print("🔑 Token:", request.auth)
    print("🎭 User role:", request.user.role)  # ✅ Added role print

    if not request.user or not request.user.is_authenticated:
        return Response({"detail": "Not authenticated"}, status=401)

    return Response({
        "username": request.user.username,
        "email": request.user.email,
        "role": request.user.role,   
    })



@api_view(["GET"])
def test_env_vars(request):
    return Response({
        "AUTO_DB_USER": os.getenv("AUTO_DB_USER"),
        "AUTO_DB_HOST": os.getenv("AUTO_DB_HOST"),
        "AUTO_DB_PASSWORD": os.getenv("AUTO_DB_PASSWORD"),  
        "Resolved DB_Auto": DB_Auto,
    })

@csrf_exempt
def run_clv_cleaned(request):
    BASE_DIR_PY = os.path.dirname(os.path.abspath(__file__))
    script_path = os.path.join(BASE_DIR_PY, "CLVcleaned.py")

    if not os.path.exists(script_path):
        return JsonResponse({"error": "Reason script not found"}, status=404)

    def generate():
        try:
            print("✅ [BACKEND] Reason Bucket triggered", flush=True)
            # yield "data: [SSE connection started]\n\n"
            process = subprocess.Popen(
                ["python", "-u", script_path],
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                bufsize=1
            )

            for line in iter(process.stdout.readline, ''):
                print(f"[ReasonBucket.py LOG] {line.strip()}", flush=True)
                yield f"data: {line.strip()}\n\n"
                time.sleep(0.1)

            process.stdout.close()
            process.wait()
            print("✅ [BACKEND] Reason Bucket completed", flush=True)

        except Exception as e:
            print(f"❌ [BACKEND ERROR] {str(e)}", file=sys.stderr, flush=True)
            yield f"data: ERROR: {str(e)}\n\n"

    response = StreamingHttpResponse(generate(), content_type='text/event-stream; charset=utf-8')
    response["Cache-Control"] = "no-cache"
    response["X-Accel-Buffering"] = "no"
    response["Access-Control-Allow-Origin"] = "*"
    return response

@csrf_exempt
def run_clv_prediction(request):
    BASE_DIR_PY = os.path.dirname(os.path.abspath(__file__))
    script_path = os.path.join(BASE_DIR_PY, "CLVprediction.py")

    if not os.path.exists(script_path):
        return JsonResponse({"error": "Reason script not found"}, status=404)

    def generate():
        try:
            print("✅ [BACKEND] Reason Bucket triggered", flush=True)
            # yield "data: [SSE connection started]\n\n"
            process = subprocess.Popen(
                ["python", "-u", script_path],
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                bufsize=1
            )

            for line in iter(process.stdout.readline, ''):
                print(f"[ReasonBucket.py LOG] {line.strip()}", flush=True)
                yield f"data: {line.strip()}\n\n"
                time.sleep(0.1)

            process.stdout.close()
            process.wait()
            print("✅ [BACKEND] Reason Bucket completed", flush=True)

        except Exception as e:
            print(f"❌ [BACKEND ERROR] {str(e)}", file=sys.stderr, flush=True)
            yield f"data: ERROR: {str(e)}\n\n"

    response = StreamingHttpResponse(generate(), content_type='text/event-stream; charset=utf-8')
    response["Cache-Control"] = "no-cache"
    response["X-Accel-Buffering"] = "no"
    response["Access-Control-Allow-Origin"] = "*"
    return response

@csrf_exempt
def run_segmentation(request):
    BASE_DIR_PY = os.path.dirname(os.path.abspath(__file__))
    script_path = os.path.join(BASE_DIR_PY, "Segmentation.py")

    if not os.path.exists(script_path):
        return JsonResponse({"error": "Reason script not found"}, status=404)

    def generate():
        try:
            print("✅ [BACKEND] Reason Bucket triggered", flush=True)
            # yield "data: [SSE connection started]\n\n"
            process = subprocess.Popen(
                ["python", "-u", script_path],
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                bufsize=1
            )

            for line in iter(process.stdout.readline, ''):
                print(f"[ReasonBucket.py LOG] {line.strip()}", flush=True)
                yield f"data: {line.strip()}\n\n"
                time.sleep(0.1)

            process.stdout.close()
            process.wait()
            print("✅ [BACKEND] Reason Bucket completed", flush=True)

        except Exception as e:
            print(f"❌ [BACKEND ERROR] {str(e)}", file=sys.stderr, flush=True)
            yield f"data: ERROR: {str(e)}\n\n"

    response = StreamingHttpResponse(generate(), content_type='text/event-stream; charset=utf-8')
    response["Cache-Control"] = "no-cache"
    response["X-Accel-Buffering"] = "no"
    response["Access-Control-Allow-Origin"] = "*"
    return response

@csrf_exempt
def run_reason_bucket(request):
    BASE_DIR_PY = os.path.dirname(os.path.abspath(__file__))
    script_path = os.path.join(BASE_DIR_PY, "ReasonBucket.py")

    if not os.path.exists(script_path):
        return JsonResponse({"error": "Reason script not found"}, status=404)

    def generate():
        try:
            print("✅ [BACKEND] Reason Bucket triggered", flush=True)
            # yield "data: [SSE connection started]\n\n"
            process = subprocess.Popen(
                ["python", "-u", script_path],
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                bufsize=1
            )

            for line in iter(process.stdout.readline, ''):
                print(f"[ReasonBucket.py LOG] {line.strip()}", flush=True)
                yield f"data: {line.strip()}\n\n"
                time.sleep(0.1)

            process.stdout.close()
            process.wait()
            print("✅ [BACKEND] Reason Bucket completed", flush=True)

        except Exception as e:
            print(f"❌ [BACKEND ERROR] {str(e)}", file=sys.stderr, flush=True)
            yield f"data: ERROR: {str(e)}\n\n"

    response = StreamingHttpResponse(generate(), content_type='text/event-stream; charset=utf-8')
    response["Cache-Control"] = "no-cache"
    response["X-Accel-Buffering"] = "no"
    response["Access-Control-Allow-Origin"] = "*"
    return response

@csrf_exempt
def run_reason_identification(request):

    BASE_DIR_PY = os.path.dirname(os.path.abspath(__file__))
    script_path = os.path.join(BASE_DIR_PY, "Reason.py")

    if not os.path.exists(script_path):
        return JsonResponse({"error": "Reason script not found"}, status=404)

    def generate():
        try:
            print("✅ [BACKEND] Reason Identification triggered", flush=True)

            # Yield immediately to keep connection open
            yield "data: [SSE connection started]\n\n"

            process = subprocess.Popen(
                ["python", "-u", script_path],
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                bufsize=1
            )

            for line in iter(process.stdout.readline, ''):
                print(f"[Reason.py LOG] {line.strip()}", flush=True)
                yield f"data: {line.strip()}\n\n"
                time.sleep(0.1)

            process.stdout.close()
            process.wait()

            print("✅ [BACKEND] Reason Identification completed", flush=True)
            yield "data: [SSE completed]\n\n"
            time.sleep(0.1)

        except Exception as e:
            error_msg = f"❌ Exception: {str(e)}"
            print(error_msg, flush=True)
            yield f"data: {error_msg}\n\n"

    response = StreamingHttpResponse(generate(), content_type='text/event-stream; charset=utf-8')
    response["Cache-Control"] = "no-cache"
    response["X-Accel-Buffering"] = "no"
    response["Access-Control-Allow-Origin"] = "*"
    return response

@csrf_exempt
def run_feature_prediction(request):
    BASE_DIR_PY = os.path.dirname(os.path.abspath(__file__))
    script_path = os.path.join(BASE_DIR_PY, "Prediction.py")

    if not os.path.exists(script_path):
        return JsonResponse({"error": "Prediction script not found"}, status=404)

    def generate():
        process = subprocess.Popen(
            ["python", "-u", script_path],
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,  # Capture both stdout + stderr
            text=True,
            bufsize=1
        )

        for line in iter(process.stdout.readline, ''):
            yield f"data: {line.strip()}\n\n"  # 👈 this is what makes it SSE
            time.sleep(0.1)

        process.stdout.close()
        process.wait()

    response = StreamingHttpResponse(generate(), content_type='text/event-stream')
    response["Cache-Control"] = "no-cache"
    response["X-Accel-Buffering"] = "no"
    response["Access-Control-Allow-Origin"] = "*"
    return response

@csrf_exempt
def get_airflow_dag_logs(request):
    """Fetches the latest DAG run logs from Airflow and returns them in JSON format."""
    try:
        dag_run_url = f"{AIRFLOW_API_URL}/{DAG_ID}/dagRuns"
        response = requests.get(dag_run_url, auth=AIRFLOW_AUTH)

        if response.status_code != 200 or not response.text.strip():
            return JsonResponse({
                "error": f"Failed to fetch DAG runs: {response.status_code} - {response.text or 'Empty response'}",
                "url": dag_run_url
            }, status=response.status_code)

        try:
            dag_runs = response.json().get("dag_runs", [])
        except ValueError:
            return JsonResponse({"error": "Invalid JSON response from Airflow", "raw": response.text}, status=500)

        if not dag_runs:
            return JsonResponse({"logs": ["No DAG runs found."]})

        latest_run = sorted(dag_runs, key=lambda x: x["execution_date"], reverse=True)[0]
        run_id = latest_run["dag_run_id"]

        # ✅ Fetch task instances
        task_instances_url = f"{AIRFLOW_API_URL}/{DAG_ID}/dagRuns/{run_id}/taskInstances"
        task_response = requests.get(task_instances_url, auth=AIRFLOW_AUTH)

        if task_response.status_code != 200 or not task_response.text.strip():
            return JsonResponse({
                "error": f"Failed to fetch task instances: {task_response.status_code} - {task_response.text or 'Empty response'}",
                "url": task_instances_url
            }, status=task_response.status_code)

        try:
            task_instances = task_response.json().get("task_instances", [])
        except ValueError:
            return JsonResponse({"error": "Invalid JSON response from Airflow task instances", "raw": task_response.text}, status=500)

        logs = []
        for task in task_instances:
            task_id = task["task_id"]
            state = task["state"]

            # ✅ Fetch actual task logs
            task_log_url = f"{AIRFLOW_API_URL}/{DAG_ID}/dagRuns/{run_id}/taskInstances/{task_id}/logs/1"
            log_response = requests.get(task_log_url, auth=AIRFLOW_AUTH)

            if log_response.status_code == 200:
                try:
                    log_text = log_response.json().get("content", "No logs available.")
                except ValueError:
                    log_text = log_response.text  # 🔥 FIX: Return raw log if not JSON
            else:
                log_text = f"Failed to fetch logs for task {task_id}"

            logs.append(f"Task {task_id}: {state} - Logs: {log_text}")

        return JsonResponse({"logs": logs})

    except requests.exceptions.ConnectionError:
        return JsonResponse({"error": "Failed to connect to Airflow. Check if it's running."}, status=500)

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


def trigger_airflow_dags(request):
    """Triggers only the first DAG. The remaining DAGs execute sequentially in Airflow."""
    dag = "data_processing_pipeline"  # This triggers all DAGs in order

    url = f"{AIRFLOW_API_URL}/{dag}/dagRuns"
    headers = {"Content-Type": "application/json"}

    response = requests.post(url, headers=headers, auth=AIRFLOW_AUTH, json={})

    if response.status_code in [200, 201]:
        return JsonResponse({"message": f"DAG {dag} triggered. Remaining DAGs will follow automatically."})
    else:
        return JsonResponse({"error": f"Failed to trigger {dag}: {response.text}"}, status=500)


@csrf_exempt
@api_view(["DELETE"])
def delete_uploaded_file(request):
    try:
        if request.method != "DELETE":
            return JsonResponse({"success": False, "message": "Invalid request method"}, status=405)

        category = request.GET.get("category", "").strip().lower()  # Normalize category input
        if not category:
            return JsonResponse({"success": False, "message": "Category not provided"}, status=400)

        # ✅ Debugging: Print UPLOAD_DIR
        print(f"🛑 DEBUG: UPLOAD_DIR = {UPLOAD_DIR}")

        # ✅ Check if UPLOAD_DIR is properly set
        if not UPLOAD_DIR:
            return JsonResponse({"success": False, "message": "UPLOAD_DIR is not set. Check settings.MEDIA_ROOT."}, status=500)

        # ✅ Ensure upload directory exists
        if not os.path.exists(UPLOAD_DIR):
            return JsonResponse({"success": False, "message": f"UPLOAD_DIR '{UPLOAD_DIR}' does not exist. Create the directory."}, status=500)

        # ✅ Convert category into a valid filename format
        file_name = f"{category}.xlsx"
        file_path = os.path.join(UPLOAD_DIR, file_name)

        # ✅ Debugging: Print full path
        print(f"📌 DEBUG: Full Path to Delete = {file_path}")

        # ✅ Check if the file exists before attempting to delete
        if os.path.exists(file_path):
            os.remove(file_path)
            return JsonResponse({"success": True, "message": f"File '{file_name}' deleted successfully."})
        else:
            available_files = os.listdir(UPLOAD_DIR) if os.path.exists(UPLOAD_DIR) else []
            return JsonResponse({
                "success": False, 
                "message": f"File '{file_name}' not found. Available files: {available_files}"
            }, status=404)

    except Exception as e:
        return JsonResponse({"success": False, "message": f"Error: {str(e)}"}, status=500)
    
@api_view(["GET"])
def get_postgres_tables(request):
    try:
        conn = psycopg2.connect(
            dbname=DB_Auto["dbname"],
            user=DB_Auto["user"],
            password=DB_Auto["password"],
            host=DB_Auto["host"],
            port=DB_Auto["port"]
        )
        cursor = conn.cursor()

        # ✅ Fetch tables from different schemas
        schemas = {
            "Stage": [],
            "Cleaned": [],
            "merged": [],
        }

        for schema in schemas.keys():
            cursor.execute(f"SELECT table_name FROM information_schema.tables WHERE table_schema = '{schema}';")
            schemas[schema] = [row[0] for row in cursor.fetchall()]

        cursor.close()
        conn.close()

        return Response(schemas, status=200)

    except psycopg2.Error as db_err:
        return Response({"error": str(db_err)}, status=500)

    except Exception as e:
        return Response({"error": str(e)}, status=500)


### 🔹 Delete PostgreSQL Tables
@api_view(["POST"])
def delete_postgres_tables(request):
    try:
        # ✅ Connect to PostgreSQL
        conn = psycopg2.connect(
            dbname=DB_Auto["dbname"],
            user=DB_Auto["user"],
            password=DB_Auto["password"],
            host=DB_Auto["host"],
            port=DB_Auto["port"]
        )
        cursor = conn.cursor()

        # ✅ Ensure schema names are correctly referenced
        schemas_to_delete = ['"Cleaned"', '"merged"']  # Use double quotes for case-sensitive names

        for schema in schemas_to_delete:
            cursor.execute(f"DROP SCHEMA IF EXISTS {schema} CASCADE;")  # Drop schema
            cursor.execute(f"CREATE SCHEMA {schema};")  # Recreate schema

        conn.commit()
        cursor.close()
        conn.close()

        # ✅ Delete Airflow logs up to the previous day
        previous_day = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%dT00:00:00Z")
        delete_logs_url = f"{AIRFLOW_API_URL}/{DAG_ID}/dagRuns"

        headers = {
            "Content-Type": "application/json"
        }

        # Get DAG runs before the previous day
        response = requests.get(delete_logs_url, headers=headers)
        response_data = response.json()

        if "dag_runs" in response_data:
            for dag_run in response_data["dag_runs"]:
                execution_date = dag_run["execution_date"]
                if execution_date < previous_day:
                    dag_run_id = dag_run["dag_run_id"]
                    delete_url = f"{delete_logs_url}/{dag_run_id}"
                    delete_response = requests.delete(delete_url, headers=headers)

        return Response({"message": "Tables deleted successfully and old Airflow logs removed."}, status=200)

    except Exception as e:
        return Response({"error": str(e)}, status=500)



@csrf_exempt
def upload_file(request):
    if request.method == "POST" and request.FILES.get("file"):
        uploaded_file = request.FILES["file"]

        # ✅ Ensure the upload directory exists
        os.makedirs(UPLOAD_DIR, exist_ok=True)

        file_path = os.path.join(UPLOAD_DIR, uploaded_file.name)

        # ✅ Stream large files to disk (Avoid Memory Issues)
        with open(file_path, "wb+") as destination:
            for chunk in uploaded_file.chunks():
                destination.write(chunk)

        return JsonResponse({"message": "File uploaded successfully", "file_path": file_path})

    return JsonResponse({"error": "No file received"}, status=400)

def get_excel_headers(file_path):
    """Efficiently reads only the headers of an Excel file."""
    try:
        wb = load_workbook(filename=file_path, read_only=True)
        sheet = wb.active
        headers = [cell.value.strip().lower() for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        wb.close()
        return headers
    except Exception as e:
        return []

@csrf_exempt
@api_view(["POST"])
@parser_classes([MultiPartParser, FormParser])
def upload_excel(request):
    try:
        file = request.FILES.get("file", None)
        category = request.data.get("category", "").strip().lower()

        if not file or not category:
            return JsonResponse({"success": False, "message": "Missing file or category"}, status=400)

        # ✅ Define standard file names for each category
        category_mapping = {
            "base": "base",
            "pr": "pr",
            "claims": "claims"
        }

        # ✅ Extract file extension
        file_extension = file.name.split(".")[-1].lower()
        allowed_extensions = ["xlsx", "csv", "xlsb"]
        
        if file_extension not in allowed_extensions:
            return JsonResponse({"success": False, "message": "Invalid file format. Only xlsx, csv, and xlsb are allowed."}, status=400)
        
        # ✅ Use mapped file name if category matches, else keep the original name
        file_name = f"{category_mapping.get(category, file.name.split('.')[0])}.{file_extension}"
        file_path = os.path.join(UPLOAD_DIR, file_name)

        # ✅ Ensure Upload Directory Exists
        os.makedirs(UPLOAD_DIR, exist_ok=True)

        # ✅ Save File with new name
        with default_storage.open(file_path, "wb+") as destination:
            for chunk in file.chunks():
                destination.write(chunk)

        return JsonResponse({"success": True, "message": f"File uploaded successfully as {file_name}"})
    
    except Exception as e:
        return JsonResponse({"success": False, "message": str(e)}, status=500)

    

# ✅ Process Uploaded Files and Store Data in PostgreSQL

@api_view(["POST"])
def process_uploaded_data(request):
    try:
        files = request.FILES  # Use request.FILES to correctly handle file objects
        selected_year = request.POST.get("year", "")  # Get selected year

        if not files:
            return JsonResponse({"success": False, "message": "No files received"}, status=400)

        # ✅ Print Available Files
        available_files = os.listdir(UPLOAD_DIR) if os.path.exists(UPLOAD_DIR) else []
        print(f"📌 Available files in upload directory before processing: {available_files}")

        # ✅ Process mapped files
        for category, filename in FILE_MAPPING.items():
            file = files.get(category)
            if not file:
                print(f"⚠️ Skipping {category}, no file uploaded.")
                continue

            file_extension = file.name.split(".")[-1].lower()
            new_filename = f"{selected_year}_{filename}" if selected_year else filename
            file_path = os.path.join(UPLOAD_DIR, new_filename)  # Use mapped filename with year prefix

            # ✅ Save file to disk
            with default_storage.open(file_path, "wb+") as destination:
                for chunk in file.chunks():
                    destination.write(chunk)

            # ✅ Read file with explicit header normalization
            print(f"📌 Reading file: {file_path}")
            if file_extension == "csv":
                df = pd.read_csv(file_path, skiprows=0)
            elif file_extension == "xlsb":
                df = pd.read_excel(file_path, engine="pyxlsb", skiprows=0)
            else:
                df = pd.read_excel(file_path, skiprows=0)

            # ✅ Normalize Headers (Trim spaces, lowercase, replace special characters with numbers)
            df.columns = (
                df.columns.str.strip()
                          .str.lower()
                          .str.replace(r"\s+", "_", regex=True)  # Replace spaces with underscores
                          .str.replace(r"[^a-z0-9_0-9]", "", regex=True)  # Remove special characters but allow numbers
            )

            print(f"✅ Processed headers: {df.columns.tolist()}")  # Debugging

            # ✅ Store in PostgreSQL
            table_name = f"{selected_year}_{category}" if selected_year else f"{category}"  # Store by category name
            store_dataframe_in_postgres(df, table_name)
            print(f"✅ Successfully stored {file_path} in {table_name}")

            # ✅ Remove File After Processing
            os.remove(file_path)
            print(f"✅ Successfully deleted {file_path}")

        # ✅ Process Feedback File if Uploaded
        # ✅ Process Feedback File if Uploaded
        feedback_file = files.get("feedback")
        if feedback_file:
            feedback_extension = feedback_file.name.split(".")[-1].lower()
            feedback_filename = f"{selected_year}_feedback.xlsx"
            feedback_path = os.path.join(UPLOAD_DIR, feedback_filename)

            with default_storage.open(feedback_path, "wb+") as destination:
                for chunk in feedback_file.chunks():
                    destination.write(chunk)

            print(f"📌 Processing feedback file: {feedback_path}")
            if feedback_extension == "csv":
                df_feedback = pd.read_csv(feedback_path)
            elif feedback_extension == "xlsb":
                df_feedback = pd.read_excel(feedback_path, engine="pyxlsb")
            else:
                df_feedback = pd.read_excel(feedback_path)

            store_dataframe_in_postgres(df_feedback, f"feedback_data_{selected_year}")
            os.remove(feedback_path)
            print(f"✅ Successfully stored feedback data and deleted {feedback_path}")


        # ✅ Process Additional Files if Uploaded
        additional_files = files.getlist("additional")
        for additional_file in additional_files:
            additional_extension = additional_file.name.split(".")[-1].lower()
            additional_filename = f"{selected_year}_{additional_file.name}"
            additional_path = os.path.join(UPLOAD_DIR, additional_filename)

            with default_storage.open(additional_path, "wb+") as destination:
                for chunk in additional_file.chunks():
                    destination.write(chunk)

            print(f"📌 Processing additional file: {additional_path}")
            if additional_extension == "csv":
                df_additional = pd.read_csv(additional_path)
            elif additional_extension == "xlsb":
                df_additional = pd.read_excel(additional_path, engine="pyxlsb")
            else:
                df_additional = pd.read_excel(additional_path)

            table_name = f"additional_{selected_year}_{os.path.splitext(additional_file.name)[0]}"
            store_dataframe_in_postgres(df_additional, table_name)
            os.remove(additional_path)
            print(f"✅ Successfully stored additional data and deleted {additional_path}")

        return Response({"success": True, "message": "Data stored successfully!"})

    except Exception as e:
        print(f"⚠️ Unexpected error: {str(e)}")
        return Response({"success": False, "message": str(e)}, status=500)


def store_dataframe_in_postgres(df, table_name):
    """✅ Creates a table in PostgreSQL and inserts data from the DataFrame."""
    from sqlalchemy import create_engine

    try:
        # ✅ PostgreSQL Connection String
        db_url = f"postgresql://{DB_Auto['user']}:{DB_Auto['password']}@{DB_Auto['host']}:{DB_Auto['port']}/{DB_Auto['dbname']}"
        engine = create_engine(db_url)

        schema_name = "Stage"

        # ✅ Debug: Print credentials used
        print(f"📌 Connecting to DB: {DB_Auto['dbname']} | User: {DB_Auto['user']} | Host: {DB_Auto['host']} | Port: {DB_Auto['port']} | Schema: {schema_name}")

        # Convert object columns to strings
        for col in df.columns:
            if df[col].dtype == "object":
                df[col] = df[col].astype(str)

        # ✅ Store Data in PostgreSQL
        df.to_sql(table_name, con=engine, schema=schema_name, if_exists="replace", index=False)

        print(f"✅ Successfully stored data in table: {schema_name}.{table_name}")

    except Exception as e:
        print(f"❌ Error storing data in PostgreSQL: {e}")
        raise



@api_view(["GET"])
def view_uploaded_data(request):
    try:
        category = request.GET.get("category", "").strip().lower()  # Convert category to lowercase
        if not category:
            return JsonResponse({"success": False, "message": "Category not provided"}, status=400)

        # ✅ Possible Filenames
        possible_filenames = [
            os.path.join(UPLOAD_DIR, f"{category}.xlsx"),
            os.path.join(UPLOAD_DIR, f"{category}.csv"),
            os.path.join(UPLOAD_DIR, f"{category}.xlsb"),
            os.path.join(UPLOAD_DIR, f"{category.capitalize()}.xlsx"),
            os.path.join(UPLOAD_DIR, f"{category.capitalize()}.csv"),
            os.path.join(UPLOAD_DIR, f"{category.capitalize()}.xlsb"),
            os.path.join(UPLOAD_DIR, f"{category.upper()}.xlsx"),
            os.path.join(UPLOAD_DIR, f"{category.upper()}.csv"),
            os.path.join(UPLOAD_DIR, f"{category.upper()}.xlsb"),
        ]

        # ✅ Check for any matching file
        file_path = next((path for path in possible_filenames if os.path.exists(path)), None)

        print(f"Searching for files: {possible_filenames}")  # ✅ Debugging
        print(f"Files in directory: {os.listdir(UPLOAD_DIR)}")  # ✅ Debugging

        if not file_path:
            return JsonResponse({"success": False, "message": f"File '{category}' not found in uploaded_files"}, status=404)

        # ✅ Read file based on extension
        file_extension = file_path.split(".")[-1].lower()
        
        if file_extension == "csv":
            df = pd.read_csv(file_path, nrows=5)  # Read CSV
        elif file_extension == "xlsb":
            df = pd.read_excel(file_path, engine="pyxlsb", nrows=5)  # Read XLSB
        else:  # Default to xlsx
            df = pd.read_excel(file_path, nrows=5)

        data_preview = df.to_dict(orient="records")

        return JsonResponse({"success": True, "data": data_preview})

    except Exception as e:
        return JsonResponse({"success": False, "message": str(e)}, status=500)






@csrf_exempt
def start_preprocessing(request):
    """ Starts the preprocessing task. """
    if request.method == "POST":
        cache.set("process_status", "Processing...", timeout=300)  # Store status in cache
        cache.set("process_steps", [], timeout=300)  # Reset steps
        return JsonResponse({"status": "Processing started"})

    return JsonResponse({"error": "Invalid request"}, status=400)



def get_status(request):
    """ Returns the current status of the processing. """
    status = cache.get("process_status", "Waiting for start...")  # Default message
    steps = cache.get("process_steps", [])
    return JsonResponse({"status": status, "steps": steps})

class GetTablesView(APIView):
    """
    Fetch all tables from the SSBI PostgreSQL database using .env config.
    """
    def get(self, request):
        try:
            import psycopg2
            conn = psycopg2.connect(
                dbname=DB_SSBI["dbname"],
                user=DB_SSBI["user"],
                password=DB_SSBI["password"],
                host=DB_SSBI["host"],
                port=DB_SSBI["port"]
            )
            cursor = conn.cursor()
            cursor.execute("SELECT table_name FROM information_schema.tables WHERE table_schema = 'prediction';")
            tables = [row[0] for row in cursor.fetchall()]
            cursor.close()
            conn.close()
            return Response({"tables": tables})
        except Exception as e:
            logger.error(f"Error fetching tables from SSBI DB: {str(e)}", exc_info=True)
            return Response({"error": str(e)}, status=500)

# ✅ API to Fetch Data from a Selected Table
class GetTableDataView(APIView):
    """
    Fetch all data and column names from the specified table using SSBI .env config.
    """
    def get(self, request):
        table_name = request.GET.get("table")
        if not table_name:
            return Response({"error": "Table name is required."}, status=400)

        try:
            import psycopg2
            conn = psycopg2.connect(
                dbname=DB_SSBI["dbname"],
                user=DB_SSBI["user"],
                password=DB_SSBI["password"],
                host=DB_SSBI["host"],
                port=DB_SSBI["port"]
            )
            cursor = conn.cursor()
            cursor.execute(f'SELECT * FROM prediction."{table_name}" LIMIT 100')
            columns = [col[0] for col in cursor.description]
            rows = cursor.fetchall()
            cursor.close()
            conn.close()

            return Response({"columns": columns, "rows": rows})
        except Exception as e:
            logger.error(f"Error fetching data from SSBI table '{table_name}': {str(e)}", exc_info=True)
            return Response({"error": str(e)}, status=500)


def get_huggingface_response(prompt):
    try:
        response = generator(prompt, max_length=50, num_return_sequences=1)
        return response[0]["generated_text"]
    except Exception as e:
        logger.error(f"Error with Hugging Face: {e}")
        return "Error: Unable to generate a response."

@api_view(["POST"])
def ai_response(request):
    """
    Generate a response using Meta's LLaMA model.
    """
    data = request.data
    prompt = data.get("prompt", "").strip()

    if not prompt:
        return Response({"error": "Prompt is required"}, status=400)

    try:
        inputs = tokenizer(prompt, return_tensors="pt").to(DEVICE)
        with torch.no_grad():
            output = model.generate(**inputs, max_length=200)
        response_text = tokenizer.decode(output[0], skip_special_tokens=True)

        return Response({"response": response_text})
    except Exception as e:
        logger.error(f"Error generating AI response: {e}")
        return Response({"error": "Failed to generate response"}, status=500)

class RegisterUserView(generics.CreateAPIView):
    queryset = User.objects.all()
    serializer_class = CreateUserSerializer
    permission_classes = [AllowAny]

# ✅ Login View (Supports username OR email)
class LoginView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        login_id = request.data.get("username")  # Can be username or email
        password = request.data.get("password")
        role = request.data.get("role")

        user = authenticate(username=login_id, password=password, role = role)

        if user is None:
            try:
                user_obj = User.objects.get(email=login_id)  # Try email lookup
                user = authenticate(username=user_obj.username, password=password, role = role)
            except User.DoesNotExist:
                return Response({"detail": "Invalid credentials."}, status=status.HTTP_401_UNAUTHORIZED)

        if user:
            refresh = RefreshToken.for_user(user)
            return Response({
                "message": f"Hi {user.username}! You successfully logged in.",
                "access_token": str(refresh.access_token),
                "refresh_token": str(refresh),
            }, status=status.HTTP_200_OK)

        return Response({"detail": "Invalid credentials."}, status=status.HTTP_401_UNAUTHORIZED)

# ✅ Logout View (With JWT Blacklist)
class LogoutView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            refresh_token = request.data.get("refresh_token")
            if refresh_token:
                token = RefreshToken(refresh_token)
                token.blacklist()  # Invalidate the token (Requires DRF-SimpleJWT blacklist enabled)
            return Response({"message": "Logged out successfully"}, status=status.HTTP_200_OK)
        except Exception as e:
          return Response({"error": "Invalid token"}, status=status.HTTP_400_BAD_REQUEST)

class UserDetailsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        """
        Fetch user details using the authenticated request.
        """
        try:
            user = request.user  # ✅ Get logged-in user from the request
            return Response({"username": user.username, "email": user.email,"role": user.role})
        
        except Exception as e:
            return Response({"error": str(e)}, status=500)
# ✅ Fetch Cleaned Data (JSON Format)
# ✅ Fetch Cleaned Data (JSON Format)


class CleanedDataView(APIView):
    """
    API to fetch all columns from PostgreSQL and return as JSON.
    """
    def get(self, request):
        try:
            logger.info("Connecting to PostgreSQL database...")
            with connections["postgres"].cursor() as cursor:
                query = 'SELECT * FROM "corrected_merged_claim_data_EF_with_xgbreasons" LIMIT 100'
                cursor.execute(query)
                
                columns = [col[0] for col in cursor.description]
                rows = cursor.fetchall()

            if not rows:
                logger.warning("No data found in the table.")
                return Response({"message": "No data found"}, status=status.HTTP_204_NO_CONTENT)

            # ✅ Convert Data to DataFrame
            data = pd.DataFrame(rows, columns=columns)

            # ✅ Replace NaN, Infinity with None (for JSON serialization)
            data = data.replace({np.nan: None, np.inf: None, -np.inf: None})

            logger.info(f"Data Retrieved: {len(data)} rows")

            return Response({"columns": columns, "data": data.to_dict(orient="records")}, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Error in CleanedDataView: {str(e)}", exc_info=True)
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
class DownloadCleanedDataCSV(APIView):
    """
    API to download PostgreSQL table as a CSV file.
    """
    def get(self, request):
        try:
            with connections["postgres"].cursor() as cursor:
                cursor.execute('SELECT * FROM public."corrected_merged_claim_data_EF_with_xgbreasons"')  # ✅ Modify table name
                columns = [col[0] for col in cursor.description]
                rows = cursor.fetchall()

            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = "attachment; filename=Cleaned_data.csv"

            writer = csv.writer(response)
            writer.writerow(columns)  # ✅ Write column headers
            writer.writerows(rows)  # ✅ Write data rows

            return response
        except Exception as e:
            return HttpResponse(f"Error: {str(e)}", content_type="text/plain", status=500)


# ✅ Fetch Power BI Access Token
class PowerBITokenView(APIView):
    """
    Returns a publicly accessible Power BI report embed link.
    """

    def get(self, request):
        embed_url = "https://app.powerbi.com/view?r=eyJrIjoiNjZhNzA4YmItOWY1NS00NGI4LTg5NGMtYmM3NDg4MjY5NWJiIiwidCI6IjRiOTExMTAzLWVmZmItNGJkZi04ZGY5LTExZDk0ZjY0YjBkNiJ9"

        return JsonResponse({
            "embedUrl": embed_url  # ✅ Only returning the embed URL
        })
class WhoView(APIView):
    """
    Returns a publicly accessible Power BI report embed link.
    """

    def get(self, request):
        embed_url = "https://app.powerbi.com/view?r=eyJrIjoiNDQ5Y2Y0NWYtZWUwOC00NjZmLTg5ZTUtMzdmZGM1YjYzMDBiIiwidCI6IjRiOTExMTAzLWVmZmItNGJkZi04ZGY5LTExZDk0ZjY0YjBkNiJ9"

        return JsonResponse({
            "embedUrl": embed_url  # ✅ Only returning the embed URL
        }
        )
class WhyView(APIView):
    """
    Returns a publicly accessible Power BI report embed link.
    """

    def get(self, request):
        embed_url = "https://app.powerbi.com/view?r=eyJrIjoiMjNlNjBlOWItNTVlMS00MmVjLWI2MjktOTlmOTM2M2YzNWM0IiwidCI6IjRiOTExMTAzLWVmZmItNGJkZi04ZGY5LTExZDk0ZjY0YjBkNiJ9"

        return JsonResponse({
            "embedUrl": embed_url  # ✅ Only returning the embed URL
        })
class HowView(APIView):
    """
    Returns a publicly accessible Power BI report embed link.
    """

    def get(self, request):
        embed_url = "https://app.powerbi.com/view?r=eyJrIjoiN2YzOTVhN2UtNWM0Ni00MjU2LThmZmQtMjM3MDQ5OGMwNmEzIiwidCI6IjRiOTExMTAzLWVmZmItNGJkZi04ZGY5LTExZDk0ZjY0YjBkNiJ9"

        return JsonResponse({
            "embedUrl": embed_url  # ✅ Only returning the embed URL
        })
        
# ✅ Fetch Predicted Data (Filtered by Month)
class FullData(APIView):
    def get(self, request):
        try:
            page = int(request.GET.get("page", 1))
            page_size = int(request.GET.get("page_size", 30))
            offset = (page - 1) * page_size

            segment = request.GET.get("segment")
            month = request.GET.get("month")
            year = request.GET.get("year")
            

            conn = psycopg2.connect(
                dbname=DB_New_Pred["dbname"],
                user=DB_New_Pred["user"],
                password=DB_New_Pred["password"],
                host=DB_New_Pred["host"],
                port=DB_New_Pred["port"]
            )
            cursor = conn.cursor()

            # Construct WHERE clause based on filters
            filters = []
            values = []

            if segment:
                filters.append('"customer_segment" = %s')
                values.append(segment)

            if month:
                filters.append('EXTRACT(MONTH FROM "policy_end_date") = %s')
                values.append(month)

            if year:
                filters.append('EXTRACT(YEAR FROM "policy_end_date") = %s')
                values.append(year)

            where_clause = f'''WHERE "predicted_status" = 'Not Renewed'{' AND ' + ' AND '.join(filters) if filters else ''}'''


            # Total count with filters
            count_query = f'''
                SELECT COUNT(*) FROM "Prediction"."2025_prediction_final_data"
                {where_clause} 
            '''
            cursor.execute(count_query, values)
            total_rows = cursor.fetchone()[0]

            # Paginated data query with filters
            data_query = f'''
                SELECT * FROM "Prediction"."2025_prediction_final_data"
                {where_clause} 
                ORDER BY "policy_no"
                OFFSET %s LIMIT %s
            '''
            cursor.execute(data_query, values + [offset, page_size])
            columns = [col[0] for col in cursor.description]
            rows = cursor.fetchall()

            data = []
            for row in rows:
                row_dict = dict(zip(columns, row))
                row_dict = {
                    k: (v.isoformat() if hasattr(v, 'isoformat') else v)
                    for k, v in row_dict.items()
                }
                data.append(row_dict)

            # Segment counts (filtered by month/year)
            filter_for_segments = []
            values_for_segments = []

            if month:
                filter_for_segments.append('EXTRACT(MONTH FROM "policy_end_date") = %s')
                values_for_segments.append(month)

            if year:
                filter_for_segments.append('EXTRACT(YEAR FROM "policy_end_date") = %s')
                values_for_segments.append(year)

            segment_where = f'WHERE {" AND ".join(filter_for_segments)}' if filter_for_segments else ''

            segment_query = f'''
                SELECT "customer_segment", COUNT(*) 
                FROM "Prediction"."2025_prediction_final_data"
                {segment_where}
                GROUP BY "customer_segment"
            '''
            cursor.execute(segment_query, values_for_segments)
            segment_counts_raw = cursor.fetchall()
            segment_counts = {row[0]: row[1] for row in segment_counts_raw}

            return Response({
                "columns": columns,
                "data": data,
                "total": total_rows,
                "page": page,
                "page_size": page_size,
                "total_pages": math.ceil(total_rows / page_size),
                "segment_counts": segment_counts
            })

        except Exception as e:
            return Response({"error": str(e)}, status=500)


class FullDataDownload(APIView):
    """
    API to download PostgreSQL table as a CSV file.
    """
    def get(self, request):
        logger.info("Connecting to PostgreSQL database...")
        try:
            
            conn = psycopg2.connect(
                 dbname=DB_New_Pred["dbname"],
                 user=DB_New_Pred["user"],
                 password=DB_New_Pred["password"],
                 host=DB_New_Pred["host"],
                 port=DB_New_Pred["port"]
                 )
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM "Prediction"."Updated Prediction with new reason"')  # ✅ Modify table name
            columns = [col[0] for col in cursor.description]
            rows = cursor.fetchall()

            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = "attachment; filename=Cleaned_data.csv"

            writer = csv.writer(response)
            writer.writerow(columns)  # ✅ Write column headers
            writer.writerows(rows)  # ✅ Write data rows

            return response
        except Exception as e:
            return HttpResponse(f"Error: {str(e)}", content_type="text/plain", status=500)
class MonthDataView(APIView):
    """
    API to fetch month-wise data from `january_FEB_Mar_Prediction`.
    """
    def get(self, request, month):
        try:
            # Validate Month Input
            try:
                month = int(month)
                if month < 1 or month > 12:
                    return Response({"error": "Invalid month number"}, status=status.HTTP_400_BAD_REQUEST)
            except ValueError:
                return Response({"error": "Month must be an integer"}, status=status.HTTP_400_BAD_REQUEST)

            table_name = "rancat_prediction_jfmamuj"
            logger.info(f"Fetching data for month {month} from table {table_name}")
            logger.info("Connecting to PostgreSQL database...")

            conn = psycopg2.connect(
                dbname=DB_Auto["dbname"],
                user=DB_Auto["user"],
                password=DB_Auto["password"],  # Or use a secure method
                host=DB_Auto["host"],
                port=DB_Auto["port"]
            )
            cursor = conn.cursor()
            query = f'SELECT * FROM prediction."{table_name}" WHERE "policy end date_MONTH" = %s LIMIT 100'
            cursor.execute(query, [month])

            columns = [col[0] for col in cursor.description]
            rows = cursor.fetchall()

            if not rows:
                logger.warning(f"No data found for month {month} in table {table_name}")
                return Response({"message": "No data found"}, status=status.HTTP_204_NO_CONTENT)

            data = pd.DataFrame(rows, columns=columns)
            data = data.replace({np.nan: None, np.inf: None, -np.inf: None})

            logger.info(f"Data Retrieved: {len(data)} rows for month {month}")
            return Response({"columns": columns, "data": data.to_dict(orient="records")}, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Error fetching data for month {month}: {str(e)}", exc_info=True)
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



# ✅ API to Download Month-Wise Data from `january_FEB_Mar_Prediction`
class MonthDataDownloadView(APIView):
    """
    API to download month-wise data from `january_FEB_Mar_Prediction` as CSV.
    """
    def get(self, request, month):
        try:
            try:
                month = int(month)
                if month < 1 or month > 12:
                    return HttpResponse("Invalid month number", content_type="text/plain", status=400)
            except ValueError:
                return HttpResponse("Month must be an integer", content_type="text/plain", status=400)

            table_name = "rancat_prediction_jfmamuj"
            logger.info(f"Downloading data for month {month} from table {table_name}")
            logger.info("Connecting to PostgreSQL database...")

            conn = psycopg2.connect(
                dbname=DB_Auto["dbname"],
                user=DB_Auto["user"],
                password=DB_Auto["password"],
                host=DB_Auto["host"],
                port=DB_Auto["port"]
            )
            cursor = conn.cursor()
            query = f'SELECT * FROM prediction."{table_name}" WHERE "policy end date_MONTH" = %s'
            cursor.execute(query, [month])

            columns = [col[0] for col in cursor.description]
            rows = cursor.fetchall()

            if not rows:
                return HttpResponse("No data found", content_type="text/plain", status=204)

            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = f'attachment; filename=month_{month}_data.csv'

            writer = csv.writer(response)
            writer.writerow(columns)
            writer.writerows(rows)

            return response

        except Exception as e:
            logger.error(f"Error downloading data for month {month}: {str(e)}", exc_info=True)
            return HttpResponse(f"Error: {str(e)}", content_type="text/plain", status=500)
        


class PastMonthDataView(APIView):
    def get(self, request, month):
        try:
            month = int(month)
            if month < 1 or month > 12:
                return Response({"error": "Invalid month number"}, status=status.HTTP_400_BAD_REQUEST)

            table_name = 'SOND_rancat45_55(Final)_segmentation'

            conn = psycopg2.connect(
                dbname=DB_PAST["dbname"],
                user=DB_PAST["user"],
                password=DB_PAST["password"],
                host=DB_PAST["host"],
                port=DB_PAST["port"]
            )
            cursor = conn.cursor()

            query = f"""
                SELECT * FROM past_data."{table_name}"
                WHERE "policy_end_date_month" = %s
                AND "policy_end_date_year" = 2024
                LIMIT 100
            """
            cursor.execute(query, [month])

            columns = [col[0] for col in cursor.description]
            rows = cursor.fetchall()

            if not rows:
                return Response({"message": "No data found"}, status=status.HTTP_204_NO_CONTENT)

            data = pd.DataFrame(rows, columns=columns)
            data = data.replace({np.nan: None, np.inf: None, -np.inf: None})

            return Response({
                "columns": columns,
                "data": data.to_dict(orient="records")
            }, status=status.HTTP_200_OK)

        except Exception as e:
            import traceback
            print("Error fetching data:", str(e))
            print(traceback.format_exc())
            return Response({
                "error": str(e),
                "traceback": traceback.format_exc()
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)




# ✅ API to Download Month-Wise Data from `january_FEB_Mar_Prediction`
class PastMonthDataDownloadView(APIView):
    """
    API to download month-wise data from `january_FEB_Mar_Prediction` as CSV.
    """
    def get(self, request, month):
        try:
            try:
                month = int(month)
                if month < 1 or month > 12:
                    return HttpResponse("Invalid month number", content_type="text/plain", status=400)
            except ValueError:
                return HttpResponse("Month must be an integer", content_type="text/plain", status=400)

            table_name = "SOND_rancat45_55(Final)_segmentation"
            logger.info(f"Downloading data for month {month} from table {table_name}")
            logger.info("Connecting to PostgreSQL database...")

            conn = psycopg2.connect(
                dbname=DB_PAST["dbname"],
                user=DB_PAST["user"],
                password=DB_PAST["password"],
                host=DB_PAST["host"],
                port=DB_PAST["port"]
            )
            cursor = conn.cursor()
            query = f"""SELECT * FROM past_data."{table_name}" WHERE "policy_end_date_month" = %s and policy_end_date_year = 2024"""
            cursor.execute(query, [month])

            columns = [col[0] for col in cursor.description]
            rows = cursor.fetchall()

            if not rows:
                return HttpResponse("No data found", content_type="text/plain", status=204)

            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = f'attachment; filename=month_{month}_data.csv'

            writer = csv.writer(response)
            writer.writerow(columns)
            writer.writerows(rows)

            return response

        except Exception as e:
            logger.error(f"Error downloading data for month {month}: {str(e)}", exc_info=True)
            return HttpResponse(f"Error: {str(e)}", content_type="text/plain", status=500)

        
class UploadExcelSSBIView(APIView):
    parser_classes = [MultiPartParser]

    def post(self, request, format=None):
        uploaded_file = request.FILES.get("file")
        if not uploaded_file:
            return Response({"error": "No file uploaded"}, status=400)

        filename = uploaded_file.name
        ext = os.path.splitext(filename)[1].lower()
        table_name = os.path.splitext(filename)[0].lower().replace(" ", "_")

        try:
            content = uploaded_file.read()

            if ext == ".csv":
                df = pd.read_csv(BytesIO(content))
            elif ext in [".xls", ".xlsx"]:
                df = pd.read_excel(BytesIO(content), engine="openpyxl")
            elif ext == ".xlsb":
                df = pd.read_excel(BytesIO(content), engine="pyxlsb")
            else:
                return Response({"error": "Unsupported file format"}, status=400)

            if df.empty:
                return Response({"error": "File is empty"}, status=400)

            df.columns = [col.strip().lower().replace(" ", "_") for col in df.columns]

            # Save to PostgreSQL
            db_url = f"postgresql://{DB_SSBI_Excel['user']}:{DB_SSBI_Excel['password']}@{DB_SSBI_Excel['host']}:{DB_SSBI_Excel['port']}/{DB_SSBI_Excel['dbname']}"
            engine = create_engine(db_url)
            df.to_sql(table_name, engine, schema="public", if_exists="replace", index=False)

            # Clean preview data to make it JSON-safe
            df_cleaned = df.replace([np.inf, -np.inf], np.nan).fillna("N/A")
            preview_rows = df_cleaned.head(50).values.tolist()

            return Response({
                "message": "File uploaded and saved.",
                "tableName": table_name,
                "columns": df.columns.tolist(),
                "rows": preview_rows,
            })

        except Exception as e:
            print("❌ Error uploading file:", e)
            print(traceback.format_exc())
            return Response({
                "error": str(e),
                "trace": traceback.format_exc()
            }, status=500)

        
class DeleteTableSSBIView(APIView):
    def delete(self, request, table_name):
        db = settings.DB_SSBI
        db_url = f"postgresql://{DB_SSBI_Excel['user']}:{DB_SSBI_Excel['password']}@{DB_SSBI_Excel['host']}:{DB_SSBI_Excel['port']}/{DB_SSBI_Excel['dbname']}"
        
        try:
            engine = create_engine(db_url)
            with engine.connect() as conn:
                conn.execute(text(f'DROP TABLE IF EXISTS public."{table_name}" CASCADE'))
                conn.commit()

            return Response({"message": f"Table '{table_name}' deleted successfully."}, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
@csrf_exempt
def fetch_postgres_tables(request):
    data = json.loads(request.body)
    conn = psycopg2.connect(
        host=data["host"],
        dbname=data["database"],
        user=data["user"],
        port=data["port"],
        password=data["password"]
    )
    cur = conn.cursor()
    schema = data.get("schema", "public")
    cur.execute(f"""
        SELECT table_name
        FROM information_schema.tables
        WHERE table_schema = %s;
    """, (schema,))
    tables = [row[0] for row in cur.fetchall()]
    return JsonResponse({"tables": tables})

@csrf_exempt
def fetch_oracle_tables(request):
    data = json.loads(request.body)
    dsn = cx_Oracle.makedsn(data["server"], data["port"], sid=data["database"])
    conn = cx_Oracle.connect(user=data["user"], password=data["password"], dsn=dsn)
    cur = conn.cursor()
    cur.execute("SELECT table_name FROM user_tables")
    tables = [row[0] for row in cur.fetchall()]
    return JsonResponse({"tables": tables})



########################################### DASHBOARD ##############################################





# def dashboard_api(request):
#     if request.method == "GET":

#         try:
#             # === DB connection ===
#             conn = psycopg2.connect(
#                 dbname=DB_New_Past["dbname"],
#                 user=DB_New_Past["user"],
#                 password=DB_New_Past["password"],
#                 host=DB_New_Past["host"],
#                 port=DB_New_Past["port"]
#             )

#             with conn.cursor() as cursor:
#                 cursor.execute("""SELECT count(distinct("policy no")) FROM public."overall_cleaned_base_and_pr_ef_policyef_with_reasons_bucket" where "Start Year"= 2024 """)
#                 churn_summary = cursor.fetchone()
#                 cursor.execute("""SELECT count(distinct("policy no")) FROM public."overall_cleaned_base_and_pr_ef_policyef_with_reasons_bucket" where "Start Year"= 2024 AND "New Customers" = 'Yes'""")
#                 churn_new = cursor.fetchone()
#                 cursor.execute("""SELECT count("policy no") FROM public."overall_cleaned_base_and_pr_ef_policyef_with_reasons_bucket" where "Start Year"= 2023 AND "Churn Label" = 'Yes'""")
#                 churn_ploicies = cursor.fetchone()
#                 cursor.execute("""SELECT count("policy no") FROM public."overall_cleaned_base_and_pr_ef_policyef_with_reasons_bucket" where "Start Year"= 2023 AND "New Customers" = 'Yes' AND "Churn Label" = 'Yes'""")
#                 churn_first_year = cursor.fetchone()

#                 cursor.execute("""
#                     SELECT 
#                         "Cleaned Branch Name 2" as branch_name,
#                         COUNT("policy no") as churned_count
#                     FROM public."overall_cleaned_base_and_pr_ef_policyef_with_reasons_bucket" 
#                     WHERE "Start Year" = 2023 
#                         AND "Churn Label" = 'Yes' 
#                         AND "Cleaned Branch Name 2" IS NOT NULL 
#                         AND "Cleaned Branch Name 2" != ''
#                     GROUP BY "Cleaned Branch Name 2"
#                     ORDER BY churned_count DESC
#                     LIMIT 5 """)
#                 top_branches_raw = cursor.fetchall()
#                 top_branches = [{"name": row[0], "value": row[1]} for row in top_branches_raw]

#                 # Product-wise Churned Policies for Pie Chart
#                 cursor.execute("""
#                     SELECT 
#                         "product name 2" as product_name,
#                         COUNT(DISTINCT "policy no") as churned_count
#                     FROM public."overall_cleaned_base_and_pr_ef_policyef_with_reasons_bucket" 
#                     WHERE "Start Year" = 2023 
#                         AND "Churn Label" = 'Yes' 
#                         AND "product name 2" IS NOT NULL 
#                         AND "product name 2" != ''
#                     GROUP BY "product name 2"
#                     ORDER BY churned_count DESC
#                 """)
#                 product_churn_raw = cursor.fetchall()
#                 product_churn = [{"name": row[0], "value": row[1]} for row in product_churn_raw]

#                 # Claim Status Distribution for Pie Chart
#                 cursor.execute("""
#                     SELECT 
#                         CASE 
#                             WHEN "Claim Status" = 'Approved' THEN 'Approved'
#                             WHEN "Claim Status" = 'Denied' THEN 'Denied'
#                             ELSE 'No Claim'
#                         END as claim_status,
#                         COUNT(DISTINCT "policy no") as policy_count
#                     FROM public."overall_cleaned_base_and_pr_ef_policyef_with_reasons_bucket" 
#                     WHERE "Start Year" = 2023 
#                         AND "Churn Label" = 'Yes'
#                     GROUP BY 
#                         CASE 
#                             WHEN "Claim Status" = 'Approved' THEN 'Approved'
#                             WHEN "Claim Status" = 'Denied' THEN 'Denied'
#                             ELSE 'No Claim'
#                         END
#                     ORDER BY policy_count DESC
#                 """)
#                 claim_status_raw = cursor.fetchall()
#                 claim_status = [{"name": row[0], "value": row[1]} for row in claim_status_raw]

#                 # Tie-up wise Churned Policies for Column Chart
#                 cursor.execute("""
#                     SELECT 
#                         "tie up" as tie_up_name,
#                         COUNT(DISTINCT "policy no") as churned_count
#                     FROM public."overall_cleaned_base_and_pr_ef_policyef_with_reasons_bucket" 
#                     WHERE "Start Year" = 2023 
#                         AND "Churn Label" = 'Yes' 
#                         AND "tie up" IS NOT NULL 
#                         AND "tie up" != ''
#                     GROUP BY "tie up"
#                     ORDER BY churned_count DESC
#                 """)
#                 tie_up_churn_raw = cursor.fetchall()
#                 tie_up_churn = [{"name": row[0], "value": row[1]} for row in tie_up_churn_raw]

#                 # Top 5 Vehicle Age Groups by Churned Policies for Bar Chart
#                 cursor.execute("""
#                     WITH vehicle_age_buckets AS (
#                         SELECT 
#                             CASE 
#                                 WHEN age >= 0 AND age <= 1 THEN '0-1'
#                                 WHEN age > 1 AND age <= 2 THEN '1-2'
#                                 WHEN age > 2 AND age <= 4 THEN '2-4'
#                                 WHEN age > 4 AND age <= 6 THEN '4-6'
#                                 WHEN age > 6 AND age <= 8 THEN '6-8'
#                                 WHEN age > 8 AND age <= 10 THEN '8-10'
#                                 WHEN age > 10 AND age <= 12 THEN '10-12'
#                                 WHEN age > 12 AND age <= 14 THEN '12-14'
#                                 WHEN age > 14 AND age <= 16 THEN '14-16'
#                                 WHEN age > 16 AND age <= 18 THEN '16-18'
#                                 WHEN age > 18 AND age <= 20 THEN '18-20'
#                                 WHEN age > 20 AND age <= 22 THEN '20-22'
#                                 WHEN age > 22 AND age <= 24 THEN '22-24'
#                                 WHEN age > 24 AND age <= 25 THEN '24-25'
#                                 WHEN age > 25 THEN '25+'
#                                 ELSE 'Unknown'
#                             END as age_bucket,
#                             COUNT(DISTINCT "policy no") as churned_count
#                         FROM public."overall_cleaned_base_and_pr_ef_policyef_with_reasons_bucket" 
#                         WHERE "Start Year" = 2023 
#                             AND "Churn Label" = 'Yes' 
#                             AND age IS NOT NULL
#                         GROUP BY 
#                             CASE 
#                                 WHEN age >= 0 AND age <= 1 THEN '0-1'
#                                 WHEN age > 1 AND age <= 2 THEN '1-2'
#                                 WHEN age > 2 AND age <= 4 THEN '2-4'
#                                 WHEN age > 4 AND age <= 6 THEN '4-6'
#                                 WHEN age > 6 AND age <= 8 THEN '6-8'
#                                 WHEN age > 8 AND age <= 10 THEN '8-10'
#                                 WHEN age > 10 AND age <= 12 THEN '10-12'
#                                 WHEN age > 12 AND age <= 14 THEN '12-14'
#                                 WHEN age > 14 AND age <= 16 THEN '14-16'
#                                 WHEN age > 16 AND age <= 18 THEN '16-18'
#                                 WHEN age > 18 AND age <= 20 THEN '18-20'
#                                 WHEN age > 20 AND age <= 22 THEN '20-22'
#                                 WHEN age > 22 AND age <= 24 THEN '22-24'
#                                 WHEN age > 24 AND age <= 25 THEN '24-25'
#                                 WHEN age > 25 THEN '25+'
#                                 ELSE 'Unknown'
#                             END
#                         ORDER BY churned_count DESC
#                         LIMIT 5
#                     )
#                     SELECT age_bucket as name, churned_count as value 
#                     FROM vehicle_age_buckets
#                 """)
#                 vehicle_age_churn_raw = cursor.fetchall()
#                 vehicle_age_churn = [{"name": row[0], "value": row[1]} for row in vehicle_age_churn_raw]

#         except Exception as e:
#             logger.error(f"Error in dashboard_api: {str(e)}", exc_info=True)
#             return JsonResponse({"error": str(e)}, status=500)
#         return JsonResponse({
#             "message": "Dashboard API is working!", 
#             "churnSummary": churn_summary[0], 
#             "churnNew": churn_new[0], 
#             "churnPloicies": churn_ploicies[0], 
#             "churnFirstYear": churn_first_year[0], 
#             "churnedrate": round((churn_ploicies[0]/churn_summary[0])*100,2) if churn_summary[0] else 0,
#             "topBranches": top_branches,  # NEW: Add top branches data
#             "productChurn": product_churn,  # NEW: Add product-wise churn data
#             "claimStatus": claim_status,  # NEW: Add claim status distribution data
#             "tieUpChurn": tie_up_churn,  # NEW: Add tie-up wise churn data
#             "vehicleAgeChurn": vehicle_age_churn  # NEW: Add vehicle age group churn data
#         })


# def dashboard_api(request):
#     if request.method == "GET":
#         try:
#             from datetime import datetime
#             current_year = datetime.now().year
#             default_year = str(current_year - 1)

#             selected_year = request.GET.get('year', default_year)   # This should match MV "Year"
#             selected_state = request.GET.get('state', 'All')

#             conn = psycopg2.connect(
#                 dbname=DB_New_Past["dbname"],
#                 user=DB_New_Past["user"],
#                 password=DB_New_Past["password"],
#                 host=DB_New_Past["host"],
#                 port=DB_New_Past["port"]
#             )

#             with conn.cursor() as cursor:
#                 # ------------------------------
#                 # Filters: use MV for year/state lists
#                 # ------------------------------
#                 cursor.execute("""
#                     SELECT DISTINCT "Year"
#                     FROM public.mv_year_state_churn
#                     ORDER BY "Year" DESC
#                 """)
#                 available_years = [str(row[0]) for row in cursor.fetchall()]

#                 cursor.execute("""
#                     SELECT DISTINCT states
#                     FROM public.mv_year_state_churn
#                     WHERE states IS NOT NULL AND states <> ''
#                     ORDER BY states
#                 """)
#                 available_states = [row[0] for row in cursor.fetchall()]

#                 # ------------------------------
#                 # Headline metrics from MV
#                 # ------------------------------
#                 # We support "All" states by summing across rows for the selected year,
#                 # and compute churn rate using the (Year-1) totals as denominator.
#                 y = int(selected_year)

#                 if selected_state == 'All':
#                     # Sum across all states for selected year
#                     cursor.execute("""
#                         SELECT
#                           COALESCE(SUM(total_policies), 0)      AS total_policies,
#                           COALESCE(SUM(new_policies), 0)         AS new_policies,
#                           COALESCE(SUM(churned_policies), 0)     AS churned_policies,
#                           COALESCE(SUM(first_year_churned), 0)   AS first_year_churned
#                         FROM public.mv_year_state_churn
#                         WHERE "Year" = %s
#                     """, (y,))
#                     row = cursor.fetchone()
#                     total_policies = row[0]
#                     new_policies = row[1]
#                     churned_policies = row[2]
#                     first_year_churn = row[3]

#                     # Denominator = total policies in (Year-1), summed across states
#                     cursor.execute("""
#                         SELECT COALESCE(SUM(total_policies), 0) AS prev_year_total
#                         FROM public.mv_year_state_churn
#                         WHERE "Year" = %s
#                     """, (y - 1,))
#                     prev_year_total = cursor.fetchone()[0] or 0

#                     churn_rate_num = round((churned_policies / prev_year_total) * 100, 1) if prev_year_total > 0 else 0.0
#                     churn_rate_label = f"{churn_rate_num:.2f}%"

#                 else:
#                     # Single state row directly from MV (includes the label)
#                     cursor.execute("""
#                         SELECT
#                           total_policies,
#                           new_policies,
#                           churned_policies,
#                           first_year_churned,
#                           churn_rate
#                         FROM public.mv_year_state_churn
#                         WHERE "Year" = %s AND states = %s
#                     """, (y, selected_state))
#                     row = cursor.fetchone()

#                     # If no row (e.g., no data for that filter), default to zeros
#                     if not row:
#                         total_policies = 0
#                         new_policies = 0
#                         churned_policies = 0
#                         first_year_churn = 0
#                         churn_rate_label = "0.00%"
#                         churn_rate_num = 0.0
#                     else:
#                         total_policies = row[0] or 0
#                         new_policies = row[1] or 0
#                         churned_policies = row[2] or 0
#                         first_year_churn = row[3] or 0
#                         churn_rate_label = row[4] or "0.00%"

#                         # For a numeric rate, recompute using the proper denominator from MV (Year-1, same state)
#                         cursor.execute("""
#                             SELECT COALESCE(SUM(total_policies), 0) AS prev_year_total
#                             FROM public.mv_year_state_churn
#                             WHERE "Year" = %s AND states = %s
#                         """, (y - 1, selected_state))
#                         prev_year_total = cursor.fetchone()[0] or 0
#                         churn_rate_num = round((churned_policies / prev_year_total) * 100, 1) if prev_year_total > 0 else 0.0

#                 # ------------------------------
#                 # Remaining breakdowns (still from base table, with previous-year churn filter)
#                 # Keep your existing logic here.
#                 # ------------------------------
#                 # Build CHURN queries (previous year) for drilldowns
#                 churn_conditions = []
#                 churn_params = []

#                 # Use Start Year = selected_year - 1 for churn distributions, as in your code
#                 churn_conditions.append('"Start Year" = %s')
#                 churn_params.append(y - 1)

#                 if selected_state != 'All':
#                     churn_conditions.append('"Cleaned State2" = %s')
#                     churn_params.append(selected_state)

#                 churn_where = 'WHERE ' + ' AND '.join(churn_conditions)

#                 # Top 8 Branches by Churn
#                 cursor.execute(f"""
#                     SELECT 
#                         "Cleaned Branch Name 2" AS branch_name,
#                         COUNT("policy no") AS churned_count
#                     FROM public."overall_cleaned_base_and_pr_ef_policyef_with_reasons_bucket"
#                     {churn_where}
#                         AND "Churn Label" = 'Yes'
#                         AND "Cleaned Branch Name 2" IS NOT NULL 
#                         AND "Cleaned Branch Name 2" != ''
#                     GROUP BY "Cleaned Branch Name 2"
#                     ORDER BY churned_count DESC
#                     LIMIT 8
#                 """, churn_params)
#                 branch_churn = [{"branch": r[0], "churn": r[1]} for r in cursor.fetchall()]

#                 # Product-wise Churned Policies
#                 cursor.execute(f"""
#                     SELECT 
#                         "product name 2" AS product_name,
#                         COUNT(DISTINCT "policy no") AS churned_count
#                     FROM public."overall_cleaned_base_and_pr_ef_policyef_with_reasons_bucket"
#                     {churn_where} AND "Churn Label" = 'Yes'
#                         AND "product name 2" IS NOT NULL 
#                         AND "product name 2" != ''
#                     GROUP BY "product name 2"
#                     ORDER BY churned_count DESC
#                 """, churn_params)
#                 product_churn = [{"name": r[0], "value": r[1]} for r in cursor.fetchall()]

#                 # Claim Status Distribution (among churned)
#                 cursor.execute(f"""
#                     SELECT 
#                         CASE 
#                             WHEN "Claim Status" = 'Approved' THEN 'Approved'
#                             WHEN "Claim Status" = 'Denied'   THEN 'Denied'
#                             ELSE 'No Claim'
#                         END AS claim_status,
#                         COUNT(DISTINCT "policy no") AS policy_count
#                     FROM public."overall_cleaned_base_and_pr_ef_policyef_with_reasons_bucket"
#                     {churn_where} AND "Churn Label" = 'Yes'
#                     GROUP BY 
#                         CASE 
#                             WHEN "Claim Status" = 'Approved' THEN 'Approved'
#                             WHEN "Claim Status" = 'Denied'   THEN 'Denied'
#                             ELSE 'No Claim'
#                         END
#                     ORDER BY policy_count DESC
#                 """, churn_params)
#                 claim_status = [{"name": r[0], "value": r[1]} for r in cursor.fetchall()]

#                 # Business Type Distribution (among churned)
#                 cursor.execute(f"""
#                     SELECT 
#                         biztype AS business_type,
#                         COUNT(DISTINCT "policy no") AS churned_count
#                     FROM public."overall_cleaned_base_and_pr_ef_policyef_with_reasons_bucket"
#                     {churn_where} AND "Churn Label" = 'Yes'
#                         AND biztype IS NOT NULL 
#                         AND biztype != ''
#                     GROUP BY biztype
#                     ORDER BY churned_count DESC
#                 """, churn_params)
#                 business_type = [{"name": r[0], "value": r[1]} for r in cursor.fetchall()]

#                 # Tie-up wise Churned Policies
#                 cursor.execute(f"""
#                     SELECT 
#                         "tie up" AS tie_up_name,
#                         COUNT(DISTINCT "policy no") AS churned_count
#                     FROM public."overall_cleaned_base_and_pr_ef_policyef_with_reasons_bucket"
#                     {churn_where} AND "Churn Label" = 'Yes'
#                         AND "tie up" IS NOT NULL 
#                         AND "tie up" != ''
#                     GROUP BY "tie up"
#                     ORDER BY churned_count DESC
#                 """, churn_params)
#                 tie_up_churn = [{"tieup": r[0], "churn": r[1]} for r in cursor.fetchall()]

#                 # Vehicle Age Groups (top 5 buckets among churned)
#                 cursor.execute(f"""
#                     WITH vehicle_age_buckets AS (
#                         SELECT 
#                             CASE 
#                                 WHEN age >= 0  AND age <= 1  THEN '0-1'
#                                 WHEN age > 1  AND age <= 2  THEN '1-2'
#                                 WHEN age > 2  AND age <= 4  THEN '2-4'
#                                 WHEN age > 4  AND age <= 6  THEN '4-6'
#                                 WHEN age > 6  AND age <= 8  THEN '6-8'
#                                 WHEN age > 8  AND age <= 10 THEN '8-10'
#                                 WHEN age > 10 AND age <= 12 THEN '10-12'
#                                 WHEN age > 12 AND age <= 14 THEN '12-14'
#                                 WHEN age > 14 AND age <= 16 THEN '14-16'
#                                 WHEN age > 16 AND age <= 18 THEN '16-18'
#                                 WHEN age > 18 AND age <= 20 THEN '18-20'
#                                 WHEN age > 20 AND age <= 22 THEN '20-22'
#                                 WHEN age > 22 AND age <= 24 THEN '22-24'
#                                 WHEN age > 24 AND age <= 25 THEN '24-25'
#                                 WHEN age > 25 THEN '25+'
#                                 ELSE 'Unknown'
#                             END AS age_bucket,
#                             COUNT(DISTINCT "policy no") AS churned_count
#                         FROM public."overall_cleaned_base_and_pr_ef_policyef_with_reasons_bucket"
#                         {churn_where} AND "Churn Label" = 'Yes'
#                             AND age IS NOT NULL
#                         GROUP BY 
#                             CASE 
#                                 WHEN age >= 0  AND age <= 1  THEN '0-1'
#                                 WHEN age > 1  AND age <= 2  THEN '1-2'
#                                 WHEN age > 2  AND age <= 4  THEN '2-4'
#                                 WHEN age > 4  AND age <= 6  THEN '4-6'
#                                 WHEN age > 6  AND age <= 8  THEN '6-8'
#                                 WHEN age > 8  AND age <= 10 THEN '8-10'
#                                 WHEN age > 10 AND age <= 12 THEN '10-12'
#                                 WHEN age > 12 AND age <= 14 THEN '12-14'
#                                 WHEN age > 14 AND age <= 16 THEN '14-16'
#                                 WHEN age > 16 AND age <= 18 THEN '16-18'
#                                 WHEN age > 18 AND age <= 20 THEN '18-20'
#                                 WHEN age > 20 AND age <= 22 THEN '20-22'
#                                 WHEN age > 22 AND age <= 24 THEN '22-24'
#                                 WHEN age > 24 AND age <= 25 THEN '24-25'
#                                 WHEN age > 25 THEN '25+'
#                                 ELSE 'Unknown'
#                             END
#                         ORDER BY churned_count DESC
#                         LIMIT 5
#                     )
#                     SELECT age_bucket, churned_count 
#                     FROM vehicle_age_buckets
#                 """, churn_params)
#                 vehicle_age_churn = [{"age": r[0], "churn": r[1]} for r in cursor.fetchall()]

#                 # Churn reasons (top 5 with % among churned)
#                 cursor.execute(f"""
#                     SELECT 
#                         "Reason Buckets" AS reason,
#                         COUNT(DISTINCT "policy no") AS count
#                     FROM public."overall_cleaned_base_and_pr_ef_policyef_with_reasons_bucket"
#                     {churn_where}
#                         AND "Churn Label" = 'Yes' 
#                         AND "Reason Buckets" IS NOT NULL 
#                         AND "Reason Buckets" != ''
#                     GROUP BY "Reason Buckets"
#                     ORDER BY count DESC
#                     LIMIT 5
#                 """, churn_params)
#                 rows = cursor.fetchall()
#                 total_for_pct = sum([r[1] for r in rows]) or 1
#                 churn_reasons = [
#                     {"reason": r[0], "count": r[1], "percentage": round((r[1] / total_for_pct) * 100)}
#                     for r in rows
#                 ]

#             conn.close()

#             return JsonResponse({
#                 "totalPolicies": total_policies,
#                 "newPolicies": new_policies,
#                 "churned": churned_policies,
#                 "firstYearChurn": first_year_churn,
#                 "churnRate": churn_rate_num,          # numeric, e.g., 50.0
#                 "churnRateLabel": churn_rate_label,   # string from MV or computed, e.g., "50.00%"
#                 "branchChurnData": branch_churn,
#                 "productChurnData": product_churn,
#                 "claimStatusChurnData": claim_status,
#                 "businessTypeChurnData": business_type,
#                 "tieupsChurnData": tie_up_churn,
#                 "vehicleAgeChurnData": vehicle_age_churn,
#                 "churnReasonsData": churn_reasons,
#                 "availableYears": available_years,
#                 "availableStates": available_states
#             })

#         except Exception as e:
#             logger.error(f"Error in dashboard_api: {str(e)}", exc_info=True)
#             return JsonResponse({"error": str(e)}, status=500)

#     else:
#         return JsonResponse({"error": "Method not allowed"}, status=405)
    

def dashboard_api(request):
    if request.method != "GET":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    try:
        from datetime import datetime
        MAX_YEAR = 2025                                 # <<< cap here
        current_year = datetime.now().year
        default_year = str(min(current_year - 1, MAX_YEAR))   # <<< clamp default

        # Incoming filters (no 'All' usage)
        incoming_year = request.GET.get('year')
        incoming_state = request.GET.get('state')

        conn = psycopg2.connect(
            dbname=DB_New_Past["dbname"],
            user=DB_New_Past["user"],
            password=DB_New_Past["password"],
            host=DB_New_Past["host"],
            port=DB_New_Past["port"]
        )

        with conn.cursor() as cursor:
            # ------------------------------------------
            # Filter lists from MV (strip any accidental 'All')
            # Restrict years to <= MAX_YEAR at SQL itself.
            # If your "year" column is TEXT, cast it: ("year")::int <= %s
            # ------------------------------------------
            cursor.execute("""
                SELECT DISTINCT "year"
                FROM public.distinct_state_year
                WHERE "year" IS NOT NULL
                  AND "year" <= %s                      -- <<< restrict in DB
                ORDER BY "year" DESC
            """, (MAX_YEAR,))
            available_years = [str(r[0]) for r in cursor.fetchall()]

            cursor.execute("""
                SELECT DISTINCT state
                FROM public.distinct_state_year
                WHERE state IS NOT NULL AND state <> ''
                ORDER BY state
            """)
            available_states = [r[0] for r in cursor.fetchall()]
            available_states = [s for s in available_states if s.strip().lower() != 'all']

            if not available_years:
                return JsonResponse({"error": "No years available"}, status=400)
            if not available_states:
                return JsonResponse({"error": "No states available"}, status=400)

            # Resolve selected year/state strictly from lists (and <= MAX_YEAR)
            selected_year = None
            if incoming_year:
                try:
                    iy = int(incoming_year)
                except ValueError:
                    iy = None
                if iy is not None and iy <= MAX_YEAR and str(iy) in available_years:
                    selected_year = str(iy)

            if not selected_year:
                selected_year = default_year if default_year in available_years else available_years[0]

            if incoming_state and incoming_state in available_states:
                selected_state = incoming_state
            else:
                selected_state = available_states[0]

            # Final numeric year (extra safety clamp)
            y = min(int(selected_year), MAX_YEAR)       # <<< clamp again

            # ------------------------------------------
            # Headline metrics (fetch-only from MV)
            # ------------------------------------------
            cursor.execute("""
                SELECT *
                FROM public.churn_1
                WHERE "Year" = %s AND "State" = %s
                LIMIT 1
            """, (y, selected_state))
            headline_row = cursor.fetchone()

            total_policies = new_policies = churned_policies = first_year_churn = 0
            churn_rate_num = None
            churn_rate_label = None

            if headline_row:
                colnames = [d[0] for d in cursor.description]
                hd = dict(zip(colnames, headline_row))
                total_policies   = hd.get("Total Policy", 0) or 0
                new_policies     = hd.get("New Policy", 0) or 0
                churned_policies = hd.get("Churn Count", 0) or 0
                first_year_churn = hd.get("First Year Churn", 0) or 0
                churn_rate_num   = hd.get("Churn Rate", None)
                churn_rate_label = hd.get('churn_rate_label', None) or hd.get('churn_rate', None)

            # ------------------------------------------
            # Visuals (fetch-only from your MVs)
            # ------------------------------------------
            cursor.execute("""
                SELECT top_branch, churn_count
                FROM public.state_year_top5_branch_prevyear_churn
                WHERE "year" = %s AND state = %s
            """, (y, selected_state))
            branch_churn = [{"branch": r[0], "churn": int(r[1])} for r in cursor.fetchall()]

            cursor.execute("""
                SELECT product_name, churn_count
                FROM public.state_year_product_prevyear_churn
                WHERE "year" = %s AND state = %s
            """, (y, selected_state))
            product_churn = [{"name": r[0], "value": int(r[1])} for r in cursor.fetchall()]

            cursor.execute("""
                SELECT claim_status, churn_count
                FROM public.state_year_claimstatus_prevyear_churn
                WHERE "year" = %s AND state = %s
            """, (y, selected_state))
            claim_status = [{"name": r[0], "value": int(r[1])} for r in cursor.fetchall()]

            cursor.execute("""
                SELECT business_type, churn_count
                FROM public.state_year_business_prevyear_churn
                WHERE "year" = %s AND state = %s
            """, (y, selected_state))
            business_type = [{"name": r[0], "value": int(r[1])} for r in cursor.fetchall()]

            cursor.execute("""
                SELECT tieups, churn_count
                FROM public.state_year_tieups_prevyear_churn
                WHERE "year" = %s AND state = %s
            """, (y, selected_state))
            tie_up_churn = [{"tieup": r[0], "churn": int(r[1])} for r in cursor.fetchall()]

            cursor.execute("""
                SELECT top_vehicle_age, churn_count
                FROM public.state_year_top5_vehicle_age_bucket_churn
                WHERE "year" = %s AND state = %s
            """, (y, selected_state))
            vehicle_age_churn = [{"age": r[0], "churn": int(r[1])} for r in cursor.fetchall()]

            # ------------------------------------------
            # Reasons (prev-year scope; y already clamped)
            # ------------------------------------------
            cursor.execute("""
                SELECT 
                    "Reason Buckets" AS reason,
                    COUNT(DISTINCT "policy no") AS cnt
                FROM public."overall_cleaned_base_and_pr_ef_policyef_with_reasons_bucket"
                WHERE "Start Year" = %s
                  AND "Churn Label" = 'Yes'
                  AND "Reason Buckets" IS NOT NULL 
                  AND "Reason Buckets" <> ''
                  AND "Cleaned State2" = %s
                GROUP BY "Reason Buckets"
                ORDER BY cnt DESC
                LIMIT 5
            """, (y - 1, selected_state))
            rs = cursor.fetchall()
            total_for_pct = sum(int(r[1]) for r in rs) or 1
            churn_reasons = [
                {"reason": r[0], "count": int(r[1]), "percentage": round((int(r[1]) / total_for_pct) * 100)}
                for r in rs
            ]

        conn.close()

        if churn_rate_num is not None:
    # Numeric field from DB → scale to percent
            churn_rate_out = f"{round(churn_rate_num * 100, 2)}%"
        else:
            churn_rate_out = churn_rate_label

        return JsonResponse({
            "totalPolicies": total_policies,
            "newPolicies": new_policies,
            "churned": churned_policies,
            "firstYearChurn": first_year_churn,
            "churnRate": churn_rate_out,
            "churnRateLabel": churn_rate_out,

            "branchChurnData": branch_churn,
            "productChurnData": product_churn,
            "claimStatusChurnData": claim_status,
            "businessTypeChurnData": business_type,
            "tieupsChurnData": tie_up_churn,
            "vehicleAgeChurnData": vehicle_age_churn,
            "churnReasonsData": churn_reasons,

            # Return only capped years
            "availableYears": available_years,
            "availableStates": available_states,
            "selectedYear": str(y),
            "selectedState": selected_state,
        })

    except Exception as e:
        logger.error(f"Error in dashboard_api: {str(e)}", exc_info=True)
        return JsonResponse({"error": str(e)}, status=500)
    
from django.http import JsonResponse
from django.views import View
import psycopg2

class DrillthroughView(View):
    def get(self, request):
        drill_type   = (request.GET.get("type") or "").strip()
        branch       = request.GET.get("branch")
        product      = request.GET.get("product")
        vehicle_age  = request.GET.get("vehicle_age")
        claim_status = request.GET.get("claim_status")
        # accept both keys for safety
        business     = (
            request.GET.get("business_type")
            or request.GET.get("business")
            or request.GET.get("biztype")
        )
        tieup        = request.GET.get("tieup")
        year_raw     = request.GET.get("year")
        state        = request.GET.get("state")
        total        = request.GET.get("total")
        new          = request.GET.get("new")
        churned      = request.GET.get("churned")
        firstyear    = request.GET.get("firstyear")

        # year-1 only for churned and firstyear; else use same year
        effective_year = None
        if year_raw:
            y = int(year_raw)
            if drill_type in {"churned", "firstyear","branch", "tieup","branch","business","product","vehicle_age","claim_status"}:
                effective_year = y - 1
            else:
                effective_year = y

        base_sql = """
            SELECT
                "policy no",
                "Cleaned Branch Name 2",
                "product name 2",
                "agebucket",
                "Claim Status",
                biztype,
                "tie up",
                "manufacturer/make",
                "Not Renewed Reasons",
                "policy end date",
                "total premium payable",
                "Cleaned State2"
            FROM public.overall_cleaned_base_and_pr_ef_policyef_with_reasons_bucket
            WHERE 1=1
        """
        params = []

        # Apply year ONCE
        if effective_year is not None:
            base_sql += ' AND "End Year" = %s'
            params.append(effective_year)

        # Exact match for state (no normalization)
        if state:
            base_sql += ' AND "Cleaned State2" = %s'
            params.append(state)

        # Type-specific filters (NO extra End Year here)
        if drill_type == "total" and total:
            pass

        elif drill_type == "new" and new:
            base_sql += ' AND "New Customers" = %s'
            params.append("Yes")

        elif drill_type == "churned" and churned:
            base_sql += ' AND "Churn Label" = %s'
            params.append("Yes")

        elif drill_type == "firstyear" and firstyear:
            base_sql += ' AND "Churn Label" = %s AND "New Customers" = %s'
            params.extend(["Yes", "Yes"])

        # Accept both 'branch' and 'branch_churn' drill types; exact match
        elif drill_type in ("branch") and branch:
            base_sql += ' AND "Cleaned Branch Name 2" = %s AND "Churn Label" = %s'
            params.extend([branch, "Yes"])

        elif drill_type == "product" and product:
            base_sql += ' AND "product name 2" = %s AND "Churn Label" = %s'
            params.extend([product, "Yes"])

        elif drill_type == "vehicle_age" and vehicle_age:
            base_sql += ' AND agebucket = %s AND "Churn Label" = %s'    
            params.extend([(vehicle_age), "Yes"])

        elif drill_type == "claim_status" and claim_status:
            base_sql += ' AND "Claim Status" = %s AND "Churn Label" = %s'
            params.extend([claim_status, "Yes"])

        elif drill_type == "business" and business:
            base_sql += ' AND biztype = %s AND "Churn Label" = %s'
            params.extend([business, "Yes"])

        elif drill_type == "tieup" and tieup:
            base_sql += ' AND "tie up" = %s AND "Churn Label" = %s'
            params.extend([tieup, "Yes"])

        try:
            conn = psycopg2.connect(
                dbname=DB_New_Past["dbname"],
                user=DB_New_Past["user"],
                password=DB_New_Past["password"],
                host=DB_New_Past["host"],
                port=DB_New_Past["port"],
            )
            cur = conn.cursor()
            cur.execute(base_sql, params)
            columns = [desc[0] for desc in cur.description]
            rows = [dict(zip(columns, row)) for row in cur.fetchall()]
            cur.close()
            conn.close()
            return JsonResponse({"count": len(rows), "results": rows}, safe=False)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)
        
# class PipelineStatusView(View):
#     def get(self, request):
#         conn = None
#         cur = None
        
#         try:
#             # Connect to PostgreSQL
#             conn = psycopg2.connect(
#                 dbname=DB_Azure_Airflow["dbname"],   
#                 user=DB_Azure_Airflow["user"],
#                 password=DB_Azure_Airflow["password"],
#                 host=DB_Azure_Airflow["host"],
#                 port=DB_Azure_Airflow["port"],
#             )
#             cur = conn.cursor()
            
#             # Query to get all necessary data
#             query = """
#                 SELECT 
#                     table_name,
#                     stage_count,
#                     dwh_loaded_cnt_base,
#                     dwh_loaded_cnt_pr,
#                     basepr_count,
#                     stage_loaded,
#                     is_base_cleaned,
#                     is_pr_cleaned,
#                     is_basepr_appended,
#                     last_updated_ts
#                 FROM pip_log.etl_metadata_logs
#                 ORDER BY table_name;
#             """
            
#             cur.execute(query)
#             rows = cur.fetchall()

#             query_log = """
#                 SELECT stage_count, appended_count ,merged_count FROM pip_log.claim_logs;
#             """
#             cur.execute(query_log)
#             row_logs = cur.fetchall()
#             # Initialize counters
#             raw_base = 0
#             raw_pr = 0
#             silver_base = 0
#             silver_pr = 0
#             gold_base = 0
#             gold_pr = 0
#             raw_log = 0
#             silver_log = 0
#             gold_log = 0


#             latest_timestamp = None
#             overall_status = "All Stages Completed"
#             failed_stage = None
            
#             # Process each row
#             for row in rows:
#                 table_name = row[0]
#                 stage_count = row[1] or 0
#                 dwh_loaded_cnt_base = row[2] or 0
#                 dwh_loaded_cnt_pr = row[3] or 0
#                 basepr_count = int(row[4]) if row[4] and row[4] != 'NO' else 0
#                 stage_loaded = row[5]
#                 is_base_cleaned = row[6]
#                 is_pr_cleaned = row[7]
#                 is_basepr_appended = row[8]
#                 last_updated = row[9]
                
#                 # Update latest timestamp
#                 if last_updated:
#                     if latest_timestamp is None or last_updated > latest_timestamp:
#                         latest_timestamp = last_updated
                
#                 # Check for failures
#                 if stage_loaded == 'NO':
#                     overall_status = "Load Failed - Raw Stage"
#                     failed_stage = "Raw"
#                 elif is_base_cleaned == 'NO' or is_pr_cleaned == 'NO':
#                     overall_status = "Load Failed - Silver Stage"
#                     failed_stage = "Silver"
#                 elif is_basepr_appended == 'NO':
#                     overall_status = "Load Failed - Gold Stage"
#                     failed_stage = "Gold"
                
#                 # Calculate counts for Base tables
#                 if table_name.lower().startswith('base'):
#                     raw_base += stage_count
#                     silver_base += dwh_loaded_cnt_base
#                     gold_base += basepr_count
                
#                 # Calculate counts for PR tables
#                 elif table_name.lower().startswith('pr'):
#                     raw_pr += stage_count
#                     silver_pr += dwh_loaded_cnt_pr
#                     gold_pr += basepr_count
            

#             for log_row in row_logs:
#                 raw_log += log_row[0] or 0
#                 silver_log += log_row[1] or 0
#                 gold_log += log_row[2] or 0

#             # Calculate total counts
#             raw_total = raw_base + raw_pr
#             silver_total = silver_base + silver_pr
#             gold_total = gold_base + gold_pr
            
#             # Format timestamp
#             if latest_timestamp:
#                 formatted_timestamp = latest_timestamp.strftime("%d.%m.%Y %H:%M:%S IST")
#             else:
#                 formatted_timestamp = datetime.now().strftime("%d.%m.%Y %H:%M:%S IST")
            
#             # Prepare response data
#             response_data = {
#                 "status": "success",
#                 "data": {
#                     "overallStatus": overall_status,
#                     "lastUpdated": formatted_timestamp,
#                     "rawRecordCount": raw_total,
#                     "silverRecordCount": silver_total,
#                     "goldenRecordCount": gold_total,
#                     # NEW: Base and PR counts from database
#                     "rawStage": {
#                         "base": raw_base,
#                         "pr": raw_pr,
#                         "log": raw_log
#                     },
#                     "silverStage": {
#                         "base": silver_base,
#                         "pr": silver_pr,
#                         "log": silver_log
#                     },
#                     "goldStage": {
#                         "base": gold_base,
#                         "pr": gold_pr,
#                         "log": gold_log
#                     },
#                     # OLD: Keep dummy data for Overview section (static)
#                     "rawStageDetails": {
#                         "columnsCleared": 41,
#                         "duplicatesRemoved": 256,
#                         "unnecessaryColumnsDropped": 3
#                     },
#                     "silverStageDetails": {
#                         "dataTypesChanged": 12,
#                         "columnsCleaned": 256,
#                         "duplicatesFound": 0
#                     },
#                     "goldStageDetails": {
#                         "baseDataAppended": "1M",
#                         "customerIdGenerated": "0.8M/0.2M",
#                         "renColumnAdded": "0.9M"
#                     }
#                 }
#             }
            
#             return JsonResponse(response_data)
            
#         except psycopg2.Error as e:
#             error_message = f"Database error: {str(e)}"
#             print(error_message)
#             traceback.print_exc()
#             return JsonResponse({
#                 "status": "error",
#                 "message": error_message
#             }, status=500)
            
#         except Exception as e:
#             error_message = f"Unexpected error: {str(e)}"
#             print(error_message)
#             traceback.print_exc()
#             return JsonResponse({
#                 "status": "error",
#                 "message": error_message
#             }, status=500)
            
#         finally:
#             # Close database connections
#             if cur:
#                 cur.close()
#             if conn:
#                 conn.close()

import os, json, secrets, base64, requests
from django.http import JsonResponse, HttpResponseBadRequest, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.core.cache import cache
from urllib.parse import urlencode
from google_auth_oauthlib.flow import Flow
from google.oauth2.credentials import Credentials as _GCreds
from google.auth.transport.requests import Request as _GRequest

# Pull paths from env
GMAIL_CREDENTIALS_FILE = os.getenv("GMAIL_CREDENTIALS_FILE", "gmail/credentials.json")
GMAIL_TOKEN_FILE       = os.getenv("GMAIL_TOKEN_FILE", "gmail/token.json")
GMAIL_REDIRECT_URI     = os.getenv("GMAIL_REDIRECT_URI")   # MUST match Google console
SCOPES = ["https://www.googleapis.com/auth/gmail.send"]
REDIRECT = os.environ["GMAIL_REDIRECT_URI"]

CACHE_PREFIX = "gmail_oauth_state:"
CACHE_TTL = 10 * 60  # 10 minutes

def _flow_for_redirect(_req):
    return Flow.from_client_secrets_file(
        GMAIL_CREDENTIALS_FILE,
        scopes=["https://www.googleapis.com/auth/gmail.send"],
        redirect_uri=REDIRECT,
    )

def _token_status():
    """Return (status, detail) for current token.json."""
    if not os.path.exists(GMAIL_TOKEN_FILE):
        return "absent", "token file not found"
    try:
        creds = _GCreds.from_authorized_user_file(GMAIL_TOKEN_FILE, SCOPES)
        if creds.expired and creds.refresh_token:
            # Try refresh in-memory (no write)
            try:
                creds.refresh(_GRequest())
                return "valid", "refreshed in-memory"
            except Exception as e:
                return "invalid", f"refresh failed: {e}"
        return "valid", "token present"
    except Exception as e:
        return "invalid", f"load failed: {e}"

# bulk_email/views_oauth.py
from django.views.decorators.csrf import csrf_exempt

import json
from urllib.parse import urlencode
from django.urls import reverse

def _load_client_info():
    with open(GMAIL_CREDENTIALS_FILE, "r", encoding="utf-8") as f:
        c = json.load(f)
    # support both "web" and "installed" secrets
    data = c.get("web") or c.get("installed") or {}
    return data.get("client_id"), data.get("client_secret")

@csrf_exempt
def gmail_auth_url(request):
    state = secrets.token_urlsafe(24)
    cache.set(CACHE_PREFIX + state, {"ok": True}, CACHE_TTL)
    params = {
        "response_type": "code",
        "client_id": json.load(open(GMAIL_CREDENTIALS_FILE))["web"]["client_id"],
        "redirect_uri": REDIRECT,
        "scope": "https://www.googleapis.com/auth/gmail.send",
        "state": state,
        "access_type": "offline",
        "include_granted_scopes": "true",
        "prompt": "consent",
    }
    auth_url = "https://accounts.google.com/o/oauth2/v2/auth?" + urlencode(params)
    return JsonResponse({"auth_url": auth_url, "state": state, "redirect_uri": REDIRECT})




def gmail_oauth_callback(request):
    """
    GET /bulk-email/gmail/callback?state=...&code=...
    Google redirects here. We exchange code -> tokens and write token.json.
    """
    try:
        state = request.GET.get("state")
        code  = request.GET.get("code")
        if not state or not code:
            return HttpResponseBadRequest("Missing state or code")

        # Validate state (helps prevent CSRF)
        if not cache.get(CACHE_PREFIX + state):
            return HttpResponseBadRequest("Unknown/expired state")

        flow = _flow_for_redirect(request)
        flow.fetch_token(code=code)
        creds = flow.credentials

        os.makedirs(os.path.dirname(GMAIL_TOKEN_FILE) or ".", exist_ok=True)
        with open(GMAIL_TOKEN_FILE, "w", encoding="utf-8") as f:
            f.write(creds.to_json())

        # Invalidate one-time state
        cache.delete(CACHE_PREFIX + state)

        return HttpResponse("✅ Gmail authorization successful. token.json saved.")
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

@csrf_exempt
def gmail_status(request):
    status, detail = _token_status()
    return JsonResponse({"status": status, "detail": detail, "token_path": GMAIL_TOKEN_FILE})

@csrf_exempt
def gmail_revoke(_request):
    """
    POST /bulk-email/gmail/revoke
    Revokes the current refresh token (if any) and deletes token.json.
    """
    try:
        if not os.path.exists(GMAIL_TOKEN_FILE):
            return JsonResponse({"ok": True, "message": "No token file present"})
        creds = _GCreds.from_authorized_user_file(GMAIL_TOKEN_FILE, SCOPES)
        token = getattr(creds, "token", None) or getattr(creds, "refresh_token", None)
        if token:
            requests.post(
                "https://oauth2.googleapis.com/revoke",
                headers={"Content-Type": "application/x-www-form-urlencoded"},
                data=urlencode({"token": token}),
                timeout=10,
            )
        try:
            os.remove(GMAIL_TOKEN_FILE)
        except Exception:
            pass
        return JsonResponse({"ok": True, "message": "Token revoked and file removed"})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

@csrf_exempt
def backend_status(request):
    if request.method == "GET":
        return JsonResponse({"message": "Backend fetched successfully !!!!!"})
    return JsonResponse({"error": "Invalid request method"}, status=405)